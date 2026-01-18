"""
Import PDC Cutters from Excel file with variant item numbers.

This command reads the 'Cutters ERP Item Numbers2.xlsx' file and creates:
- InventoryItem records for each base cutter (auto-increment code)
- ItemVariant records for each variant with its unique ERP item number
- ItemAttributeValue records for cutter attributes

Excel Structure (DATA sheet):
- MN: Material Number (SAP reference - stored as mat_number)
- Product name: Item name
- Cutter Type: e.g., CT97, OBS ERC (stored as hdbs_code attribute)
- Size: e.g., 1313, 1608
- Chamfer: e.g., 10C, 18C, U
- Family: e.g., P - Premium
- Cutter Shape: e.g., Round, Trifex
- Category: (usually empty)
- ENO As New Cutter: Item number for NEW-EO variant
- ENO Ground Cutter: Item number for USED-GRD variant
- ARDT Reclaim Cutter: Item number for USED-RCL variant
- Retrofit Cutter: Item number for NEW-RET variant
- LSTK Reclaim Cutter: Item number for client reclaim (NEW-CLI)
- New Stock: Item number for NEW-PUR variant

Usage:
    python manage.py import_cutters_excel                    # Preview
    python manage.py import_cutters_excel --confirm          # Import
    python manage.py import_cutters_excel --file other.xlsx  # Custom file
"""
import os
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import transaction
from django.core.exceptions import ValidationError
from apps.inventory.models import (
    InventoryItem, InventoryCategory, ItemVariant, VariantCase,
    CategoryAttribute, ItemAttributeValue
)


class Command(BaseCommand):
    help = "Import PDC Cutters from Excel file with variant item numbers"

    # Mapping of Excel columns to variant case codes
    VARIANT_COLUMN_MAP = {
        'ENO As New Cutter': 'NEW-EO',
        'ENO Ground Cutter': 'USED-GRD',
        'ENO Ground Cutter ': 'USED-GRD',  # With trailing space
        'ARDT Reclaim  Cutter': 'USED-RCL',  # Note double space
        'ARDT Reclaim Cutter': 'USED-RCL',
        'Retrofit Cutter': 'NEW-RET',
        'LSTK Reclaim Cutter': 'NEW-CLI',  # Mapped to client reclaim
        'New Stock': 'NEW-PUR',
    }

    # Mapping of Excel columns to attribute codes
    ATTRIBUTE_COLUMN_MAP = {
        'Cutter Type': 'hdbs_code',      # HDBS Code (e.g., CT109)
        'Size': 'cutter_size',
        'Chamfer': 'chamfer',
        'Family': 'family',
        'Cutter Shape': 'cutter_shape',
    }

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='docs/Cutters ERP Item Numbers2.xlsx',
            help='Path to Excel file (default: docs/Cutters ERP Item Numbers2.xlsx)',
        )
        parser.add_argument(
            '--confirm',
            action='store_true',
            help='Actually import the data (without this flag, just preview)',
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            help='Skip items that already exist instead of updating',
        )

    def get_next_item_code(self, category):
        """Generate next auto-increment code for the category."""
        prefix = "CUT"
        # Find the highest existing number
        existing = InventoryItem.objects.filter(
            category=category,
            code__startswith=f"{prefix}-"
        ).values_list('code', flat=True)

        max_num = 0
        for code in existing:
            try:
                num = int(code.split('-')[1])
                if num > max_num:
                    max_num = num
            except (IndexError, ValueError):
                continue

        return f"{prefix}-{max_num + 1:04d}"

    def handle(self, *args, **options):
        import openpyxl

        file_path = options['file']
        confirm = options['confirm']
        skip_existing = options['skip_existing']

        self.stdout.write("=" * 60)
        self.stdout.write("Import PDC Cutters from Excel")
        self.stdout.write("=" * 60)

        # Check file exists
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        # Load Excel file
        self.stdout.write(f"\nLoading: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb['DATA']
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error loading Excel: {e}"))
            return

        # Get headers
        headers = [cell.value for cell in ws[1]]
        self.stdout.write(f"Found {len(headers)} columns")

        # Find PDC Cutters category
        pdc_category = InventoryCategory.objects.filter(code__in=['CT-PDC', 'CUT-PDC']).first()
        if not pdc_category:
            self.stdout.write(self.style.ERROR("PDC Cutters category not found (CT-PDC or CUT-PDC)"))
            return
        self.stdout.write(f"Category: {pdc_category.code}")

        # Get variant cases
        variant_cases = {vc.code: vc for vc in VariantCase.objects.all()}
        self.stdout.write(f"Variant cases: {list(variant_cases.keys())}")

        # Get category attributes for PDC category (keyed by attribute code)
        category_attributes = {}
        for ca in CategoryAttribute.objects.filter(category=pdc_category).select_related('attribute'):
            if ca.attribute:
                category_attributes[ca.attribute.code] = ca
        self.stdout.write(f"Category attributes: {list(category_attributes.keys())}")

        # Process rows
        row_count = 0
        items_created = 0
        items_updated = 0
        items_skipped = 0
        variants_created = 0
        errors = []

        # Track next code number for new items
        next_code_num = 1
        existing_codes = set(InventoryItem.objects.filter(
            category=pdc_category,
            code__startswith="CUT-"
        ).values_list('code', flat=True))

        # Find max existing number
        for code in existing_codes:
            try:
                num = int(code.split('-')[1])
                if num >= next_code_num:
                    next_code_num = num + 1
            except (IndexError, ValueError):
                continue

        self.stdout.write(f"Next code number: CUT-{next_code_num:04d}")
        self.stdout.write("\nProcessing rows...")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row[0]:
                continue

            row_count += 1
            row_dict = dict(zip(headers, row))

            mn = str(row_dict.get('MN', '')).strip()
            product_name = str(row_dict.get('Product name', '')).strip()

            if not mn or not product_name:
                errors.append(f"Row {row_num}: Missing MN or Product name")
                continue

            if not confirm:
                # Preview mode
                if row_count <= 5:
                    self.stdout.write(f"  Row {row_num}: MN={mn}, Name={product_name[:40]}")
                    # Count non-empty variant columns
                    variant_count = sum(1 for col in self.VARIANT_COLUMN_MAP.keys()
                                       if row_dict.get(col))
                    self.stdout.write(f"    Variants: {variant_count}")
                elif row_count == 6:
                    self.stdout.write("  ...")
            else:
                # Import mode
                try:
                    with transaction.atomic():
                        # Check if item exists by mat_number (MN)
                        existing_item = InventoryItem.objects.filter(
                            category=pdc_category,
                            mat_number=mn
                        ).first()

                        if existing_item and skip_existing:
                            items_skipped += 1
                            continue

                        # Create or update item
                        if existing_item:
                            item = existing_item
                            items_updated += 1
                        else:
                            # Generate auto-increment code
                            item_code = f"CUT-{next_code_num:04d}"
                            next_code_num += 1

                            item = InventoryItem(
                                code=item_code,
                                category=pdc_category,
                                has_variants=True,
                            )
                            items_created += 1

                        item.name = product_name
                        item.mat_number = mn  # Store MN as SAP reference
                        item.is_active = True
                        item.save()

                        # Set attributes
                        for excel_col, attr_code in self.ATTRIBUTE_COLUMN_MAP.items():
                            value = row_dict.get(excel_col)
                            if value and attr_code in category_attributes:
                                cat_attr = category_attributes[attr_code]
                                ItemAttributeValue.objects.update_or_create(
                                    item=item,
                                    attribute=cat_attr,
                                    defaults={
                                        'text_value': str(value).strip(),
                                    }
                                )

                        # Create variants
                        for excel_col, variant_code in self.VARIANT_COLUMN_MAP.items():
                            erp_item_number = row_dict.get(excel_col)
                            if erp_item_number and str(erp_item_number).strip():
                                erp_item_number = str(erp_item_number).strip()

                                # Get variant case
                                variant_case = variant_cases.get(variant_code)
                                if not variant_case:
                                    continue

                                # Check if variant already exists for this item+case
                                existing_variant = ItemVariant.objects.filter(
                                    base_item=item,
                                    variant_case=variant_case
                                ).first()

                                if existing_variant:
                                    # Update ERP item number if different
                                    if existing_variant.erp_item_no != erp_item_number:
                                        existing_variant.erp_item_no = erp_item_number
                                        existing_variant.save()
                                else:
                                    # Create new variant
                                    # Code is auto-generated by model, erp_item_no is unique
                                    try:
                                        ItemVariant.objects.create(
                                            base_item=item,
                                            variant_case=variant_case,
                                            erp_item_no=erp_item_number,
                                            is_active=True,
                                        )
                                        variants_created += 1
                                    except ValidationError as e:
                                        errors.append(f"Row {row_num} variant {variant_code}: {e}")

                except Exception as e:
                    errors.append(f"Row {row_num}: {e}")

        # Summary
        self.stdout.write("\n" + "=" * 60)
        self.stdout.write("Summary")
        self.stdout.write("=" * 60)
        self.stdout.write(f"Total rows processed: {row_count}")

        if not confirm:
            # Count variants
            total_variants = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                row_dict = dict(zip(headers, row))
                total_variants += sum(1 for col in self.VARIANT_COLUMN_MAP.keys()
                                     if row_dict.get(col) and str(row_dict.get(col)).strip())

            self.stdout.write(f"Items to create: {row_count}")
            self.stdout.write(f"Variants to create: {total_variants}")
            self.stdout.write("\n" + self.style.WARNING("DRY RUN - No data was imported"))
            self.stdout.write("Run with --confirm to actually import the data:")
            self.stdout.write(f"  python manage.py import_cutters_excel --confirm")
        else:
            self.stdout.write(f"Items created: {items_created}")
            self.stdout.write(f"Items updated: {items_updated}")
            self.stdout.write(f"Items skipped: {items_skipped}")
            self.stdout.write(f"Variants created: {variants_created}")

            if errors:
                self.stdout.write(self.style.WARNING(f"\nErrors ({len(errors)}):"))
                for error in errors[:10]:
                    self.stdout.write(f"  - {error}")
                if len(errors) > 10:
                    self.stdout.write(f"  ... and {len(errors) - 10} more")

            self.stdout.write("\n" + self.style.SUCCESS("Import complete!"))

        self.stdout.write("=" * 60)
