"""
Import cutter stock quantities from ERP On-hand Excel file.

This command reads stock quantities from the ERP On-hand.xlsx file and updates
the VariantStock model with the current quantities.

On-hand.xlsx Column Mapping (headers in row 6):
- Column A (1): Item number (e.g., CT-0001, ENO-CT-0005, RCLM-ARDT-0001)
- Column E (5): Color - HDBS MAT Number used to match our InventoryItem
- Column H (8): Warehouse (Store, R Warehous, Shopfloor, Transit)
- Column I (9): Location
- Column J (10): Available physical (quantity)

ERP Item Prefix → Variant Case Mapping:
- CT-*         → NEW-PUR  (New Purchased)
- ENO-CT-*     → NEW-EO   (New E&O / As New)
- RCLM-ARDT-*  → USED-RCL (ARDT Reclaimed)
- RCLM-*       → NEW-CLI  (LSTK/Client Reclaimed)
- RTRO-*       → NEW-RET  (Retrofit as New)

Usage:
    python manage.py import_stock_from_onhand                    # Preview mode
    python manage.py import_stock_from_onhand --confirm          # Apply changes
    python manage.py import_stock_from_onhand --file path/to.xlsx --confirm
"""

import os
from decimal import Decimal
from collections import defaultdict
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class Command(BaseCommand):
    help = 'Import cutter stock quantities from ERP On-hand Excel file'

    # On-hand.xlsx column mapping (1-indexed as used by openpyxl)
    COLUMN_MAP = {
        'item_number': 1,    # Column A - ERP Item Number (CT-0001, etc.)
        'item_group': 2,     # Column B - Item group
        'configuration': 3,  # Column C - Configuration
        'size': 4,           # Column D - Size
        'color': 5,          # Column E - Color (HDBS MAT Number!)
        'style': 6,          # Column F - Style
        'site': 7,           # Column G - Site
        'warehouse': 8,      # Column H - Warehouse
        'location': 9,       # Column I - Location
        'qty': 10,           # Column J - Available physical
    }

    # ERP Item Prefix → Variant Case mapping
    PREFIX_TO_VARIANT = {
        'ENO-CT': 'NEW-EO',      # ENO As New Cutter
        'RCLM-ARDT': 'USED-RCL', # ARDT Reclaim Cutter
        'RCLM-': 'NEW-CLI',      # LSTK/Client Reclaim Cutter
        'RTRO-': 'NEW-RET',      # Retrofit as New
        'CT-': 'NEW-PUR',        # New Stock (Purchased)
    }

    # Warehouses to include (filter out Transit, etc.)
    INCLUDE_WAREHOUSES = ['Store', 'R Warehous', 'Shopfloor']

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='docs/On-hand.xlsx',
            help='Path to ERP On-hand Excel file (default: docs/On-hand.xlsx)'
        )
        parser.add_argument(
            '--confirm',
            action='store_true',
            help='Actually apply changes (default is preview mode)'
        )
        parser.add_argument(
            '--header-row',
            type=int,
            default=6,
            help='Row number containing headers (default: 6 for On-hand.xlsx)'
        )
        parser.add_argument(
            '--include-transit',
            action='store_true',
            help='Include Transit warehouse quantities (default: exclude)'
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Show detailed per-item import info'
        )

    def get_variant_case_for_item(self, item_number):
        """Determine variant case based on ERP item number prefix."""
        item_upper = item_number.upper()

        # Check in specific order (longest/most specific first)
        if item_upper.startswith('ENO-CT'):
            return 'NEW-EO'
        elif item_upper.startswith('RCLM-ARDT'):
            return 'USED-RCL'
        elif item_upper.startswith('RCLM-'):
            return 'NEW-CLI'
        elif item_upper.startswith('RTRO-'):
            return 'NEW-RET'
        elif item_upper.startswith('CT-') or item_upper.startswith('CT0'):
            return 'NEW-PUR'
        return None

    def handle(self, *args, **options):
        if not HAS_OPENPYXL:
            self.stderr.write(self.style.ERROR('openpyxl is required. Install with: pip install openpyxl'))
            return

        file_path = options['file']
        confirm = options['confirm']
        header_row = options['header_row']
        include_transit = options['include_transit']
        verbose = options['verbose']

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        self.stdout.write(f'Loading Excel file: {file_path}')

        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        sheet_name = ws.title

        self.stdout.write(f'Processing sheet: {sheet_name}')
        self.stdout.write(f'Header row: {header_row}, Data starts at row: {header_row + 1}')

        # Import models
        from apps.inventory.models import (
            InventoryItem, ItemVariant, VariantCase, VariantStock,
            InventoryLocation, ItemAttributeValue
        )

        # Get default location (main warehouse)
        default_location = InventoryLocation.objects.filter(is_default=True).first()
        if not default_location:
            default_location = InventoryLocation.objects.first()
            if not default_location:
                self.stderr.write(self.style.ERROR('No inventory locations found. Create one first.'))
                return

        self.stdout.write(f'Using location: {default_location.name}')

        # Get variant cases
        variant_cases = {vc.code: vc for vc in VariantCase.objects.all()}
        self.stdout.write(f'Available variant cases: {list(variant_cases.keys())}')

        # Build HDBS code to item mapping
        hdbs_to_item = {}

        # Method 1: Direct mat_number field
        for item in InventoryItem.objects.filter(mat_number__isnull=False).exclude(mat_number=''):
            hdbs_to_item[str(item.mat_number).strip()] = item

        # Method 2: hdbs_code attribute
        for attr_val in ItemAttributeValue.objects.filter(
            attribute__attribute__code__in=['hdbs_code', 'hdbs', 'hdbs_mat'],
            text_value__isnull=False
        ).exclude(text_value='').select_related('item'):
            hdbs_code = str(attr_val.text_value).strip()
            if hdbs_code not in hdbs_to_item:
                hdbs_to_item[hdbs_code] = attr_val.item

        self.stdout.write(f'Found {len(hdbs_to_item)} items with HDBS codes')

        # Include warehouses
        warehouses_to_include = set(self.INCLUDE_WAREHOUSES)
        if include_transit:
            warehouses_to_include.add('Transit')
        self.stdout.write(f'Including warehouses: {warehouses_to_include}')

        # First pass: Aggregate quantities by (HDBS_code, variant_case)
        # This handles multiple rows for same item in different warehouses/locations
        aggregated = defaultdict(lambda: {
            'qty': Decimal('0'),
            'erp_items': set(),
            'warehouses': set(),
        })

        skipped = {
            'no_item_number': 0,
            'no_hdbs_code': 0,
            'not_cutter': 0,
            'unknown_prefix': 0,
            'excluded_warehouse': 0,
            'zero_qty': 0,
        }

        # Process rows (data starts after header row)
        self.stdout.write(f'Reading data rows...')
        for row_idx in range(header_row + 1, ws.max_row + 1):
            item_number = ws.cell(row=row_idx, column=self.COLUMN_MAP['item_number']).value
            if not item_number:
                skipped['no_item_number'] += 1
                continue

            item_number = str(item_number).strip()

            # Only process cutter-related items
            item_upper = item_number.upper()
            is_cutter = any(p in item_upper for p in ['CT-', 'CT0', 'ENO-CT', 'RCLM', 'RTRO'])
            if not is_cutter:
                skipped['not_cutter'] += 1
                continue

            # Get HDBS code from Color column
            hdbs_code = ws.cell(row=row_idx, column=self.COLUMN_MAP['color']).value
            if not hdbs_code:
                skipped['no_hdbs_code'] += 1
                continue
            hdbs_code = str(hdbs_code).strip()

            # Get warehouse
            warehouse = ws.cell(row=row_idx, column=self.COLUMN_MAP['warehouse']).value or ''
            warehouse = str(warehouse).strip()

            if warehouse and warehouse not in warehouses_to_include:
                skipped['excluded_warehouse'] += 1
                continue

            # Get quantity
            qty_raw = ws.cell(row=row_idx, column=self.COLUMN_MAP['qty']).value
            try:
                qty = Decimal(str(qty_raw or 0))
            except:
                qty = Decimal('0')

            if qty <= 0:
                skipped['zero_qty'] += 1
                continue

            # Determine variant case from item prefix
            variant_code = self.get_variant_case_for_item(item_number)
            if not variant_code:
                skipped['unknown_prefix'] += 1
                continue

            # Aggregate
            key = (hdbs_code, variant_code)
            aggregated[key]['qty'] += qty
            aggregated[key]['erp_items'].add(item_number)
            aggregated[key]['warehouses'].add(warehouse)

        self.stdout.write(f'Aggregated {len(aggregated)} unique (HDBS, variant) combinations')
        self.stdout.write(f'Skipped rows:')
        for reason, count in skipped.items():
            if count > 0:
                self.stdout.write(f'  - {reason}: {count}')

        # Statistics
        stats = {
            'combinations_processed': 0,
            'items_matched': 0,
            'items_not_found': 0,
            'variants_created': 0,
            'variants_updated': 0,
            'stock_records_created': 0,
            'stock_records_updated': 0,
            'total_quantity_imported': Decimal('0'),
        }
        not_found_items = []
        matched_items = []

        if confirm:
            self.stdout.write(self.style.WARNING('CONFIRM mode - Changes will be applied'))
        else:
            self.stdout.write(self.style.NOTICE('PREVIEW mode - No changes will be made'))

        self.stdout.write(f'\nProcessing {len(aggregated)} aggregated combinations...\n')

        with transaction.atomic():
            for (hdbs_code, variant_code), data in aggregated.items():
                stats['combinations_processed'] += 1
                qty = data['qty']
                erp_items = data['erp_items']

                # Find matching inventory item by HDBS code
                item = hdbs_to_item.get(hdbs_code)
                if not item:
                    stats['items_not_found'] += 1
                    not_found_items.append({
                        'hdbs_code': hdbs_code,
                        'variant_code': variant_code,
                        'erp_items': list(erp_items),
                        'qty': qty,
                    })
                    continue

                stats['items_matched'] += 1
                stats['total_quantity_imported'] += qty

                matched_items.append({
                    'hdbs_code': hdbs_code,
                    'item_code': item.code,
                    'variant_code': variant_code,
                    'qty': qty,
                })

                variant_case = variant_cases.get(variant_code)
                if not variant_case:
                    self.stderr.write(self.style.WARNING(f'  Variant case not found: {variant_code}'))
                    continue

                if confirm:
                    # Find or create variant
                    variant, variant_created = ItemVariant.objects.get_or_create(
                        base_item=item,
                        variant_case=variant_case,
                        defaults={
                            'code': f'{item.code}-{variant_code}',
                        }
                    )
                    if variant_created:
                        stats['variants_created'] += 1
                    else:
                        stats['variants_updated'] += 1

                    # For NEW-CLI (LSTK Reclaim), set account to LSTK
                    if variant_code == 'NEW-CLI' and not variant.account:
                        variant.account = 'LSTK'
                        variant.save(update_fields=['account'])

                    # Find or create stock record
                    stock, stock_created = VariantStock.objects.get_or_create(
                        variant=variant,
                        location=default_location,
                        defaults={
                            'quantity_on_hand': qty,
                            'quantity_available': qty,
                        }
                    )
                    if stock_created:
                        stats['stock_records_created'] += 1
                    else:
                        stock.quantity_on_hand = qty
                        stock.quantity_available = qty
                        stock.last_movement_date = timezone.now()
                        stock.save()
                        stats['stock_records_updated'] += 1

                    if verbose:
                        self.stdout.write(f'  {item.code} [{variant_code}]: {qty}')
                else:
                    # Preview mode - just count
                    stats['variants_updated'] += 1

            if not confirm:
                # Rollback in preview mode
                transaction.set_rollback(True)

        # Print summary
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(self.style.SUCCESS('IMPORT SUMMARY'))
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(f"Combinations processed: {stats['combinations_processed']}")
        self.stdout.write(f"Items matched:          {stats['items_matched']}")
        self.stdout.write(self.style.WARNING(f"Items not found:        {stats['items_not_found']}"))

        if confirm:
            self.stdout.write(f"Variants created:       {stats['variants_created']}")
            self.stdout.write(f"Variants updated:       {stats['variants_updated']}")
            self.stdout.write(f"Stock records created:  {stats['stock_records_created']}")
            self.stdout.write(f"Stock records updated:  {stats['stock_records_updated']}")

        self.stdout.write(self.style.SUCCESS(f"Total quantity:         {stats['total_quantity_imported']}"))

        # Show quantity by variant case
        qty_by_variant = defaultdict(Decimal)
        for item in matched_items:
            qty_by_variant[item['variant_code']] += item['qty']

        self.stdout.write('')
        self.stdout.write('Quantity by Variant Case:')
        for vc, qty in sorted(qty_by_variant.items()):
            self.stdout.write(f'  {vc}: {qty}')

        if not_found_items and verbose:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(f'Items not found in system ({len(not_found_items)} items):'))
            for item in not_found_items[:20]:
                self.stdout.write(f"  HDBS {item['hdbs_code']} [{item['variant_code']}]: {item['qty']} ({', '.join(item['erp_items'][:3])})")
            if len(not_found_items) > 20:
                self.stdout.write(f"  ... and {len(not_found_items) - 20} more")

        if not confirm:
            self.stdout.write('')
            self.stdout.write(self.style.NOTICE('This was a PREVIEW. Run with --confirm to apply changes.'))
