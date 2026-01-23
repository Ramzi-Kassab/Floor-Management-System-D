"""
Import cutter stock quantities from ERP Excel file.

This command reads stock quantities from the Cutter Inventory Excel file
and updates the VariantStock model with the current quantities.

Excel Column Mapping:
- Column B (MN): HDBS MAT Number (hdbs_code attribute)
- Column H: ENO As New Cutter -> NEW-EO variant
- Column I: ENO Ground Cutter -> USED-GRD variant (GRD-EO)
- Column J: ARDT Reclaim Cutter -> USED-RCL variant
- Column K: LSTK Reclaim Cutter -> NEW-CLI variant (with LSTK account)
- Column L: New Stock -> NEW-PUR variant

Usage:
    python manage.py import_stock_from_excel                    # Preview mode
    python manage.py import_stock_from_excel --confirm          # Apply changes
    python manage.py import_stock_from_excel --file path/to.xlsx --confirm
"""

import os
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class Command(BaseCommand):
    help = 'Import cutter stock quantities from ERP Excel file'

    # Excel column mapping (0-indexed)
    COLUMN_MAP = {
        'hdbs_code': 1,      # Column B - MN (HDBS MAT Number)
        'product_name': 2,   # Column C - Product name
        'cutter_type': 3,    # Column D - Cutter type
        'size': 4,           # Column E - Size
        'chamfer': 5,        # Column F - Chamfer
        'family': 6,         # Column G - Family
        'NEW-EO': 7,         # Column H - ENO As New Cutter
        'USED-GRD': 8,       # Column I - ENO Ground Cutter
        'USED-RCL': 9,       # Column J - ARDT Reclaim Cutter
        'NEW-CLI': 10,       # Column K - LSTK Reclaim Cutter
        'NEW-PUR': 11,       # Column L - New Stock
    }

    # Variant cases to process
    VARIANT_CASES = ['NEW-EO', 'USED-GRD', 'USED-RCL', 'NEW-CLI', 'NEW-PUR']

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='docs/Cutter Inventory 01-15-2026.xlsx',
            help='Path to Excel file (default: docs/Cutter Inventory 01-15-2026.xlsx)'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='Sheet name to process (default: first sheet with "Cutters Inventory" in name)'
        )
        parser.add_argument(
            '--confirm',
            action='store_true',
            help='Actually apply changes (default is preview mode)'
        )
        parser.add_argument(
            '--skip-header',
            type=int,
            default=1,
            help='Number of header rows to skip (default: 1)'
        )

    def handle(self, *args, **options):
        if not HAS_OPENPYXL:
            self.stderr.write(self.style.ERROR('openpyxl is required. Install with: pip install openpyxl'))
            return

        file_path = options['file']
        sheet_name = options['sheet']
        confirm = options['confirm']
        skip_header = options['skip_header']

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        self.stdout.write(f'Loading Excel file: {file_path}')

        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Find the right sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                self.stderr.write(self.style.ERROR(f'Sheet "{sheet_name}" not found. Available: {wb.sheetnames}'))
                return
            ws = wb[sheet_name]
        else:
            # Auto-detect: first sheet with "Cutters Inventory" in name
            ws = None
            for name in wb.sheetnames:
                if 'cutters inventory' in name.lower():
                    ws = wb[name]
                    sheet_name = name
                    break
            if not ws:
                ws = wb.active
                sheet_name = ws.title

        self.stdout.write(f'Processing sheet: {sheet_name}')

        # Import models
        from apps.inventory.models import (
            InventoryItem, ItemVariant, VariantCase, VariantStock,
            InventoryLocation, ItemAttributeValue
        )

        # Get default location (main warehouse)
        default_location = InventoryLocation.objects.filter(is_default=True).first()
        if not default_location:
            default_location = InventoryLocation.objects.first()
            if not default_location:
                self.stderr.write(self.style.ERROR('No inventory locations found. Create one first.'))
                return

        self.stdout.write(f'Using location: {default_location.name}')

        # Get variant cases
        variant_cases = {vc.code: vc for vc in VariantCase.objects.all()}
        missing_cases = [c for c in self.VARIANT_CASES if c not in variant_cases]
        if missing_cases:
            self.stderr.write(self.style.WARNING(f'Warning: Missing variant cases: {missing_cases}'))

        # Build HDBS code to item mapping
        # Method 1: Direct mat_number field
        hdbs_to_item = {}
        for item in InventoryItem.objects.filter(mat_number__isnull=False).exclude(mat_number=''):
            hdbs_to_item[item.mat_number] = item

        # Method 2: hdbs_code attribute
        for attr_val in ItemAttributeValue.objects.filter(
            attribute__attribute__code__in=['hdbs_code', 'hdbs', 'hdbs_mat'],
            text_value__isnull=False
        ).exclude(text_value='').select_related('item'):
            if attr_val.text_value not in hdbs_to_item:
                hdbs_to_item[attr_val.text_value] = attr_val.item

        self.stdout.write(f'Found {len(hdbs_to_item)} items with HDBS codes')

        # Statistics
        stats = {
            'rows_processed': 0,
            'items_matched': 0,
            'items_not_found': 0,
            'variants_created': 0,
            'variants_updated': 0,
            'stock_records_created': 0,
            'stock_records_updated': 0,
            'total_quantity_imported': Decimal('0'),
        }
        not_found_items = []

        # Process rows
        rows = list(ws.iter_rows(min_row=skip_header + 1, values_only=True))
        total_rows = len(rows)

        if confirm:
            self.stdout.write(self.style.WARNING('CONFIRM mode - Changes will be applied'))
        else:
            self.stdout.write(self.style.NOTICE('PREVIEW mode - No changes will be made'))

        self.stdout.write(f'Processing {total_rows} rows...\n')

        with transaction.atomic():
            for row_idx, row in enumerate(rows, start=skip_header + 1):
                if not row or not row[self.COLUMN_MAP['hdbs_code']]:
                    continue

                hdbs_code = str(row[self.COLUMN_MAP['hdbs_code']]).strip()
                product_name = str(row[self.COLUMN_MAP['product_name']] or '').strip()

                stats['rows_processed'] += 1

                # Find matching inventory item
                item = hdbs_to_item.get(hdbs_code)
                if not item:
                    stats['items_not_found'] += 1
                    not_found_items.append({
                        'hdbs_code': hdbs_code,
                        'product_name': product_name,
                        'row': row_idx,
                    })
                    continue

                stats['items_matched'] += 1

                # Process each variant case
                for case_code in self.VARIANT_CASES:
                    col_idx = self.COLUMN_MAP.get(case_code)
                    if col_idx is None:
                        continue

                    qty_raw = row[col_idx]
                    if qty_raw is None or qty_raw == '':
                        continue

                    try:
                        qty = Decimal(str(qty_raw))
                    except:
                        continue

                    if qty <= 0:
                        continue

                    stats['total_quantity_imported'] += qty

                    variant_case = variant_cases.get(case_code)
                    if not variant_case:
                        continue

                    if confirm:
                        # Find or create variant
                        variant, variant_created = ItemVariant.objects.get_or_create(
                            base_item=item,
                            variant_case=variant_case,
                            defaults={
                                'code': f'{item.code}-{case_code}',
                            }
                        )
                        if variant_created:
                            stats['variants_created'] += 1
                        else:
                            stats['variants_updated'] += 1

                        # For NEW-CLI (LSTK Reclaim), set account to LSTK
                        if case_code == 'NEW-CLI' and not variant.account:
                            variant.account = 'LSTK'
                            variant.save(update_fields=['account'])

                        # Find or create stock record
                        stock, stock_created = VariantStock.objects.get_or_create(
                            variant=variant,
                            location=default_location,
                            defaults={
                                'quantity_on_hand': qty,
                                'quantity_available': qty,
                            }
                        )
                        if stock_created:
                            stats['stock_records_created'] += 1
                        else:
                            stock.quantity_on_hand = qty
                            stock.quantity_available = qty
                            stock.last_movement_date = timezone.now()
                            stock.save()
                            stats['stock_records_updated'] += 1
                    else:
                        # Preview mode - just count
                        stats['variants_updated'] += 1

            if not confirm:
                # Rollback in preview mode
                transaction.set_rollback(True)

        # Print summary
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(self.style.SUCCESS('IMPORT SUMMARY'))
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(f"Rows processed:        {stats['rows_processed']}")
        self.stdout.write(f"Items matched:         {stats['items_matched']}")
        self.stdout.write(self.style.WARNING(f"Items not found:       {stats['items_not_found']}"))
        self.stdout.write(f"Variants created:      {stats['variants_created']}")
        self.stdout.write(f"Variants updated:      {stats['variants_updated']}")
        self.stdout.write(f"Stock records created: {stats['stock_records_created']}")
        self.stdout.write(f"Stock records updated: {stats['stock_records_updated']}")
        self.stdout.write(self.style.SUCCESS(f"Total quantity:        {stats['total_quantity_imported']}"))

        if not_found_items:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(f'Items not found in system ({len(not_found_items)} items):'))
            for item in not_found_items[:20]:
                self.stdout.write(f"  Row {item['row']}: {item['hdbs_code']} - {item['product_name']}")
            if len(not_found_items) > 20:
                self.stdout.write(f"  ... and {len(not_found_items) - 20} more")

        if not confirm:
            self.stdout.write('')
            self.stdout.write(self.style.NOTICE('This was a PREVIEW. Run with --confirm to apply changes.'))
