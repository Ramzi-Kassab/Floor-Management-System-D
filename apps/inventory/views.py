"""
ARDT FMS - Inventory App Views
Version: 5.5
"""

import io
import json
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .forms import (
    CategoryAttributeForm,
    InventoryCategoryForm,
    InventoryItemForm,
    InventoryLocationForm,
    InventoryStockForm,
    InventoryTransactionForm,
    ItemVariantForm,
    StockAdjustmentForm,
)
from .models import (
    CategoryAttribute,
    InventoryCategory,
    InventoryItem,
    InventoryLocation,
    InventoryStock,
    InventoryTransaction,
    ItemAttributeValue,
    ItemVariant,
)


# =============================================================================
# Category Views
# =============================================================================


class CategoryListView(LoginRequiredMixin, ListView):
    """List all inventory categories."""

    model = InventoryCategory
    template_name = "inventory/category_list.html"
    context_object_name = "categories"

    def get_queryset(self):
        return InventoryCategory.objects.filter(parent__isnull=True).prefetch_related("children")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Inventory Categories"
        return context


class CategoryCreateView(LoginRequiredMixin, CreateView):
    """Create inventory category with inheritance support."""

    model = InventoryCategory
    form_class = InventoryCategoryForm
    template_name = "inventory/category_form.html"
    success_url = reverse_lazy("inventory:category_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        category = form.instance

        # Copy attributes from parent category if exists
        if category.parent:
            self._copy_parent_attributes(category)

            # Inherit name_template from parent if not set
            if not category.name_template and category.parent.name_template:
                category.name_template = category.parent.name_template
                category.save(update_fields=["name_template"])

        messages.success(self.request, f"Category '{category.name}' created successfully.")
        return response

    def _copy_parent_attributes(self, category):
        """Copy attributes from parent category, marking them as inherited."""
        for parent_attr in category.parent.attributes.all():
            CategoryAttribute.objects.create(
                category=category,
                code=parent_attr.code,
                name=parent_attr.name,
                attribute_type=parent_attr.attribute_type,
                unit=parent_attr.unit,
                min_value=parent_attr.min_value,
                max_value=parent_attr.max_value,
                options=parent_attr.options,
                is_required=parent_attr.is_required,
                is_used_in_name=parent_attr.is_used_in_name,
                display_order=parent_attr.display_order,
                is_inherited=True,
            )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Category"
        context["form_title"] = "Create Inventory Category"

        # Pass parent category for code prefix guidance
        parent_id = self.request.GET.get("parent")
        if parent_id:
            try:
                context["parent_category"] = InventoryCategory.objects.get(pk=parent_id)
            except InventoryCategory.DoesNotExist:
                pass

        return context


class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    """Update inventory category with sync to children support."""

    model = InventoryCategory
    form_class = InventoryCategoryForm
    template_name = "inventory/category_form.html"
    success_url = reverse_lazy("inventory:category_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        category = form.instance

        # Handle sync to children if requested
        sync_fields = self.request.POST.get("sync_to_children", "").split(",")
        sync_fields = [f.strip() for f in sync_fields if f.strip()]

        if sync_fields:
            synced_count = self._sync_to_children(category, sync_fields)
            if synced_count > 0:
                messages.info(
                    self.request,
                    f"Changes synced to {synced_count} child categor{'y' if synced_count == 1 else 'ies'}."
                )

        messages.success(self.request, f"Category '{category.name}' updated successfully.")
        return response

    def _sync_to_children(self, category, fields):
        """Sync specified fields to all child categories."""
        children = category.children.all()
        if not children.exists():
            return 0

        update_data = {}
        if "item_type" in fields:
            update_data["item_type"] = category.item_type
        if "name_template" in fields:
            update_data["name_template"] = category.name_template
        if "is_active" in fields:
            update_data["is_active"] = category.is_active

        if update_data:
            return children.update(**update_data)
        return 0

    def _get_inherited_attributes(self, category):
        """Get attributes inherited from parent category chain."""
        inherited = []
        parent = category.parent
        while parent:
            for attr in parent.attributes.all():
                inherited.append(attr)
            parent = parent.parent
        return inherited

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.name}"
        context["form_title"] = "Edit Category"

        # Add parent category for guidance
        if self.object.parent:
            context["parent_category"] = self.object.parent

        # Add child categories
        context["child_categories"] = list(self.object.children.all())

        # Add inherited attributes from parent chain
        context["inherited_attributes"] = self._get_inherited_attributes(self.object)

        return context


class CategoryDetailView(LoginRequiredMixin, DetailView):
    """View category details with attributes, children, and items."""

    model = InventoryCategory
    template_name = "inventory/category_detail.html"
    context_object_name = "category"

    def _get_inherited_attributes(self, category):
        """Get attributes inherited from parent category chain."""
        inherited = []
        parent = category.parent
        while parent:
            for attr in parent.attributes.all():
                inherited.append(attr)
            parent = parent.parent
        return inherited

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"{self.object.code} - {self.object.name}"

        # Child categories
        context["child_categories"] = self.object.children.all()

        # Own attributes (not inherited)
        context["own_attributes"] = self.object.attributes.filter(is_inherited=False)

        # Inherited attributes from parent chain
        context["inherited_attributes"] = self._get_inherited_attributes(self.object)

        # Items in this category
        context["items"] = self.object.items.all()[:10]  # Limit to 10 for preview
        context["items_count"] = self.object.items.count()

        return context


class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    """Delete inventory category."""

    model = InventoryCategory
    template_name = "inventory/category_confirm_delete.html"
    success_url = reverse_lazy("inventory:category_list")

    def form_valid(self, form):
        messages.success(self.request, f"Category '{self.object.name}' deleted successfully.")
        return super().form_valid(form)


# =============================================================================
# Category Attribute Views
# =============================================================================


class CategoryAttributeCreateView(LoginRequiredMixin, CreateView):
    """Create attribute for a category."""

    model = CategoryAttribute
    form_class = CategoryAttributeForm
    template_name = "inventory/category_attribute_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.category = get_object_or_404(InventoryCategory, pk=kwargs["category_pk"])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.category = self.category
        messages.success(self.request, f"Attribute '{form.instance.name}' created.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inventory:category_update", kwargs={"pk": self.category.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Add Attribute"
        context["form_title"] = f"Add Attribute to {self.category.name}"
        context["category"] = self.category
        return context


class CategoryAttributeUpdateView(LoginRequiredMixin, UpdateView):
    """Update category attribute."""

    model = CategoryAttribute
    form_class = CategoryAttributeForm
    template_name = "inventory/category_attribute_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.category = get_object_or_404(InventoryCategory, pk=kwargs["category_pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        return get_object_or_404(CategoryAttribute, pk=self.kwargs["pk"], category=self.category)

    def form_valid(self, form):
        messages.success(self.request, f"Attribute '{form.instance.name}' updated.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inventory:category_update", kwargs={"pk": self.category.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.name}"
        context["form_title"] = f"Edit Attribute: {self.object.name}"
        context["category"] = self.category
        return context


class CategoryAttributeDeleteView(LoginRequiredMixin, DeleteView):
    """Delete category attribute."""

    model = CategoryAttribute
    template_name = "inventory/category_attribute_confirm_delete.html"

    def dispatch(self, request, *args, **kwargs):
        self.category = get_object_or_404(InventoryCategory, pk=kwargs["category_pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        return get_object_or_404(CategoryAttribute, pk=self.kwargs["pk"], category=self.category)

    def get_success_url(self):
        return reverse_lazy("inventory:category_update", kwargs={"pk": self.category.pk})

    def form_valid(self, form):
        messages.success(self.request, f"Attribute '{self.object.name}' deleted.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["category"] = self.category
        return context


# =============================================================================
# Location Views
# =============================================================================


class LocationListView(LoginRequiredMixin, ListView):
    """List all inventory locations."""

    model = InventoryLocation
    template_name = "inventory/location_list.html"
    context_object_name = "locations"
    paginate_by = 25

    def get_queryset(self):
        qs = InventoryLocation.objects.select_related("warehouse")

        warehouse = self.request.GET.get("warehouse")
        if warehouse:
            qs = qs.filter(warehouse_id=warehouse)

        search = self.request.GET.get("q")
        if search:
            qs = qs.filter(Q(code__icontains=search) | Q(name__icontains=search))

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Inventory Locations"
        context["search_query"] = self.request.GET.get("q", "")
        from apps.sales.models import Warehouse

        context["warehouses"] = Warehouse.objects.filter(is_active=True)
        context["current_warehouse"] = self.request.GET.get("warehouse", "")
        return context


class LocationCreateView(LoginRequiredMixin, CreateView):
    """Create inventory location."""

    model = InventoryLocation
    form_class = InventoryLocationForm
    template_name = "inventory/location_form.html"
    success_url = reverse_lazy("inventory:location_list")

    def form_valid(self, form):
        messages.success(self.request, f"Location '{form.instance.code}' created successfully.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Location"
        context["form_title"] = "Create Inventory Location"
        return context


class LocationUpdateView(LoginRequiredMixin, UpdateView):
    """Update inventory location."""

    model = InventoryLocation
    form_class = InventoryLocationForm
    template_name = "inventory/location_form.html"
    success_url = reverse_lazy("inventory:location_list")

    def form_valid(self, form):
        messages.success(self.request, f"Location '{form.instance.code}' updated successfully.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.code}"
        context["form_title"] = "Edit Location"
        return context


# =============================================================================
# Item Views
# =============================================================================


class ItemListView(LoginRequiredMixin, ListView):
    """List all inventory items."""

    model = InventoryItem
    template_name = "inventory/item_list.html"
    context_object_name = "items"
    paginate_by = 25

    def get_queryset(self):
        qs = InventoryItem.objects.select_related("category", "category__parent", "primary_supplier").annotate(
            current_stock=Sum("stock_records__quantity_on_hand")
        )

        # Filter by category (including children)
        category = self.request.GET.get("category")
        if category:
            # Include items in subcategories
            qs = qs.filter(Q(category_id=category) | Q(category__parent_id=category))

        # Filter by type
        item_type = self.request.GET.get("type")
        if item_type:
            qs = qs.filter(item_type=item_type)

        # Filter by supplier
        supplier = self.request.GET.get("supplier")
        if supplier:
            qs = qs.filter(primary_supplier_id=supplier)

        # Filter by active status
        active = self.request.GET.get("active")
        if active == "1":
            qs = qs.filter(is_active=True)
        elif active == "0":
            qs = qs.filter(is_active=False)

        # Filter by low stock
        low_stock = self.request.GET.get("low_stock")
        if low_stock:
            qs = qs.filter(current_stock__lte=F("min_stock"))

        # Search
        search = self.request.GET.get("q")
        if search:
            qs = qs.filter(
                Q(code__icontains=search)
                | Q(name__icontains=search)
                | Q(description__icontains=search)
                | Q(supplier_part_number__icontains=search)
            )

        return qs.order_by("code")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Inventory Items"

        # Get categories with hierarchy (parent categories only for filter)
        context["categories"] = InventoryCategory.objects.filter(is_active=True, parent__isnull=True).prefetch_related("children")

        context["type_choices"] = InventoryItem.ItemType.choices

        # Get suppliers for filter
        from apps.supplychain.models import Supplier
        context["suppliers"] = Supplier.objects.filter(is_active=True).order_by("name")

        # Current filter values
        context["current_category"] = self.request.GET.get("category", "")
        context["current_type"] = self.request.GET.get("type", "")
        context["current_supplier"] = self.request.GET.get("supplier", "")
        context["current_active"] = self.request.GET.get("active", "")
        context["search_query"] = self.request.GET.get("q", "")
        context["low_stock_filter"] = self.request.GET.get("low_stock", "")

        # Summary statistics
        all_items = InventoryItem.objects.all()
        context["total_items"] = all_items.count()
        context["active_items"] = all_items.filter(is_active=True).count()
        context["low_stock_count"] = InventoryItem.objects.annotate(
            stock=Sum("stock_records__quantity_on_hand")
        ).filter(stock__lte=F("min_stock")).count()

        return context


class ItemDetailView(LoginRequiredMixin, DetailView):
    """View inventory item details."""

    model = InventoryItem
    template_name = "inventory/item_detail.html"
    context_object_name = "item"

    def get_queryset(self):
        return InventoryItem.objects.select_related("category", "primary_supplier", "created_by")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = self.object
        context["page_title"] = f"{item.code} - {item.name}"

        # Stock by location
        context["stock_by_location"] = item.stock_records.select_related("location", "location__warehouse")

        # Recent transactions
        context["recent_transactions"] = item.transactions.select_related("created_by", "from_location", "to_location").order_by(
            "-transaction_date"
        )[:10]

        # Total stock
        context["total_stock"] = item.total_stock

        # Is low stock?
        context["is_low_stock"] = item.total_stock <= item.min_stock

        # Attribute values
        context["attribute_values"] = item.attribute_values.select_related("attribute").order_by(
            "attribute__display_order", "attribute__name"
        )

        # Variant stock total
        variant_stock = 0
        for variant in item.variants.all():
            variant_stock += variant.total_stock
        context["variant_stock"] = variant_stock

        return context


class ItemCreateView(LoginRequiredMixin, CreateView):
    """Create inventory item."""

    model = InventoryItem
    form_class = InventoryItemForm
    template_name = "inventory/item_form.html"

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        # Save attribute values
        self._save_attribute_values(self.object)
        messages.success(self.request, f"Item '{form.instance.code}' created successfully.")
        return response

    def _save_attribute_values(self, item):
        """Save attribute values from form submission."""
        if not item.category:
            return

        for key, value in self.request.POST.items():
            if key.startswith("attr_"):
                try:
                    attr_id = int(key.replace("attr_", ""))
                    attribute = CategoryAttribute.objects.get(pk=attr_id, category=item.category)

                    # Create or update attribute value
                    attr_value, _ = ItemAttributeValue.objects.get_or_create(
                        item=item,
                        attribute=attribute
                    )

                    # Set value based on type
                    if attribute.attribute_type == CategoryAttribute.AttributeType.NUMBER:
                        if value:
                            attr_value.number_value = Decimal(value)
                        else:
                            attr_value.number_value = None
                    elif attribute.attribute_type == CategoryAttribute.AttributeType.BOOLEAN:
                        attr_value.boolean_value = value in ["on", "true", "True", "1"]
                    elif attribute.attribute_type == CategoryAttribute.AttributeType.DATE:
                        if value:
                            from datetime import datetime
                            attr_value.date_value = datetime.strptime(value, "%Y-%m-%d").date()
                        else:
                            attr_value.date_value = None
                    else:
                        attr_value.text_value = value or ""

                    attr_value.save()
                except (ValueError, CategoryAttribute.DoesNotExist):
                    pass

    def get_success_url(self):
        return reverse_lazy("inventory:item_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Item"
        context["form_title"] = "Create Inventory Item"
        # Add categories with hierarchy for dropdown
        context["categories"] = InventoryCategory.objects.filter(
            is_active=True, parent__isnull=True
        ).prefetch_related("children")
        context["type_choices"] = InventoryItem.ItemType.choices
        context["existing_attribute_values"] = json.dumps({})
        return context


class ItemUpdateView(LoginRequiredMixin, UpdateView):
    """Update inventory item."""

    model = InventoryItem
    form_class = InventoryItemForm
    template_name = "inventory/item_form.html"

    def form_valid(self, form):
        response = super().form_valid(form)
        # Save attribute values
        self._save_attribute_values(self.object)
        messages.success(self.request, f"Item '{form.instance.code}' updated successfully.")
        return response

    def _save_attribute_values(self, item):
        """Save attribute values from form submission."""
        if not item.category:
            return

        for key, value in self.request.POST.items():
            if key.startswith("attr_"):
                try:
                    attr_id = int(key.replace("attr_", ""))
                    attribute = CategoryAttribute.objects.get(pk=attr_id, category=item.category)

                    # Create or update attribute value
                    attr_value, _ = ItemAttributeValue.objects.get_or_create(
                        item=item,
                        attribute=attribute
                    )

                    # Set value based on type
                    if attribute.attribute_type == CategoryAttribute.AttributeType.NUMBER:
                        if value:
                            attr_value.number_value = Decimal(value)
                        else:
                            attr_value.number_value = None
                    elif attribute.attribute_type == CategoryAttribute.AttributeType.BOOLEAN:
                        attr_value.boolean_value = value in ["on", "true", "True", "1"]
                    elif attribute.attribute_type == CategoryAttribute.AttributeType.DATE:
                        if value:
                            from datetime import datetime
                            attr_value.date_value = datetime.strptime(value, "%Y-%m-%d").date()
                        else:
                            attr_value.date_value = None
                    else:
                        attr_value.text_value = value or ""

                    attr_value.save()
                except (ValueError, CategoryAttribute.DoesNotExist):
                    pass

    def get_success_url(self):
        return reverse_lazy("inventory:item_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.code}"
        context["form_title"] = "Edit Inventory Item"
        # Add categories with hierarchy for dropdown
        context["categories"] = InventoryCategory.objects.filter(
            is_active=True, parent__isnull=True
        ).prefetch_related("children")
        context["type_choices"] = InventoryItem.ItemType.choices

        # Get existing attribute values for this item
        existing_values = {}
        for attr_value in self.object.attribute_values.select_related("attribute"):
            attr = attr_value.attribute
            if attr.attribute_type == CategoryAttribute.AttributeType.NUMBER:
                existing_values[attr.code] = float(attr_value.number_value) if attr_value.number_value else ""
            elif attr.attribute_type == CategoryAttribute.AttributeType.BOOLEAN:
                existing_values[attr.code] = attr_value.boolean_value or False
            elif attr.attribute_type == CategoryAttribute.AttributeType.DATE:
                existing_values[attr.code] = str(attr_value.date_value) if attr_value.date_value else ""
            else:
                existing_values[attr.code] = attr_value.text_value or ""
        context["existing_attribute_values"] = json.dumps(existing_values)
        return context


# =============================================================================
# Transaction Views
# =============================================================================


class TransactionListView(LoginRequiredMixin, ListView):
    """List all inventory transactions."""

    model = InventoryTransaction
    template_name = "inventory/transaction_list.html"
    context_object_name = "transactions"
    paginate_by = 50

    def get_queryset(self):
        qs = InventoryTransaction.objects.select_related("item", "from_location", "to_location", "created_by")

        # Filter by type
        trans_type = self.request.GET.get("type")
        if trans_type:
            qs = qs.filter(transaction_type=trans_type)

        # Filter by item
        item = self.request.GET.get("item")
        if item:
            qs = qs.filter(item_id=item)

        # Search
        search = self.request.GET.get("q")
        if search:
            qs = qs.filter(Q(transaction_number__icontains=search) | Q(reference_number__icontains=search))

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Inventory Transactions"
        context["type_choices"] = InventoryTransaction.TransactionType.choices
        context["current_type"] = self.request.GET.get("type", "")
        context["search_query"] = self.request.GET.get("q", "")
        return context


class TransactionDetailView(LoginRequiredMixin, DetailView):
    """View transaction details."""

    model = InventoryTransaction
    template_name = "inventory/transaction_detail.html"
    context_object_name = "transaction"

    def get_queryset(self):
        return InventoryTransaction.objects.select_related("item", "from_location", "to_location", "created_by")


class TransactionCreateView(LoginRequiredMixin, CreateView):
    """Create inventory transaction."""

    model = InventoryTransaction
    form_class = InventoryTransactionForm
    template_name = "inventory/transaction_form.html"
    success_url = reverse_lazy("inventory:transaction_list")

    def form_valid(self, form):
        transaction = form.save(commit=False)
        transaction.created_by = self.request.user
        transaction.transaction_date = timezone.now()
        transaction.transaction_number = self.generate_transaction_number()

        # Calculate total cost
        transaction.total_cost = transaction.quantity * transaction.unit_cost

        transaction.save()

        # Update stock levels
        self.update_stock(transaction)

        messages.success(self.request, f"Transaction '{transaction.transaction_number}' created successfully.")
        return redirect(self.success_url)

    def generate_transaction_number(self):
        """Generate unique transaction number."""
        prefix = "TXN"
        today = timezone.now().strftime("%Y%m%d")
        last = InventoryTransaction.objects.filter(transaction_number__startswith=f"{prefix}-{today}").order_by("-id").first()
        if last:
            try:
                last_num = int(last.transaction_number.split("-")[-1])
                next_num = last_num + 1
            except (ValueError, IndexError):
                next_num = 1
        else:
            next_num = 1
        return f"{prefix}-{today}-{str(next_num).zfill(4)}"

    def update_stock(self, transaction):
        """Update stock levels based on transaction."""
        if transaction.transaction_type == "RECEIPT":
            # Add to destination
            stock, _ = InventoryStock.objects.get_or_create(
                item=transaction.item, location=transaction.to_location, defaults={"quantity_on_hand": 0}
            )
            stock.quantity_on_hand = F("quantity_on_hand") + transaction.quantity
            stock.last_movement_date = timezone.now()
            stock.save()

        elif transaction.transaction_type == "ISSUE":
            # Remove from source
            try:
                stock = InventoryStock.objects.get(item=transaction.item, location=transaction.from_location)
                stock.quantity_on_hand = F("quantity_on_hand") - transaction.quantity
                stock.last_movement_date = timezone.now()
                stock.save()
            except InventoryStock.DoesNotExist:
                pass

        elif transaction.transaction_type == "TRANSFER":
            # Remove from source
            try:
                from_stock = InventoryStock.objects.get(item=transaction.item, location=transaction.from_location)
                from_stock.quantity_on_hand = F("quantity_on_hand") - transaction.quantity
                from_stock.last_movement_date = timezone.now()
                from_stock.save()
            except InventoryStock.DoesNotExist:
                pass

            # Add to destination
            to_stock, _ = InventoryStock.objects.get_or_create(
                item=transaction.item, location=transaction.to_location, defaults={"quantity_on_hand": 0}
            )
            to_stock.quantity_on_hand = F("quantity_on_hand") + transaction.quantity
            to_stock.last_movement_date = timezone.now()
            to_stock.save()

        elif transaction.transaction_type == "ADJUSTMENT":
            # Adjustment can be positive or negative
            if transaction.to_location:
                stock, _ = InventoryStock.objects.get_or_create(
                    item=transaction.item, location=transaction.to_location, defaults={"quantity_on_hand": 0}
                )
                stock.quantity_on_hand = F("quantity_on_hand") + transaction.quantity
                stock.last_movement_date = timezone.now()
                stock.save()
            elif transaction.from_location:
                try:
                    stock = InventoryStock.objects.get(item=transaction.item, location=transaction.from_location)
                    stock.quantity_on_hand = F("quantity_on_hand") - transaction.quantity
                    stock.last_movement_date = timezone.now()
                    stock.save()
                except InventoryStock.DoesNotExist:
                    pass

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "New Transaction"
        context["form_title"] = "Create Inventory Transaction"
        return context


# =============================================================================
# Stock Views
# =============================================================================


class StockListView(LoginRequiredMixin, ListView):
    """List all stock records."""

    model = InventoryStock
    template_name = "inventory/stock_list.html"
    context_object_name = "stock_records"
    paginate_by = 50

    def get_queryset(self):
        qs = InventoryStock.objects.select_related("item", "location", "location__warehouse").filter(quantity_on_hand__gt=0)

        # Filter by warehouse
        warehouse = self.request.GET.get("warehouse")
        if warehouse:
            qs = qs.filter(location__warehouse_id=warehouse)

        # Filter by item
        item = self.request.GET.get("item")
        if item:
            qs = qs.filter(item_id=item)

        return qs.order_by("item__code", "location__code")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Stock Levels"
        from apps.sales.models import Warehouse

        context["warehouses"] = Warehouse.objects.filter(is_active=True)
        context["current_warehouse"] = self.request.GET.get("warehouse", "")
        return context


class StockAdjustView(LoginRequiredMixin, View):
    """Quick stock adjustment view."""

    def post(self, request, pk):
        stock = get_object_or_404(InventoryStock, pk=pk)
        form = StockAdjustmentForm(request.POST)

        if form.is_valid():
            adjustment_type = form.cleaned_data["adjustment_type"]
            quantity = form.cleaned_data["quantity"]
            reason = form.cleaned_data["reason"]
            notes = form.cleaned_data["notes"]

            # Create transaction
            trans_type = "ADJUSTMENT"
            transaction = InventoryTransaction.objects.create(
                transaction_number=f"ADJ-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                transaction_type=trans_type,
                transaction_date=timezone.now(),
                item=stock.item,
                from_location=stock.location if adjustment_type == "REMOVE" else None,
                to_location=stock.location if adjustment_type == "ADD" else None,
                quantity=quantity,
                unit=stock.item.unit,
                reason=reason,
                notes=notes,
                created_by=request.user,
            )

            # Update stock
            if adjustment_type == "ADD":
                stock.quantity_on_hand = F("quantity_on_hand") + quantity
            else:
                stock.quantity_on_hand = F("quantity_on_hand") - quantity
            stock.last_movement_date = timezone.now()
            stock.save()

            messages.success(request, f"Stock adjusted successfully. Transaction: {transaction.transaction_number}")
        else:
            messages.error(request, "Invalid adjustment data.")

        return redirect("inventory:item_detail", pk=stock.item.pk)


# =============================================================================
# Import/Export Views
# =============================================================================


class ItemExportView(LoginRequiredMixin, View):
    """Export inventory items to Excel."""

    def get(self, request):
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Items"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "Code*",
            "Name*",
            "Item Type*",
            "Category Code",
            "Description",
            "Unit",
            "Standard Cost",
            "Currency",
            "Min Stock",
            "Max Stock",
            "Reorder Point",
            "Reorder Qty",
            "Lead Time Days",
            "Supplier Code",
            "Supplier Part#",
            "Serialized (Y/N)",
            "Lot Controlled (Y/N)",
            "Active (Y/N)",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Get items to export
        items = InventoryItem.objects.select_related("category", "primary_supplier").order_by("code")

        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=item.code).border = thin_border
            ws.cell(row=row, column=2, value=item.name).border = thin_border
            ws.cell(row=row, column=3, value=item.item_type).border = thin_border
            ws.cell(row=row, column=4, value=item.category.code if item.category else "").border = thin_border
            ws.cell(row=row, column=5, value=item.description or "").border = thin_border
            ws.cell(row=row, column=6, value=item.unit).border = thin_border
            ws.cell(row=row, column=7, value=float(item.standard_cost)).border = thin_border
            ws.cell(row=row, column=8, value=item.currency).border = thin_border
            ws.cell(row=row, column=9, value=float(item.min_stock)).border = thin_border
            ws.cell(row=row, column=10, value=float(item.max_stock) if item.max_stock else "").border = thin_border
            ws.cell(row=row, column=11, value=float(item.reorder_point)).border = thin_border
            ws.cell(row=row, column=12, value=float(item.reorder_quantity)).border = thin_border
            ws.cell(row=row, column=13, value=item.lead_time_days or "").border = thin_border
            ws.cell(row=row, column=14, value=item.primary_supplier.code if item.primary_supplier else "").border = thin_border
            ws.cell(row=row, column=15, value=item.supplier_part_number or "").border = thin_border
            ws.cell(row=row, column=16, value="Y" if item.is_serialized else "N").border = thin_border
            ws.cell(row=row, column=17, value="Y" if item.is_lot_controlled else "N").border = thin_border
            ws.cell(row=row, column=18, value="Y" if item.is_active else "N").border = thin_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"inventory_items_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response


class ItemImportTemplateView(LoginRequiredMixin, View):
    """Download Excel template for importing items."""

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        required_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers with descriptions
        headers = [
            ("Code*", "Unique item code (required)"),
            ("Name*", "Item name (required)"),
            ("Item Type*", "RAW_MATERIAL, COMPONENT, CONSUMABLE, TOOL, SPARE_PART, FINISHED_GOOD"),
            ("Category Code", "Category code (must exist)"),
            ("Description", "Item description"),
            ("Unit", "EA, KG, M, L, etc. (default: EA)"),
            ("Standard Cost", "Standard unit cost (default: 0)"),
            ("Currency", "SAR, USD, EUR (default: SAR)"),
            ("Min Stock", "Minimum stock level (default: 0)"),
            ("Max Stock", "Maximum stock level"),
            ("Reorder Point", "When to reorder (default: 0)"),
            ("Reorder Qty", "How much to order (default: 1)"),
            ("Lead Time Days", "Supplier lead time"),
            ("Supplier Code", "Supplier code (must exist)"),
            ("Supplier Part#", "Part number at supplier"),
            ("Serialized (Y/N)", "Track by serial number (default: N)"),
            ("Lot Controlled (Y/N)", "Track by lot/batch (default: N)"),
            ("Active (Y/N)", "Is item active (default: Y)"),
        ]

        # Write headers
        for col, (header, desc) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            # Add comment with description
            ws.cell(row=2, column=col, value=desc).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Add example rows
        examples = [
            ("CUT-001", "Diamond Cutter 13mm", "COMPONENT", "CUTTERS", "Standard diamond cutter", "EA", 250.00, "SAR", 10, 100, 15, 20, 14, "", "", "N", "Y", "Y"),
            ("MAT-001", "Tungsten Carbide Powder", "RAW_MATERIAL", "MATRIX", "Grade A matrix material", "KG", 150.00, "SAR", 5, 50, 10, 15, 21, "", "", "N", "Y", "Y"),
            ("CON-001", "Brazing Alloy", "CONSUMABLE", "", "High temp brazing alloy", "KG", 85.00, "SAR", 2, 20, 5, 10, 7, "", "", "N", "N", "Y"),
        ]

        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column].width = adjusted_width

        # Set row height for description row
        ws.row_dimensions[2].height = 40

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="inventory_import_template.xlsx"'

        return response


class ItemImportView(LoginRequiredMixin, View):
    """Import inventory items from Excel."""

    def get(self, request):
        context = {
            "page_title": "Import Inventory Items",
            "categories": InventoryCategory.objects.filter(is_active=True),
            "type_choices": InventoryItem.ItemType.choices,
        }
        return render(request, "inventory/item_import.html", context)

    def post(self, request):
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file to import.")
            return redirect("inventory:item_import")

        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls)")
            return redirect("inventory:item_import")

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active

            # Get category and supplier lookups
            categories = {c.code.upper(): c for c in InventoryCategory.objects.all()}
            from apps.supplychain.models import Supplier
            suppliers = {s.code.upper(): s for s in Supplier.objects.all()}

            valid_types = [t[0] for t in InventoryItem.ItemType.choices]

            created_count = 0
            updated_count = 0
            errors = []

            # Skip header row(s) - find first data row
            start_row = 1
            for row in ws.iter_rows(min_row=1, max_row=5):
                cell_value = str(row[0].value or "").strip()
                if cell_value and cell_value not in ["Code*", "Code", "Unique item code (required)"]:
                    break
                start_row += 1

            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
                # Skip empty rows
                if not row or not row[0]:
                    continue

                code = str(row[0]).strip() if row[0] else None
                name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                item_type = str(row[2]).strip().upper() if len(row) > 2 and row[2] else None

                # Validate required fields
                if not code or not name or not item_type:
                    errors.append(f"Row {row_idx}: Missing required field (Code, Name, or Item Type)")
                    continue

                if item_type not in valid_types:
                    errors.append(f"Row {row_idx}: Invalid item type '{item_type}'")
                    continue

                # Parse optional fields
                category_code = str(row[3]).strip().upper() if len(row) > 3 and row[3] else None
                description = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                unit = str(row[5]).strip().upper() if len(row) > 5 and row[5] else "EA"

                try:
                    standard_cost = Decimal(str(row[6])) if len(row) > 6 and row[6] else Decimal("0")
                except (InvalidOperation, ValueError):
                    standard_cost = Decimal("0")

                currency = str(row[7]).strip().upper() if len(row) > 7 and row[7] else "SAR"

                try:
                    min_stock = Decimal(str(row[8])) if len(row) > 8 and row[8] else Decimal("0")
                except (InvalidOperation, ValueError):
                    min_stock = Decimal("0")

                try:
                    max_stock = Decimal(str(row[9])) if len(row) > 9 and row[9] else None
                except (InvalidOperation, ValueError):
                    max_stock = None

                try:
                    reorder_point = Decimal(str(row[10])) if len(row) > 10 and row[10] else Decimal("0")
                except (InvalidOperation, ValueError):
                    reorder_point = Decimal("0")

                try:
                    reorder_quantity = Decimal(str(row[11])) if len(row) > 11 and row[11] else Decimal("1")
                except (InvalidOperation, ValueError):
                    reorder_quantity = Decimal("1")

                try:
                    lead_time_days = int(row[12]) if len(row) > 12 and row[12] else None
                except (ValueError, TypeError):
                    lead_time_days = None

                supplier_code = str(row[13]).strip().upper() if len(row) > 13 and row[13] else None
                supplier_part = str(row[14]).strip() if len(row) > 14 and row[14] else ""
                is_serialized = str(row[15]).strip().upper() in ["Y", "YES", "TRUE", "1"] if len(row) > 15 and row[15] else False
                is_lot_controlled = str(row[16]).strip().upper() in ["Y", "YES", "TRUE", "1"] if len(row) > 16 and row[16] else False
                is_active = str(row[17]).strip().upper() not in ["N", "NO", "FALSE", "0"] if len(row) > 17 and row[17] else True

                # Lookup category
                category = categories.get(category_code) if category_code else None
                if category_code and not category:
                    errors.append(f"Row {row_idx}: Category '{category_code}' not found")
                    continue

                # Lookup supplier
                supplier = suppliers.get(supplier_code) if supplier_code else None
                if supplier_code and not supplier:
                    errors.append(f"Row {row_idx}: Supplier '{supplier_code}' not found (skipping supplier)")
                    supplier = None

                # Create or update item
                item, created = InventoryItem.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name,
                        "item_type": item_type,
                        "category": category,
                        "description": description,
                        "unit": unit,
                        "standard_cost": standard_cost,
                        "currency": currency,
                        "min_stock": min_stock,
                        "max_stock": max_stock,
                        "reorder_point": reorder_point,
                        "reorder_quantity": reorder_quantity,
                        "lead_time_days": lead_time_days,
                        "primary_supplier": supplier,
                        "supplier_part_number": supplier_part,
                        "is_serialized": is_serialized,
                        "is_lot_controlled": is_lot_controlled,
                        "is_active": is_active,
                    },
                )

                if created:
                    item.created_by = request.user
                    item.save()
                    created_count += 1
                else:
                    updated_count += 1

            # Report results
            if created_count > 0 or updated_count > 0:
                messages.success(request, f"Import complete: {created_count} items created, {updated_count} items updated.")

            if errors:
                error_msg = f"Import had {len(errors)} errors. First 5: " + "; ".join(errors[:5])
                if len(errors) > 5:
                    error_msg += f"... and {len(errors) - 5} more."
                messages.warning(request, error_msg)

            if created_count == 0 and updated_count == 0 and not errors:
                messages.info(request, "No items were imported. Please check your file format.")

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

        return redirect("inventory:item_list")


# =============================================================================
# API Views
# =============================================================================


class CategoryAttributesAPIView(LoginRequiredMixin, View):
    """API endpoint to get category attributes."""

    def get(self, request, category_pk):
        category = get_object_or_404(InventoryCategory, pk=category_pk)

        attributes = []
        for attr in category.attributes.all().order_by("display_order", "name"):
            attributes.append({
                "id": attr.id,
                "code": attr.code,
                "name": attr.name,
                "attribute_type": attr.attribute_type,
                "unit": attr.unit,
                "min_value": float(attr.min_value) if attr.min_value else None,
                "max_value": float(attr.max_value) if attr.max_value else None,
                "options": attr.options,
                "is_required": attr.is_required,
                "is_used_in_name": attr.is_used_in_name,
            })

        return JsonResponse({
            "category_id": category.pk,
            "category_code": category.code,
            "item_type": category.item_type,
            "name_template": category.name_template,
            "attributes": attributes,
        })


class CategoryGenerateCodeAPIView(LoginRequiredMixin, View):
    """API endpoint to generate next item code for a category."""

    def post(self, request, category_pk):
        category = get_object_or_404(InventoryCategory, pk=category_pk)
        code = category.generate_next_code()
        return JsonResponse({"code": code})


# =============================================================================
# Item Variant Views
# =============================================================================


class ItemVariantCreateView(LoginRequiredMixin, CreateView):
    """Create item variant."""

    model = ItemVariant
    form_class = ItemVariantForm
    template_name = "inventory/item_variant_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.item = get_object_or_404(InventoryItem, pk=kwargs["item_pk"])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.base_item = self.item
        form.instance.created_by = self.request.user
        messages.success(self.request, f"Variant '{form.instance.code}' created.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inventory:item_detail", kwargs={"pk": self.item.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Add Variant - {self.item.code}"
        context["form_title"] = f"Add Variant for {self.item.name}"
        context["item"] = self.item
        return context


class ItemVariantUpdateView(LoginRequiredMixin, UpdateView):
    """Update item variant."""

    model = ItemVariant
    form_class = ItemVariantForm
    template_name = "inventory/item_variant_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.item = get_object_or_404(InventoryItem, pk=kwargs["item_pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        return get_object_or_404(ItemVariant, pk=self.kwargs["pk"], base_item=self.item)

    def form_valid(self, form):
        messages.success(self.request, f"Variant '{form.instance.code}' updated.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inventory:item_detail", kwargs={"pk": self.item.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit Variant - {self.object.code}"
        context["form_title"] = f"Edit Variant: {self.object.code}"
        context["item"] = self.item
        return context


class ItemVariantDeleteView(LoginRequiredMixin, DeleteView):
    """Delete item variant."""

    model = ItemVariant
    template_name = "inventory/item_variant_confirm_delete.html"

    def dispatch(self, request, *args, **kwargs):
        self.item = get_object_or_404(InventoryItem, pk=kwargs["item_pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        return get_object_or_404(ItemVariant, pk=self.kwargs["pk"], base_item=self.item)

    def get_success_url(self):
        return reverse_lazy("inventory:item_detail", kwargs={"pk": self.item.pk})

    def form_valid(self, form):
        messages.success(self.request, f"Variant '{self.object.code}' deleted.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["item"] = self.item
        return context
