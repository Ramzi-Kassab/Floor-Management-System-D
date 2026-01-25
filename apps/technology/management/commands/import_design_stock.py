"""
Import design (RM) stock quantities from ERP Excel file.

This command reads stock quantities for Raw Material (RM) items from the ERP
Items export file and updates the Design model with current quantities.

Expected Columns:
- Item number: ERP item number (filter for RM-* items)
- Search name: MAT number (or extract from Product name)
- Product name: e.g., "8 1/2:GT65DH:1234567" - used to extract MAT if Search name empty
- Physical inventory: Current stock quantity (On Hand)
- Ordered in total: Quantity on order

Usage:
    python manage.py import_design_stock                    # Preview mode
    python manage.py import_design_stock --confirm          # Apply changes
    python manage.py import_design_stock --file path/to.xlsx --confirm
"""

import os
from collections import defaultdict
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class Command(BaseCommand):
    help = 'Import design (RM) stock quantities from ERP Excel file'

    # Supports both Items file and On-hand file formats
    HEADER_PATTERNS = {
        'item_number': ['item number'],
        'product_name': ['product name'],
        'search_name': ['search name', 'color'],  # Color column in On-hand file contains MAT
        'physical_inventory': ['physical inventory', 'available physical'],  # On-hand uses "Available physical"
        'ordered_total': ['ordered in total'],
        'warehouse': ['warehouse'],
    }

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='docs/Items_639021531472517099.xlsx',
            help='Path to ERP Items Excel file'
        )
        parser.add_argument(
            '--confirm',
            action='store_true',
            help='Actually apply changes (default is preview mode)'
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Show detailed per-item import info'
        )

    def find_header_row(self, ws):
        """Auto-detect the header row by searching for 'Item number' column."""
        max_search_rows = 20

        for row_idx in range(1, min(max_search_rows + 1, ws.max_row + 1)):
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if cell_str in self.HEADER_PATTERNS['item_number']:
                        column_map = self._map_columns(ws, row_idx)
                        return row_idx, column_map

        return None, None

    def _map_columns(self, ws, header_row):
        """Map column names to their positions based on header row."""
        column_map = {}

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if not cell_value:
                continue

            cell_str = str(cell_value).strip().lower()

            for field_name, patterns in self.HEADER_PATTERNS.items():
                if cell_str in patterns or any(p in cell_str for p in patterns):
                    if field_name not in column_map:
                        column_map[field_name] = col_idx
                        break

        return column_map

    def parse_product_name(self, product_name):
        """
        Parse product name to extract size, HDBS type, and MAT number.
        Format: "8 1/2:GT65DH:1234567"
        """
        if not product_name:
            return None, None, None

        product_name = str(product_name).strip()

        # Try colon separator first
        if ':' in product_name:
            parts = [p.strip() for p in product_name.split(':')]
            if len(parts) >= 3:
                return parts[0], parts[1], parts[2]
            elif len(parts) == 2:
                return parts[0], parts[1], None

        # Fallback to space separator
        parts = product_name.split()
        if len(parts) >= 3:
            for i, part in enumerate(parts):
                if part.isdigit() and len(part) >= 6:
                    size_parts = parts[:i-1] if i > 1 else []
                    hdbs_type = parts[i-1] if i > 0 else None
                    return ' '.join(size_parts), hdbs_type, part

        return None, None, None

    def handle(self, *args, **options):
        if not HAS_OPENPYXL:
            self.stderr.write(self.style.ERROR('openpyxl is required. Install with: pip install openpyxl'))
            return

        file_path = options['file']
        confirm = options['confirm']
        verbose = options['verbose']

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'File not found: {file_path}'))
            return

        self.stdout.write(f'Loading Excel file: {file_path}')

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        sheet_name = ws.title

        self.stdout.write(f'Processing sheet: {sheet_name}')

        header_row, column_map = self.find_header_row(ws)

        if header_row is None:
            self.stderr.write(self.style.ERROR(
                'Could not find header row. Looking for "Item number" column in first 20 rows.'
            ))
            return

        self.stdout.write(self.style.SUCCESS(f'Found headers at row {header_row}'))
        self.stdout.write(f'Column mapping: {column_map}')

        # Validate required columns
        has_mat_source = 'search_name' in column_map or 'product_name' in column_map
        has_qty_source = 'physical_inventory' in column_map

        if not has_mat_source:
            self.stderr.write(self.style.ERROR(
                f'Missing MAT number source. Need "Search name" or "Product name" column.'
            ))
            return

        if not has_qty_source:
            self.stderr.write(self.style.ERROR(
                f'Missing quantity column. Need "Physical inventory" column.'
            ))
            return

        data_start_row = header_row + 1
        self.stdout.write(f'Data starts at row: {data_start_row}')

        from apps.technology.models import Design

        # Build MAT -> Design mapping
        mat_to_design = {}
        for design in Design.objects.all():
            if design.mat_no:
                mat_to_design[str(design.mat_no).strip()] = design

        self.stdout.write(f'Found {len(mat_to_design)} designs with MAT numbers')

        # Statistics
        stats = {
            'rows_processed': 0,
            'rm_rows': 0,
            'mats_matched': 0,
            'mats_not_found': 0,
            'designs_updated': 0,
            'total_on_hand': 0,
            'total_on_order': 0,
            'skipped_not_rm': 0,
            'skipped_no_mat': 0,
        }
        not_found_items = []
        matched_items = []

        if confirm:
            self.stdout.write(self.style.WARNING('CONFIRM mode - Changes will be applied'))
        else:
            self.stdout.write(self.style.NOTICE('PREVIEW mode - No changes will be made'))

        self.stdout.write(f'\nProcessing rows...\n')

        with transaction.atomic():
            for row_idx in range(data_start_row, ws.max_row + 1):
                item_number = ws.cell(row=row_idx, column=column_map['item_number']).value
                if not item_number:
                    continue

                item_number = str(item_number).strip()
                stats['rows_processed'] += 1

                # Only process RM items
                if not item_number.upper().startswith('RM'):
                    stats['skipped_not_rm'] += 1
                    continue

                stats['rm_rows'] += 1

                # Get MAT number
                mat_no = None
                hdbs_type = None

                if 'search_name' in column_map:
                    search_name = ws.cell(row=row_idx, column=column_map['search_name']).value
                    if search_name:
                        mat_no = str(search_name).strip()

                if not mat_no and 'product_name' in column_map:
                    product_name = ws.cell(row=row_idx, column=column_map['product_name']).value
                    if product_name:
                        _, hdbs_type, mat_no = self.parse_product_name(product_name)

                if not mat_no:
                    stats['skipped_no_mat'] += 1
                    continue

                # Get warehouse value - only 'Store' counts as on-hand
                warehouse = ''
                if 'warehouse' in column_map:
                    warehouse_raw = ws.cell(row=row_idx, column=column_map['warehouse']).value
                    warehouse = str(warehouse_raw or '').strip()

                # Get quantities - only count on-hand if warehouse is 'Store'
                qty_on_hand = 0
                qty_on_order = 0

                if warehouse.lower() == 'store':
                    if 'physical_inventory' in column_map:
                        qty_raw = ws.cell(row=row_idx, column=column_map['physical_inventory']).value
                        try:
                            qty_on_hand = int(float(str(qty_raw or 0)))
                        except (ValueError, TypeError):
                            qty_on_hand = 0

                # On order - leaving for now as requested
                # if 'ordered_total' in column_map:
                #     qty_raw = ws.cell(row=row_idx, column=column_map['ordered_total']).value
                #     try:
                #         qty_on_order = int(float(str(qty_raw or 0)))
                #     except (ValueError, TypeError):
                #         qty_on_order = 0

                # Match to design
                design = mat_to_design.get(mat_no)
                if not design:
                    stats['mats_not_found'] += 1
                    not_found_items.append({
                        'mat_no': mat_no,
                        'hdbs_type': hdbs_type or 'N/A',
                        'erp_item': item_number,
                        'qty_on_hand': qty_on_hand,
                        'qty_on_order': qty_on_order,
                    })
                    continue

                stats['mats_matched'] += 1
                stats['total_on_hand'] += qty_on_hand
                stats['total_on_order'] += qty_on_order

                matched_items.append({
                    'mat_no': mat_no,
                    'hdbs_type': design.hdbs_type,
                    'qty_on_hand': qty_on_hand,
                    'qty_on_order': qty_on_order,
                    'old_on_hand': design.quantity_on_hand,
                    'old_on_order': design.quantity_on_order,
                })

                if confirm:
                    design.quantity_on_hand = qty_on_hand
                    design.quantity_on_order = qty_on_order
                    design.stock_last_updated = timezone.now()
                    design.save(update_fields=['quantity_on_hand', 'quantity_on_order', 'stock_last_updated'])
                    stats['designs_updated'] += 1

                    if verbose:
                        self.stdout.write(f'  {design.mat_no} ({design.hdbs_type}): OH={qty_on_hand}, OO={qty_on_order}')
                else:
                    stats['designs_updated'] += 1

            if not confirm:
                transaction.set_rollback(True)

        # Print summary
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(self.style.SUCCESS('IMPORT SUMMARY'))
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(f"Rows processed:        {stats['rows_processed']}")
        self.stdout.write(f"RM rows found:         {stats['rm_rows']}")
        self.stdout.write(f"Skipped (not RM):      {stats['skipped_not_rm']}")
        self.stdout.write(f"Skipped (no MAT):      {stats['skipped_no_mat']}")
        self.stdout.write(f"MATs matched:          {stats['mats_matched']}")
        self.stdout.write(self.style.WARNING(f"MATs not found:        {stats['mats_not_found']}"))

        if confirm:
            self.stdout.write(f"Designs updated:       {stats['designs_updated']}")

        self.stdout.write(self.style.SUCCESS(f"Total On Hand:         {stats['total_on_hand']}"))
        self.stdout.write(self.style.SUCCESS(f"Total On Order:        {stats['total_on_order']}"))

        # Show matched items sample
        if matched_items and verbose:
            self.stdout.write('')
            self.stdout.write('Matched designs (sample):')
            for item in matched_items[:15]:
                change_oh = f" (was {item['old_on_hand']})" if item['old_on_hand'] != item['qty_on_hand'] else ""
                change_oo = f" (was {item['old_on_order']})" if item['old_on_order'] != item['qty_on_order'] else ""
                self.stdout.write(f"  {item['mat_no']} ({item['hdbs_type']}): OH={item['qty_on_hand']}{change_oh}, OO={item['qty_on_order']}{change_oo}")
            if len(matched_items) > 15:
                self.stdout.write(f"  ... and {len(matched_items) - 15} more")

        # Show not found items
        if not_found_items:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(f'MATs not found in system ({len(not_found_items)} items):'))
            for item in not_found_items[:20]:
                self.stdout.write(f"  MAT {item['mat_no']} ({item['hdbs_type']}): OH={item['qty_on_hand']}, OO={item['qty_on_order']} ({item['erp_item']})")
            if len(not_found_items) > 20:
                self.stdout.write(f"  ... and {len(not_found_items) - 20} more")

        if not confirm:
            self.stdout.write('')
            self.stdout.write(self.style.NOTICE('This was a PREVIEW. Run with --confirm to apply changes.'))
