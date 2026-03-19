"""
ARDT FMS - Technology Views
Version: 5.4 - Sprint 3

Views for Design, BOM, and Cutter Layout management.
"""

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models
from django.http import JsonResponse
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, DeleteView, DetailView, ListView, TemplateView, UpdateView, View

from django.http import JsonResponse

from .forms import BOMForm, BOMLineForm, BitSizeForm, BreakerSlotForm, ConnectionForm, DesignCutterLayoutForm, DesignForm, HDBSTypeForm, SMITypeForm
from .models import BOM, BOMLine, BitSize, BitType, BreakerSlot, Connection, ConnectionSize, ConnectionType, Design, DesignCutterLayout, HDBSType, IADCCode, SMIType


# =============================================================================
# DESIGN VIEWS
# =============================================================================


class DesignListView(LoginRequiredMixin, ListView):
    """List all designs with filtering."""

    model = Design
    template_name = "technology/design_list.html"
    context_object_name = "designs"
    paginate_by = 25

    PAGE_SIZE_OPTIONS = [10, 25, 50, 100, 'all']

    def get_paginate_by(self, queryset):
        """Allow dynamic page size from URL parameter."""
        page_size = self.request.GET.get('page_size', '25')
        if page_size == 'all':
            return None  # No pagination - show all
        try:
            return int(page_size)
        except (ValueError, TypeError):
            return self.paginate_by

    def get_queryset(self):
        queryset = Design.objects.select_related(
            "size", "connection_ref", "connection_ref__connection_type",
            "connection_ref__connection_size", "breaker_slot", "iadc_code_ref"
        ).order_by("-updated_at")

        # General search (MAT No., Ref MAT No., HDBS Type, SMI Type)
        search = self.request.GET.get("q")
        if search:
            queryset = queryset.filter(
                Q(mat_no__icontains=search) |
                Q(ref_mat_no__icontains=search) |
                Q(hdbs_type__icontains=search) |
                Q(smi_type__icontains=search)
            )

        # Type search (HDBS Type or SMI Type)
        type_search = self.request.GET.get("type")
        if type_search:
            queryset = queryset.filter(
                Q(hdbs_type__icontains=type_search) |
                Q(smi_type__icontains=type_search)
            )

        # Size filter
        size_id = self.request.GET.get("size")
        if size_id:
            queryset = queryset.filter(size_id=size_id)

        # Status filter
        status = self.request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)

        # Category filter
        category = self.request.GET.get("category")
        if category:
            queryset = queryset.filter(category=category)

        # Order Level filter
        order_level = self.request.GET.get("order_level")
        if order_level:
            queryset = queryset.filter(order_level=order_level)

        # Sorting
        sort = self.request.GET.get("sort", "-updated_at")
        if sort:
            queryset = queryset.order_by(sort)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Designs"
        context["search_query"] = self.request.GET.get("q", "")
        context["type_search"] = self.request.GET.get("type", "")
        context["current_size"] = self.request.GET.get("size", "")
        context["current_status"] = self.request.GET.get("status", "")
        context["current_category"] = self.request.GET.get("category", "")
        context["current_order_level"] = self.request.GET.get("order_level", "")
        context["current_sort"] = self.request.GET.get("sort", "-updated_at")
        context["category_choices"] = Design.Category.choices
        context["status_choices"] = Design.Status.choices
        context["order_level_choices"] = Design.OrderLevel.choices
        context["sizes"] = BitSize.objects.filter(is_active=True).order_by("size_decimal")
        # Count drafts for the banner (only show if not already filtering by status)
        if not self.request.GET.get("status"):
            context["draft_count"] = Design.objects.filter(status=Design.Status.DRAFT).count()
        else:
            context["draft_count"] = 0
        # Page size options
        context["page_size"] = self.request.GET.get("page_size", "25")
        context["page_size_options"] = self.PAGE_SIZE_OPTIONS
        context["total_count"] = self.get_queryset().count()
        return context


class DesignDetailView(LoginRequiredMixin, DetailView):
    """View design details with BOM and cutter layouts."""

    model = Design
    template_name = "technology/design_detail.html"
    context_object_name = "design"

    def get_queryset(self):
        return Design.objects.select_related(
            "size", "connection_ref", "connection_ref__connection_type",
            "connection_ref__connection_size", "breaker_slot", "iadc_code_ref",
            "formation_type_ref", "application_ref", "created_by"
        ).prefetch_related("special_technologies")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Design {self.object.hdbs_type}"
        context["boms"] = self.object.boms.select_related("created_by").order_by("-created_at")
        context["cutter_layouts"] = self.object.cutter_layouts.order_by("blade_number", "position_number")
        context["work_orders_count"] = self.object.work_orders.count()

        # Pocket statistics
        pocket_configs = self.object.pocket_configs.all()
        pocket_configs_count = pocket_configs.count()
        pocket_configs_total = sum(c.count for c in pocket_configs)
        pockets = self.object.pockets.all()
        pocket_assignments_count = pockets.count()

        context["pocket_configs_count"] = pocket_configs_count
        context["pocket_configs_total"] = pocket_configs_total
        context["pocket_assignments_count"] = pocket_assignments_count

        # 3-Tab completion status
        # Tab 1: Grid - complete when all pockets are assigned
        grid_complete = pocket_configs_count > 0 and pocket_assignments_count == pocket_configs_total

        # Tab 2: Locations - complete when all pockets have blade_location
        locations_with_value = sum(1 for p in pockets if p.blade_location)
        locations_complete = pocket_assignments_count > 0 and locations_with_value == pocket_assignments_count

        # Tab 3: Engagements - complete when all pockets have engagement_order
        engagements_with_value = sum(1 for p in pockets if p.engagement_order)
        engagements_complete = pocket_assignments_count > 0 and engagements_with_value == pocket_assignments_count

        context["grid_complete"] = grid_complete
        context["locations_complete"] = locations_complete
        context["engagements_complete"] = engagements_complete

        # Overall pocket layout complete only when all 3 tabs are done
        context["pocket_layout_complete"] = grid_complete and locations_complete and engagements_complete
        return context


class DesignCreateView(LoginRequiredMixin, CreateView):
    """Create a new design."""

    model = Design
    form_class = DesignForm
    template_name = "technology/design_form.html"
    success_url = reverse_lazy("technology:design_list")

    def get_next_layout_number(self):
        """Generate the next pocket layout number."""
        from django.db.models import Max
        from django.db.models.functions import Cast
        from django.db.models import IntegerField

        # Try to get max numeric layout number
        max_num = Design.objects.filter(
            pocket_layout_number__regex=r'^\d+$'
        ).annotate(
            num=Cast('pocket_layout_number', IntegerField())
        ).aggregate(Max('num'))['num__max']

        return str((max_num or 0) + 1)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Design"
        context["submit_text"] = "Create Design"
        context["next_layout_number"] = self.get_next_layout_number()
        # Pass 'from' parameter to template for back link customization
        context["from_page"] = self.request.GET.get("from", "")
        return context

    def get_initial(self):
        initial = super().get_initial()
        # Pre-fill with next layout number
        initial['pocket_layout_number'] = self.get_next_layout_number()
        return initial

    def get_success_url(self):
        """Return to the page that initiated the design creation."""
        from_page = self.request.POST.get("from_page") or self.request.GET.get("from", "")

        if from_page == "bom_create":
            # Return to BOM create with the new design pre-selected
            return reverse_lazy("technology:bom_create") + f"?new_design={self.object.pk}"
        elif from_page == "cutter_map":
            # Return to cutter map with the new design
            return f"/cutter-map/?design_id={self.object.pk}&design_mat={self.object.mat_no}&design_hdbs={self.object.hdbs_type}&design_size={self.object.size}&from=design_create"

        # Default: go to design detail
        return reverse_lazy("technology:design_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # Auto-generate layout number if not provided
        if not form.instance.pocket_layout_number:
            form.instance.pocket_layout_number = self.get_next_layout_number()
        messages.success(self.request, f"Design {form.instance.hdbs_type} created successfully.")
        return super().form_valid(form)


class DesignUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing design."""

    model = Design
    form_class = DesignForm
    template_name = "technology/design_form.html"

    def get_success_url(self):
        return reverse_lazy("technology:design_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit Design {self.object.hdbs_type}"
        context["submit_text"] = "Update Design"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Design {self.object.hdbs_type} updated successfully.")
        return super().form_valid(form)


class DesignDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a design."""

    model = Design
    template_name = "technology/design_confirm_delete.html"
    success_url = reverse_lazy("technology:design_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Delete Design {self.object.hdbs_type}"
        # Check for related objects that would be deleted
        context["bom_count"] = self.object.boms.count()
        context["pocket_config_count"] = self.object.pocket_configs.count()
        context["pocket_count"] = self.object.pockets.count()
        return context

    def form_valid(self, form):
        design = self.object
        design_name = design.hdbs_type

        # Delete in correct order due to protected foreign keys:
        # 1. DesignPocket references DesignPocketConfig via PROTECT
        # 2. DesignPocketConfig references Design via CASCADE
        # So we must delete pockets first, then pocket_configs, then the design

        # Delete pockets first (they have protected FK to pocket_configs)
        design.pockets.all().delete()

        # Delete pocket configs
        design.pocket_configs.all().delete()

        # Now delete the design (this will cascade delete BOMs, etc.)
        messages.success(self.request, f"Design {design_name} deleted successfully.")
        return super().form_valid(form)


class DesignPocketsView(LoginRequiredMixin, DetailView):
    """Pockets Layout and Features configuration view."""

    model = Design
    template_name = "technology/design_pockets.html"
    context_object_name = "design"

    def get_queryset(self):
        return Design.objects.select_related('size').prefetch_related(
            'pocket_configs__pocket_size',
            'pocket_configs__pocket_shape',
            'pockets__pocket_config'
        )

    def get_context_data(self, **kwargs):
        from .models import PocketSize, PocketShape, DesignPocketConfig

        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Pockets Layout - {self.object.hdbs_type}"

        # Reference data for dropdowns
        context["pocket_sizes"] = PocketSize.objects.filter(is_active=True)
        context["pocket_shapes"] = PocketShape.objects.filter(is_active=True)
        context["length_choices"] = DesignPocketConfig.LengthType.choices

        # Existing configurations for this design
        context["pocket_configs"] = self.object.pocket_configs.select_related(
            'pocket_size', 'pocket_shape'
        ).order_by('order')

        # Existing pockets for this design
        context["pockets"] = self.object.pockets.select_related(
            'pocket_config__pocket_size'
        ).order_by('blade_number', 'row_number', 'position_in_row')

        # Calculate totals for validation
        config_total = sum(c.count for c in context["pocket_configs"])
        entered_total = context["pockets"].count()
        context["config_total"] = config_total
        context["entered_total"] = entered_total

        # Group pockets by blade and row for grid display
        pockets_grid = {}
        for pocket in context["pockets"]:
            blade = pocket.blade_number
            row = pocket.row_number
            if blade not in pockets_grid:
                pockets_grid[blade] = {}
            if row not in pockets_grid[blade]:
                pockets_grid[blade][row] = []
            pockets_grid[blade][row].append(pocket)
        context["pockets_grid"] = pockets_grid

        # Generate color palette for pocket configs
        colors = ['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6',
                  '#EC4899', '#06B6D4', '#84CC16', '#F97316', '#6366F1']
        for i, config in enumerate(context["pocket_configs"]):
            if not config.color_code:
                config.display_color = colors[i % len(colors)]
            else:
                config.display_color = config.color_code

        # Check for unmatched cutters in related BOMs
        from .models import BOM, BOMLine
        unmatched_cutters = []
        boms = BOM.objects.filter(design=self.object)
        for bom in boms:
            unmatched_lines = BOMLine.objects.filter(bom=bom, inventory_item__isnull=True)
            for line in unmatched_lines:
                unmatched_cutters.append({
                    'hdbs_code': line.hdbs_code,
                    'size': line.cutter_size,
                    'type': line.cutter_type,
                    'bom_code': bom.code
                })
        context["unmatched_cutters"] = unmatched_cutters
        context["has_unmatched_cutters"] = len(unmatched_cutters) > 0

        return context


class PocketConfigCreateView(LoginRequiredMixin, View):
    """Add a pocket configuration to a design."""

    def post(self, request, pk):
        from .models import Design, DesignPocketConfig, PocketSize, PocketShape

        design = get_object_or_404(Design, pk=pk)

        pocket_size_id = request.POST.get('pocket_size')
        length_type = request.POST.get('length_type')
        pocket_shape_id = request.POST.get('pocket_shape')
        count = request.POST.get('count')

        if pocket_size_id and length_type and pocket_shape_id and count:
            # Get the next order number
            last_config = design.pocket_configs.order_by('-order').first()
            next_order = (last_config.order + 1) if last_config else 1

            DesignPocketConfig.objects.create(
                design=design,
                order=next_order,
                pocket_size_id=pocket_size_id,
                length_type=length_type,
                pocket_shape_id=pocket_shape_id,
                count=int(count)
            )
            messages.success(request, "Pocket configuration added successfully.")
        else:
            messages.error(request, "Please fill in all required fields.")

        return redirect('technology:design_pockets', pk=pk)


class PocketConfigDeleteView(LoginRequiredMixin, View):
    """Delete a pocket configuration from a design."""

    def post(self, request, pk, config_pk):
        from .models import DesignPocketConfig

        config = get_object_or_404(DesignPocketConfig, pk=config_pk, design_id=pk)
        config.delete()
        messages.success(request, "Pocket configuration deleted.")
        return redirect('technology:design_pockets', pk=pk)


class DesignPocketsUpdateInfoView(LoginRequiredMixin, View):
    """Update pocket-related fields (rows count, pocket layout number) from pockets page."""

    def post(self, request, pk):
        design = get_object_or_404(Design, pk=pk)

        # Update pocket_rows_count
        rows_count = request.POST.get('pocket_rows_count')
        if rows_count:
            try:
                design.pocket_rows_count = int(rows_count)
            except ValueError:
                pass

        # Update pocket_layout_number
        layout_number = request.POST.get('pocket_layout_number')
        if layout_number is not None:
            design.pocket_layout_number = layout_number if layout_number else None

        design.save(update_fields=['pocket_rows_count', 'pocket_layout_number'])

        # Always return JSON - this is an API endpoint
        return JsonResponse({
            'success': True,
            'pocket_rows_count': design.pocket_rows_count,
            'pocket_layout_number': design.pocket_layout_number or ''
        })


class PocketConfigReorderView(LoginRequiredMixin, View):
    """Reorder pocket configurations (move up/down)."""

    def post(self, request, pk):
        from .models import DesignPocketConfig

        config_id = request.POST.get('config_id')
        direction = request.POST.get('direction')  # 'up' or 'down'

        if not config_id or direction not in ('up', 'down'):
            messages.error(request, "Invalid request.")
            return redirect('technology:design_pockets', pk=pk)

        config = get_object_or_404(DesignPocketConfig, pk=config_id, design_id=pk)
        configs = list(DesignPocketConfig.objects.filter(design_id=pk).order_by('order'))

        current_index = None
        for i, c in enumerate(configs):
            if c.pk == config.pk:
                current_index = i
                break

        if current_index is None:
            return redirect('technology:design_pockets', pk=pk)

        if direction == 'up' and current_index > 0:
            # Swap with previous
            configs[current_index], configs[current_index - 1] = configs[current_index - 1], configs[current_index]
        elif direction == 'down' and current_index < len(configs) - 1:
            # Swap with next
            configs[current_index], configs[current_index + 1] = configs[current_index + 1], configs[current_index]

        # Update order values
        for i, c in enumerate(configs, start=1):
            c.order = i
            c.save(update_fields=['order'])

        return redirect('technology:design_pockets', pk=pk)


class PocketConfigUpdateRowView(LoginRequiredMixin, View):
    """Update the row number of a pocket configuration."""

    def post(self, request, pk):
        from .models import DesignPocketConfig

        config_id = request.POST.get('config_id')
        row_number = request.POST.get('row_number')

        if not config_id or not row_number:
            messages.error(request, "Invalid request.")
            return redirect('technology:design_pockets', pk=pk)

        config = get_object_or_404(DesignPocketConfig, pk=config_id, design_id=pk)

        try:
            config.row_number = int(row_number)
            config.save(update_fields=['row_number'])
        except ValueError:
            messages.error(request, "Invalid row number.")

        return redirect('technology:design_pockets', pk=pk)


class DesignPocketsGridSaveView(LoginRequiredMixin, View):
    """Save grid cell assignments to database."""

    def post(self, request, pk):
        import json
        from .models import DesignPocket, DesignPocketConfig

        design = get_object_or_404(Design, pk=pk)

        try:
            data = json.loads(request.body)
            grid_data = data.get('gridData', {})
            row_separators = data.get('rowSeparators', [])

            # Delete existing pocket assignments
            DesignPocket.objects.filter(design=design).delete()

            # Save new assignments
            for key, config_id in grid_data.items():
                blade, col = key.split('_')
                blade = int(blade)
                col = int(col)

                # Calculate row and position in row based on separators
                row = 1
                row_start = 1
                for sep in sorted(row_separators):
                    if col > sep:
                        row += 1
                        row_start = sep + 1

                position_in_row = col - row_start + 1

                config = DesignPocketConfig.objects.filter(pk=config_id, design=design).first()
                if config:
                    DesignPocket.objects.create(
                        design=design,
                        blade_number=blade,
                        row_number=row,
                        position_in_row=position_in_row,
                        position_in_blade=col,
                        pocket_config=config
                    )

            return JsonResponse({'success': True, 'message': 'Grid saved'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    def get(self, request, pk):
        """Load grid data from database."""
        from .models import DesignPocket
        from collections import defaultdict

        design = get_object_or_404(Design, pk=pk)
        pockets = list(DesignPocket.objects.filter(design=design).select_related('pocket_config'))

        if not pockets:
            return JsonResponse({
                'success': True,
                'gridData': {},
                'rowSeparators': [],
                'locationData': {},
                'engagementData': {}
            })

        # --- Row-aligned virtual column calculation ---
        # Different blades have different cutter counts per row.
        # We need to pad shorter blades so all R2 cutters appear after
        # the R1 separator, all R3 after R2 separator, etc.

        # Step 1: Find max cutters per row across ALL blades
        blade_row_counts = defaultdict(int)
        for p in pockets:
            blade_row_counts[(p.blade_number, p.row_number)] += 1

        all_blades = sorted(set(p.blade_number for p in pockets))
        all_rows = sorted(set(p.row_number for p in pockets))

        row_max = {}
        for row in all_rows:
            row_max[row] = max(
                blade_row_counts.get((blade, row), 0)
                for blade in all_blades
            )

        # Step 2: Calculate row start offsets (virtual column where each row begins)
        row_start = {}
        offset = 1
        for row in all_rows:
            row_start[row] = offset
            offset += row_max[row]

        # Step 3: Row separators — cumulative end of each row except last
        row_separators = []
        cumulative = 0
        for row in all_rows[:-1]:
            cumulative += row_max[row]
            row_separators.append(cumulative)

        # Step 4: Build grid data using virtual columns
        # Virtual column = row_start[row_number] + position_in_row - 1
        grid_data = {}
        location_data = {}
        engagement_data = {}
        for pocket in pockets:
            vcol = row_start[pocket.row_number] + pocket.position_in_row - 1
            key = f"{pocket.blade_number}_{vcol}"
            grid_data[key] = pocket.pocket_config_id
            if pocket.blade_location:
                location_data[key] = pocket.blade_location
            if pocket.engagement_order:
                engagement_data[key] = pocket.engagement_order

        return JsonResponse({
            'success': True,
            'gridData': grid_data,
            'rowSeparators': sorted(row_separators),
            'locationData': location_data,
            'engagementData': engagement_data
        })


class DesignPocketsLocationSaveView(LoginRequiredMixin, View):
    """Save pocket location assignments to database."""

    def post(self, request, pk):
        import json
        from .models import DesignPocket

        design = get_object_or_404(Design, pk=pk)

        try:
            data = json.loads(request.body)
            location_data = data.get('locationData', {})

            # Validate sequential order within each blade AND row
            # Location order: C < N < T < S < G
            # Each row can have its own independent sequence
            location_order = {'C': 1, 'N': 2, 'T': 3, 'S': 4, 'G': 5}

            # Get existing pockets to know which row each column belongs to
            pockets = DesignPocket.objects.filter(design=design)
            col_to_row = {}
            for pocket in pockets:
                col_to_row[f"{pocket.blade_number}_{pocket.position_in_blade}"] = pocket.row_number

            # Group by blade and row
            for blade_num in range(1, design.no_of_blades + 1):
                # Group locations by row
                rows_locations = {}
                for key, loc in location_data.items():
                    b, col = key.split('_')
                    if int(b) == blade_num and loc:
                        row_num = col_to_row.get(key, 1)
                        if row_num not in rows_locations:
                            rows_locations[row_num] = []
                        rows_locations[row_num].append((int(col), loc))

                # Validate each row independently
                for row_num, row_locations in rows_locations.items():
                    # Sort by column position
                    row_locations.sort(key=lambda x: x[0])

                    # Check sequence within this row
                    last_order = 0
                    for col, loc in row_locations:
                        current_order = location_order.get(loc, 0)
                        if current_order < last_order:
                            return JsonResponse({
                                'success': False,
                                'message': f'Invalid sequence on Blade {blade_num}, Row {row_num}: {loc} cannot come after previous location'
                            }, status=400)
                        last_order = current_order

            # Update pocket locations
            for key, location in location_data.items():
                blade, col = key.split('_')
                blade = int(blade)
                col = int(col)

                DesignPocket.objects.filter(
                    design=design,
                    blade_number=blade,
                    position_in_blade=col
                ).update(blade_location=location if location else None)

            return JsonResponse({'success': True, 'message': 'Locations saved'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    def get(self, request, pk):
        """Load location data from database."""
        from .models import DesignPocket

        design = get_object_or_404(Design, pk=pk)
        pockets = DesignPocket.objects.filter(design=design)

        location_data = {}
        for pocket in pockets:
            key = f"{pocket.blade_number}_{pocket.position_in_blade}"
            if pocket.blade_location:
                location_data[key] = pocket.blade_location

        return JsonResponse({
            'success': True,
            'locationData': location_data
        })


class DesignPocketsEngagementSaveView(LoginRequiredMixin, View):
    """Save engagement order assignments to database."""

    def post(self, request, pk):
        import json
        from .models import DesignPocket

        design = get_object_or_404(Design, pk=pk)

        try:
            data = json.loads(request.body)
            engagement_data = data.get('engagementData', {})

            # Validate uniqueness
            values = [v for v in engagement_data.values() if v]
            if len(values) != len(set(values)):
                return JsonResponse({
                    'success': False,
                    'message': 'Duplicate engagement numbers found'
                }, status=400)

            # Update engagement orders
            for key, order in engagement_data.items():
                blade, col = key.split('_')
                blade = int(blade)
                col = int(col)

                DesignPocket.objects.filter(
                    design=design,
                    blade_number=blade,
                    position_in_blade=col
                ).update(engagement_order=order if order else None)

            return JsonResponse({'success': True, 'message': 'Engagements saved'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    def get(self, request, pk):
        """Load engagement data from database."""
        from .models import DesignPocket

        design = get_object_or_404(Design, pk=pk)
        pockets = DesignPocket.objects.filter(design=design)

        engagement_data = {}
        for pocket in pockets:
            key = f"{pocket.blade_number}_{pocket.position_in_blade}"
            if pocket.engagement_order:
                engagement_data[key] = pocket.engagement_order

        return JsonResponse({
            'success': True,
            'engagementData': engagement_data
        })


class DesignPocketsResetView(LoginRequiredMixin, View):
    """Reset all pocket data for a design (pockets and pocket configs)."""

    def post(self, request, pk):
        from .models import DesignPocket, DesignPocketConfig

        design = get_object_or_404(Design, pk=pk)

        try:
            # Delete in correct order (pockets reference pocket_configs via protected FK)
            pockets_deleted = DesignPocket.objects.filter(design=design).count()
            DesignPocket.objects.filter(design=design).delete()

            configs_deleted = DesignPocketConfig.objects.filter(design=design).count()
            DesignPocketConfig.objects.filter(design=design).delete()

            return JsonResponse({
                'success': True,
                'message': f'Reset complete. Deleted {pockets_deleted} pockets and {configs_deleted} configurations.',
                'pockets_deleted': pockets_deleted,
                'configs_deleted': configs_deleted
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)


class PocketsLayoutListView(LoginRequiredMixin, ListView):
    """List all designs with pockets layout summary."""

    model = Design
    template_name = "technology/pockets_layout_list.html"
    context_object_name = "designs"
    paginate_by = 25

    def get_queryset(self):
        from django.db.models import Count

        queryset = Design.objects.select_related(
            "size", "created_by"
        ).annotate(
            config_count=Count('pocket_configs')
        ).order_by("-updated_at")

        search = self.request.GET.get("q")
        if search:
            queryset = queryset.filter(
                Q(mat_no__icontains=search) |
                Q(hdbs_type__icontains=search) |
                Q(smi_type__icontains=search)
            )

        size_id = self.request.GET.get("size")
        if size_id:
            queryset = queryset.filter(size_id=size_id)

        status = self.request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Pockets Layout"
        context["search_query"] = self.request.GET.get("q", "")
        context["current_size"] = self.request.GET.get("size", "")
        context["current_status"] = self.request.GET.get("status", "")
        context["status_choices"] = Design.Status.choices
        context["sizes"] = BitSize.objects.filter(is_active=True).order_by("size_decimal")
        return context


# =============================================================================
# BOM VIEWS
# =============================================================================


class BOMListView(LoginRequiredMixin, ListView):
    """List all BOMs with material availability status."""

    model = BOM
    template_name = "technology/bom_list.html"
    context_object_name = "boms"
    paginate_by = 50  # Higher limit since DataTables handles pagination

    def get_queryset(self):
        from .filters import BOMFilter
        queryset = BOM.objects.select_related(
            "design", "design__size", "design__iadc_code_ref", "created_by",
            "smi_type", "smi_type__hdbs_type", "parent_bom"
        ).prefetch_related(
            "lines__inventory_item__variants__variant_case",
            "lines__inventory_item__variants__stock_records",
            "drillbits_brazing",  # Drill bits linked via brazing_bom
            "drillbits_system",   # Drill bits linked via system_bom
            "drill_bits",         # Drill bits linked via legacy bom FK
        ).order_by("-created_at")

        # Apply django-filter
        self.filterset = BOMFilter(self.request.GET, queryset=queryset)
        return self.filterset.qs

    def get_context_data(self, **kwargs):
        from apps.inventory.models import VariantStock
        from django.db.models import Sum, DecimalField, Value
        from django.db.models.functions import Coalesce

        context = super().get_context_data(**kwargs)
        context["page_title"] = "Bills of Materials"
        context["filter"] = self.filterset
        context["sizes"] = BitSize.objects.filter(is_active=True).order_by("size_decimal")

        NEW_VARIANT_CODES = {"NEW-PUR", "NEW-EO", "NEW-RET"}

        # Add material availability status for each BOM
        boms_with_status = []
        for bom in context["boms"]:
            all_available = True
            line_count = 0
            for line in bom.lines.all():
                line_count += 1
                if line.inventory_item:
                    # Sum new variant stock (same logic as cutter inventory)
                    total_new = 0
                    for variant in line.inventory_item.variants.all():
                        if variant.variant_case and variant.variant_case.code in NEW_VARIANT_CODES:
                            stock = variant.stock_records.aggregate(
                                total=Coalesce(Sum("quantity_on_hand"), Value(0), output_field=DecimalField())
                            )["total"]
                            total_new += int(stock)
                    if total_new < line.quantity:
                        all_available = False
                else:
                    all_available = False

            # Collect linked drill bits (brazing, system, and legacy bom FK)
            linked_drillbits = {}  # {pk: serial_number} to dedup
            for db in bom.drillbits_brazing.all():
                linked_drillbits[db.pk] = db.serial_number
            for db in bom.drillbits_system.all():
                linked_drillbits[db.pk] = db.serial_number
            for db in bom.drill_bits.all():
                linked_drillbits[db.pk] = db.serial_number

            # Build list of dicts with pk and serial for template links
            linked_sns_list = [{'pk': pk, 'sn': sn} for pk, sn in linked_drillbits.items()]
            linked_sns_list.sort(key=lambda x: x['sn'])

            boms_with_status.append({
                'bom': bom,
                'materials_ready': all_available if line_count > 0 else None,
                'line_count': line_count,
                'linked_sns': [d['sn'] for d in linked_sns_list],  # Backwards compat for filters
                'linked_drillbits': linked_sns_list,  # With PKs for links
            })

        context["boms_with_status"] = boms_with_status
        context["smi_types"] = SMIType.objects.filter(is_active=True).select_related('hdbs_type', 'size').order_by('smi_name')
        context["iadc_codes"] = IADCCode.objects.filter(is_active=True).order_by('code')
        context["status_choices"] = BOM.Status.choices
        return context


class BOMDetailView(LoginRequiredMixin, DetailView):
    """View BOM details with line items and inventory availability."""

    model = BOM
    template_name = "technology/bom_detail.html"
    context_object_name = "bom"

    def get_queryset(self):
        return BOM.objects.select_related("design", "created_by")

    def get_context_data(self, **kwargs):
        from apps.inventory.models import StockBalance
        from django.db.models import Sum

        context = super().get_context_data(**kwargs)
        context["page_title"] = f"BOM {self.object.code}"

        # Get lines with inventory items
        lines = self.object.lines.select_related(
            "inventory_item", "inventory_item__category", "inventory_item__primary_supplier"
        ).order_by("line_number")

        # Enrich lines with stock information
        lines_with_stock = []
        all_available = True

        for line in lines:
            item = line.inventory_item
            required = line.quantity

            # Get total stock for this item (only if inventory_item exists)
            if item:
                stock_data = StockBalance.objects.filter(
                    item=item
                ).aggregate(
                    total_on_hand=Sum('qty_on_hand'),
                    total_available=Sum('qty_available')
                )
                on_hand = stock_data['total_on_hand'] or 0
                available = stock_data['total_available'] or 0
            else:
                # No linked inventory item - show as unavailable
                on_hand = 0
                available = 0

            # Determine availability status
            if available >= required:
                status = 'available'
            elif available > 0:
                status = 'partial'
                all_available = False
            else:
                status = 'unavailable'
                all_available = False

            lines_with_stock.append({
                'line': line,
                'on_hand': on_hand,
                'available': available,
                'required': required,
                'shortage': max(0, required - available),
                'status': status,
            })

        context["lines"] = lines_with_stock
        context["all_materials_available"] = all_available
        context["total_cost"] = self.object.total_cost
        context["line_form"] = BOMLineForm()

        # Get Cutter Layout (CL) data from Design's pockets
        design = self.object.design
        cutter_layout = None
        if design:
            from .models import DesignPocket
            pockets = DesignPocket.objects.filter(design=design).select_related(
                'pocket_config', 'pocket_config__pocket_size'
            ).order_by('blade_number', 'position_in_blade')

            if pockets.exists():
                # Build color map from BOM lines
                line_colors = {}
                for line in lines:
                    if line.color_code:
                        line_colors[line.line_number] = line.color_code

                # Organize pockets by blade
                blades_data = {}
                for pocket in pockets:
                    blade_num = pocket.blade_number
                    if blade_num not in blades_data:
                        blades_data[blade_num] = []

                    # Get color from pocket_config order → BOM line
                    color = '#4A4A4A'  # default gray
                    if pocket.pocket_config:
                        config_order = pocket.pocket_config.order
                        color = pocket.pocket_config.color_code or line_colors.get(config_order, '#4A4A4A')

                    blades_data[blade_num].append({
                        'position': pocket.position_in_blade,
                        'row': pocket.row_number,
                        'location': pocket.get_blade_location_display() if pocket.blade_location else '',
                        'location_code': pocket.blade_location or '',
                        'config': pocket.pocket_config,
                        'color': color,
                        'size': pocket.pocket_config.pocket_size.code if pocket.pocket_config and pocket.pocket_config.pocket_size else '',
                    })

                cutter_layout = {
                    'blade_count': design.no_of_blades or len(blades_data),
                    'blades': blades_data,
                    'total_pockets': pockets.count(),
                }

        context["cutter_layout"] = cutter_layout
        return context


class BOMCreateView(LoginRequiredMixin, CreateView):
    """Create a new BOM."""

    model = BOM
    form_class = BOMForm
    template_name = "technology/bom_form.html"
    success_url = reverse_lazy("technology:bom_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create BOM"
        context["submit_text"] = "Create BOM"
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f"BOM {form.instance.code} created successfully.")
        return super().form_valid(form)


class BOMUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing BOM."""

    model = BOM
    form_class = BOMForm
    template_name = "technology/bom_form.html"

    def get_success_url(self):
        return reverse_lazy("technology:bom_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit BOM {self.object.code}"
        context["submit_text"] = "Update BOM"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"BOM {self.object.code} updated successfully.")
        return super().form_valid(form)


class BOMDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a BOM."""

    model = BOM
    success_url = reverse_lazy("technology:bom_list")

    def post(self, request, *args, **kwargs):
        bom = self.get_object()
        bom_code = bom.code
        bom.delete()
        messages.success(request, f"BOM {bom_code} deleted successfully.")
        return redirect(self.success_url)


class BOMCloneView(LoginRequiredMixin, View):
    """Clone an existing BOM with all details, ready for editing."""

    def _generate_clone_code(self, source_code):
        """Generate a unique clone code by appending -C1, -C2, etc."""
        base = source_code.rstrip('0123456789').rstrip('-C')
        if base == source_code:
            base = source_code
        i = 1
        while True:
            candidate = f"{base}-C{i}"
            if not BOM.objects.filter(code=candidate).exists():
                return candidate
            i += 1

    def _do_clone(self, request, pk, new_code=None):
        import copy
        source_bom = get_object_or_404(BOM, pk=pk)

        if not new_code:
            new_code = self._generate_clone_code(source_bom.code)
        elif BOM.objects.filter(code=new_code).exists():
            return None, f"BOM with code {new_code} already exists"

        # Deep-copy source_data so mutations don't affect the original
        source_data = copy.deepcopy(source_bom.source_data) if source_bom.source_data else {}

        new_bom = BOM.objects.create(
            design=source_bom.design,
            code=new_code,
            name=f"{source_bom.name} (Clone)" if source_bom.name else "",
            revision="A",
            status=BOM.Status.DRAFT,
            bom_type=source_bom.bom_type,
            system_mat_no=source_bom.system_mat_no,
            smi_type=source_bom.smi_type,
            source_type=source_bom.source_type,
            source_mat_number=source_bom.source_mat_number,
            source_sn_number=source_bom.source_sn_number,
            source_revision_level=source_bom.source_revision_level,
            source_software_version=source_bom.source_software_version,
            source_data=source_data,
            notes=f"Cloned from {source_bom.code}",
            created_by=request.user,
        )

        # Clone all BOM lines
        for line in source_bom.lines.all():
            BOMLine.objects.create(
                bom=new_bom,
                line_number=line.line_number,
                order_number=line.order_number,
                inventory_item=line.inventory_item,
                quantity=line.quantity,
                unit=line.unit,
                color_code=line.color_code,
                cutter_size=line.cutter_size,
                cutter_chamfer=line.cutter_chamfer,
                cutter_type=line.cutter_type,
                hdbs_code=line.hdbs_code,
                family_number=line.family_number,
                unit_cost=line.unit_cost,
                position=line.position,
                is_optional=line.is_optional,
                is_phantom=line.is_phantom,
                notes=line.notes,
            )

        return new_bom, None

    def get(self, request, pk):
        """GET: auto-clone with generated code and redirect to cutter map for editing."""
        new_bom, error = self._do_clone(request, pk)
        if error:
            messages.error(request, error)
            return redirect(reverse("technology:bom_detail", args=[pk]))

        messages.success(request, f'BOM cloned as "{new_bom.code}" (Draft). You can now edit it.')

        # If BOM has source_data, open in cutter map for editing; otherwise go to BOM detail
        if new_bom.source_data:
            return redirect(f"{reverse('cutter_map:bom_view', args=[new_bom.pk])}?edit=1")
        else:
            return redirect(reverse("technology:bom_detail", args=[new_bom.pk]))

    def post(self, request, pk):
        """POST: clone with a specific new_code (API call)."""
        import json
        try:
            data = json.loads(request.body)
            new_code = data.get("new_code", "").strip()
        except json.JSONDecodeError:
            new_code = request.POST.get("new_code", "").strip()

        if not new_code:
            return JsonResponse({"success": False, "error": "New BOM code is required"}, status=400)

        new_bom, error = self._do_clone(request, pk, new_code)
        if error:
            return JsonResponse({"success": False, "error": error}, status=400)

        return JsonResponse({
            "success": True,
            "redirect_url": reverse("cutter_map:bom_view", args=[new_bom.pk])
        })


class BOMLineCreateView(LoginRequiredMixin, View):
    """Add a line to a BOM."""

    def post(self, request, pk):
        bom = get_object_or_404(BOM, pk=pk)
        form = BOMLineForm(request.POST)
        if form.is_valid():
            line = form.save(commit=False)
            line.bom = bom
            line.save()
            messages.success(request, "Line added successfully.")
        else:
            messages.error(request, "Failed to add line.")
        return redirect("technology:bom_detail", pk=pk)


class BOMLineDeleteView(LoginRequiredMixin, View):
    """Delete a line from a BOM."""

    def post(self, request, pk, line_pk):
        line = get_object_or_404(BOMLine, pk=line_pk, bom_id=pk)
        line.delete()
        messages.success(request, "Line deleted.")
        return redirect("technology:bom_detail", pk=pk)


# =============================================================================
# BOM BUILDER VIEWS - REMOVED (Jan 21, 2026)
# Manual BOM Builder replaced by Cutter Map PDF Generator at /cutter-map/bom/<id>/
# =============================================================================


# =============================================================================
# BOM PDF IMPORT VIEWS
# =============================================================================


class BOMPDFImportView(LoginRequiredMixin, View):
    """Upload and parse a BOM PDF for import."""

    def get(self, request, pk):
        bom = get_object_or_404(BOM, pk=pk)
        return self._render(request, bom)

    def post(self, request, pk):
        bom = get_object_or_404(BOM, pk=pk)

        if 'pdf_file' not in request.FILES:
            messages.error(request, "Please select a PDF file to upload.")
            return self._render(request, bom)

        pdf_file = request.FILES['pdf_file']

        # Validate file type
        if not pdf_file.name.lower().endswith('.pdf'):
            messages.error(request, "Please upload a PDF file.")
            return self._render(request, bom)

        # Parse PDF
        from apps.technology.services.pdf_parser import parse_bom_pdf

        try:
            pdf_bytes = pdf_file.read()
            parsed_data = parse_bom_pdf(file_bytes=pdf_bytes)

            if parsed_data.errors:
                for err in parsed_data.errors:
                    messages.warning(request, err)

            # Store parsed data in session for confirmation
            request.session['parsed_bom_data'] = {
                'header': {
                    'sn_number': parsed_data.header.sn_number,
                    'mat_number': parsed_data.header.mat_number,
                    'date_created': parsed_data.header.date_created.isoformat() if parsed_data.header.date_created else None,
                    'revision_level': parsed_data.header.revision_level,
                    'software_version': parsed_data.header.software_version,
                },
                'bom_lines': [
                    {
                        'order_number': line.order_number,
                        'size': line.size,
                        'chamfer': line.chamfer,
                        'cutter_type': line.cutter_type,
                        'count': line.count,
                        'mat_number': line.mat_number,
                        'family_number': line.family_number,
                        'color_code': line.color_code,
                    }
                    for line in parsed_data.bom_lines
                ],
                'cutter_positions': [
                    {
                        'blade_number': pos.blade_number,
                        'row_number': pos.row_number,
                        'location': pos.location,
                        'position_in_location': pos.position_in_location,
                        'cutter_type': pos.cutter_type,
                        'order_number': pos.order_number,
                        'chamfer': pos.chamfer,
                    }
                    for pos in parsed_data.cutter_positions
                ],
                'raw_text': parsed_data.raw_text[:5000] if parsed_data.raw_text else '',  # First 5000 chars for debugging
            }

            # Save the PDF file to BOM - reset file position first
            pdf_file.seek(0)
            bom.source_pdf_file = pdf_file
            bom.save(update_fields=['source_pdf_file'])

            return redirect('technology:bom_pdf_import_confirm', pk=pk)

        except Exception as e:
            messages.error(request, f"Error parsing PDF: {str(e)}")
            return self._render(request, bom)

    def _render(self, request, bom):
        from django.shortcuts import render
        return render(request, 'technology/bom_pdf_import.html', {
            'bom': bom,
            'page_title': f"Import PDF - {bom.code}",
        })


class BOMPDFImportConfirmView(LoginRequiredMixin, View):
    """Confirm and apply parsed PDF data to BOM."""

    def get(self, request, pk):
        import json

        bom = get_object_or_404(BOM, pk=pk)

        parsed_data = request.session.get('parsed_bom_data')

        # If no parsed data in session but BOM has a PDF, auto-parse it
        if not parsed_data and bom.source_pdf_file:
            from apps.technology.services.pdf_parser import parse_bom_pdf
            try:
                result = parse_bom_pdf(file_path=bom.source_pdf_file.path)
                parsed_data = {
                    'header': {
                        'sn_number': result.header.sn_number,
                        'mat_number': result.header.mat_number,
                        'date_created': result.header.date_created.isoformat() if result.header.date_created else None,
                        'revision_level': result.header.revision_level,
                        'software_version': result.header.software_version,
                    },
                    'bom_lines': [
                        {
                            'order_number': line.order_number,
                            'size': line.size,
                            'chamfer': line.chamfer,
                            'cutter_type': line.cutter_type,
                            'count': line.count,
                            'mat_number': line.mat_number,
                            'family_number': line.family_number,
                            'color_code': line.color_code,
                        }
                        for line in result.bom_lines
                    ],
                    'cutter_positions': [
                        {
                            'blade_number': pos.blade_number,
                            'row_number': pos.row_number,
                            'location': pos.location,
                            'position_in_location': pos.position_in_location,
                            'cutter_type': pos.cutter_type,
                            'order_number': pos.order_number,
                            'chamfer': pos.chamfer,
                        }
                        for pos in result.cutter_positions
                    ],
                    'raw_text': result.raw_text[:5000] if result.raw_text else '',
                }
                # Store in session for subsequent requests
                request.session['parsed_bom_data'] = parsed_data
                if result.errors:
                    for err in result.errors:
                        messages.warning(request, err)
            except Exception as e:
                messages.error(request, f"Error parsing PDF: {str(e)}")
                return redirect('technology:bom_pdf_import', pk=pk)

        if not parsed_data:
            messages.warning(request, "No parsed data found. Please upload a PDF first.")
            return redirect('technology:bom_pdf_import', pk=pk)

        # Serialize data for JavaScript
        bom_lines_json = json.dumps(parsed_data.get('bom_lines', []))
        cutter_positions_json = json.dumps(parsed_data.get('cutter_positions', []))

        from django.shortcuts import render
        return render(request, 'technology/bom_pdf_import_confirm.html', {
            'bom': bom,
            'parsed_data': parsed_data,
            'bom_lines_json': bom_lines_json,
            'cutter_positions_json': cutter_positions_json,
            'page_title': f"Confirm Import - {bom.code}",
        })

    def post(self, request, pk):
        import json
        from datetime import datetime as dt

        bom = get_object_or_404(BOM, pk=pk)

        parsed_data = request.session.get('parsed_bom_data')
        if not parsed_data:
            messages.error(request, "No parsed data found.")
            return redirect('technology:bom_pdf_import', pk=pk)

        try:
            header = parsed_data['header']

            # Check if we have modified data from the interactive form
            bom_lines_json = request.POST.get('bom_lines_json', '')
            cutter_positions_json = request.POST.get('cutter_positions_json', '')

            if bom_lines_json:
                try:
                    bom_lines_data = json.loads(bom_lines_json)
                except json.JSONDecodeError:
                    bom_lines_data = parsed_data['bom_lines']
            else:
                bom_lines_data = parsed_data['bom_lines']

            if cutter_positions_json:
                try:
                    cutter_positions_data = json.loads(cutter_positions_json)
                except json.JSONDecodeError:
                    cutter_positions_data = parsed_data.get('cutter_positions', [])
            else:
                cutter_positions_data = parsed_data.get('cutter_positions', [])

            # Update BOM with header info
            bom.source_type = BOM.SourceType.PDF_IMPORT
            bom.source_mat_number = header.get('mat_number', '')
            bom.source_sn_number = header.get('sn_number', '')
            bom.source_revision_level = header.get('revision_level', '')
            bom.source_software_version = header.get('software_version', '')

            if header.get('date_created'):
                try:
                    bom.source_date_created = dt.fromisoformat(header['date_created'])
                except (ValueError, TypeError):
                    pass

            bom.save()

            # Clear existing lines if requested
            if request.POST.get('clear_existing') == 'yes':
                bom.lines.all().delete()
                # Also clear cutter positions
                BOMCutterPosition.objects.filter(bom=bom).delete()

            # Get max line number
            max_line = bom.lines.aggregate(max_line=models.Max("line_number"))
            next_line_num = (max_line["max_line"] or 0) + 1

            # Create BOM lines and map order numbers to BOMLine objects
            lines_created = 0
            order_to_line = {}  # Map order_number to BOMLine for grid positions

            for line_data in bom_lines_data:
                bom_line = BOMLine.objects.create(
                    bom=bom,
                    line_number=next_line_num,
                    order_number=line_data['order_number'],
                    quantity=line_data.get('count', 0),
                    color_code=line_data.get('color_code', '#4A4A4A'),
                    cutter_size=line_data.get('size', ''),
                    cutter_chamfer=line_data.get('chamfer', ''),
                    cutter_type=line_data.get('cutter_type', ''),
                    hdbs_code=line_data.get('mat_number', ''),
                    family_number=line_data.get('family_number', ''),
                )
                order_to_line[line_data['order_number']] = bom_line
                next_line_num += 1
                lines_created += 1

            # Import cutter grid positions if requested
            positions_created = 0
            if request.POST.get('import_grid') == 'yes' and cutter_positions_data:
                for pos_data in cutter_positions_data:
                    order_num = pos_data.get('order_number', 1)
                    bom_line = order_to_line.get(order_num)

                    BOMCutterPosition.objects.create(
                        bom=bom,
                        bom_line=bom_line,
                        blade_number=pos_data.get('blade_number', 1),
                        row_number=pos_data.get('row_number', 1),
                        blade_location=pos_data.get('location', 'CONE'),
                        position_in_location=pos_data.get('position_in_location', 1),
                        cutter_type_display=pos_data.get('cutter_type', ''),
                        order_number_display=order_num,
                        chamfer_display=pos_data.get('chamfer', ''),
                    )
                    positions_created += 1

            # Clear session data
            del request.session['parsed_bom_data']

            if positions_created > 0:
                messages.success(request, f"Successfully imported {lines_created} lines and {positions_created} grid positions from PDF.")
            else:
                messages.success(request, f"Successfully imported {lines_created} lines from PDF.")

            return redirect('technology:bom_detail', pk=pk)

        except Exception as e:
            messages.error(request, f"Error applying import: {str(e)}")
            return redirect('technology:bom_pdf_import_confirm', pk=pk)


class BOMPDFServeView(LoginRequiredMixin, View):
    """Serve BOM PDF file without X-Frame-Options to allow embedding."""

    def get(self, request, pk):
        from django.http import FileResponse
        import os

        bom = get_object_or_404(BOM, pk=pk)

        if not bom.source_pdf_file:
            from django.http import HttpResponseNotFound
            return HttpResponseNotFound("No PDF file available")

        file_path = bom.source_pdf_file.path
        if not os.path.exists(file_path):
            from django.http import HttpResponseNotFound
            return HttpResponseNotFound("PDF file not found")

        response = FileResponse(
            open(file_path, 'rb'),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
        # Allow embedding in same origin
        response['X-Frame-Options'] = 'SAMEORIGIN'
        return response


class BOMPDFExportView(LoginRequiredMixin, View):
    """Export BOM to PDF."""

    def get(self, request, pk):
        from django.http import HttpResponse

        bom = get_object_or_404(BOM, pk=pk)

        try:
            from apps.technology.services.pdf_generator import generate_bom_pdf

            pdf_bytes = generate_bom_pdf(bom)

            # Create response
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            filename = f"BOM_{bom.code}_{bom.revision}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            messages.error(request, f"Error generating PDF: {str(e)}")
            return redirect('technology:bom_detail', pk=pk)


# =============================================================================
# CUTTER LAYOUT VIEWS
# =============================================================================


class CutterLayoutCreateView(LoginRequiredMixin, View):
    """Add a cutter layout to a design."""

    def post(self, request, design_pk):
        design = get_object_or_404(Design, pk=design_pk)
        form = DesignCutterLayoutForm(request.POST)
        if form.is_valid():
            layout = form.save(commit=False)
            layout.design = design
            layout.save()
            messages.success(request, "Cutter layout added.")
        else:
            messages.error(request, "Failed to add cutter layout.")
        return redirect("technology:design_detail", pk=design_pk)


class CutterLayoutDeleteView(LoginRequiredMixin, View):
    """Delete a cutter layout from a design."""

    def post(self, request, design_pk, layout_pk):
        layout = get_object_or_404(DesignCutterLayout, pk=layout_pk, design_id=design_pk)
        layout.delete()
        messages.success(request, "Cutter layout deleted.")
        return redirect("technology:design_detail", pk=design_pk)


# =============================================================================
# API VIEWS (for modal pickers)
# =============================================================================


class APIConnectionsView(LoginRequiredMixin, View):
    """API endpoint for connection selection modal."""

    def get(self, request):
        # Get filter parameters
        conn_id = request.GET.get('id', '')
        q = request.GET.get('q', '')
        type_id = request.GET.get('type', '')
        size_id = request.GET.get('size', '')
        can_replace = request.GET.get('can_replace', '')

        # Base queryset
        queryset = Connection.objects.filter(is_active=True).select_related(
            'connection_type', 'connection_size', 'upper_section_type'
        )

        # Apply filters
        if conn_id:
            # Filter by specific ID (for edit mode initialization)
            queryset = queryset.filter(id=conn_id)
        else:
            if q:
                queryset = queryset.filter(mat_no__icontains=q)
            if type_id:
                queryset = queryset.filter(connection_type_id=type_id)
            if size_id:
                queryset = queryset.filter(connection_size_id=size_id)
            if can_replace == 'true':
                queryset = queryset.filter(can_replace_in_ksa=True)
            elif can_replace == 'false':
                queryset = queryset.filter(can_replace_in_ksa=False)

        # Build response
        connections = []
        for conn in queryset[:100]:
            connections.append({
                'id': conn.id,
                'mat_no': conn.mat_no,
                'type': conn.connection_type.code,
                'type_name': conn.connection_type.name,
                'size': conn.connection_size.size_inches,
                'upper_section': conn.upper_section_type.name if conn.upper_section_type else None,
                'can_replace_in_ksa': conn.can_replace_in_ksa,
                'special_features': conn.special_features,
            })

        # Build filter options
        types = ConnectionType.objects.filter(is_active=True).values('id', 'code', 'name')
        sizes = ConnectionSize.objects.filter(is_active=True).values('id', 'size_inches')

        return JsonResponse({
            'connections': connections,
            'filters': {
                'types': [{'id': t['id'], 'code': t['code'], 'name': t['name']} for t in types],
                'sizes': [{'id': s['id'], 'size': s['size_inches']} for s in sizes],
            }
        })


class APIBreakerSlotsView(LoginRequiredMixin, View):
    """API endpoint for breaker slot selection modal."""

    def get(self, request):
        # Get filter parameters
        slot_id = request.GET.get('id', '')
        q = request.GET.get('q', '')
        material = request.GET.get('material', '')
        min_width = request.GET.get('min_width', '')
        max_width = request.GET.get('max_width', '')

        # Base queryset
        queryset = BreakerSlot.objects.filter(is_active=True)

        # Apply filters
        if slot_id:
            # Filter by specific ID (for edit mode initialization)
            queryset = queryset.filter(id=slot_id)
        else:
            if q:
                queryset = queryset.filter(mat_no__icontains=q)
            if material:
                queryset = queryset.filter(material=material)
            if min_width:
                queryset = queryset.filter(slot_width__gte=float(min_width))
            if max_width:
                queryset = queryset.filter(slot_width__lte=float(max_width))

        # Build response
        breaker_slots = []
        for slot in queryset[:100]:
            breaker_slots.append({
                'id': slot.id,
                'mat_no': slot.mat_no,
                'slot_width': str(slot.slot_width),
                'slot_depth': str(slot.slot_depth),
                'slot_length': str(slot.slot_length) if slot.slot_length else None,
                'material': slot.material,
                'material_display': slot.get_material_display(),
                'hardness': slot.hardness,
                'compatible_sizes': list(slot.compatible_sizes.values_list('size_display', flat=True)),
            })

        # Build filter options - material choices from model
        materials = [{'code': code, 'name': name} for code, name in BreakerSlot.Material.choices]

        # Bit sizes for compatible sizes checkboxes in quick create
        bit_sizes = [{'id': bs.id, 'display': bs.size_display} for bs in BitSize.objects.filter(is_active=True)]

        return JsonResponse({
            'breaker_slots': breaker_slots,
            'filters': {
                'materials': materials,
                'sizes': bit_sizes,
            }
        })


# =============================================================================
# CONNECTION CRUD VIEWS
# =============================================================================


class ConnectionListView(LoginRequiredMixin, ListView):
    """List all connections with filtering."""

    model = Connection
    template_name = "technology/connection_list.html"
    context_object_name = "connections"
    paginate_by = 25

    def get_queryset(self):
        queryset = Connection.objects.select_related(
            'connection_type', 'connection_size', 'upper_section_type'
        ).order_by('mat_no')

        search = self.request.GET.get("q")
        if search:
            queryset = queryset.filter(
                Q(mat_no__icontains=search) |
                Q(connection_type__code__icontains=search) |
                Q(connection_type__name__icontains=search)
            )

        type_filter = self.request.GET.get("type")
        if type_filter:
            queryset = queryset.filter(connection_type_id=type_filter)

        size_filter = self.request.GET.get("size")
        if size_filter:
            queryset = queryset.filter(connection_size_id=size_filter)

        ksa_filter = self.request.GET.get("ksa")
        if ksa_filter == "yes":
            queryset = queryset.filter(can_replace_in_ksa=True)
        elif ksa_filter == "no":
            queryset = queryset.filter(can_replace_in_ksa=False)

        if not self.request.GET.get("show_inactive"):
            queryset = queryset.filter(is_active=True)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["connection_types"] = ConnectionType.objects.filter(is_active=True)
        context["connection_sizes"] = ConnectionSize.objects.filter(is_active=True)
        return context


class ConnectionDetailView(LoginRequiredMixin, DetailView):
    """View connection details including related designs."""

    model = Connection
    template_name = "technology/connection_detail.html"
    context_object_name = "connection"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get designs using this connection
        context["related_designs"] = self.object.designs.select_related('size').order_by('hdbs_type')
        return context


class ConnectionCreateView(LoginRequiredMixin, CreateView):
    """Create a new connection."""

    model = Connection
    form_class = ConnectionForm
    template_name = "technology/connection_form.html"
    success_url = reverse_lazy("technology:connection_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Connection"
        context["submit_text"] = "Create Connection"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Connection {form.instance.mat_no} created successfully.")
        return super().form_valid(form)


class ConnectionUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing connection."""

    model = Connection
    form_class = ConnectionForm
    template_name = "technology/connection_form.html"

    def get_success_url(self):
        return reverse_lazy("technology:connection_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit Connection {self.object.mat_no}"
        context["submit_text"] = "Update Connection"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Connection {self.object.mat_no} updated successfully.")
        return super().form_valid(form)


class ConnectionDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a connection."""

    model = Connection
    template_name = "technology/connection_confirm_delete.html"
    success_url = reverse_lazy("technology:connection_list")

    def form_valid(self, form):
        messages.success(self.request, f"Connection {self.object.mat_no} deleted.")
        return super().form_valid(form)


# =============================================================================
# BREAKER SLOT CRUD VIEWS
# =============================================================================


class BreakerSlotListView(LoginRequiredMixin, ListView):
    """List all breaker slots with filtering."""

    model = BreakerSlot
    template_name = "technology/breaker_slot_list.html"
    context_object_name = "breaker_slots"
    paginate_by = 25

    def get_queryset(self):
        queryset = BreakerSlot.objects.prefetch_related('compatible_sizes').order_by('mat_no')

        search = self.request.GET.get("q")
        if search:
            queryset = queryset.filter(mat_no__icontains=search)

        material_filter = self.request.GET.get("material")
        if material_filter:
            queryset = queryset.filter(material=material_filter)

        if not self.request.GET.get("show_inactive"):
            queryset = queryset.filter(is_active=True)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["material_choices"] = BreakerSlot.Material.choices
        return context


class BreakerSlotDetailView(LoginRequiredMixin, DetailView):
    """View breaker slot details including related designs."""

    model = BreakerSlot
    template_name = "technology/breaker_slot_detail.html"
    context_object_name = "breaker_slot"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get designs using this breaker slot
        context["related_designs"] = self.object.designs.select_related('size').order_by('hdbs_type')
        return context


class BreakerSlotCreateView(LoginRequiredMixin, CreateView):
    """Create a new breaker slot."""

    model = BreakerSlot
    form_class = BreakerSlotForm
    template_name = "technology/breaker_slot_form.html"
    success_url = reverse_lazy("technology:breaker_slot_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Breaker Slot"
        context["submit_text"] = "Create Breaker Slot"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Breaker Slot {form.instance.mat_no} created successfully.")
        return super().form_valid(form)


class BreakerSlotUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing breaker slot."""

    model = BreakerSlot
    form_class = BreakerSlotForm
    template_name = "technology/breaker_slot_form.html"

    def get_success_url(self):
        return reverse_lazy("technology:breaker_slot_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit Breaker Slot {self.object.mat_no}"
        context["submit_text"] = "Update Breaker Slot"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Breaker Slot {self.object.mat_no} updated successfully.")
        return super().form_valid(form)


class BreakerSlotDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a breaker slot."""

    model = BreakerSlot
    template_name = "technology/breaker_slot_confirm_delete.html"
    success_url = reverse_lazy("technology:breaker_slot_list")

    def form_valid(self, form):
        messages.success(self.request, f"Breaker Slot {self.object.mat_no} deleted.")
        return super().form_valid(form)


# =============================================================================
# QUICK CREATE API VIEWS (for modal quick-create without leaving the page)
# =============================================================================


class APIConnectionCreateView(LoginRequiredMixin, View):
    """API endpoint to create a new connection from the design form modal."""

    def post(self, request):
        import json

        try:
            data = json.loads(request.body)

            # Validate required fields
            mat_no = data.get('mat_no', '').strip()
            connection_type_id = data.get('connection_type')
            connection_size_id = data.get('connection_size')

            if not mat_no:
                return JsonResponse({'success': False, 'error': 'MAT No. is required'}, status=400)
            if not connection_type_id:
                return JsonResponse({'success': False, 'error': 'Connection Type is required'}, status=400)
            if not connection_size_id:
                return JsonResponse({'success': False, 'error': 'Connection Size is required'}, status=400)

            # Check if MAT No. already exists
            if Connection.objects.filter(mat_no=mat_no).exists():
                return JsonResponse({'success': False, 'error': f'Connection with MAT No. {mat_no} already exists'}, status=400)

            # Create the connection
            connection = Connection.objects.create(
                mat_no=mat_no,
                connection_type_id=connection_type_id,
                connection_size_id=connection_size_id,
                special_features=data.get('special_features', ''),
                can_replace_in_ksa=data.get('can_replace_in_ksa', False),
                remarks=data.get('remarks', ''),
                is_active=data.get('is_active', True)
            )

            return JsonResponse({
                'success': True,
                'connection': {
                    'id': connection.id,
                    'mat_no': connection.mat_no,
                    'type': connection.connection_type.code,
                    'type_name': connection.connection_type.name,
                    'size': connection.connection_size.size_inches,
                    'can_replace_in_ksa': connection.can_replace_in_ksa,
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class APIBreakerSlotCreateView(LoginRequiredMixin, View):
    """API endpoint to create a new breaker slot from the design form modal."""

    def post(self, request):
        import json

        try:
            data = json.loads(request.body)

            # Validate required fields
            mat_no = data.get('mat_no', '').strip()
            slot_width = data.get('slot_width')
            slot_depth = data.get('slot_depth')
            material = data.get('material')

            if not mat_no:
                return JsonResponse({'success': False, 'error': 'MAT No. is required'}, status=400)
            if not slot_width:
                return JsonResponse({'success': False, 'error': 'Slot Width is required'}, status=400)
            if not slot_depth:
                return JsonResponse({'success': False, 'error': 'Slot Depth is required'}, status=400)
            if not material:
                return JsonResponse({'success': False, 'error': 'Material is required'}, status=400)

            # Check if MAT No. already exists
            if BreakerSlot.objects.filter(mat_no=mat_no).exists():
                return JsonResponse({'success': False, 'error': f'Breaker Slot with MAT No. {mat_no} already exists'}, status=400)

            # Create the breaker slot
            breaker_slot = BreakerSlot.objects.create(
                mat_no=mat_no,
                slot_width=slot_width,
                slot_depth=slot_depth,
                slot_length=data.get('slot_length') or None,
                material=material,
                hardness=data.get('hardness', ''),
                remarks=data.get('remarks', ''),
                is_active=data.get('is_active', True)
            )

            # Add compatible sizes if provided
            compatible_sizes = data.get('compatible_sizes', [])
            if compatible_sizes:
                breaker_slot.compatible_sizes.set(compatible_sizes)

            return JsonResponse({
                'success': True,
                'breaker_slot': {
                    'id': breaker_slot.id,
                    'mat_no': breaker_slot.mat_no,
                    'slot_width': str(breaker_slot.slot_width),
                    'slot_depth': str(breaker_slot.slot_depth),
                    'material': breaker_slot.material,
                    'material_display': breaker_slot.get_material_display(),
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class APIDesignSaveDraftView(LoginRequiredMixin, View):
    """API endpoint to save a design as draft from the design form."""

    def post(self, request):
        import json

        try:
            data = json.loads(request.body)

            # Validate minimum required fields
            mat_no = data.get('mat_no', '').strip()
            hdbs_type = data.get('hdbs_type', '').strip()

            if not mat_no and not hdbs_type:
                return JsonResponse({
                    'success': False,
                    'error': 'At least MAT No. or HDBS Type is required to save a draft'
                }, status=400)

            # Check for existing draft with same MAT No
            design_id = data.get('design_id')  # For updating existing draft
            if design_id:
                # Update existing design
                try:
                    design = Design.objects.get(id=design_id)
                except Design.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Design not found'}, status=404)
            else:
                # Create new design
                design = Design()

            # Set all provided fields
            design.mat_no = mat_no or f"DRAFT-{Design.objects.count() + 1}"
            design.hdbs_type = hdbs_type or "DRAFT"
            design.status = Design.Status.DRAFT

            # Optional fields
            if data.get('category'):
                design.category = data['category']
            if data.get('size'):
                design.size_id = data['size']
            if data.get('smi_type'):
                design.smi_type = data['smi_type']
            if data.get('ref_mat_no'):
                design.ref_mat_no = data['ref_mat_no']
            if data.get('ardt_item_no'):
                design.ardt_item_no = data['ardt_item_no']
            if data.get('body_material'):
                design.body_material = data['body_material']
            if data.get('series'):
                design.series = data['series']
            if data.get('no_of_blades'):
                design.no_of_blades = data['no_of_blades']
            if data.get('cutter_size'):
                design.cutter_size = data['cutter_size']
            if data.get('total_pockets_count'):
                design.total_pockets_count = data['total_pockets_count']
            if data.get('gage_length'):
                design.gage_length = data['gage_length']
            if data.get('gage_relief'):
                design.gage_relief = data['gage_relief']
            if data.get('nozzle_count'):
                design.nozzle_count = data['nozzle_count']
            if data.get('nozzle_bore_size'):
                design.nozzle_bore_size = data['nozzle_bore_size']
            if data.get('nozzle_config'):
                design.nozzle_config = data['nozzle_config']
            if data.get('port_count'):
                design.port_count = data['port_count']
            if data.get('port_size'):
                design.port_size = data['port_size']
            if data.get('order_level'):
                design.order_level = data['order_level']
            if data.get('iadc_code_ref'):
                design.iadc_code_ref_id = data['iadc_code_ref']
            if data.get('connection_ref'):
                design.connection_ref_id = data['connection_ref']
            if data.get('breaker_slot'):
                design.breaker_slot_id = data['breaker_slot']
            if data.get('formation_type_ref'):
                design.formation_type_ref_id = data['formation_type_ref']
            if data.get('application_ref'):
                design.application_ref_id = data['application_ref']
            if data.get('revision'):
                design.revision = data['revision']
            if data.get('description'):
                design.description = data['description']
            if data.get('notes'):
                design.notes = data['notes']

            design.save()

            return JsonResponse({
                'success': True,
                'design': {
                    'id': design.id,
                    'mat_no': design.mat_no,
                    'hdbs_type': design.hdbs_type,
                    'status': design.status,
                },
                'message': 'Design saved as draft'
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# BIT SIZE CRUD VIEWS (Simple list of sizes)
# =============================================================================


class BitSizeListView(LoginRequiredMixin, ListView):
    """List all bit sizes - simple list."""

    model = BitSize
    template_name = "technology/bit_size_list.html"
    context_object_name = "sizes"
    paginate_by = 50

    def get_queryset(self):
        queryset = BitSize.objects.order_by('size_decimal')

        search = self.request.GET.get("q")
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(size_display__icontains=search) |
                Q(size_inches__icontains=search) |
                Q(description__icontains=search)
            )

        if not self.request.GET.get("show_inactive"):
            queryset = queryset.filter(is_active=True)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Bit Sizes"
        return context


class BitSizeCreateView(LoginRequiredMixin, CreateView):
    """Create a new bit size."""

    model = BitSize
    form_class = BitSizeForm
    template_name = "technology/bit_size_form.html"
    success_url = reverse_lazy("technology:bit_size_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Add Bit Size"
        context["submit_text"] = "Add Size"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Bit Size {form.instance.size_display} added successfully.")
        return super().form_valid(form)


class BitSizeUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing bit size."""

    model = BitSize
    form_class = BitSizeForm
    template_name = "technology/bit_size_form.html"
    success_url = reverse_lazy("technology:bit_size_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit Size {self.object.size_display}"
        context["submit_text"] = "Save Changes"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Bit Size {self.object.size_display} updated successfully.")
        return super().form_valid(form)


class BitSizeDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a bit size."""

    model = BitSize
    template_name = "technology/bit_size_confirm_delete.html"
    success_url = reverse_lazy("technology:bit_size_list")

    def form_valid(self, form):
        try:
            messages.success(self.request, f"Bit Size {self.object.size_display} deleted.")
            return super().form_valid(form)
        except Exception:
            messages.error(
                self.request,
                f"Cannot delete {self.object.size_display} — it is referenced by designs, "
                f"SMI types, or other records. Reassign those records first, or use the "
                f"merge_duplicate_sizes management command."
            )
            return redirect("technology:bit_size_list")


# =============================================================================
# HDBS TYPE CRUD VIEWS (Internal type naming with SMI children)
# =============================================================================


class HDBSTypeListView(LoginRequiredMixin, ListView):
    """List all HDBS types with their SMI names."""

    model = HDBSType
    template_name = "technology/hdbs_type_list.html"
    context_object_name = "hdbs_types"

    def get_queryset(self):
        # Load all records — filtering/pagination handled client-side
        return HDBSType.objects.prefetch_related(
            'smi_types', 'smi_types__size', 'sizes'
        ).order_by('hdbs_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Bit Types (HDBS/SMI)"
        context["sizes"] = BitSize.objects.filter(is_active=True).order_by('size_decimal')

        # Build flat rows for the table: each row = {hdbs, size, smi}
        flat_rows = []
        for hdbs in context["hdbs_types"]:
            sizes = list(hdbs.sizes.filter(is_active=True).order_by('size_decimal'))
            smi_types = list(hdbs.smi_types.select_related('size').all())

            if sizes:
                for size in sizes:
                    # Find SMI types for this specific size OR no size (applies to all)
                    matching_smi = [s for s in smi_types if s.size_id == size.id or s.size_id is None]
                    if matching_smi:
                        for smi in matching_smi:
                            flat_rows.append({'hdbs': hdbs, 'size': size, 'smi': smi})
                    else:
                        # No SMI for this size
                        flat_rows.append({'hdbs': hdbs, 'size': size, 'smi': None})
            else:
                # HDBS has no sizes assigned
                if smi_types:
                    for smi in smi_types:
                        flat_rows.append({'hdbs': hdbs, 'size': None, 'smi': smi})
                else:
                    # No sizes and no SMI types
                    flat_rows.append({'hdbs': hdbs, 'size': None, 'smi': None})

        context["flat_rows"] = flat_rows
        return context


class HDBSTypeDetailView(LoginRequiredMixin, DetailView):
    """View HDBS type details with SMI names and related designs."""

    model = HDBSType
    template_name = "technology/hdbs_type_detail.html"
    context_object_name = "hdbs_type"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get SMI types for this HDBS
        smi_types = self.object.smi_types.select_related('size').order_by('size__size_decimal', 'smi_name')
        context["smi_types"] = smi_types

        # Build flat table rows: each row = (size, smi)
        # If HDBS has sizes, show one row per size with matching SMI types
        # SMI types with size=None apply to all sizes
        type_size_smi_rows = []
        sizes = self.object.sizes.filter(is_active=True).order_by('size_decimal')

        if sizes.exists():
            for size in sizes:
                # Find SMI types for this specific size OR no size (applies to all)
                size_smi_types = [s for s in smi_types if s.size_id == size.id or s.size_id is None]
                if size_smi_types:
                    for smi in size_smi_types:
                        type_size_smi_rows.append({'size': size, 'smi': smi})
                else:
                    # No SMI for this size - show row with empty SMI
                    type_size_smi_rows.append({'size': size, 'smi': None})
        else:
            # HDBS has no sizes assigned - show all SMI types without size
            for smi in smi_types:
                type_size_smi_rows.append({'size': None, 'smi': smi})
            if not smi_types:
                type_size_smi_rows.append({'size': None, 'smi': None})

        context["type_size_smi_rows"] = type_size_smi_rows

        # Get designs using this HDBS type - both legacy field and junction table
        legacy_design_ids = Design.objects.filter(
            hdbs_type__icontains=self.object.hdbs_name
        ).values_list('id', flat=True)
        junction_design_ids = self.object.design_assignments.filter(
            is_current=True
        ).values_list('design_id', flat=True)

        all_design_ids = set(legacy_design_ids) | set(junction_design_ids)
        context["related_designs"] = Design.objects.filter(
            id__in=all_design_ids
        ).select_related('size').order_by('mat_no')[:30]

        # Current design assignments via junction table
        context["design_hdbs_assignments"] = self.object.design_assignments.filter(
            is_current=True
        ).select_related('design', 'design__size', 'assigned_by').order_by('design__mat_no')
        return context


class HDBSTypeCreateView(LoginRequiredMixin, CreateView):
    """Create a new HDBS type."""

    model = HDBSType
    form_class = HDBSTypeForm
    template_name = "technology/hdbs_type_form.html"

    def get_success_url(self):
        # If coming from design or bom form, stay on page with success message
        from_page = self.request.GET.get('from')
        if from_page in ('design', 'bom'):
            return f"{reverse_lazy('technology:hdbs_type_create')}?from={from_page}&created=1"
        return reverse_lazy("technology:hdbs_type_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Add HDBS Type"
        context["submit_text"] = "Add Type"
        context["just_created"] = self.request.GET.get('created') == '1'
        context["from_page"] = self.request.GET.get('from')
        return context

    def form_valid(self, form):
        from_page = self.request.GET.get('from')
        if from_page == 'bom':
            messages.success(self.request, f"HDBS Type {form.instance.hdbs_name} created successfully. You can now close this tab and return to your BOM form.")
        else:
            messages.success(self.request, f"HDBS Type {form.instance.hdbs_name} created successfully. You can now close this tab and return to your Design form.")
        return super().form_valid(form)


class HDBSTypeUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing HDBS type."""

    model = HDBSType
    form_class = HDBSTypeForm
    template_name = "technology/hdbs_type_form.html"

    def get_success_url(self):
        return reverse_lazy("technology:hdbs_type_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit HDBS Type {self.object.hdbs_name}"
        context["submit_text"] = "Save Changes"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"HDBS Type {self.object.hdbs_name} updated successfully.")
        return super().form_valid(form)


class HDBSTypeDeleteView(LoginRequiredMixin, DeleteView):
    """Delete an HDBS type."""

    model = HDBSType
    template_name = "technology/hdbs_type_confirm_delete.html"
    success_url = reverse_lazy("technology:hdbs_type_list")

    def form_valid(self, form):
        messages.success(self.request, f"HDBS Type {self.object.hdbs_name} deleted.")
        return super().form_valid(form)


# =============================================================================
# SMI TYPE CRUD VIEWS (Client-facing naming linked to HDBS)
# =============================================================================


class SMITypeCreateView(LoginRequiredMixin, CreateView):
    """Create a new SMI type linked to an HDBS type."""

    model = SMIType
    form_class = SMITypeForm
    template_name = "technology/smi_type_form.html"

    def get_hdbs_type(self):
        """Get the HDBS type from URL or query param."""
        hdbs_pk = self.kwargs.get('hdbs_pk') or self.request.GET.get('hdbs')
        if hdbs_pk:
            return get_object_or_404(HDBSType, pk=hdbs_pk)
        return None

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        hdbs_type = self.get_hdbs_type()
        if hdbs_type:
            # Only show sizes that belong to this HDBS type
            if hdbs_type.sizes.exists():
                form.fields['size'].queryset = hdbs_type.sizes.filter(is_active=True).order_by('size_decimal')
            else:
                # If HDBS has no sizes, show all active sizes
                form.fields['size'].queryset = BitSize.objects.filter(is_active=True).order_by('size_decimal')
        return form

    def get_initial(self):
        initial = super().get_initial()
        # Pre-select HDBS type if provided in URL
        hdbs_pk = self.kwargs.get('hdbs_pk') or self.request.GET.get('hdbs')
        if hdbs_pk:
            initial['hdbs_type'] = hdbs_pk
        # Pre-select size if provided in query string
        size_pk = self.request.GET.get('size')
        if size_pk:
            initial['size'] = size_pk
        return initial

    def get_success_url(self):
        return reverse_lazy("technology:hdbs_type_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Add SMI Type"
        context["submit_text"] = "Add SMI Type"
        hdbs_type = self.get_hdbs_type()
        if hdbs_type:
            context["hdbs_type"] = hdbs_type
        return context

    def form_valid(self, form):
        messages.success(self.request, f"SMI Type {form.instance.smi_name} added successfully.")
        return super().form_valid(form)


class SMITypeCreateStandaloneView(LoginRequiredMixin, CreateView):
    """Create a new SMI type without pre-selected HDBS."""

    model = SMIType
    form_class = SMITypeForm
    template_name = "technology/smi_type_form.html"

    def get_success_url(self):
        # If coming from design or bom form, stay on page with success message
        from_page = self.request.GET.get('from')
        if from_page in ('design', 'bom'):
            return f"{reverse_lazy('technology:smi_type_create_standalone')}?from={from_page}&created=1"
        return reverse_lazy("technology:hdbs_type_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Add SMI Type"
        context["submit_text"] = "Add SMI Type"
        context["just_created"] = self.request.GET.get('created') == '1'
        context["from_page"] = self.request.GET.get('from')
        return context

    def form_valid(self, form):
        from_page = self.request.GET.get('from')
        if from_page == 'bom':
            messages.success(self.request, f"SMI Type {form.instance.smi_name} added successfully. You can now close this tab and return to your BOM form.")
        elif from_page == 'design':
            messages.success(self.request, f"SMI Type {form.instance.smi_name} added successfully. You can now close this tab and return to your Design form.")
        else:
            messages.success(self.request, f"SMI Type {form.instance.smi_name} added successfully.")
        return super().form_valid(form)


class SMITypeUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing SMI type."""

    model = SMIType
    form_class = SMITypeForm
    template_name = "technology/smi_type_form.html"

    def get_success_url(self):
        return reverse_lazy("technology:hdbs_type_detail", kwargs={"pk": self.object.hdbs_type.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit SMI Type {self.object.smi_name}"
        context["submit_text"] = "Save Changes"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"SMI Type {self.object.smi_name} updated successfully.")
        return super().form_valid(form)


class SMITypeDeleteView(LoginRequiredMixin, DeleteView):
    """Delete an SMI type."""

    model = SMIType
    template_name = "technology/smi_type_confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy("technology:hdbs_type_detail", kwargs={"pk": self.object.hdbs_type.pk})

    def form_valid(self, form):
        messages.success(self.request, f"SMI Type {self.object.smi_name} deleted.")
        return super().form_valid(form)


# =============================================================================
# API VIEWS FOR QUICK CREATE (Types)
# =============================================================================


class APIHDBSTypesView(LoginRequiredMixin, View):
    """API endpoint to get HDBS types for dropdowns."""

    def get(self, request):
        # Filter by active status unless show_inactive is set
        show_inactive = request.GET.get('show_inactive')
        if show_inactive:
            hdbs_types = HDBSType.objects.all()
        else:
            hdbs_types = HDBSType.objects.filter(is_active=True)

        # Search filter
        q = request.GET.get('q', '').strip()
        if q:
            hdbs_types = hdbs_types.filter(
                Q(hdbs_name__icontains=q) |
                Q(description__icontains=q) |
                Q(smi_types__smi_name__icontains=q)
            ).distinct()

        # Filter by size - show only HDBS types that have this size selected
        # OR have no sizes (meaning they apply to all sizes)
        size_id = request.GET.get('size', '').strip()
        if size_id:
            hdbs_types = hdbs_types.filter(
                Q(sizes__id=size_id) | Q(sizes__isnull=True)
            ).distinct()

        hdbs_types = hdbs_types.prefetch_related('sizes', 'smi_types').order_by('hdbs_name')

        # Build response with SMI types filtered by size if provided
        data = {'hdbs_types': []}
        for t in hdbs_types:
            # Get SMI types, optionally filtered by size
            if show_inactive:
                smi_qs = t.smi_types.all()
            else:
                smi_qs = t.smi_types.filter(is_active=True)

            # If size is specified, filter SMI types to those matching the size or with no size
            if size_id:
                smi_qs = smi_qs.filter(Q(size_id=size_id) | Q(size__isnull=True))

            smi_list = [
                {'id': s.id, 'smi_name': s.smi_name, 'is_active': s.is_active, 'size_id': s.size_id}
                for s in smi_qs.order_by('smi_name')
            ]

            # If size filter is applied, only show that size; otherwise show all sizes
            if size_id:
                sizes_list = [{'id': s.id, 'display': s.size_display} for s in t.sizes.filter(is_active=True, id=size_id)]
            else:
                sizes_list = [{'id': s.id, 'display': s.size_display} for s in t.sizes.filter(is_active=True)]

            data['hdbs_types'].append({
                'id': t.id,
                'hdbs_name': t.hdbs_name,
                'description': t.description or '',
                'is_active': t.is_active,
                'sizes': sizes_list,
                'smi_types': smi_list
            })

        return JsonResponse(data)


class APIHDBSTypeCreateView(LoginRequiredMixin, View):
    """API endpoint to quick create an HDBS type (optionally with SMI type)."""

    def post(self, request):
        import json
        try:
            data = json.loads(request.body)
            hdbs_name = data.get('hdbs_name', '').strip()

            if not hdbs_name:
                return JsonResponse({'success': False, 'error': 'HDBS Name is required'}, status=400)

            if HDBSType.objects.filter(hdbs_name=hdbs_name).exists():
                return JsonResponse({'success': False, 'error': 'HDBS Type already exists'}, status=400)

            hdbs_type = HDBSType.objects.create(
                hdbs_name=hdbs_name,
                description=data.get('description', ''),
                is_active=True
            )

            # Add sizes if provided
            size_ids = data.get('sizes', [])
            if size_ids:
                hdbs_type.sizes.set(size_ids)

            response_data = {
                'success': True,
                'hdbs_type': {
                    'id': hdbs_type.id,
                    'hdbs_name': hdbs_type.hdbs_name,
                }
            }

            # Also create SMI type if smi_name is provided
            smi_name = data.get('smi_name', '').strip()
            size_id = data.get('size_id')
            if smi_name and size_id:
                try:
                    size = BitSize.objects.get(pk=size_id)
                    # Add the size to HDBS type if not already added
                    hdbs_type.sizes.add(size)

                    smi_type = SMIType.objects.create(
                        smi_name=smi_name,
                        hdbs_type=hdbs_type,
                        size=size,
                        description=data.get('smi_description', ''),
                        is_active=True
                    )
                    response_data['smi_type'] = {
                        'id': smi_type.id,
                        'smi_name': smi_type.smi_name,
                    }
                except BitSize.DoesNotExist:
                    # Size not found, skip SMI creation but don't fail
                    pass

            return JsonResponse(response_data)

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class APISMITypeCreateView(LoginRequiredMixin, View):
    """API endpoint to quick create an SMI type (uses size from design form)."""

    def post(self, request):
        import json
        try:
            data = json.loads(request.body)
            smi_name = data.get('smi_name', '').strip()
            hdbs_type_id = data.get('hdbs_type_id')
            size_id = data.get('size_id')

            if not smi_name:
                return JsonResponse({'success': False, 'error': 'SMI Name is required'}, status=400)

            if not hdbs_type_id:
                return JsonResponse({'success': False, 'error': 'HDBS Type is required'}, status=400)

            if not size_id:
                return JsonResponse({'success': False, 'error': 'Please select a Size in the design form first'}, status=400)

            hdbs_type = get_object_or_404(HDBSType, pk=hdbs_type_id)
            size = get_object_or_404(BitSize, pk=size_id)

            if SMIType.objects.filter(smi_name=smi_name, hdbs_type=hdbs_type, size=size).exists():
                return JsonResponse({'success': False, 'error': 'SMI Type already exists for this HDBS and size'}, status=400)

            # Add size to HDBS type's compatible sizes if not already there
            hdbs_type.sizes.add(size)

            smi_type = SMIType.objects.create(
                smi_name=smi_name,
                hdbs_type=hdbs_type,
                size=size,
                description=data.get('description', ''),
                is_active=True
            )

            return JsonResponse({
                'success': True,
                'smi_type': {
                    'id': smi_type.id,
                    'smi_name': smi_type.smi_name,
                    'hdbs_type_id': hdbs_type.id,
                    'hdbs_name': hdbs_type.hdbs_name,
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# SMI TYPE FILTER API (for BOM creation workflow)
# =============================================================================


class APISMITypesFilterView(LoginRequiredMixin, View):
    """
    API endpoint to get SMI Types filtered by HDBS Type and Size.
    Used in the BOM creation workflow to show only relevant SMI options.
    """

    def get(self, request):
        hdbs_type_id = request.GET.get('hdbs_type_id')
        size_id = request.GET.get('size_id')

        queryset = SMIType.objects.filter(is_active=True)

        if hdbs_type_id:
            queryset = queryset.filter(hdbs_type_id=hdbs_type_id)

        if size_id:
            queryset = queryset.filter(size_id=size_id)

        smi_types = []
        for smi in queryset.select_related('hdbs_type', 'size'):
            smi_types.append({
                'id': smi.id,
                'smi_name': smi.smi_name,
                'hdbs_type_id': smi.hdbs_type_id,
                'hdbs_name': smi.hdbs_type.hdbs_name,
                'size_id': smi.size_id,
                'size_display': str(smi.size) if smi.size else '',
            })

        return JsonResponse({
            'success': True,
            'smi_types': smi_types,
        })


# =============================================================================
# COMBINED BOM CREATE/BUILDER VIEW
# =============================================================================


class BOMCreateWithBuilderView(LoginRequiredMixin, TemplateView):
    """
    BOM creation page - Design selection for Cutter Map workflow.

    This view shows a list of L3/L4 Designs for the user to select,
    then redirects to the Cutter Map app to import PDF and create the BOM.
    """

    template_name = "technology/bom_create_builder.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create BOM"

        # Get designs for selection (L3/L4/L5.5) - include all statuses
        context["designs"] = Design.objects.filter(
            order_level__in=["3", "4", "5.5"]
        ).select_related("size").prefetch_related(
            "boms"
        ).order_by("-created_at")

        # SMI Types and IADC Codes for BOM creation
        context["smi_types"] = SMIType.objects.filter(is_active=True).select_related('hdbs_type', 'size').order_by('smi_name')
        context["iadc_codes"] = IADCCode.objects.filter(is_active=True).order_by('code')
        context["status_choices"] = BOM.Status.choices

        # Handle pre-selection from drill bit detail page
        design_id = self.request.GET.get('design_id')
        drillbit_id = self.request.GET.get('drillbit_id')

        if design_id:
            context["preselect_design_id"] = design_id
        if drillbit_id:
            context["preselect_drillbit_id"] = drillbit_id
            # Get drill bit details for display
            try:
                from apps.workorders.models import DrillBit
                drillbit = DrillBit.objects.select_related('design').get(pk=drillbit_id)
                context["preselect_drillbit"] = drillbit
            except (DrillBit.DoesNotExist, ValueError):
                pass

        return context


class APIDesignsFilterView(LoginRequiredMixin, View):
    """
    API endpoint to get Designs filtered by HDBS Type and/or Size.
    Used in the BOM creation workflow to show matching designs.
    """

    def get(self, request):
        hdbs_type = request.GET.get('hdbs_type')
        size_id = request.GET.get('size_id')

        queryset = Design.objects.filter(
            order_level__in=["3", "4", "5.5"],
            status__in=[Design.Status.DRAFT, Design.Status.ACTIVE]
        )

        if hdbs_type:
            queryset = queryset.filter(hdbs_type__icontains=hdbs_type)

        if size_id:
            queryset = queryset.filter(size_id=size_id)

        designs = []
        for d in queryset.select_related("size")[:20]:
            designs.append({
                'id': d.id,
                'mat_no': d.mat_no,
                'hdbs_type': d.hdbs_type,
                'size': str(d.size) if d.size else '',
                'order_level': d.order_level,
                'display': f"{d.mat_no} - {d.hdbs_type} ({d.size})" if d.size else f"{d.mat_no} - {d.hdbs_type}",
            })

        return JsonResponse({
            'success': True,
            'designs': designs,
        })


# =============================================================================
# DESIGN IMPORT VIEW
# =============================================================================


class DesignImportView(LoginRequiredMixin, TemplateView):
    """
    Web interface for importing Designs from ERP Items Excel file.

    Reads the Items Excel file and imports RM (Raw Material) entries as Designs.
    Only imports designs where the MAT number is in the L3 or L4 list.

    Excel Format:
    - Item number: RM-{FC|RC}-{IB|MB|SB|IT|MT}-NNNN
    - Product name: Size : HDBS Type : MAT Number
    - Search name: MAT number (used for matching)
    """

    template_name = "technology/design_import.html"

    # L3 MAT numbers - designs without cutters, upper section separate
    L3_MATS = {
        '1112857', '1134878', '1140080', '1141517', '1141549', '1141552',
        '1145822', '1145823', '1145824', '1145825', '1145826', '1145877',
        '1146009', '1153090', '1158490', '1160847', '1186940', '1187881',
        '1188084', '1188149', '1188197', '1191006', '1192557', '1197910',
        '1220553', '1221158', '1223029', '1223749', '1227237', '1228690',
        '1228723', '1228748', '1228782', '1228848', '1228896', '1229187',
        '1231398', '1234376', '1245350', '1245357', '1245431', '1246680',
        '1248668', '1251085', '1270578', '1284241', '2013684', '2016137',
        '2016920', '1146317'
    }

    # L4 MAT numbers - designs without cutters, upper section welded/machined
    L4_MATS = {
        '1145821', '1149944', '1150007', '1158011', '1192119', '1198387',
        '1198515', '1201053', '1208569', '1212881', '1217589', '1224225',
        '1228739', '1229062', '1231285', '1238258', '1239176', '1239239',
        '1239257', '1239258', '1240803', '1246020', '1255267', '1255451',
        '1257458', '1258119', '1259957', '1261236', '1262047', '1267027',
        '1267122', '1267809', '1269498', '1270617', '1270865', '1272920',
        '1281273', '1282368', '1283567', '1290130', '2000401', '2001502',
        '2009096', '2013542', '2017993', '2019988', '2022318', '2022350',
        '2025595', '2027359', '2028159', '2028515'
    }

    # Size mapping: normalize various formats to standard display and decimal
    SIZE_MAP = {
        # Fractional sizes
        '3 5/8': ('3 5/8"', '3.625'),
        '3 7/8': ('3 7/8"', '3.875'),
        '3 3/4': ('3 3/4"', '3.750'),
        '5 1/4': ('5 1/4"', '5.250'),
        '5 7/8': ('5 7/8"', '5.875'),
        '6 1/8': ('6 1/8"', '6.125'),
        '8 1/2': ('8 1/2"', '8.500'),
        '8 3/8': ('8 3/8"', '8.375'),
        '12 1/4': ('12 1/4"', '12.250'),
        # Whole numbers
        '3': ('3"', '3.000'),
        '5': ('5"', '5.000'),
        '6': ('6"', '6.000'),
        '8': ('8"', '8.000'),
        '12': ('12"', '12.000'),
        '16': ('16"', '16.000'),
        '17': ('17"', '17.000'),
        '22': ('22"', '22.000'),
        # Decimal formats
        '3.625': ('3 5/8"', '3.625'),
        '3.875': ('3 7/8"', '3.875'),
        '5.875': ('5 7/8"', '5.875'),
        '6.125': ('6 1/8"', '6.125'),
        '8.375': ('8 3/8"', '8.375'),
        '8.5': ('8 1/2"', '8.500'),
        '8.500': ('8 1/2"', '8.500'),
        '12.25': ('12 1/4"', '12.250'),
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Import Designs from Excel"

        # Get available Excel files in docs folder
        import os
        docs_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'docs')
        excel_files = []
        if os.path.exists(docs_path):
            for f in os.listdir(docs_path):
                if f.endswith('.xlsx') and ('item' in f.lower() or 'Item' in f):
                    excel_files.append(f)
        context["excel_files"] = sorted(excel_files)

        # Counts
        context["total_l3"] = len(self.L3_MATS)
        context["total_l4"] = len(self.L4_MATS)
        context["total_designs"] = Design.objects.count()

        # Count existing designs by order level
        context["existing_l3"] = Design.objects.filter(order_level='3').count()
        context["existing_l4"] = Design.objects.filter(order_level='4').count()

        return context

    def normalize_size(self, size_str):
        """Normalize size string to standard format."""
        import re
        from decimal import Decimal

        if not size_str:
            return None

        # Remove quotes and extra spaces, trim
        size_str = size_str.strip().replace('"', '').replace("'", '').strip()

        # Direct lookup
        if size_str in self.SIZE_MAP:
            return self.SIZE_MAP[size_str]

        # Try to parse decimal
        try:
            decimal_val = Decimal(size_str)
            # Round to 3 decimal places
            decimal_val = decimal_val.quantize(Decimal('0.001'))
            str_val = str(decimal_val)
            if str_val in self.SIZE_MAP:
                return self.SIZE_MAP[str_val]
            # Create a new entry
            return (f'{decimal_val}"', str(decimal_val))
        except:
            pass

        return None

    def parse_product_name(self, product_name):
        """
        Parse Product name to extract Size, HDBS Type, and MAT number.
        Primary format uses ':' as separator: 'Size : HDBS Type : MAT Number'
        """
        import re

        if not product_name:
            return None, None, None

        normalized = product_name.strip()

        # Primary method: split by colon (with optional surrounding spaces)
        if ':' in normalized:
            parts = [p.strip() for p in normalized.split(':')]
            if len(parts) >= 3:
                size_str = parts[0].replace('"', '').replace("'", '').strip()
                hdbs_type = re.sub(r'-\d+$', '', parts[1].strip())  # Remove suffix like -1
                mat_no = parts[2].strip()
                if mat_no.isdigit() and len(mat_no) >= 6:
                    return size_str, hdbs_type, mat_no
            elif len(parts) == 2:
                # Might be 'Size : HDBS MAT' format
                size_str = parts[0].replace('"', '').replace("'", '').strip()
                rest = parts[1].strip().split()
                if len(rest) >= 2:
                    mat_no = rest[-1]
                    hdbs_type = re.sub(r'-\d+$', '', rest[-2])
                    if mat_no.isdigit() and len(mat_no) >= 6:
                        return size_str, hdbs_type, mat_no

        # Fallback: handle legacy format without colons (space separated)
        normalized = re.sub(r'(["\'])\s*([A-Za-z])', r'\1 \2', normalized)
        parts = normalized.split()

        if len(parts) < 2:
            return None, None, None

        # Find the MAT number (last numeric part, 6-7 digits)
        mat_no = None
        for i in range(len(parts) - 1, -1, -1):
            if parts[i].isdigit() and len(parts[i]) >= 6:
                mat_no = parts[i]
                parts = parts[:i]
                break

        if not mat_no:
            return None, None, None

        # Find HDBS type (alphanumeric, before the MAT number)
        hdbs_type = None
        for i in range(len(parts) - 1, -1, -1):
            part = parts[i]
            clean_part = re.sub(r'-\d+$', '', part)
            if re.match(r'^[A-Za-z]{2}[A-Za-z0-9]+$', clean_part):
                hdbs_type = clean_part
                parts = parts[:i]
                break

        if not hdbs_type:
            return None, None, mat_no

        # Remaining parts are the size
        size_str = ' '.join(parts)
        size_str = size_str.replace('"', '').replace("'", '').strip()

        return size_str, hdbs_type, mat_no

    def get_category_and_body(self, item_number):
        """
        Determine category and body material from item number.
        Format: RM-{FC|RC}-{IB|MB|SB|IT|MT}-NNNN
        """
        parts = item_number.split('-')
        if len(parts) < 3:
            return None, None

        bit_type = parts[1]  # FC or RC
        body_type = parts[2]  # IB, MB, SB, IT, MT

        if bit_type == 'FC':
            category = 'FC'
            if body_type == 'MB':
                body_material = 'M'
            elif body_type == 'SB':
                body_material = 'S'
            elif body_type == 'IB':
                body_material = 'M'
            else:
                body_material = ''
        elif bit_type == 'RC':
            if body_type == 'IT':
                category = 'TCI'
            elif body_type == 'MT':
                category = 'MT'
            else:
                category = 'TCI'
            body_material = ''
        else:
            category = 'FC'
            body_material = ''

        return category, body_material

    def get_or_create_size(self, size_str, dry_run=False):
        """Get or create BitSize record."""
        from decimal import Decimal

        normalized = self.normalize_size(size_str)
        if not normalized:
            return None

        size_display, size_code = normalized
        size_decimal = Decimal(size_code)

        if dry_run:
            existing = BitSize.objects.filter(code=size_code).first()
            return existing

        size, created = BitSize.objects.get_or_create(
            code=size_code,
            defaults={
                'size_decimal': size_decimal,
                'size_display': size_display,
                'size_inches': size_display,
            }
        )
        return size

    def post(self, request, *args, **kwargs):
        """Handle file upload, preview, and import."""
        import os
        import re
        from decimal import Decimal

        action = request.POST.get('action', 'preview')

        # Determine file source
        uploaded_file = request.FILES.get('excel_file')
        selected_file = request.POST.get('selected_file', '')

        try:
            import pandas as pd
        except ImportError:
            return JsonResponse({'success': False, 'error': 'pandas is required'}, status=500)

        if uploaded_file:
            df = pd.read_excel(uploaded_file)
            file_name = uploaded_file.name
        elif selected_file:
            docs_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'docs')
            file_path = os.path.join(docs_path, selected_file)
            if not os.path.exists(file_path):
                return JsonResponse({'success': False, 'error': f'File not found: {selected_file}'}, status=400)
            df = pd.read_excel(file_path)
            file_name = selected_file
        else:
            return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

        # Check required columns
        required_cols = ['Item number', 'Product name', 'Search name']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return JsonResponse({
                'success': False,
                'error': f'Missing required columns: {missing}'
            }, status=400)

        # Filter RM rows
        rm_rows = df[df['Item number'].str.startswith('RM-', na=False)]

        # Get all valid MAT numbers
        all_mats = self.L3_MATS | self.L4_MATS

        # Filter by MAT numbers in the provided list
        rm_filtered = rm_rows[rm_rows['Search name'].astype(str).isin(all_mats)]

        dry_run = action == 'preview'

        stats = {
            'rows_total': len(rm_rows),
            'rows_matched': 0,  # Will count unique MATs
            'rows_with_duplicates': 0,
            'l3_count': 0,
            'l4_count': 0,
            'fc_count': 0,
            'rc_count': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }

        preview_data = []
        errors = []

        # First pass: collect all HDBS types per MAT number to detect duplicates
        mat_to_entries = {}
        for idx, row in rm_filtered.iterrows():
            item_number = str(row['Item number']).strip()
            product_name = str(row['Product name']).strip()
            search_name = str(row['Search name']).strip()

            size_str, hdbs_type, mat_no = self.parse_product_name(product_name)
            if not mat_no:
                mat_no = search_name

            if not mat_no:
                errors.append({
                    'item_number': item_number,
                    'product_name': product_name,
                    'error': 'Could not extract MAT number'
                })
                stats['errors'] += 1
                continue

            if mat_no not in mat_to_entries:
                mat_to_entries[mat_no] = []
            mat_to_entries[mat_no].append({
                'hdbs_type': hdbs_type,
                'item_number': item_number,
                'size_str': size_str,
            })

        stats['rows_matched'] = len(mat_to_entries)
        stats['rows_with_duplicates'] = sum(1 for v in mat_to_entries.values() if len(v) > 1)

        # Second pass: process unique MAT numbers
        for mat_no, entries in mat_to_entries.items():
            # Use first entry as primary
            primary = entries[0]
            item_number = primary['item_number']
            hdbs_type = primary['hdbs_type']
            size_str = primary['size_str']

            # Collect alternative HDBS types
            alt_hdbs_types = []
            if len(entries) > 1:
                for entry in entries[1:]:
                    if entry['hdbs_type'] and entry['hdbs_type'] != hdbs_type:
                        alt_hdbs_types.append(entry['hdbs_type'])

            # Determine order level
            if mat_no in self.L3_MATS:
                order_level = '3'
                stats['l3_count'] += 1
            elif mat_no in self.L4_MATS:
                order_level = '4'
                stats['l4_count'] += 1
            else:
                order_level = ''

            # Get category and body material
            category, body_material = self.get_category_and_body(item_number)

            if category == 'FC':
                stats['fc_count'] += 1
            else:
                stats['rc_count'] += 1

            # Extract series from HDBS type
            series = ''
            if hdbs_type:
                match = re.match(r'^([A-Za-z]{2,3})', hdbs_type)
                if match:
                    series = match.group(1).upper()

            # Check if exists
            existing = Design.objects.filter(mat_no=mat_no).first()

            # Build notes for alternative HDBS types
            notes = ''
            if alt_hdbs_types:
                notes = f"Alternative HDBS types found: {', '.join(alt_hdbs_types)}"

            preview_item = {
                'item_number': item_number,
                'mat_no': mat_no,
                'size': size_str or '-',
                'hdbs_type': hdbs_type or '-',
                'alt_hdbs_types': alt_hdbs_types,  # Show alternatives in preview
                'category': category or '-',
                'body_material': body_material or '-',
                'order_level': f'L{order_level}' if order_level else '-',
                'series': series or '-',
                'status': 'update' if existing else 'create',
                'has_duplicates': len(entries) > 1,
            }
            preview_data.append(preview_item)

            if dry_run:
                if existing:
                    stats['updated'] += 1
                else:
                    stats['created'] += 1
                continue

            # Get or create size
            size_obj = None
            if size_str:
                size_obj = self.get_or_create_size(size_str)

            if existing:
                # Update existing design - trim all values
                existing.hdbs_type = (hdbs_type or existing.hdbs_type or '').strip()
                existing.category = (category or existing.category or '').strip()
                existing.body_material = (body_material or existing.body_material or '').strip()
                existing.order_level = (order_level or existing.order_level or '').strip()
                existing.status = 'ACTIVE'
                if size_obj:
                    existing.size = size_obj
                if series:
                    existing.series = series.strip()
                # Store alternative HDBS types in notes
                if notes:
                    existing.notes = notes
                existing.save()
                stats['updated'] += 1
            else:
                # Create new design - trim all values
                try:
                    design = Design.objects.create(
                        mat_no=mat_no.strip(),
                        hdbs_type=(hdbs_type or '').strip(),
                        category=(category or 'FC').strip(),
                        body_material=(body_material or '').strip(),
                        order_level=(order_level or '').strip(),
                        size=size_obj,
                        series=(series or '').strip(),
                        status='ACTIVE',
                        description=f"Imported from {item_number}",
                        notes=notes,  # Store alternative HDBS types
                    )
                    stats['created'] += 1
                except Exception as e:
                    errors.append({
                        'item_number': item_number,
                        'mat_no': mat_no,
                        'error': str(e)
                    })
                    stats['errors'] += 1

        return JsonResponse({
            'success': True,
            'dry_run': dry_run,
            'file_name': file_name,
            'stats': stats,
            'preview': preview_data[:100],  # Limit preview to 100 items
            'errors': errors[:20],  # Limit errors to 20
        })



# =============================================================================
# DESIGN STOCK IMPORT FROM EXCEL (RM Items from ERP Items file)
# =============================================================================


class DesignStockImportView(LoginRequiredMixin, TemplateView):
    """
    Web interface for importing design (RM) stock quantities from ERP Excel files.

    Expected columns:
    - Item number: ERP item number (filter for RM-* items)
    - Search name: MAT number (or extract from Product name)
    - Product name: e.g., "8 1/2:GT65DH:1234567" - used to extract MAT if Search name empty
    - Physical inventory: Current stock quantity
    - Ordered in total: Quantity on order
    """

    template_name = "technology/design_stock_import.html"

    # Header names for auto-detection (case-insensitive)
    # Supports both Items file and On-hand file formats
    HEADER_PATTERNS = {
        'item_number': ['item number'],
        'product_name': ['product name'],
        'search_name': ['search name', 'color'],  # Color column in On-hand file contains MAT
        'physical_inventory': ['physical inventory', 'available physical'],  # On-hand uses "Available physical"
        'ordered_total': ['ordered in total'],
        'warehouse': ['warehouse'],
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Import Design Stock from Excel"

        # Get available Excel files in docs folder
        import os
        docs_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'docs')
        excel_files = []
        if os.path.exists(docs_path):
            for f in os.listdir(docs_path):
                if f.endswith('.xlsx'):
                    excel_files.append(f)
        context["excel_files"] = sorted(excel_files)

        # Stats
        context["total_designs"] = Design.objects.count()
        context["designs_with_stock"] = Design.objects.filter(quantity_on_hand__gt=0).count()

        return context

    def _find_header_row(self, ws):
        """Auto-detect header row by searching for 'Item number' column."""
        max_search_rows = 20

        for row_idx in range(1, min(max_search_rows + 1, ws.max_row + 1)):
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if cell_str in self.HEADER_PATTERNS['item_number']:
                        column_map = self._map_columns(ws, row_idx)
                        return row_idx, column_map

        return None, None

    def _map_columns(self, ws, header_row):
        """Map column names to their positions."""
        column_map = {}

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if not cell_value:
                continue

            cell_str = str(cell_value).strip().lower()

            for field_name, patterns in self.HEADER_PATTERNS.items():
                if cell_str in patterns or any(p in cell_str for p in patterns):
                    if field_name not in column_map:
                        column_map[field_name] = col_idx
                        break

        return column_map

    def _parse_product_name(self, product_name):
        """
        Parse product name to extract size, HDBS type, and MAT number.
        Format examples:
        - "8 1/2:GT65DH:1234567" -> ("8 1/2", "GT65DH", "1234567")
        - "6 1/8 : MMD54H : 1145821" -> ("6 1/8", "MMD54H", "1145821")
        """
        if not product_name:
            return None, None, None

        product_name = str(product_name).strip()

        # Try colon separator first (primary)
        if ':' in product_name:
            parts = [p.strip() for p in product_name.split(':')]
            if len(parts) >= 3:
                return parts[0], parts[1], parts[2]
            elif len(parts) == 2:
                return parts[0], parts[1], None

        # Fallback to space separator
        parts = product_name.split()
        if len(parts) >= 3:
            # Try to identify MAT number (usually all digits, 6-7 chars)
            for i, part in enumerate(parts):
                if part.isdigit() and len(part) >= 6:
                    # Found MAT, HDBS type should be before it
                    size_parts = parts[:i-1] if i > 1 else []
                    hdbs_type = parts[i-1] if i > 0 else None
                    return ' '.join(size_parts), hdbs_type, part

        return None, None, None

    def post(self, request, *args, **kwargs):
        """Handle file upload, preview, and import."""
        import os
        from collections import defaultdict
        from django.db import transaction
        from django.utils import timezone

        try:
            import openpyxl
        except ImportError:
            return JsonResponse({'success': False, 'error': 'openpyxl is required'}, status=500)

        action = request.POST.get('action', 'preview')

        # Determine file source
        uploaded_file = request.FILES.get('excel_file')
        selected_file = request.POST.get('selected_file', '')

        if uploaded_file:
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            file_name = uploaded_file.name
        elif selected_file:
            docs_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'docs')
            file_path = os.path.join(docs_path, selected_file)
            if not os.path.exists(file_path):
                return JsonResponse({'success': False, 'error': f'File not found: {selected_file}'}, status=400)
            wb = openpyxl.load_workbook(file_path, data_only=True)
            file_name = selected_file
        else:
            return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

        ws = wb.active
        sheet_name = ws.title

        header_row, column_map = self._find_header_row(ws)

        if header_row is None:
            return JsonResponse({
                'success': False,
                'error': 'Could not find header row. Looking for "Item number" column in first 20 rows.'
            }, status=400)

        # Check for required columns - need either search_name or product_name for MAT
        has_mat_source = 'search_name' in column_map or 'product_name' in column_map
        has_qty_source = 'physical_inventory' in column_map

        if not has_mat_source:
            return JsonResponse({
                'success': False,
                'error': f'Missing MAT number source. Need "Search name" or "Product name" column. Found columns: {list(column_map.keys())}'
            }, status=400)

        if not has_qty_source:
            return JsonResponse({
                'success': False,
                'error': f'Missing quantity column. Need "Physical inventory" column. Found columns: {list(column_map.keys())}'
            }, status=400)

        data_start_row = header_row + 1

        # Build MAT -> Design mapping
        mat_to_design = {}
        for design in Design.objects.all():
            if design.mat_no:
                mat_to_design[str(design.mat_no).strip()] = design

        stats = {
            'rows_processed': 0,
            'rm_rows': 0,
            'mats_matched': 0,
            'mats_not_found': 0,
            'total_on_hand': 0,
            'total_on_order': 0,
            'skipped_not_rm': 0,
            'skipped_no_mat': 0,
        }
        not_found = []
        preview_data = []

        # Process rows
        for row_idx in range(data_start_row, ws.max_row + 1):
            item_number = ws.cell(row=row_idx, column=column_map['item_number']).value
            if not item_number:
                continue

            item_number = str(item_number).strip()
            stats['rows_processed'] += 1

            # Only process RM items
            if not item_number.upper().startswith('RM'):
                stats['skipped_not_rm'] += 1
                continue

            stats['rm_rows'] += 1

            # Check warehouse - ONLY process 'Store' rows
            warehouse = ''
            if 'warehouse' in column_map:
                warehouse_raw = ws.cell(row=row_idx, column=column_map['warehouse']).value
                warehouse = str(warehouse_raw or '').strip()

            # Skip non-Store rows entirely
            if warehouse.lower() != 'store':
                continue

            # Get MAT number - try Search name first, then parse from Product name
            mat_no = None
            hdbs_type = None

            if 'search_name' in column_map:
                search_name = ws.cell(row=row_idx, column=column_map['search_name']).value
                if search_name:
                    mat_no = str(search_name).strip()

            if not mat_no and 'product_name' in column_map:
                product_name = ws.cell(row=row_idx, column=column_map['product_name']).value
                if product_name:
                    _, hdbs_type, mat_no = self._parse_product_name(product_name)

            if not mat_no:
                stats['skipped_no_mat'] += 1
                continue

            # Get quantity from Physical inventory
            qty_on_hand = 0
            qty_on_order = 0

            if 'physical_inventory' in column_map:
                qty_raw = ws.cell(row=row_idx, column=column_map['physical_inventory']).value
                try:
                    qty_on_hand = int(float(str(qty_raw or 0)))
                except (ValueError, TypeError):
                    qty_on_hand = 0

            # Match to design
            design = mat_to_design.get(mat_no)
            if not design:
                stats['mats_not_found'] += 1
                not_found.append({
                    'mat_no': mat_no,
                    'hdbs_type': hdbs_type or 'N/A',
                    'erp_item': item_number,
                    'qty_on_hand': qty_on_hand,
                    'qty_on_order': qty_on_order,
                })
                continue

            stats['mats_matched'] += 1
            stats['total_on_hand'] += qty_on_hand
            stats['total_on_order'] += qty_on_order

            preview_data.append({
                'mat_no': mat_no,
                'hdbs_type': design.hdbs_type,
                'size': str(design.size) if design.size else '',
                'category': design.category,
                'current_on_hand': design.quantity_on_hand,
                'current_on_order': design.quantity_on_order,
                'new_on_hand': qty_on_hand,
                'new_on_order': qty_on_order,
                'erp_item': item_number,
                'matched': True,
            })

        if action == 'preview':
            return JsonResponse({
                'success': True,
                'file_name': file_name,
                'sheet_name': sheet_name,
                'header_row': header_row,
                'column_map': {k: v for k, v in column_map.items()},
                'stats': stats,
                'preview': preview_data[:100],
                'not_found': not_found[:30],
            })

        elif action == 'import':
            import_stats = {
                'designs_updated': 0,
                'total_on_hand': 0,
                'total_on_order': 0,
            }

            with transaction.atomic():
                for row_data in preview_data:
                    design = mat_to_design.get(row_data['mat_no'])
                    if not design:
                        continue

                    design.quantity_on_hand = row_data['new_on_hand']
                    design.quantity_on_order = row_data['new_on_order']
                    design.stock_last_updated = timezone.now()
                    design.save(update_fields=['quantity_on_hand', 'quantity_on_order', 'stock_last_updated'])

                    import_stats['designs_updated'] += 1
                    import_stats['total_on_hand'] += row_data['new_on_hand']
                    import_stats['total_on_order'] += row_data['new_on_order']

            return JsonResponse({
                'success': True,
                'stats': {
                    'mats_matched': stats['mats_matched'],
                    'designs_updated': import_stats['designs_updated'],
                    'total_on_hand': import_stats['total_on_hand'],
                    'total_on_order': import_stats['total_on_order'],
                },
            })

        return JsonResponse({'success': False, 'error': 'Invalid action'}, status=400)


# =============================================================================
# BOM LIST API ENDPOINTS
# =============================================================================


@login_required
def api_boms_list(request):
    """Return BOMs for a given design_id as JSON. Used by WO create form BOM selector."""
    design_id = request.GET.get('design_id')
    if not design_id:
        return JsonResponse({'boms': []})
    boms = BOM.objects.filter(design_id=design_id).order_by('-status', 'code')
    result = []
    for b in boms:
        result.append({
            'id': b.pk,
            'code': b.code or b.system_mat_no or str(b),
            'system_mat': b.system_mat_no or '',
            'status': b.status,
            'status_display': b.get_status_display(),
        })
    return JsonResponse({'boms': result})


class APIBOMUpdateFieldView(LoginRequiredMixin, View):
    """Inline update of BOM fields (smi_type, status) or Design fields (iadc_code_ref)."""

    def post(self, request, pk):
        import json
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

        bom = get_object_or_404(BOM, pk=pk)
        field = data.get('field')
        value = data.get('value')

        if field == 'smi_type':
            if value:
                smi = get_object_or_404(SMIType, pk=value)
                bom.smi_type = smi
            else:
                bom.smi_type = None
            bom.save(update_fields=['smi_type', 'updated_at'])
            return JsonResponse({'success': True, 'display': bom.smi_type.smi_name if bom.smi_type else '-'})

        elif field == 'status':
            if value not in dict(BOM.Status.choices):
                return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
            bom.status = value
            bom.save(update_fields=['status', 'updated_at'])
            return JsonResponse({'success': True, 'display': bom.get_status_display()})

        elif field == 'iadc':
            design = bom.design
            if value:
                iadc = get_object_or_404(IADCCode, pk=value)
                design.iadc_code_ref = iadc
            else:
                design.iadc_code_ref = None
            design.save(update_fields=['iadc_code_ref'])
            return JsonResponse({'success': True, 'display': design.iadc_code_ref.code if design.iadc_code_ref else '-'})

        return JsonResponse({'success': False, 'error': 'Unknown field'}, status=400)


class APIBOMMaterialsDetailView(LoginRequiredMixin, View):
    """Return detailed materials inventory for a BOM, including variant stock breakdown."""

    # Variant codes and which count as "new" for shortage calculation
    VARIANT_CODES = ["NEW-PUR", "NEW-RET", "NEW-EO", "USED-GRD", "USED-RCL", "NEW-CLI", "USED-CLI"]
    NEW_VARIANT_CODES = {"NEW-PUR", "NEW-EO", "NEW-RET"}

    def get(self, request, pk):
        from apps.inventory.models import VariantStock
        from django.db.models import Sum, DecimalField, Value
        from django.db.models.functions import Coalesce
        from decimal import Decimal

        bom = get_object_or_404(
            BOM.objects.select_related('design', 'design__size'),
            pk=pk
        )
        lines = bom.lines.select_related(
            'inventory_item', 'inventory_item__category'
        ).prefetch_related(
            'inventory_item__variants__variant_case',
            'inventory_item__variants__stock_records',
        ).order_by('line_number')

        items = []
        min_builds = None

        for line in lines:
            item = line.inventory_item
            required = int(line.quantity)

            if item:
                # Get variant stock breakdown (same pattern as cutter inventory page)
                variants = {}
                total_stock = 0
                total_new = 0
                for variant in item.variants.all():
                    if variant.variant_case and variant.variant_case.code in self.VARIANT_CODES:
                        stock = variant.stock_records.aggregate(
                            total=Coalesce(Sum("quantity_on_hand"), Value(0), output_field=DecimalField())
                        )["total"]
                        qty = int(stock)
                        code = variant.variant_case.code
                        variants[code] = qty
                        total_stock += qty
                        if code in self.NEW_VARIANT_CODES:
                            total_new += qty

                # Shortage is based on new variants only (ENO New, Retrofit, New Stock)
                available_new = total_new
                shortage = max(0, required - available_new)
                possible = available_new // required if required > 0 else 0
                if min_builds is None or possible < min_builds:
                    min_builds = possible

                items.append({
                    'line_number': line.line_number,
                    'order_number': line.order_number,
                    'mat_number': item.mat_number or item.code,
                    'description': str(item),
                    'color_code': line.color_code or '',
                    'cutter_size': line.cutter_size or '',
                    'cutter_type': line.cutter_type or '',
                    'chamfer': line.cutter_chamfer or '',
                    'required': required,
                    'available_new': available_new,
                    'total_stock': total_stock,
                    'shortage': shortage,
                    'possible_builds': possible,
                    'variants': variants,
                })
            else:
                items.append({
                    'line_number': line.line_number,
                    'order_number': line.order_number,
                    'mat_number': line.hdbs_code or f'Line {line.line_number}',
                    'description': 'No inventory item linked',
                    'color_code': line.color_code or '',
                    'cutter_size': line.cutter_size or '',
                    'cutter_type': line.cutter_type or '',
                    'chamfer': line.cutter_chamfer or '',
                    'required': required,
                    'available_new': 0,
                    'total_stock': 0,
                    'shortage': required,
                    'possible_builds': 0,
                    'variants': {},
                })
                if min_builds is None or 0 < min_builds:
                    min_builds = 0

        has_shortage = any(i['shortage'] > 0 for i in items)

        return JsonResponse({
            'success': True,
            'bom_code': bom.code,
            'design_mat': bom.design.mat_no or '-',
            'items': items,
            'total_lines': len(items),
            'has_shortage': has_shortage,
            'possible_builds': min_builds if min_builds is not None else 0,
            'variant_codes': self.VARIANT_CODES,
            'new_variant_codes': list(self.NEW_VARIANT_CODES),
        })
