"""
ARDT FMS - Reports Views
Version: 5.4
Comprehensive reporting suite with Excel export.
"""

import io
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Avg, Count, F, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import ListView, TemplateView, View

from apps.inventory.models import InventoryItem, InventoryStock, InventoryTransaction
from apps.maintenance.models import Equipment, MaintenanceRequest, MaintenanceWorkOrder
from apps.quality.models import Inspection, NCR
from apps.supplychain.models import CAPA, PurchaseOrder, PurchaseRequisition, Receipt, Supplier, Vendor
from apps.workorders.models import DrillBit, WorkOrder

from .models import ReportExportLog


# =============================================================================
# Excel Export Mixin
# =============================================================================


class ExcelExportMixin:
    """Mixin to add Excel export capability to views."""

    def export_to_excel(self, queryset, columns, filename, sheet_name="Report"):
        """
        Export queryset to Excel file.

        Args:
            queryset: Django queryset to export
            columns: List of tuples (field_name, header_name)
            filename: Name of the file to download
            sheet_name: Name of the Excel sheet
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("openpyxl is not installed. Please install it with: pip install openpyxl", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (field, header) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, (field, header) in enumerate(columns, 1):
                # Handle nested fields (e.g., "customer.name")
                value = obj
                for part in field.split("."):
                    if value is None:
                        break
                    if hasattr(value, part):
                        value = getattr(value, part)
                    elif hasattr(value, "get"):
                        value = value.get(part)
                    else:
                        value = None

                # Handle callable (properties)
                if callable(value):
                    value = value()

                # Format specific types
                if hasattr(value, "strftime"):
                    value = value.strftime("%Y-%m-%d %H:%M") if hasattr(value, "hour") else value.strftime("%Y-%m-%d")

                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                cell.border = thin_border

        # Auto-adjust column widths
        for col_idx, (field, header) in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(header)
            for row in range(2, min(ws.max_row + 1, 100)):  # Check first 100 rows
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'

        # Log export
        ReportExportLog.objects.create(
            report_type=sheet_name,
            export_format="EXCEL",
            record_count=queryset.count() if hasattr(queryset, "count") else len(queryset),
            exported_by=self.request.user if hasattr(self, "request") else None,
        )

        return response


# =============================================================================
# Reports Dashboard
# =============================================================================


class ReportsDashboardView(LoginRequiredMixin, TemplateView):
    """Main reports dashboard with links to all reports."""

    template_name = "reports/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Reports & Analytics"

        # Quick stats
        today = timezone.now().date()
        month_start = today.replace(day=1)

        context["stats"] = {
            "work_orders_this_month": WorkOrder.objects.filter(created_at__date__gte=month_start).count(),
            "active_work_orders": WorkOrder.objects.filter(status__in=["PLANNED", "IN_PROGRESS"]).count(),
            "low_stock_items": InventoryStock.objects.filter(quantity_on_hand__lte=F("item__reorder_point")).count(),
            "open_ncrs": NCR.objects.filter(status__in=["OPEN", "IN_PROGRESS"]).count(),
            "pending_maintenance": MaintenanceRequest.objects.filter(status="PENDING").count(),
            "pending_pos": PurchaseOrder.objects.filter(status="PENDING").count(),
        }

        return context


# =============================================================================
# Work Order Reports
# =============================================================================


class WorkOrderReportView(LoginRequiredMixin, ExcelExportMixin, ListView):
    """Work order report with filters and Excel export."""

    model = WorkOrder
    template_name = "reports/workorder_report.html"
    context_object_name = "work_orders"
    paginate_by = 50

    def get_queryset(self):
        qs = WorkOrder.objects.select_related("customer", "drill_bit", "assigned_to", "design").order_by("-created_at")

        # Date range filter
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Status filter
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        # Customer filter
        customer = self.request.GET.get("customer")
        if customer:
            qs = qs.filter(customer_id=customer)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Work Order Report"
        context["status_choices"] = WorkOrder.Status.choices

        # Summary stats
        qs = self.get_queryset()
        context["summary"] = {
            "total": qs.count(),
            "completed": qs.filter(status="COMPLETED").count(),
            "in_progress": qs.filter(status="IN_PROGRESS").count(),
            "overdue": qs.filter(due_date__lt=timezone.now().date(), status__in=["PLANNED", "IN_PROGRESS"]).count(),
        }

        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            columns = [
                ("wo_number", "WO Number"),
                ("customer.name", "Customer"),
                ("drill_bit.serial_number", "Bit Serial"),
                ("work_type", "Type"),
                ("status", "Status"),
                ("priority", "Priority"),
                ("assigned_to.username", "Assigned To"),
                ("created_at", "Created"),
                ("due_date", "Due Date"),
                ("completed_at", "Completed"),
            ]
            return self.export_to_excel(self.get_queryset(), columns, "work_order_report", "Work Orders")
        return super().get(request, *args, **kwargs)


# =============================================================================
# Inventory Reports
# =============================================================================


class InventoryReportView(LoginRequiredMixin, ExcelExportMixin, ListView):
    """Inventory stock level report with Excel export."""

    model = InventoryStock
    template_name = "reports/inventory_report.html"
    context_object_name = "stock_items"
    paginate_by = 50

    def get_queryset(self):
        qs = InventoryStock.objects.select_related("item", "item__category", "location").order_by("item__name")

        # Category filter
        category = self.request.GET.get("category")
        if category:
            qs = qs.filter(item__category_id=category)

        # Location filter
        location = self.request.GET.get("location")
        if location:
            qs = qs.filter(location_id=location)

        # Low stock filter
        low_stock = self.request.GET.get("low_stock")
        if low_stock == "true":
            qs = qs.filter(quantity_on_hand__lte=F("item__reorder_point"))

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Inventory Report"

        # Summary
        qs = self.get_queryset()
        context["summary"] = {
            "total_items": qs.count(),
            "low_stock": qs.filter(quantity_on_hand__lte=F("item__reorder_point")).count(),
            "out_of_stock": qs.filter(quantity_on_hand=0).count(),
            "total_value": qs.aggregate(
                total=Sum(F("quantity_on_hand") * F("item__standard_cost"), default=0)
            )["total"] or 0,
        }

        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            columns = [
                ("item.code", "Item Code"),
                ("item.name", "Item Name"),
                ("item.category.name", "Category"),
                ("location.name", "Location"),
                ("quantity_on_hand", "Qty On Hand"),
                ("quantity_reserved", "Qty Reserved"),
                ("quantity_available", "Qty Available"),
                ("item.reorder_point", "Reorder Point"),
                ("item.standard_cost", "Unit Cost"),
            ]
            return self.export_to_excel(self.get_queryset(), columns, "inventory_report", "Inventory")
        return super().get(request, *args, **kwargs)


class LowStockAlertView(LoginRequiredMixin, ExcelExportMixin, ListView):
    """Low stock alert report."""

    model = InventoryStock
    template_name = "reports/low_stock_alert.html"
    context_object_name = "alerts"

    def get_queryset(self):
        return (
            InventoryStock.objects.filter(quantity_on_hand__lte=F("item__reorder_point"))
            .select_related("item", "item__category", "location")
            .order_by("quantity_on_hand")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Low Stock Alerts"
        context["critical_count"] = self.get_queryset().filter(quantity_on_hand=0).count()
        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            columns = [
                ("item.code", "Item Code"),
                ("item.name", "Item Name"),
                ("location.name", "Location"),
                ("quantity_on_hand", "Current Stock"),
                ("item.reorder_point", "Reorder Point"),
                ("quantity_on_hand", "Shortage"),
            ]
            return self.export_to_excel(self.get_queryset(), columns, "low_stock_alerts", "Low Stock")
        return super().get(request, *args, **kwargs)


# =============================================================================
# Quality Reports
# =============================================================================


class QualityReportView(LoginRequiredMixin, ExcelExportMixin, ListView):
    """Quality metrics report (Inspections and NCRs)."""

    model = NCR
    template_name = "reports/quality_report.html"
    context_object_name = "ncrs"
    paginate_by = 50

    def get_queryset(self):
        qs = NCR.objects.select_related("work_order", "inspection", "reported_by").order_by("-created_at")

        # Date range
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Severity filter
        severity = self.request.GET.get("severity")
        if severity:
            qs = qs.filter(severity=severity)

        # Status filter
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Quality Report"
        context["severity_choices"] = NCR.Severity.choices
        context["status_choices"] = NCR.Status.choices

        # Summary
        qs = self.get_queryset()
        context["summary"] = {
            "total_ncrs": qs.count(),
            "open": qs.filter(status="OPEN").count(),
            "critical": qs.filter(severity="CRITICAL").count(),
            "closed_this_month": qs.filter(
                status="CLOSED",
                closed_at__month=timezone.now().month,
            ).count(),
        }

        # Inspection stats
        context["inspection_stats"] = {
            "total": Inspection.objects.count(),
            "passed": Inspection.objects.filter(result="PASS").count(),
            "failed": Inspection.objects.filter(result="FAIL").count(),
        }

        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            columns = [
                ("ncr_number", "NCR Number"),
                ("work_order.wo_number", "Work Order"),
                ("ncr_type", "Type"),
                ("severity", "Severity"),
                ("status", "Status"),
                ("description", "Description"),
                ("reported_by.username", "Reported By"),
                ("created_at", "Created"),
                ("closed_at", "Closed Date"),
            ]
            return self.export_to_excel(self.get_queryset(), columns, "quality_report", "NCRs")
        return super().get(request, *args, **kwargs)


# =============================================================================
# Maintenance Reports
# =============================================================================


class MaintenanceReportView(LoginRequiredMixin, ExcelExportMixin, ListView):
    """Maintenance report with equipment and MWO tracking."""

    model = MaintenanceWorkOrder
    template_name = "reports/maintenance_report.html"
    context_object_name = "mwos"
    paginate_by = 50

    def get_queryset(self):
        qs = MaintenanceWorkOrder.objects.select_related("equipment", "request", "assigned_to").order_by("-created_at")

        # Date range
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Status filter
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        # Work type filter
        work_type = self.request.GET.get("work_type")
        if work_type:
            qs = qs.filter(work_type=work_type)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Maintenance Report"
        context["status_choices"] = MaintenanceWorkOrder.Status.choices

        # Summary
        qs = self.get_queryset()
        context["summary"] = {
            "total_mwos": qs.count(),
            "completed": qs.filter(status="COMPLETED").count(),
            "in_progress": qs.filter(status="IN_PROGRESS").count(),
            "pending_requests": MaintenanceRequest.objects.filter(status="PENDING").count(),
        }

        # Equipment stats
        context["equipment_stats"] = {
            "total": Equipment.objects.count(),
            "operational": Equipment.objects.filter(status="OPERATIONAL").count(),
            "under_maintenance": Equipment.objects.filter(status="MAINTENANCE").count(),
        }

        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            columns = [
                ("mwo_number", "MWO Number"),
                ("equipment.code", "Equipment"),
                ("work_type", "Work Type"),
                ("status", "Status"),
                ("priority", "Priority"),
                ("assigned_to.username", "Assigned To"),
                ("scheduled_date", "Scheduled"),
                ("completed_date", "Completed"),
            ]
            return self.export_to_excel(self.get_queryset(), columns, "maintenance_report", "Maintenance")
        return super().get(request, *args, **kwargs)


class EquipmentHealthReportView(LoginRequiredMixin, ExcelExportMixin, ListView):
    """Equipment health scoring report."""

    model = Equipment
    template_name = "reports/equipment_health.html"
    context_object_name = "equipment_list"

    def get_queryset(self):
        return (
            Equipment.objects.select_related("category")
            .annotate(
                mwo_count=Count("work_orders"),
                completed_mwos=Count("work_orders", filter=Q(work_orders__status="COMPLETED")),
            )
            .order_by("-mwo_count")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Equipment Health Report"

        # Calculate health scores
        equipment_list = []
        for eq in self.get_queryset():
            # Simple health score: 100 - (total MWOs * 5), minimum 0
            health_score = max(0, 100 - (eq.mwo_count * 5))
            equipment_list.append(
                {
                    "equipment": eq,
                    "health_score": health_score,
                    "status_class": "text-green-600" if health_score >= 70 else ("text-yellow-600" if health_score >= 40 else "text-red-600"),
                }
            )

        context["equipment_health"] = equipment_list
        return context


# =============================================================================
# Supply Chain Reports
# =============================================================================


class SupplyChainReportView(LoginRequiredMixin, ExcelExportMixin, ListView):
    """Supply chain report with PO/PR/GRN tracking."""

    model = PurchaseOrder
    template_name = "reports/supplychain_report.html"
    context_object_name = "purchase_orders"
    paginate_by = 50

    def get_queryset(self):
        qs = PurchaseOrder.objects.select_related("vendor", "created_by").order_by("-created_at")

        # Date range
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Status filter
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        # Supplier filter
        supplier = self.request.GET.get("supplier")
        if supplier:
            qs = qs.filter(supplier_id=supplier)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Supply Chain Report"
        context["status_choices"] = PurchaseOrder.Status.choices
        context["suppliers"] = Supplier.objects.filter(is_active=True).order_by("name")

        # Summary
        qs = self.get_queryset()
        context["summary"] = {
            "total_pos": qs.count(),
            "pending": qs.filter(status="PENDING").count(),
            "total_value": qs.aggregate(total=Sum("total_amount"))["total"] or 0,
            "pending_prs": PurchaseRequisition.objects.filter(status="PENDING").count(),
        }

        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            columns = [
                ("po_number", "PO Number"),
                ("vendor.name", "Supplier"),
                ("status", "Status"),
                ("order_date", "Order Date"),
                ("expected_date", "Expected Date"),
                ("total_amount", "Total Amount"),
                ("created_by.username", "Created By"),
            ]
            return self.export_to_excel(self.get_queryset(), columns, "supply_chain_report", "Purchase Orders")
        return super().get(request, *args, **kwargs)
