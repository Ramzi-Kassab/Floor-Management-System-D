"""
Job Card Excel Parser

Parses individual Job Card Excel files (one per work order) and extracts
all data needed for ERP item creation, including cutter BOM with variant
breakdown, evaluation summaries, and ERP item number lookup.

Job Card Structure (29 sheets):
  - "Data" sheet: Core fields, status flags, cutter BOM, modified cutters
  - "Evaluation" / "Eval-LSTK": Evaluation matrices (6-blade grids)
  - "Engineer Eval": Engineer/Tech Rep evaluation
  - "QC Evaluation": Quality Control evaluation
  - "Die Check" / "Final Die Check": Die check records
  - "Final QC" / "Final Inspection": Final quality checks

Data Sheet Layout:
  B1-B25:  WO header (WO#, serial, size, type, L5 MAT, date, account, etc.)
  D2-G10:  Modified cutters table (quick cutter swaps)
  D12-M28: Cutter BOM table (group, qty, size, part#, desc, new/rclm/comment)
  N12-O28: Original BOM reference (cutters as BOM, qty as BOM)
  I2-I9:   Decision flags (rerun, inspection, scrap, build-up checkboxes)
  L2-L5:   Decision flag values (1 = checked)
  D30+:    Special process fields (USR, pin size, cerebro, salvage)
           NOTE: Row positions are NOT fixed — they shift when BOM has more lines.
           Labels are searched dynamically by text content.
"""
import os
import re
import logging
from decimal import Decimal, InvalidOperation

import openpyxl

logger = logging.getLogger(__name__)


# =============================================================================
# CONSTANTS
# =============================================================================

ACCOUNT_TO_ITEM_GROUP = {
    'LSTK': 'RPR-FC-LST',
    'ARAMCO': 'RPR-FC-AR',
    'SAUDI ARAMCO': 'RPR-FC-AR',
    'HALLIBURTON': 'RPR-FC-HDB',
    'HDBS': 'RPR-FC-HDB',
    'HAL_REGIONAL': 'RPR-FC-REG',
    'REGIONAL': 'RPR-FC-REG',
    'ARDT': 'RPR-ARDT',
    'RC-LSTK': 'RPR-RC-LST',
    'RC-ARAMCO': 'RPR-RC-AR',
    'RC-REGIONAL': 'RPR-RC-REG',
    # These don't get items:
    # UR, WFD, SUB → movement journal only
    # L3, L4 → handled by other team
}

# Normalize account names (variations → canonical form)
ACCOUNT_ALIASES = {
    'SAUDI ARAMCO': 'ARAMCO',
    'SA': 'ARAMCO',
    'HAL': 'HALLIBURTON',
    'HDB': 'HALLIBURTON',
    'HDBS': 'HALLIBURTON',
    'REGIONAL': 'HAL_REGIONAL',
    'HAL REGIONAL': 'HAL_REGIONAL',
}

# Column M remark patterns → variant case codes
# Longest match first to avoid partial matches (e.g., "ARDT RCLM" before "RCLM")
REMARK_TO_VARIANT = {
    'ARDT RCLM': 'USED-RCL',     # ARDT Reclaim → RCLM-ARDT-*
    'ARDT RCL': 'USED-RCL',
    'ENO GRD': 'NEW-EO',         # ENO Ground — treat as ENO (not yet in ERP)
    'GROUND': 'NEW-EO',          # Ground — treat as ENO
    'GRD': 'NEW-EO',             # Ground — treat as ENO
    'ENO': 'NEW-EO',             # ENO As New → ENO-CT-*
    'RCLM': 'CLI-RCL',           # Plain RCLM = LSTK Reclaim (Client Reclaimed)
    'RTRO': 'NEW-RET',           # Retrofit → RTRO-*
    'RTR': 'NEW-RET',
    'RETROFIT': 'NEW-RET',
    'NEW': 'NEW-PUR',            # Explicit new
}


# =============================================================================
# DERIVATION FUNCTIONS
# =============================================================================

def parse_size_to_inches(raw_size):
    """Convert size string to Decimal inches.

    Handles:
      - Numeric: 6.125 → 6.125
      - Fraction string: '3 3/4"' → 3.75
      - Whole string: '12"' → 12
      - Combined: '12 1/4"' → 12.25
      - Slash only: '3/4"' → 0.75
    """
    if raw_size is None:
        return None

    if isinstance(raw_size, (int, float)):
        return Decimal(str(raw_size))

    s = str(raw_size).strip().strip('"').strip("'").strip('"').strip()
    if not s:
        return None

    try:
        return Decimal(s)
    except InvalidOperation:
        pass

    # Try "whole fraction" pattern: "3 3/4", "12 1/4"
    m = re.match(r'^(\d+)\s+(\d+)/(\d+)$', s)
    if m:
        whole = int(m.group(1))
        num = int(m.group(2))
        den = int(m.group(3))
        if den > 0:
            return Decimal(str(whole)) + Decimal(str(num)) / Decimal(str(den))

    # Try fraction only: "3/4"
    m = re.match(r'^(\d+)/(\d+)$', s)
    if m:
        num = int(m.group(1))
        den = int(m.group(2))
        if den > 0:
            return Decimal(str(num)) / Decimal(str(den))

    # Try whole number only: "12"
    m = re.match(r'^(\d+)$', s)
    if m:
        return Decimal(m.group(1))

    # Try "N x N" pattern (e.g., "8 1/2 x 3 1/4" — take first part)
    parts = re.split(r'\s*[xX×]\s*', s)
    if len(parts) > 1:
        return parse_size_to_inches(parts[0])

    logger.warning(f"Could not parse size: {repr(raw_size)}")
    return None


def derive_body_material(smi_type):
    """Determine body material from SMI Type.

    If type starts or ends with 's' (case-insensitive) → SB (Steel Body)
    Otherwise → MB (Matrix Body)
    """
    t = str(smi_type or '').strip()
    if t and (t[0].lower() == 's' or t[-1].lower() == 's'):
        return 'SB'
    return 'MB'


def derive_crush_shear(smi_type):
    """Detect Crush & Shear from SMI Type.

    True if type starts or ends with 'CS' (case-insensitive).
    """
    t = str(smi_type or '').strip().upper()
    if not t:
        return False
    return t.startswith('CS') or t.endswith('CS')


def derive_original_l5_mat(l5_mat_full):
    """Strip M and everything after it from L5 MAT.

    Examples:
      '1224750M'  → '1224750'
      '1224750M1' → '1224750'
      '1251085M2' → '1251085'
      '1224750'   → '1224750' (no M, unchanged)
    """
    return re.split(r'[Mm]', str(l5_mat_full or '').strip())[0]


# =============================================================================
# LABEL-BASED CELL SEARCH (for shifted layouts)
# =============================================================================

def _find_label_value(ws, label_text, search_col_range=range(3, 10),
                      search_row_range=range(28, 45), value_offset='below'):
    """Find a label in the worksheet and return the value from an adjacent cell.

    Job Card layouts can shift when the BOM table has more/fewer rows. Instead
    of hardcoding cell references, search for label text and read the value
    from the cell below (or to the right of) the label.

    Args:
        ws: openpyxl worksheet
        label_text: Text to search for (case-insensitive substring match)
        search_col_range: Column range to search (default C-I)
        search_row_range: Row range to search (default 28-44)
        value_offset: 'below' = cell one row below the label (same column)
                      'right' = cell one column to the right (same row)

    Returns:
        Value from the adjacent cell, or None if label not found
    """
    label_lower = label_text.lower()
    for row in search_row_range:
        for col in search_col_range:
            cell_val = ws.cell(row=row, column=col).value
            if cell_val and label_lower in str(cell_val).lower():
                if value_offset == 'below':
                    return ws.cell(row=row + 1, column=col).value
                elif value_offset == 'right':
                    return ws.cell(row=row, column=col + 1).value
                else:
                    return ws.cell(row=row + 1, column=col).value
    return None


def _is_checkbox_checked(value):
    """Check if a cell value represents a checked checkbox.

    Returns True only for explicit truthy values: 1, 'yes', 'true', 'x'.
    Returns False for label text, None, 0, empty strings, etc.
    """
    if value is None:
        return False
    v = str(value).strip()
    if not v:
        return False
    # Numeric check
    try:
        return int(float(v)) == 1
    except (ValueError, TypeError):
        pass
    return v.lower() in ('yes', 'true', 'x')


# =============================================================================
# CUTTER VARIANT PARSING
# =============================================================================

def parse_cutter_variants(new_qty, reclaim_qty, additional_comment, part_no):
    """Parse columns K, L, M into a list of variant allocations.

    RULE: If column M has bracketed remarks like [1 ARDT RCLM][2 RCLM],
    it is the COMPLETE authoritative breakdown — columns K and L are IGNORED.
    Only if M is empty, fall back to K (New) and L (LSTK Reclaim).

    Returns list of dicts:
      [{'variant_case': 'NEW-PUR', 'qty': 1, 'part_no': '717748'}, ...]
    """
    variants = []
    comment = str(additional_comment or '').strip()

    # Parse bracketed remarks — supports BOTH formats:
    #   [RTRO 18], [NEW 5], [ ENO 29]   → text first, qty last
    #   [18 RTRO], [5 NEW], [29 ENO]    → qty first, text last
    # Try text-first format first (more common in ARAMCO cards)
    remark_items_text_first = re.findall(r'\[\s*([^\]\d]+?)\s+(\d+)\s*\]', comment)
    remark_items_qty_first = re.findall(r'\[\s*(\d+)\s+([^\]]+?)\s*\]', comment)

    # Use whichever format matched (normalize to [(qty, type), ...])
    if remark_items_text_first:
        remark_items = [(qty, rtype) for rtype, qty in remark_items_text_first]
    elif remark_items_qty_first:
        remark_items = remark_items_qty_first  # Already (qty, type) order
    else:
        remark_items = []

    if remark_items:
        # --- Column M is authoritative: ignore K and L ---
        for qty_str, remark_type in remark_items:
            qty = int(qty_str)
            remark_upper = remark_type.strip().upper()
            # Match against known variant types (longest match first)
            matched = False
            for key in sorted(REMARK_TO_VARIANT.keys(), key=len, reverse=True):
                if key in remark_upper:
                    variants.append({
                        'variant_case': REMARK_TO_VARIANT[key],
                        'qty': qty,
                        'part_no': part_no,
                    })
                    matched = True
                    break
            if not matched:
                variants.append({
                    'variant_case': 'UNKNOWN',
                    'qty': qty,
                    'part_no': part_no,
                    'remark': remark_type.strip(),
                })
    else:
        # --- Legacy fallback: use K and L ---
        # Column L might contain bracket remarks like "[RTRO 18]" in some
        # job cards — if so, parse them as remark_items (authoritative,
        # overrides column K just like column M brackets do).
        rq_str = str(reclaim_qty or '').strip()
        if '[' in rq_str:
            # Column L has bracket notation — authoritative, ignore K
            # Support both [TEXT QTY] and [QTY TEXT] formats
            tf = re.findall(r'\[\s*([^\]\d]+?)\s+(\d+)\s*\]', rq_str)
            qf = re.findall(r'\[\s*(\d+)\s+([^\]]+?)\s*\]', rq_str)
            if tf:
                extra_remarks = [(qty, rt) for rt, qty in tf]
            elif qf:
                extra_remarks = qf
            else:
                extra_remarks = []
            for qty_str, remark_type in extra_remarks:
                qty = int(qty_str)
                remark_upper = remark_type.strip().upper()
                matched = False
                for key in sorted(REMARK_TO_VARIANT.keys(), key=len, reverse=True):
                    if key in remark_upper:
                        variants.append({
                            'variant_case': REMARK_TO_VARIANT[key],
                            'qty': qty,
                            'part_no': part_no,
                        })
                        matched = True
                        break
                if not matched:
                    variants.append({
                        'variant_case': 'UNKNOWN',
                        'qty': qty,
                        'part_no': part_no,
                        'remark': remark_type.strip(),
                    })
        else:
            # Plain numeric fallback for K and L
            try:
                nq = int(new_qty) if new_qty and str(new_qty).strip() else 0
            except (ValueError, TypeError):
                nq = 0
            if nq > 0:
                variants.append({
                    'variant_case': 'NEW-PUR',
                    'qty': nq,
                    'part_no': part_no,
                })

            try:
                rq = int(reclaim_qty) if reclaim_qty and rq_str else 0
            except (ValueError, TypeError):
                rq = 0
            if rq > 0:
                variants.append({
                    'variant_case': 'CLI-RCL',  # Column L = LSTK Reclaim (Client Reclaimed)
                    'qty': rq,
                    'part_no': part_no,
                })

    return variants


def lookup_erp_item_number(part_no, variant_case_code):
    """Find the ERP item number for a cutter by HDBS MAT # and variant case.

    Queries InventoryItem by mat_number, then finds the matching ItemVariant.
    Returns the erp_item_no string, or None if not found.
    """
    try:
        from apps.inventory.models import ItemVariant
        variant = ItemVariant.objects.select_related(
            'base_item', 'variant_case'
        ).filter(
            base_item__mat_number=str(part_no),
            variant_case__code=variant_case_code,
            is_active=True,
        ).first()
        if variant:
            return variant.erp_item_no or None
        return None
    except Exception as e:
        logger.warning(f"Error looking up ERP item for {part_no}/{variant_case_code}: {e}")
        return None


# =============================================================================
# EVALUATION EXTRACTION
# =============================================================================

# Known evaluation sheet names and their types
EVAL_SHEET_MAP = {
    'Evaluation': 'RECEIVING',
    'Eval-LSTK': 'ARDT',
    'Eval & Quot-AR': 'ARDT',       # ARAMCO evaluation sheet
    'Engineer Eval': 'ENGINEER',
    'QC Evaluation': 'QC',
    'Die Check': 'DIE_CHECK',
    'Final Die Check': 'FINAL_DIE_CHECK',
    'Final QC': 'FINAL_QC',
    'Final Inspection': 'FINAL_INSPECTION',
    'Rework': 'REWORK',             # Rework evaluation
}

# Valid cutter evaluation action codes (from the Job Card system)
VALID_ACTIONS = {'O', 'X', 'R', 'S', 'F', 'L', 'P', 'V', 'I'}


def _extract_eval_grid(ws, start_row=3, end_row=33, start_col=3, end_col=14):
    """Extract evaluation grid data from an evaluation sheet.

    The evaluation grid is a 6-blade matrix where each blade has positions
    from Gauge to Cone. The grid layout varies by sheet but generally:
    - Rows = position (Gauge, Shoulder, Nose, Cone) per blade
    - Columns = blade number (1-6), with sub-columns for each position

    This function extracts a summary: total actions by type.

    Args:
        ws: openpyxl worksheet
        start_row: First row of grid data
        end_row: Last row of grid data
        start_col: First column (C=3)
        end_col: Last column (N=14)

    Returns:
        dict with action_counts and total_positions
    """
    action_counts = {}
    total_positions = 0
    filled_positions = 0

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue

            cell_str = str(val).strip().upper()
            if not cell_str:
                continue

            # Valid action codes are single letters or letter+digit (e.g. "O", "X2", "R1")
            # Skip multi-word text like headers ("Internal Evaluation Sheet")
            # Max valid length: 2 chars (letter + optional digit)
            if len(cell_str) > 3:
                continue

            total_positions += 1

            # Extract the action code (first letter), ignoring suffix numbers
            # e.g., "O1" -> "O", "X2" -> "X", "R" -> "R"
            action_code = cell_str[0] if cell_str else ''
            if action_code in VALID_ACTIONS:
                action_counts[action_code] = action_counts.get(action_code, 0) + 1
                filled_positions += 1

    return {
        'action_counts': action_counts,
        'total_positions': total_positions,
        'filled_positions': filled_positions,
    }


def _extract_eval_decision(ws):
    """Extract the evaluation decision from common cell locations.

    Different eval sheets store the decision (REPAIR, RERUN, SCRAP, etc.)
    in different cells. We check multiple known locations.

    Returns:
        str: Decision text or empty string
    """
    # Common locations for decision field
    decision_cells = ['B35', 'B36', 'D35', 'D36', 'A35', 'A36']
    for cell_ref in decision_cells:
        try:
            val = ws[cell_ref].value
            if val:
                text = str(val).strip()
                # Check if it looks like a decision keyword
                for keyword in ('REPAIR', 'RERUN', 'RE-RUN', 'SCRAP', 'DEBRAZE',
                                'RETROFIT', 'NEW BUILD', 'BODY RETROFIT',
                                'CUTTER RETROFIT'):
                    if keyword in text.upper():
                        return keyword.replace('RE-RUN', 'RERUN')
        except Exception:
            pass
    return ''


def _extract_eval_remarks(ws):
    """Extract remarks/notes from common cell locations.

    Returns:
        str: Remarks text or empty string
    """
    remark_cells = ['D36', 'D37', 'D38', 'B37', 'B38']
    remarks = []
    for cell_ref in remark_cells:
        try:
            val = ws[cell_ref].value
            if val:
                text = str(val).strip()
                if text and len(text) > 2:  # Skip single chars/numbers
                    remarks.append(text)
        except Exception:
            pass
    return '; '.join(remarks) if remarks else ''


def _extract_evaluations(wb):
    """Extract evaluation summaries from all evaluation sheets in a Job Card.

    Scans known evaluation sheet names, extracts action grids, decisions,
    and remarks from each. Returns a list of evaluation summary dicts.

    Args:
        wb: openpyxl workbook

    Returns:
        list of dicts, each with:
          - eval_type: str (RECEIVING, ARDT, ENGINEER, QC, etc.)
          - sheet_name: str (actual sheet name in workbook)
          - action_counts: dict (e.g., {'O': 15, 'X': 5, 'R': 3})
          - total_positions: int
          - filled_positions: int
          - decision: str (e.g., 'REPAIR', 'SCRAP')
          - remarks: str
    """
    evaluations = []

    for sheet_name, eval_type in EVAL_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Extract the evaluation grid
        grid = _extract_eval_grid(ws)

        # Skip sheets with no data in the grid
        if grid['filled_positions'] == 0:
            continue

        # Extract decision and remarks
        decision = _extract_eval_decision(ws)
        remarks = _extract_eval_remarks(ws)

        evaluations.append({
            'eval_type': eval_type,
            'sheet_name': sheet_name,
            'action_counts': grid['action_counts'],
            'total_positions': grid['total_positions'],
            'filled_positions': grid['filled_positions'],
            'decision': decision,
            'remarks': remarks,
        })

        logger.debug(
            f"Extracted {eval_type} evaluation from '{sheet_name}': "
            f"{grid['filled_positions']} actions, decision={decision or 'none'}"
        )

    return evaluations


def _extract_eval_lstk_buildup_counts(wb):
    """Extract build-up counts specifically from Eval-LSTK sheet.

    The Eval-LSTK sheet has build-up totals at:
      R34 = Pocket Build Up count
      U34 = Fin Build Up count
      X34 = Impact Arrestor Build Up count

    Returns:
        dict with pocket_buildup, fin_buildup, ia_buildup counts
    """
    result = {'pocket_buildup': 0, 'fin_buildup': 0, 'ia_buildup': 0}
    if 'Eval-LSTK' not in wb.sheetnames:
        return result

    lstk_ws = wb['Eval-LSTK']
    cell_map = {
        'R34': 'pocket_buildup',
        'U34': 'fin_buildup',
        'X34': 'ia_buildup',
    }
    for cell_ref, key in cell_map.items():
        try:
            val = lstk_ws[cell_ref].value
            if val:
                result[key] = int(val)
        except (ValueError, TypeError):
            pass

    return result


# =============================================================================
# MAIN PARSER
# =============================================================================

def parse_job_card(file_path):
    """Parse a Job Card Excel file and return a dict for ERPJobData creation.

    Reads the 'Data' sheet for core fields, cutter BOM, modified cutters,
    job status flags, and USR. Also checks 'Evaluation' and 'Eval-LSTK'
    sheets for hardfacing/build-up detection.

    Args:
        file_path: Path to the Job Card .xlsx file

    Returns:
        dict with all fields needed to create an ERPJobData record
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Data']

    # --- Core fields (B column) ---
    wo_number = ws['B1'].value
    serial = ws['B2'].value
    raw_size = ws['B3'].value
    smi_type = str(ws['B4'].value or '').strip()
    l5_mat_full = str(ws['B5'].value or '').strip()
    date_received = ws['B6'].value
    account = str(ws['B7'].value or '').strip()
    l3_l4_mat = ws['B8'].value
    evaluated_by = ws['B11'].value
    reviewed_by = ws['B14'].value
    contract_number = ws['B24'].value
    vendor_number = ws['B25'].value

    # --- Special Process Fields (label-based search) ---
    # Labels can shift when BOM table has more/fewer rows.
    # Search for label text, then read the value cell below the label.
    usr_checkbox_val = _find_label_value(ws, 'upper section replacement',
                                         value_offset='below')
    usr_pin_size = _find_label_value(ws, 'pin size', value_offset='below')
    cerebro_val = _find_label_value(ws, 'cerebro', value_offset='below')
    salvage_val = _find_label_value(ws, 'track salvage', value_offset='below')
    # Fallback: also try 'salvage cutters' as an alternate label
    if salvage_val is None:
        salvage_val = _find_label_value(ws, 'salvage cutters', value_offset='below')

    has_cerebro = _is_checkbox_checked(cerebro_val)
    track_salvage = _is_checkbox_checked(salvage_val)

    # --- Job status flags ---
    # Column L values (set to 1 when checked):
    l2 = ws['L2'].value
    l3 = ws['L3'].value
    l4 = ws['L4'].value
    l5 = ws['L5'].value

    # Column I flags (checkboxes/labels for decision flags):
    # I2 = Rerun by ARDT, I3 = Rerun by Engineer
    # I4 = Initial Bit Inspection, I5 = Scrap
    # I7 = ARDT Matrix Build up, I8 = Engineer Matrix Build up
    i7_build_up = ws['I7'].value  # ARDT Matrix Build up checkbox
    i8_build_up = ws['I8'].value  # Engineer Matrix Build up checkbox

    is_rerun = (l2 == 1 or l3 == 1)
    is_inspection_only = (l4 == 1)
    is_scrap = (l5 == 1)

    # Build-up flags from I column (direct checkbox approach)
    has_buildup_from_flags = False
    for val in [i7_build_up, i8_build_up]:
        if val is not None:
            v = str(val).strip()
            if v == '1' or v.lower() in ('yes', 'true', 'x'):
                has_buildup_from_flags = True
                break

    # --- Normalize account name ---
    account = ACCOUNT_ALIASES.get(account.upper(), account) if account else account

    # --- Parse size ---
    size_inches = parse_size_to_inches(raw_size)

    # --- Parse cutter BOM (rows 12-28) with variant info ---
    # Table 2: D12-M28 = Final cutter bill
    # Table 3: N12-O28 = Original BOM reference
    cutter_bom = []
    for row in range(12, 29):
        group = ws.cell(row=row, column=4).value    # D = group #
        qty = ws.cell(row=row, column=5).value       # E = BOM qty
        size = ws.cell(row=row, column=6).value      # F = cutter size
        part_no = ws.cell(row=row, column=7).value   # G = part #
        desc = ws.cell(row=row, column=8).value      # H = description
        new_qty = ws.cell(row=row, column=11).value   # K = new qty
        rcl_qty = ws.cell(row=row, column=12).value   # L = reclaim qty
        comment = ws.cell(row=row, column=13).value   # M = additional comment
        # Original BOM reference (Table 3)
        bom_cutters = ws.cell(row=row, column=14).value  # N = cutters as BOM
        bom_qty = ws.cell(row=row, column=15).value      # O = qty as BOM

        # Skip empty rows (no part# or placeholder spaces)
        if not part_no or not str(part_no).strip():
            continue
        size_str = str(size or '').strip()
        if size_str == '' or size_str == ' ':
            continue

        pn = str(part_no).strip()
        variants = parse_cutter_variants(new_qty, rcl_qty, comment, pn)

        # Lookup ERP item numbers for each variant
        for v in variants:
            v['erp_item_no'] = lookup_erp_item_number(pn, v['variant_case'])

        # Safe int parsing for qty fields (may contain bracket text)
        def _safe_int(val):
            if not val:
                return 0
            s = str(val).strip()
            if not s:
                return 0
            try:
                return int(s)
            except (ValueError, TypeError):
                return 0

        entry = {
            'group': _safe_int(group) or (len(cutter_bom) + 1),
            'qty': _safe_int(qty),
            'size': size_str,
            'part_no': pn,
            'description': str(desc or '').strip(),
            'new_qty': _safe_int(new_qty),
            'reclaim_qty': _safe_int(rcl_qty),
            'additional_comment': str(comment or '').strip(),
            'variants': variants,
        }

        # Include original BOM reference if available
        if bom_cutters and str(bom_cutters).strip():
            entry['bom_cutters_original'] = str(bom_cutters).strip()
        if bom_qty and str(bom_qty).strip():
            try:
                entry['bom_qty_original'] = int(bom_qty)
            except (ValueError, TypeError):
                entry['bom_qty_original'] = str(bom_qty).strip()

        cutter_bom.append(entry)

    # --- Parse modified cutters (rows 4-10) ---
    modified_cutters = []
    for row in range(4, 11):
        group = ws.cell(row=row, column=4).value    # D
        qty = ws.cell(row=row, column=5).value       # E
        part_no = ws.cell(row=row, column=6).value   # F
        replaces = ws.cell(row=row, column=7).value  # G
        if part_no and str(part_no).strip():
            modified_cutters.append({
                'group': int(group) if group else None,
                'qty': int(qty) if qty else None,
                'part_no': str(part_no).strip(),
                'replaces_group': int(replaces) if replaces else None,
            })

    # --- Detect hardfacing / build-up (multi-method) ---
    has_hardfacing = has_buildup_from_flags  # Start with I7/I8 flag detection

    # Method 2: Check Evaluation sheet D36 for "build" keyword
    if not has_hardfacing and 'Evaluation' in wb.sheetnames:
        eval_ws = wb['Evaluation']
        remarks = str(eval_ws['D36'].value or '').lower()
        if 'build' in remarks:
            has_hardfacing = True

    # Method 3: Check Eval-LSTK sheet R34, U34, X34 for non-zero build-up counts
    if not has_hardfacing and 'Eval-LSTK' in wb.sheetnames:
        lstk_ws = wb['Eval-LSTK']
        for cell_ref in ['R34', 'U34', 'X34']:
            val = lstk_ws[cell_ref].value
            try:
                if val and int(val) > 0:
                    has_hardfacing = True
                    break
            except (ValueError, TypeError):
                pass

    # --- Extract evaluation summaries ---
    evaluations = _extract_evaluations(wb)

    # --- Extract price based on account ---
    price = None
    account_upper = (account or '').upper()
    if account_upper == 'ARAMCO' and 'Eval & Quot-AR' in wb.sheetnames:
        price = wb['Eval & Quot-AR']['Z56'].value
    elif account_upper == 'LSTK' and 'Quotation' in wb.sheetnames:
        price = wb['Quotation']['Y58'].value
    elif account_upper == 'HALLIBURTON' and 'Qut. HALL.' in wb.sheetnames:
        price = wb['Qut. HALL.']['AD56'].value
    # RC-LSTK: no price sheet (leave None)

    # Sanitize price to numeric
    if price is not None:
        try:
            price = float(str(price).replace(',', '').replace('$', '').strip())
        except (ValueError, TypeError):
            price = None

    wb.close()

    # --- Handle date_received ---
    from datetime import datetime, date
    if isinstance(date_received, datetime):
        date_received = date_received.date()
    elif not isinstance(date_received, date):
        date_received = None

    # --- USR detection: checkbox value OR pin size value ---
    has_usr = _is_checkbox_checked(usr_checkbox_val)
    if not has_usr and usr_pin_size is not None:
        # Pin size present = USR needed (but ignore label text)
        pin_str = str(usr_pin_size).strip()
        if pin_str and pin_str.lower() not in ('pin size', 'pin mat no.', ''):
            has_usr = True

    # --- Build result ---
    return {
        'work_order_number': str(wo_number or '').strip(),
        'serial_number': str(serial or '').strip(),
        'size_raw': str(raw_size or '').strip(),
        'size_inches': size_inches,
        'smi_type': smi_type,
        'l5_mat_full': l5_mat_full,
        'l5_mat_original': derive_original_l5_mat(l5_mat_full),
        'date_received': date_received,
        'account': account,
        'contract_number': str(contract_number or '').strip(),
        'vendor_number': str(vendor_number or '').strip(),
        'l3_l4_mat': str(l3_l4_mat or '').strip(),
        'evaluated_by': str(evaluated_by or '').strip(),
        'reviewed_by': str(reviewed_by or '').strip(),
        'body_material': derive_body_material(smi_type),
        'item_group': ACCOUNT_TO_ITEM_GROUP.get(account, ''),
        'size_class': 'JUMBO' if size_inches and size_inches >= 12 else 'AB',
        'has_port': bool(size_inches and size_inches < 4),
        'has_usr': has_usr,
        'has_hardfacing': has_hardfacing,
        'has_crush_shear': derive_crush_shear(smi_type),
        'is_rerun': is_rerun,
        'is_inspection_only': is_inspection_only,
        'is_scrap': is_scrap,
        # --- Cutter data ---
        'cutter_bom_data': cutter_bom,
        'modified_cutters_data': modified_cutters,
        # --- Process flags ---
        'has_cerebro': has_cerebro,
        'track_salvage': track_salvage,
        'usr_pin_size': str(usr_pin_size or '').strip(),
        # --- Evaluation summaries ---
        'evaluations': evaluations,
        # --- Pricing ---
        'price': price,
        # --- Source ---
        'source_file': os.path.basename(file_path),
    }


# =============================================================================
# BITS TRACKING PARSER
# =============================================================================

def parse_bits_tracking_row(row_data):
    """Parse a single row from a BITS TRACKING sheet into ERPJobData fields.

    BITS TRACKING is a separate Excel file with 11 account sheets
    (LSTK, ARAMCO, HALLIBURTON, etc.) tracking all work orders.
    Each row represents one WO with columns:
      Col 1 (A): WO Number
      Col 2 (B): Serial Number
      Col 3 (C): Size (raw)
      Col 4 (D): Type (SMI Type)
      Col 5 (E): L5 MAT
      Col 7 (G): Date Received
      Col 12 (L): USR flag
      Col 13 (M): Build-up notes
      Col 28 (AB): ERP Item Number

    Args:
        row_data: dict with column letters or indices as keys

    Returns:
        dict with fields compatible with ERPJobData creation
    """
    def _get(key, default=''):
        val = row_data.get(key, default)
        return str(val).strip() if val is not None else default

    raw_size = _get('C') or _get('size') or _get(3, '')
    size_inches = parse_size_to_inches(raw_size) if raw_size else None
    smi_type = _get('D') or _get('type') or _get(4, '')
    l5_mat = _get('E') or _get('l5_mat') or _get(5, '')
    account = _get('account', '')  # Usually known from sheet name

    # USR detection from column L
    usr_val = _get('L') or _get('usr') or _get(12, '')
    has_usr = bool(usr_val and usr_val.lower() not in ('', 'no', 'false', '0'))

    # Build-up from column M
    buildup_val = _get('M') or _get('buildup') or _get(13, '')
    has_hardfacing = 'build' in buildup_val.lower() if buildup_val else False

    return {
        'work_order_number': _get('A') or _get('wo_number') or _get(1, ''),
        'serial_number': _get('B') or _get('serial') or _get(2, ''),
        'size_raw': raw_size,
        'size_inches': size_inches,
        'smi_type': smi_type,
        'l5_mat_full': l5_mat,
        'l5_mat_original': derive_original_l5_mat(l5_mat),
        'account': account,
        'body_material': derive_body_material(smi_type),
        'item_group': ACCOUNT_TO_ITEM_GROUP.get(account, ''),
        'size_class': 'JUMBO' if size_inches and size_inches >= 12 else 'AB',
        'has_port': bool(size_inches and size_inches < 4),
        'has_usr': has_usr,
        'has_hardfacing': has_hardfacing,
        'has_crush_shear': derive_crush_shear(smi_type),
        'item_number': _get('AB') or _get('item_number') or _get(28, ''),
    }


# =============================================================================
# BATCH PARSING
# =============================================================================

def parse_job_card_safe(file_path):
    """Parse a Job Card with error handling.

    Wraps parse_job_card() to catch and log errors without crashing.

    Args:
        file_path: Path to the Job Card .xlsx file

    Returns:
        tuple: (parsed_data_dict, error_string_or_None)
    """
    try:
        data = parse_job_card(file_path)
        return data, None
    except KeyError as e:
        error = f"Missing sheet or cell in {os.path.basename(file_path)}: {e}"
        logger.error(error)
        return None, error
    except Exception as e:
        error = f"Error parsing {os.path.basename(file_path)}: {type(e).__name__}: {e}"
        logger.error(error, exc_info=True)
        return None, error


def parse_job_cards_batch(file_paths):
    """Parse multiple Job Card files and return results.

    Args:
        file_paths: list of file paths

    Returns:
        dict with 'success' (list of parsed data), 'errors' (list of error dicts)
    """
    success = []
    errors = []

    for fp in file_paths:
        data, error = parse_job_card_safe(fp)
        if data:
            success.append(data)
        else:
            errors.append({
                'file': os.path.basename(fp),
                'error': error,
            })

    logger.info(
        f"Batch parse complete: {len(success)} success, {len(errors)} errors "
        f"out of {len(file_paths)} files"
    )
    return {'success': success, 'errors': errors}
