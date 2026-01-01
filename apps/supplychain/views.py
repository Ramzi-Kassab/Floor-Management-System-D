"""
ARDT FMS - Supply Chain Views
Sprint 6: Supply Chain & Finance Integration
"""

import io

from django.conf import settings as django_settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.mail import EmailMessage
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from decimal import Decimal
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, DetailView, ListView, TemplateView, UpdateView, View

from .forms import (
    CAPAForm,
    PRApprovalForm,
    PurchaseOrderForm,
    PurchaseOrderLineForm,
    PurchaseRequisitionForm,
    PurchaseRequisitionLineForm,
    ReceiptForm,
    ReceiptLineForm,
    SupplierForm,
    VendorForm,
)
from .models import (
    CAPA,
    PurchaseOrder,
    PurchaseOrderLine,
    PurchaseRequisition,
    PurchaseRequisitionLine,
    Receipt,
    ReceiptLine,
    Supplier,
    Vendor,
)

User = get_user_model()


# =============================================================================
# Supplier Views (Legacy - use Vendor for new implementations)
# =============================================================================


class SupplierListView(LoginRequiredMixin, ListView):
    """List all suppliers (legacy)."""

    model = Supplier
    template_name = "supplychain/supplier_list.html"
    context_object_name = "suppliers"

    def get_queryset(self):
        qs = Supplier.objects.all()

        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q))

        active = self.request.GET.get("active")
        if active == "1":
            qs = qs.filter(status="ACTIVE")
        elif active == "0":
            qs = qs.exclude(status="ACTIVE")

        return qs.order_by("name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Suppliers"
        context["search_query"] = self.request.GET.get("q", "")
        return context


class SupplierDetailView(LoginRequiredMixin, DetailView):
    """Supplier detail view (legacy)."""

    model = Supplier
    template_name = "supplychain/supplier_detail.html"
    context_object_name = "supplier"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.object.code
        return context


class SupplierCreateView(LoginRequiredMixin, CreateView):
    """Create a new supplier (legacy)."""

    model = Supplier
    form_class = SupplierForm
    template_name = "supplychain/supplier_form.html"
    success_url = reverse_lazy("supplychain:supplier_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "New Supplier"
        context["form_title"] = "Create Supplier"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Supplier created successfully.")
        return super().form_valid(form)


class SupplierUpdateView(LoginRequiredMixin, UpdateView):
    """Update a supplier (legacy)."""

    model = Supplier
    form_class = SupplierForm
    template_name = "supplychain/supplier_form.html"

    def get_success_url(self):
        return reverse_lazy("supplychain:supplier_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.code}"
        context["form_title"] = f"Edit Supplier - {self.object.code}"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Supplier updated successfully.")
        return super().form_valid(form)


# =============================================================================
# Vendor Views (Sprint 6)
# =============================================================================


class VendorListView(LoginRequiredMixin, ListView):
    """List all vendors."""

    model = Vendor
    template_name = "supplychain/vendor_list.html"
    context_object_name = "vendors"

    def get_queryset(self):
        qs = Vendor.objects.annotate(po_count=Count("purchase_orders"))

        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(Q(vendor_id__icontains=q) | Q(company_name__icontains=q))

        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        return qs.order_by("company_name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Vendors"
        context["search_query"] = self.request.GET.get("q", "")
        context["status_choices"] = Vendor.Status.choices
        return context


class VendorDetailView(LoginRequiredMixin, DetailView):
    """Vendor detail view."""

    model = Vendor
    template_name = "supplychain/vendor_detail.html"
    context_object_name = "vendor"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.object.vendor_id
        context["contacts"] = self.object.contacts.all()
        context["recent_pos"] = self.object.purchase_orders.order_by("-created_at")[:10]
        return context


class VendorCreateView(LoginRequiredMixin, CreateView):
    """Create a new vendor."""

    model = Vendor
    form_class = VendorForm
    template_name = "supplychain/vendor_form.html"
    success_url = reverse_lazy("supplychain:vendor_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "New Vendor"
        context["form_title"] = "Create Vendor"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Vendor created successfully.")
        return super().form_valid(form)


class VendorUpdateView(LoginRequiredMixin, UpdateView):
    """Update a vendor."""

    model = Vendor
    form_class = VendorForm
    template_name = "supplychain/vendor_form.html"

    def get_success_url(self):
        return reverse_lazy("supplychain:vendor_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.vendor_id}"
        context["form_title"] = f"Edit Vendor - {self.object.company_name}"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Vendor updated successfully.")
        return super().form_valid(form)


# =============================================================================
# Purchase Requisition Views
# =============================================================================


class PRListView(LoginRequiredMixin, ListView):
    """List purchase requisitions."""

    model = PurchaseRequisition
    template_name = "supplychain/pr_list.html"
    context_object_name = "requisitions"

    def get_queryset(self):
        qs = PurchaseRequisition.objects.select_related(
            "requested_by", "approved_by"
        ).prefetch_related(
            "lines__inventory_item"
        ).annotate(line_count=Count("lines"))

        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Purchase Requisitions"
        context["status_choices"] = PurchaseRequisition.Status.choices
        context["current_status"] = self.request.GET.get("status", "")
        return context


class PRDetailView(LoginRequiredMixin, DetailView):
    """PR detail view."""

    model = PurchaseRequisition
    template_name = "supplychain/pr_detail.html"
    context_object_name = "pr"

    def get_queryset(self):
        return PurchaseRequisition.objects.select_related(
            "requested_by", "approved_by"
        ).prefetch_related("lines__inventory_item")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.object.requisition_number
        return context


class PRCreateView(LoginRequiredMixin, CreateView):
    """Create a new PR."""

    model = PurchaseRequisition
    form_class = PurchaseRequisitionForm
    template_name = "supplychain/pr_form.html"

    def get_success_url(self):
        return reverse_lazy("supplychain:pr_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "New Purchase Requisition"
        context["form_title"] = "Create Purchase Requisition"
        return context

    def form_valid(self, form):
        form.instance.requested_by = self.request.user
        form.instance.request_date = timezone.now().date()
        messages.success(self.request, "Purchase Requisition created successfully.")
        return super().form_valid(form)


class PRSubmitView(LoginRequiredMixin, View):
    """Submit a PR for approval."""

    def post(self, request, pk):
        pr = get_object_or_404(
            PurchaseRequisition.objects.select_related("requested_by"),
            pk=pk,
            status=PurchaseRequisition.Status.DRAFT,
        )

        # Check that PR has lines before submitting
        if not pr.lines.exists():
            messages.error(request, "Cannot submit a requisition without any line items.")
            return redirect("supplychain:pr_detail", pk=pr.pk)

        pr.submit()
        messages.success(request, f"Requisition {pr.requisition_number} submitted for approval.")
        return redirect("supplychain:pr_detail", pk=pr.pk)


class PRApproveView(LoginRequiredMixin, View):
    """Approve or reject a PR."""

    def get(self, request, pk):
        pr = get_object_or_404(PurchaseRequisition, pk=pk)
        form = PRApprovalForm()
        return self.render(request, pr, form)

    def post(self, request, pk):
        pr = get_object_or_404(PurchaseRequisition, pk=pk)
        form = PRApprovalForm(request.POST)

        if form.is_valid():
            action = form.cleaned_data["action"]
            if action == "approve":
                pr.status = PurchaseRequisition.Status.APPROVED
                pr.approved_by = request.user
                messages.success(request, "PR approved successfully.")
            else:
                pr.status = PurchaseRequisition.Status.REJECTED
                messages.info(request, "PR rejected.")
            pr.save()
            return redirect("supplychain:pr_detail", pk=pr.pk)

        return self.render(request, pr, form)

    def render(self, request, pr, form):
        from django.shortcuts import render

        return render(
            request,
            "supplychain/pr_approve.html",
            {
                "pr": pr,
                "form": form,
                "page_title": f"Review {pr.requisition_number}",
            },
        )


class PRAddLineView(LoginRequiredMixin, CreateView):
    """Add line to PR."""

    model = PurchaseRequisitionLine
    form_class = PurchaseRequisitionLineForm
    template_name = "supplychain/pr_line_form.html"

    def get_context_data(self, **kwargs):
        from apps.inventory.models import InventoryCategory

        context = super().get_context_data(**kwargs)
        context["pr"] = get_object_or_404(PurchaseRequisition, pk=self.kwargs["pk"])
        context["page_title"] = "Add Line"
        context["categories"] = InventoryCategory.objects.filter(is_active=True).order_by("name")
        return context

    def form_valid(self, form):
        pr = get_object_or_404(PurchaseRequisition, pk=self.kwargs["pk"])
        form.instance.requisition = pr
        last_line = pr.lines.order_by("-line_number").first()
        form.instance.line_number = (last_line.line_number + 1) if last_line else 1

        messages.success(self.request, "Line added successfully.")
        response = super().form_valid(form)

        # Handle "Add & Add Another" button
        if "add_another" in self.request.POST:
            return redirect("supplychain:pr_add_line", pk=self.kwargs["pk"])

        return response

    def get_success_url(self):
        return reverse_lazy("supplychain:pr_detail", kwargs={"pk": self.kwargs["pk"]})


class PREditLineView(LoginRequiredMixin, UpdateView):
    """Edit PR line."""

    model = PurchaseRequisitionLine
    form_class = PurchaseRequisitionLineForm
    template_name = "supplychain/pr_line_form.html"
    pk_url_kwarg = "line_pk"

    def get_queryset(self):
        # Only allow editing lines from DRAFT PRs
        return PurchaseRequisitionLine.objects.filter(
            requisition__status="DRAFT"
        ).select_related("requisition")

    def get_context_data(self, **kwargs):
        from apps.inventory.models import InventoryCategory

        context = super().get_context_data(**kwargs)
        context["pr"] = self.object.requisition
        context["page_title"] = "Edit Line"
        context["is_edit"] = True
        context["categories"] = InventoryCategory.objects.filter(is_active=True).order_by("name")
        return context

    def form_valid(self, form):
        messages.success(self.request, "Line updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("supplychain:pr_detail", kwargs={"pk": self.object.requisition.pk})


class PRDeleteLineView(LoginRequiredMixin, View):
    """Delete PR line."""

    def post(self, request, pk, line_pk):
        line = get_object_or_404(
            PurchaseRequisitionLine,
            pk=line_pk,
            requisition_id=pk,
            requisition__status="DRAFT"
        )
        pr_pk = line.requisition.pk
        line_number = line.line_number
        line.delete()

        # Renumber remaining lines
        remaining_lines = PurchaseRequisitionLine.objects.filter(
            requisition_id=pr_pk, line_number__gt=line_number
        ).order_by("line_number")
        for i, ln in enumerate(remaining_lines, start=line_number):
            ln.line_number = i
            ln.save(update_fields=["line_number"])

        messages.success(request, "Line deleted successfully.")
        return redirect("supplychain:pr_detail", pk=pr_pk)


class PRLinesExportView(LoginRequiredMixin, View):
    """Export PR lines to Excel."""

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # Get filter parameters
        status = request.GET.get("status", "")
        pr_id = request.GET.get("pr", "")

        # Query lines
        lines = PurchaseRequisitionLine.objects.select_related(
            "requisition", "requisition__requested_by", "inventory_item"
        ).order_by("requisition__requisition_number", "line_number")

        if status:
            lines = lines.filter(requisition__status=status)
        if pr_id:
            lines = lines.filter(requisition_id=pr_id)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "PR Lines"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Headers
        headers = [
            "PR Number", "PR Title", "Status", "Department", "Requested By",
            "Line #", "Item Code", "Item Name", "Description",
            "Qty", "UoM", "Est. Price", "Subtotal", "Notes"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data rows
        for row_num, line in enumerate(lines, 2):
            data = [
                line.requisition.requisition_number,
                line.requisition.title,
                line.requisition.get_status_display(),
                str(line.requisition.department) if line.requisition.department else "",
                line.requisition.requested_by.get_full_name() or line.requisition.requested_by.username,
                line.line_number,
                line.inventory_item.code if line.inventory_item else "",
                line.inventory_item.name if line.inventory_item else "",
                line.item_description or "",
                float(line.quantity_requested) if line.quantity_requested else 0,
                line.unit_of_measure or "",
                float(line.estimated_unit_price) if line.estimated_unit_price else 0,
                float(line.estimated_total) if line.estimated_total else 0,
                line.notes or ""
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                if col in [10, 12, 13]:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right")

        # Adjust column widths
        column_widths = [15, 30, 12, 15, 20, 8, 12, 30, 30, 10, 8, 12, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"pr_lines_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# =============================================================================
# Purchase Order Views
# =============================================================================


class PRConvertToPOView(LoginRequiredMixin, View):
    """Convert approved PR to PO."""

    def get(self, request, pk):
        pr = get_object_or_404(PurchaseRequisition, pk=pk, status=PurchaseRequisition.Status.APPROVED)
        # Get vendors (new model) - exclude suspended, disqualified, inactive
        vendors = Vendor.objects.exclude(
            status__in=[Vendor.Status.SUSPENDED, Vendor.Status.DISQUALIFIED, Vendor.Status.INACTIVE]
        ).order_by("name")
        # Get suppliers (legacy model) - only active ones
        suppliers = Supplier.objects.filter(is_active=True).order_by("name")
        return render(request, "supplychain/pr_convert_to_po.html", {
            "pr": pr,
            "vendors": vendors,
            "suppliers": suppliers,
            "page_title": f"Convert {pr.requisition_number} to PO"
        })

    def post(self, request, pk):
        pr = get_object_or_404(PurchaseRequisition, pk=pk, status=PurchaseRequisition.Status.APPROVED)
        selection = request.POST.get("supplier")  # Combined field

        if not selection:
            messages.error(request, "Please select a vendor or supplier.")
            return redirect("supplychain:pr_convert_to_po", pk=pk)

        vendor = None
        # Check if selection is a vendor (prefixed with "vendor_") or supplier (plain ID)
        if selection.startswith("vendor_"):
            vendor_id = selection.replace("vendor_", "")
            vendor = get_object_or_404(Vendor, pk=vendor_id)
        else:
            # Supplier selected - find or create corresponding Vendor
            supplier = get_object_or_404(Supplier, pk=selection)
            # Try to find existing vendor with same code, or create one
            vendor, created = Vendor.objects.get_or_create(
                vendor_code=supplier.code,
                defaults={
                    'name': supplier.name,
                    'contact_name': supplier.contact_person,
                    'email': supplier.email,
                    'phone': supplier.phone,
                    'address_line_1': supplier.address[:200] if supplier.address else '',
                    'country': supplier.country,
                    'status': Vendor.Status.ACTIVE,
                }
            )

        # Create PO
        po = PurchaseOrder.objects.create(
            vendor=vendor,
            order_date=timezone.now().date(),
            required_date=pr.required_date or timezone.now().date(),
            requisition=pr,
            work_order=pr.work_order,
            created_by=request.user,
            status=PurchaseOrder.Status.DRAFT,
            internal_notes=f"Created from {pr.requisition_number}"
        )

        # Copy lines
        for pr_line in pr.lines.all():
            PurchaseOrderLine.objects.create(
                purchase_order=po,
                line_number=pr_line.line_number,
                item_description=pr_line.item_description or (pr_line.inventory_item.name if pr_line.inventory_item else ""),
                part_number=pr_line.part_number or "",
                inventory_item=pr_line.inventory_item,
                quantity_ordered=pr_line.quantity_requested,
                unit_of_measure=pr_line.unit_of_measure,
                unit_price=pr_line.estimated_unit_price or Decimal('0.00'),
                required_date=pr.required_date or timezone.now().date()
            )

        # Calculate totals
        po.subtotal_amount = sum(
            (line.quantity_ordered * line.unit_price) for line in po.lines.all()
        )
        po.total_amount = po.subtotal_amount
        po.save()

        # Update PR
        pr.status = PurchaseRequisition.Status.CONVERTED_TO_PO
        pr.converted_to_po = po
        pr.converted_at = timezone.now()
        pr.save()

        messages.success(request, f"Created {po.po_number} from {pr.requisition_number}")
        return redirect("supplychain:po_detail", pk=po.pk)


class POListView(LoginRequiredMixin, ListView):
    """List purchase orders."""

    model = PurchaseOrder
    template_name = "supplychain/po_list.html"
    context_object_name = "orders"

    def get_queryset(self):
        qs = PurchaseOrder.objects.select_related("vendor", "created_by").annotate(line_count=Count("lines"))

        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        vendor = self.request.GET.get("vendor")
        if vendor:
            qs = qs.filter(vendor_id=vendor)

        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Purchase Orders"
        context["status_choices"] = PurchaseOrder.Status.choices
        context["current_status"] = self.request.GET.get("status", "")
        context["vendors"] = Vendor.objects.filter(status="ACTIVE")
        return context


class PODetailView(LoginRequiredMixin, DetailView):
    """PO detail view."""

    model = PurchaseOrder
    template_name = "supplychain/po_detail.html"
    context_object_name = "po"

    def get_queryset(self):
        return PurchaseOrder.objects.select_related("vendor", "created_by")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.object.po_number
        context["lines"] = self.object.lines.all()
        context["receipts"] = self.object.receipts.order_by("-created_at")
        return context


class POCreateView(LoginRequiredMixin, CreateView):
    """Create a new PO."""

    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = "supplychain/po_form.html"

    def get_success_url(self):
        return reverse_lazy("supplychain:po_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "New Purchase Order"
        context["form_title"] = "Create Purchase Order"
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, "Purchase Order created successfully.")
        return super().form_valid(form)


class POUpdateView(LoginRequiredMixin, UpdateView):
    """Update a PO."""

    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = "supplychain/po_form.html"

    def get_success_url(self):
        return reverse_lazy("supplychain:po_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.po_number}"
        context["form_title"] = f"Edit PO - {self.object.po_number}"
        return context

    def form_valid(self, form):
        messages.success(self.request, "PO updated successfully.")
        return super().form_valid(form)


class POAddLineView(LoginRequiredMixin, CreateView):
    """Add line to PO."""

    model = PurchaseOrderLine
    form_class = PurchaseOrderLineForm
    template_name = "supplychain/po_line_form.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["po"] = get_object_or_404(PurchaseOrder, pk=self.kwargs["pk"])
        context["page_title"] = "Add Line"
        return context

    def form_valid(self, form):
        po = get_object_or_404(PurchaseOrder, pk=self.kwargs["pk"])
        form.instance.purchase_order = po
        last_line = po.lines.order_by("-line_number").first()
        form.instance.line_number = (last_line.line_number + 1) if last_line else 1

        messages.success(self.request, "Line added successfully.")
        response = super().form_valid(form)

        # Update PO total
        total = po.lines.aggregate(total=Sum(F("quantity") * F("unit_price")))["total"] or 0
        po.total_amount = total
        po.save()

        return response

    def get_success_url(self):
        return reverse_lazy("supplychain:po_detail", kwargs={"pk": self.kwargs["pk"]})


class POApproveView(LoginRequiredMixin, View):
    """Approve a PO (changes status from DRAFT to APPROVED)."""

    def post(self, request, pk):
        po = get_object_or_404(PurchaseOrder, pk=pk)

        if po.status not in [PurchaseOrder.Status.DRAFT, PurchaseOrder.Status.PENDING_APPROVAL]:
            messages.error(request, "Only draft or pending approval POs can be approved.")
            return redirect("supplychain:po_detail", pk=pk)

        po.status = PurchaseOrder.Status.APPROVED
        po.approved_by = request.user
        po.approved_at = timezone.now()
        po.save()

        messages.success(request, f"{po.po_number} has been approved.")
        return redirect("supplychain:po_detail", pk=pk)


class POSendView(LoginRequiredMixin, View):
    """Mark PO as sent to vendor."""

    def post(self, request, pk):
        po = get_object_or_404(PurchaseOrder, pk=pk)

        if po.status not in [PurchaseOrder.Status.DRAFT, PurchaseOrder.Status.APPROVED]:
            messages.error(request, "Only draft or approved POs can be sent.")
            return redirect("supplychain:po_detail", pk=pk)

        po.status = PurchaseOrder.Status.SENT
        po.sent_to_vendor_at = timezone.now()
        po.save()

        messages.success(request, f"{po.po_number} has been marked as sent to vendor.")
        return redirect("supplychain:po_detail", pk=pk)


# =============================================================================
# Receipt Views (formerly Goods Receipt / GRN)
# =============================================================================


class ReceiptListView(LoginRequiredMixin, ListView):
    """List goods receipts."""

    model = Receipt
    template_name = "supplychain/grn_list.html"
    context_object_name = "receipts"

    def get_queryset(self):
        return Receipt.objects.select_related(
            "purchase_order",
            "vendor",
            "received_by"
        ).order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Goods Receipts"
        return context


class ReceiptDetailView(LoginRequiredMixin, DetailView):
    """Receipt detail view."""

    model = Receipt
    template_name = "supplychain/grn_detail.html"
    context_object_name = "grn"

    def get_queryset(self):
        return Receipt.objects.select_related("purchase_order", "vendor", "received_by", "inspected_by")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.object.receipt_number
        context["lines"] = self.object.lines.select_related("po_line")
        return context


class ReceiptCreateView(LoginRequiredMixin, CreateView):
    """Create a new Receipt."""

    model = Receipt
    form_class = ReceiptForm
    template_name = "supplychain/grn_form.html"

    def get_form(self):
        form = super().get_form()
        # Only show SENT or PARTIALLY_RECEIVED POs (APPROVED also for direct processing)
        form.fields["purchase_order"].queryset = PurchaseOrder.objects.filter(
            status__in=[
                PurchaseOrder.Status.APPROVED,
                PurchaseOrder.Status.SENT,
                PurchaseOrder.Status.PARTIALLY_RECEIVED
            ]
        ).select_related("vendor").order_by("-po_number")
        return form

    def get_initial(self):
        initial = super().get_initial()
        po_pk = self.request.GET.get("po")
        if po_pk:
            initial["purchase_order"] = po_pk
        initial["receipt_date"] = timezone.now().date()
        return initial

    def get_success_url(self):
        return reverse_lazy("supplychain:grn_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "New Goods Receipt"
        context["form_title"] = "Create Goods Receipt"
        return context

    def form_valid(self, form):
        form.instance.received_by = self.request.user
        # Set vendor from the PO
        form.instance.vendor = form.instance.purchase_order.vendor
        messages.success(self.request, "Goods Receipt created successfully.")
        return super().form_valid(form)


class ReceiptAddLineView(LoginRequiredMixin, CreateView):
    """Add line to Receipt."""

    model = ReceiptLine
    form_class = ReceiptLineForm
    template_name = "supplychain/grn_line_form.html"

    def get_form(self):
        form = super().get_form()
        receipt = get_object_or_404(Receipt, pk=self.kwargs["pk"])
        form.fields["po_line"].queryset = receipt.purchase_order.lines.all()
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["grn"] = get_object_or_404(Receipt, pk=self.kwargs["pk"])
        context["page_title"] = "Add Receipt Line"
        return context

    def form_valid(self, form):
        receipt = get_object_or_404(Receipt, pk=self.kwargs["pk"])
        form.instance.receipt = receipt

        # Set line number
        last_line = receipt.lines.order_by("-line_number").first()
        form.instance.line_number = (last_line.line_number + 1) if last_line else 1

        # Default accepted to received if not set
        if not form.instance.quantity_accepted:
            form.instance.quantity_accepted = form.instance.quantity_received
            form.instance.inspection_status = 'PASSED'

        messages.success(self.request, "Receipt line added successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("supplychain:grn_detail", kwargs={"pk": self.kwargs["pk"]})


class GRNConfirmView(LoginRequiredMixin, View):
    """Confirm a goods receipt - updates inventory and PO quantities."""

    def post(self, request, pk):
        from apps.inventory.models import InventoryTransaction, Stock, InventoryLocation

        receipt = get_object_or_404(
            Receipt.objects.select_related("purchase_order", "vendor"),
            pk=pk,
            status=Receipt.Status.DRAFT
        )

        if not receipt.lines.exists():
            messages.error(request, "Cannot confirm a receipt with no lines.")
            return redirect("supplychain:grn_detail", pk=pk)

        # Get default receiving location
        default_location = InventoryLocation.objects.filter(
            is_active=True
        ).first()

        # Process each line
        for line in receipt.lines.select_related("po_line__inventory_item"):
            po_line = line.po_line

            # Update PO line received quantity
            po_line.quantity_received += line.quantity_accepted
            po_line.save()

            # Update inventory if linked to inventory item
            if po_line.inventory_item and default_location:
                item = po_line.inventory_item

                # Create inventory transaction
                InventoryTransaction.objects.create(
                    item=item,
                    transaction_type=InventoryTransaction.TransactionType.RECEIPT,
                    quantity=line.quantity_accepted,
                    to_location=default_location,
                    reference_type="Receipt",
                    reference_id=str(receipt.pk),
                    notes=f"GRN {receipt.receipt_number} - PO {receipt.purchase_order.po_number}"
                )

                # Update or create stock record
                stock, created = Stock.objects.get_or_create(
                    item=item,
                    location=default_location,
                    defaults={"quantity_on_hand": Decimal("0")}
                )
                stock.quantity_on_hand += line.quantity_accepted
                stock.quantity_available = float(stock.quantity_on_hand) - float(stock.quantity_reserved)
                stock.last_movement_date = timezone.now()
                stock.save()

            # Mark line as passed inspection
            line.inspection_status = 'PASSED'
            line.save()

        # Update receipt status
        receipt.status = Receipt.Status.COMPLETED
        receipt.quality_acceptable = True
        receipt.inspected_by = request.user
        receipt.inspection_date = timezone.now().date()
        receipt.save()

        # Update PO status
        po = receipt.purchase_order
        if po.is_fully_received:
            po.status = PurchaseOrder.Status.COMPLETED
        else:
            po.status = PurchaseOrder.Status.PARTIALLY_RECEIVED
        po.save()

        messages.success(request, f"Receipt {receipt.receipt_number} confirmed. Inventory updated.")
        return redirect("supplychain:grn_detail", pk=pk)


# Legacy view aliases for backward compatibility
GRNListView = ReceiptListView
GRNDetailView = ReceiptDetailView
GRNCreateView = ReceiptCreateView
GRNAddLineView = ReceiptAddLineView


# =============================================================================
# CAPA Views
# =============================================================================


class CAPAListView(LoginRequiredMixin, ListView):
    """List CAPAs."""

    model = CAPA
    template_name = "supplychain/capa_list.html"
    context_object_name = "capas"

    def get_queryset(self):
        qs = CAPA.objects.select_related("ncr", "assigned_to")

        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "CAPAs"
        context["status_choices"] = CAPA.Status.choices
        context["current_status"] = self.request.GET.get("status", "")
        return context


class CAPADetailView(LoginRequiredMixin, DetailView):
    """CAPA detail view."""

    model = CAPA
    template_name = "supplychain/capa_detail.html"
    context_object_name = "capa"

    def get_queryset(self):
        return CAPA.objects.select_related("ncr", "assigned_to")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.object.capa_number
        return context


class CAPACreateView(LoginRequiredMixin, CreateView):
    """Create a new CAPA."""

    model = CAPA
    form_class = CAPAForm
    template_name = "supplychain/capa_form.html"

    def get_initial(self):
        initial = super().get_initial()
        ncr_pk = self.request.GET.get("ncr")
        if ncr_pk:
            initial["ncr"] = ncr_pk
        return initial

    def get_success_url(self):
        return reverse_lazy("supplychain:capa_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "New CAPA"
        context["form_title"] = "Create CAPA"
        return context

    def form_valid(self, form):
        messages.success(self.request, "CAPA created successfully.")
        return super().form_valid(form)


class CAPAUpdateView(LoginRequiredMixin, UpdateView):
    """Update a CAPA."""

    model = CAPA
    form_class = CAPAForm
    template_name = "supplychain/capa_form.html"

    def get_success_url(self):
        return reverse_lazy("supplychain:capa_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit {self.object.capa_number}"
        context["form_title"] = f"Edit CAPA - {self.object.capa_number}"
        return context

    def form_valid(self, form):
        messages.success(self.request, "CAPA updated successfully.")
        return super().form_valid(form)


# =============================================================================
# PO PDF and Email Views
# =============================================================================


class POPDFView(LoginRequiredMixin, View):
    """Generate PDF for a Purchase Order."""

    def get(self, request, pk):
        from xhtml2pdf import pisa

        po = get_object_or_404(
            PurchaseOrder.objects.select_related("vendor", "created_by"),
            pk=pk
        )
        lines = po.lines.all()

        # Render HTML template
        html_content = render_to_string("supplychain/po_pdf.html", {
            "po": po,
            "lines": lines,
            "company_name": getattr(django_settings, "COMPANY_NAME", "ARDT FMS"),
        })

        # Create PDF
        buffer = io.BytesIO()
        pisa_status = pisa.CreatePDF(html_content, dest=buffer)

        if pisa_status.err:
            return HttpResponse("Error generating PDF", status=500)

        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{po.po_number}.pdf"'
        return response


class POEmailView(LoginRequiredMixin, View):
    """Email PO to vendor."""

    def get(self, request, pk):
        """Show confirmation page."""
        from django.shortcuts import render

        po = get_object_or_404(
            PurchaseOrder.objects.select_related("vendor"),
            pk=pk
        )

        if not po.vendor.primary_email:
            messages.error(request, "Vendor does not have an email address configured.")
            return redirect("supplychain:po_detail", pk=pk)

        return render(request, "supplychain/po_email_confirm.html", {
            "po": po,
            "page_title": f"Email {po.po_number}",
        })

    def post(self, request, pk):
        """Send the email with PDF attachment."""
        from xhtml2pdf import pisa

        po = get_object_or_404(
            PurchaseOrder.objects.select_related("vendor", "created_by"),
            pk=pk
        )
        lines = po.lines.all()

        if not po.vendor.primary_email:
            messages.error(request, "Vendor does not have an email address.")
            return redirect("supplychain:po_detail", pk=pk)

        # Generate PDF
        html_content = render_to_string("supplychain/po_pdf.html", {
            "po": po,
            "lines": lines,
            "company_name": getattr(django_settings, "COMPANY_NAME", "ARDT FMS"),
        })

        buffer = io.BytesIO()
        pisa.CreatePDF(html_content, dest=buffer)
        buffer.seek(0)
        pdf_content = buffer.read()

        # Compose email
        subject = f"Purchase Order {po.po_number}"
        body = render_to_string("supplychain/po_email_body.html", {
            "po": po,
            "sender_name": request.user.get_full_name() or request.user.username,
        })

        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=django_settings.DEFAULT_FROM_EMAIL,
            to=[po.vendor.primary_email],
        )
        email.attach(f"{po.po_number}.pdf", pdf_content, "application/pdf")

        try:
            email.send()
            # Update PO status to SENT if it's still in DRAFT
            if po.status == PurchaseOrder.Status.DRAFT:
                po.status = PurchaseOrder.Status.SENT
                po.save()
            messages.success(request, f"Purchase Order sent to {po.vendor.primary_email}")
        except Exception as e:
            messages.error(request, f"Failed to send email: {str(e)}")

        return redirect("supplychain:po_detail", pk=pk)
