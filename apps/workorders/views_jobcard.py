"""
ARDT FMS - Job Card Views
Version: 1.0

Enhanced views for Job Card functionality including:
- Work Order Dashboard with modern UI
- Cutter Evaluation Matrix
- Router Sheet with QR tracking
- QC Forms (LPT, API Thread, E-Checklist)
- Instruction Rules management
"""

import json as _json
from datetime import datetime, timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Q, Count, Sum, F, Max
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import (
    CreateView, DeleteView, DetailView, ListView, UpdateView, TemplateView, View
)

from .models import (
    WorkOrder, DrillBit, BitEvent, Location,
    CutterEvaluationMatrix, CutterEvaluationEntry, ReceivingInspection,
    ReceivingInspectionAttachment,
    RouterSheetEntry, EvaluationChecklist,
    LPTReport, APIThreadInspection,
    DieCheckReport, StandaloneLPTReport, StandaloneThreadReport,
    InstructionRule, InstructionRuleCondition,
    ProcessRoute, ProcessRouteOperation,
    RepairEvaluation, WorkOrderCost, ProductionPlanEntry,
    PlannerSettings, PlannerHoliday,
    EvaluationRoute, EvaluationRouteStep
)
from apps.notifications.services import notify, create_form_revision
from .utils import generate_work_order_qr, generate_drill_bit_qr


# =============================================================================
# WORK ORDER DASHBOARD - Modern UI
# =============================================================================

class WorkOrderDashboardView(LoginRequiredMixin, TemplateView):
    """
    Main Work Order Dashboard with summary cards and quick filters.
    Similar to cutter inventory dashboard.
    """
    template_name = "workorders/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Summary statistics
        total_wo = WorkOrder.objects.count()
        context['total_work_orders'] = total_wo

        # Status breakdown
        context['draft_count'] = WorkOrder.objects.filter(status=WorkOrder.Status.DRAFT).count()
        context['in_progress_count'] = WorkOrder.objects.filter(status=WorkOrder.Status.IN_PROGRESS).count()
        context['qc_pending_count'] = WorkOrder.objects.filter(status=WorkOrder.Status.QC_PENDING).count()
        context['completed_count'] = WorkOrder.objects.filter(status=WorkOrder.Status.COMPLETED).count()

        # Overdue work orders
        today = timezone.now().date()
        context['overdue_count'] = WorkOrder.objects.filter(
            due_date__lt=today
        ).exclude(
            status__in=[WorkOrder.Status.COMPLETED, WorkOrder.Status.CANCELLED]
        ).count()

        # Due this week
        week_end = today + timedelta(days=7)
        context['due_this_week'] = WorkOrder.objects.filter(
            due_date__gte=today,
            due_date__lte=week_end
        ).exclude(
            status__in=[WorkOrder.Status.COMPLETED, WorkOrder.Status.CANCELLED]
        ).count()

        # Recent work orders
        context['recent_work_orders'] = WorkOrder.objects.select_related(
            'customer', 'drill_bit', 'assigned_to'
        ).order_by('-created_at')[:10]

        # Drill bit statistics
        context['total_bits'] = DrillBit.objects.count()
        context['bits_in_production'] = DrillBit.objects.filter(
            status=DrillBit.Status.IN_PRODUCTION
        ).count()
        context['bits_at_rig'] = DrillBit.objects.filter(
            physical_status=DrillBit.PhysicalStatus.AT_RIG
        ).count()

        context['page_title'] = 'Work Orders Dashboard'
        return context


class WorkOrderListEnhancedView(LoginRequiredMixin, ListView):
    """
    Enhanced Work Order List with Excel-like filtering.
    """
    model = WorkOrder
    template_name = "workorders/workorder_list_enhanced.html"
    context_object_name = "work_orders"
    paginate_by = None  # Default to all for column filters

    def get_paginate_by(self, queryset):
        """Allow page size to be changed via query parameter."""
        page_size = self.request.GET.get('page_size', 'all')
        if page_size == 'all':
            return None
        try:
            page_size = int(page_size)
            if page_size in [25, 50, 100, 200]:
                return page_size
        except (ValueError, TypeError):
            pass
        return None

    def get_queryset(self):
        queryset = WorkOrder.objects.select_related(
            "customer", "drill_bit", "drill_bit__design", "drill_bit__design__size",
            "assigned_to", "design", "design__size", "bom", "account"
        ).order_by("-created_at")

        # Filter by status
        status = self.request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)

        # Filter by wo_type
        wo_type = self.request.GET.get("wo_type")
        if wo_type:
            queryset = queryset.filter(wo_type=wo_type)

        # Filter by priority
        priority = self.request.GET.get("priority")
        if priority:
            queryset = queryset.filter(priority=priority)

        # Filter by customer
        customer_id = self.request.GET.get("customer")
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)

        # Filter by account
        account_id = self.request.GET.get("account")
        if account_id:
            queryset = queryset.filter(account_id=account_id)

        # Search
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(wo_number__icontains=search)
                | Q(customer__name__icontains=search)
                | Q(drill_bit__serial_number__icontains=search)
                | Q(drss_no__icontains=search)
                | Q(brazing_mat_no__icontains=search)
            )

        # Date range filter
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Work Orders"
        context["status_choices"] = WorkOrder.Status.choices
        context["wo_type_choices"] = WorkOrder.WOType.choices
        context["priority_choices"] = WorkOrder.Priority.choices

        # Get unique customers for filter
        from apps.sales.models import Customer, Account
        context["customers"] = Customer.objects.filter(
            work_orders__isnull=False
        ).distinct().order_by('name')
        context["accounts"] = Account.objects.filter(is_active=True).order_by('sort_order')
        context["current_account"] = self.request.GET.get("account", "")

        # Current filters
        context["current_status"] = self.request.GET.get("status", "")
        context["current_wo_type"] = self.request.GET.get("wo_type", "")
        context["current_priority"] = self.request.GET.get("priority", "")
        context["current_customer"] = self.request.GET.get("customer", "")
        context["current_search"] = self.request.GET.get("search", "")
        context["date_from"] = self.request.GET.get("date_from", "")
        context["date_to"] = self.request.GET.get("date_to", "")

        # Page size
        paginate_by = self.get_paginate_by(None)
        context["page_size"] = 'all' if paginate_by is None else paginate_by
        context["total_count"] = self.get_queryset().count()

        # Summary statistics (from ALL work orders, not filtered)
        all_wos = WorkOrder.objects.all()
        status_counts = dict(
            all_wos.values_list('status')
            .annotate(c=Count('pk'))
            .values_list('status', 'c')
        )
        today = timezone.now().date()
        overdue_count = all_wos.filter(
            due_date__lt=today,
            status__in=[
                WorkOrder.Status.DRAFT, WorkOrder.Status.PLANNED,
                WorkOrder.Status.RELEASED, WorkOrder.Status.IN_PROGRESS,
                WorkOrder.Status.QC_PENDING, WorkOrder.Status.ON_HOLD,
            ]
        ).count()
        context["stats"] = {
            "total": all_wos.count(),
            "in_progress": status_counts.get(WorkOrder.Status.IN_PROGRESS, 0),
            "qc_pending": status_counts.get(WorkOrder.Status.QC_PENDING, 0),
            "completed": status_counts.get(WorkOrder.Status.COMPLETED, 0),
            "on_hold": status_counts.get(WorkOrder.Status.ON_HOLD, 0),
            "overdue": overdue_count,
            "draft": status_counts.get(WorkOrder.Status.DRAFT, 0),
            "released": status_counts.get(WorkOrder.Status.RELEASED, 0),
        }

        # ── Build evaluation step pipeline per WO ──
        wo_list = list(context.get('work_orders', context.get('object_list', [])))
        wo_pks = [wo.pk for wo in wo_list]

        if wo_pks:
            # Bulk fetch all evaluations for all visible WOs
            all_evals = list(
                CutterEvaluationMatrix.objects
                .filter(work_order_id__in=wo_pks)
                .select_related('evaluated_by', 'approved_by')
                .order_by('work_order_id', 'evaluation_type')
            )
            # Build map: wo_pk -> {eval_type: eval_obj}
            eval_map = {}
            for ev in all_evals:
                eval_map.setdefault(ev.work_order_id, {})[ev.evaluation_type] = ev

            # Bulk fetch standalone reports
            die_checks = {}
            for dc in DieCheckReport.objects.filter(work_order_id__in=wo_pks).order_by('created_at'):
                die_checks.setdefault(dc.work_order_id, []).append(dc)
            lpt_reports = {}
            for lpt in StandaloneLPTReport.objects.filter(work_order_id__in=wo_pks).order_by('created_at'):
                lpt_reports.setdefault(lpt.work_order_id, []).append(lpt)
            thread_reports = {}
            for tr in StandaloneThreadReport.objects.filter(work_order_id__in=wo_pks).order_by('created_at'):
                thread_reports.setdefault(tr.work_order_id, []).append(tr)

            # Bulk fetch evaluation routes by account+bit_type
            try:
                routes = list(
                    EvaluationRoute.objects.filter(is_active=True)
                    .select_related('account')
                    .prefetch_related('steps')
                )
                # Build route lookup: (account_id, bit_type) -> route
                route_lookup = {}
                for r in routes:
                    key = (r.account_id, r.bit_type)
                    if r.is_default or key not in route_lookup:
                        route_lookup[key] = r
            except Exception:
                route_lookup = {}

            # For each WO, build its step pipeline
            for wo in wo_list:
                wo_evals = eval_map.get(wo.pk, {})
                wo_die_checks = die_checks.get(wo.pk, [])
                wo_lpt = lpt_reports.get(wo.pk, [])
                wo_thread = thread_reports.get(wo.pk, [])

                # Determine route
                bit_type = 'NEW' if wo.wo_type in [WorkOrder.WOType.FC_NEW, WorkOrder.WOType.RC_NEW] else 'USED'
                route = route_lookup.get((wo.account_id, bit_type)) if wo.account_id else None

                steps = []
                if route:
                    for rs in route.steps.all().order_by('order'):
                        ev = wo_evals.get(rs.evaluation_type)
                        step_info = self._build_step_info(
                            rs.evaluation_type, rs.get_evaluation_type_display(),
                            ev, wo, rs.is_required, rs.is_conditional,
                            wo_die_checks, wo_lpt, wo_thread,
                        )
                        steps.append(step_info)
                else:
                    # Legacy: show evaluations that actually exist
                    for eval_type, ev in wo_evals.items():
                        step_info = self._build_step_info(
                            eval_type, ev.get_evaluation_type_display(),
                            ev, wo, True, False,
                            wo_die_checks, wo_lpt, wo_thread,
                        )
                        steps.append(step_info)

                wo.eval_steps = steps
                wo.eval_total = len(steps)
                wo.eval_done = sum(1 for s in steps if s['status'] == 'COMPLETED' or s['status'] == 'APPROVED')
                wo.has_route = route is not None

                # Determine evaluation outcome from first completed eval with a decision
                wo.eval_outcome = ''
                for s in steps:
                    if s.get('decision'):
                        wo.eval_outcome = s['decision']
                        break

        # ── Serialize WO data as JSON for v2 JS table ──
        import json
        wo_json_list = []
        for wo in wo_list:
            # Build bit description
            bit_desc = ''
            if wo.drill_bit:
                bit_desc = wo.drill_bit.serial_number
                if wo.design and wo.design.mat_no:
                    bit_desc += f' — {wo.design.mat_no}'
            elif wo.description:
                bit_desc = wo.description

            # Size
            size_str = ''
            if wo.drill_bit and wo.drill_bit.design and wo.drill_bit.design.size:
                size_str = str(wo.drill_bit.design.size)
            elif wo.design and wo.design.size:
                size_str = str(wo.design.size)

            # Progress from eval steps
            pct = 0
            if hasattr(wo, 'eval_total') and wo.eval_total > 0:
                pct = int((wo.eval_done / wo.eval_total) * 100)

            # Map status to v2 codes
            status_map = {
                'DRAFT': 'pd', 'PLANNED': 'pd', 'RELEASED': 'pd',
                'IN_PROGRESS': 'ip', 'ON_HOLD': 'hd',
                'QC_PENDING': 'rv', 'QC_PASSED': 'rv', 'QC_FAILED': 'rv',
                'COMPLETED': 'dn', 'CANCELLED': 'hd',
            }
            v2_status = status_map.get(wo.status, 'pd')

            # Map priority
            prio_map = {'LOW': 'L', 'NORMAL': 'M', 'HIGH': 'H', 'URGENT': 'H', 'CRITICAL': 'H'}
            v2_prio = prio_map.get(wo.priority, 'M')

            # Build route steps for JS
            route_steps = []
            for s in getattr(wo, 'eval_steps', []):
                # Map step status to v2 state
                state_map = {
                    'PENDING': 'pending', 'DRAFT': 'pending',
                    'IN_PROGRESS': 'active', 'COMPLETED': 'done', 'APPROVED': 'done',
                }
                route_steps.append({
                    'key': s['type_code'].lower().replace('_', ''),
                    'label': s['type_label'],
                    'state': state_map.get(s['status'], 'pending'),
                    'url': s.get('url', ''),
                    'tech': s.get('evaluated_by', ''),
                    'decision': s.get('decision', ''),
                    'notes': '',
                    'isBranch': s['type_code'] in ('PDC_EVAL', 'ARDT', 'ENGINEER'),
                })

            # Eval outcome mapping
            outcome_map = {
                'Repair': 'full-repair', 'Rerun': 'light-dress',
                'Scrap': 'scrap', 'Debraze': 'scrap',
                'Cutter Retrofit': 'full-repair', 'New Build': 'full-repair',
                'Body Retrofit': 'full-repair',
            }
            eval_outcome = outcome_map.get(getattr(wo, 'eval_outcome', ''), '')

            wo_json_list.append({
                'pk': wo.pk,
                'id': wo.wo_number,
                'title': bit_desc or wo.wo_number,
                'serial': wo.drill_bit.serial_number if wo.drill_bit else '',
                'customer': wo.account.code if wo.account else (wo.customer.name if wo.customer else '—'),
                'prio': v2_prio,
                'status': v2_status,
                'statusDisplay': wo.get_status_display(),
                'assigned': (wo.assigned_to.get_short_name() or wo.assigned_to.username) if wo.assigned_to else '—',
                'due': wo.due_date.strftime('%b %d') if wo.due_date else '—',
                'dueRaw': wo.due_date.isoformat() if wo.due_date else '',
                'hrs': 0,
                'pct': pct,
                'size': size_str,
                'type': wo.get_wo_type_display(),
                'notes': wo.notes or '',
                'evalOutcome': eval_outcome,
                'route': route_steps,
                'detailUrl': f'/work-orders/enhanced/{wo.pk}/',
                'isOverdue': wo.is_overdue if hasattr(wo, 'is_overdue') else False,
            })

        context['wo_json'] = json.dumps(wo_json_list)

        return context

    def _build_step_info(self, eval_type, type_label, ev, wo, is_required, is_conditional,
                         wo_die_checks, wo_lpt, wo_thread):
        """Build a step info dict for the evaluation pipeline."""
        # Determine status
        if ev:
            if ev.status == CutterEvaluationMatrix.Status.APPROVED:
                status = 'APPROVED'
            elif ev.is_complete or ev.status == CutterEvaluationMatrix.Status.COMPLETED:
                status = 'COMPLETED'
            elif ev.status == CutterEvaluationMatrix.Status.IN_PROGRESS:
                status = 'IN_PROGRESS'
            else:
                status = 'DRAFT'
        else:
            status = 'PENDING'

        # Build URL for the step
        url = ''
        if ev:
            if eval_type == 'PDC_EVAL':
                url = reverse('workorders:pre_repair_eval_edit', args=[wo.pk, ev.pk])
            elif eval_type in ('DIE_CHECK', 'FINAL_DIE_CHECK'):
                # Link to die check report if exists
                dc = next((d for d in wo_die_checks if d.evaluation_id == ev.pk), None)
                if dc:
                    url = reverse('workorders:die_check_edit', args=[wo.pk, dc.pk, ev.pk])
                else:
                    url = reverse('workorders:die_check_create', args=[wo.pk, ev.pk])
            else:
                url = reverse('workorders:cutter_evaluation_edit', args=[wo.pk, ev.pk])
        else:
            # Auto-create URL
            url = reverse('workorders:evaluation_auto', args=[wo.pk, eval_type])

        # Standalone reports linked to this eval
        linked_reports = []
        if ev:
            for dc in wo_die_checks:
                if dc.evaluation_id == ev.pk:
                    linked_reports.append({
                        'type': 'Die Check',
                        'result': dc.result,
                        'complete': dc.is_complete,
                    })
            for lpt in wo_lpt:
                if lpt.evaluation_id == ev.pk:
                    linked_reports.append({
                        'type': 'LPT',
                        'result': lpt.result,
                        'complete': lpt.is_complete,
                    })
            for tr in wo_thread:
                if tr.evaluation_id == ev.pk:
                    linked_reports.append({
                        'type': 'Thread',
                        'result': tr.result,
                        'complete': tr.is_complete,
                    })

        return {
            'type_code': eval_type,
            'type_label': type_label,
            'status': status,
            'is_required': is_required,
            'is_conditional': is_conditional,
            'url': url,
            'decision': ev.get_decision_display() if ev and ev.decision else '',
            'evaluated_by': (ev.evaluated_by.get_short_name() or ev.evaluated_by.username) if ev and ev.evaluated_by else '',
            'evaluated_at': ev.evaluated_at if ev else None,
            'created_at': ev.created_at if ev else None,
            'updated_at': ev.updated_at if ev else None,
            'linked_reports': linked_reports,
        }


class WorkOrderDetailEnhancedView(LoginRequiredMixin, DetailView):
    """
    Enhanced Work Order Detail with Job Card tabs.
    """
    model = WorkOrder
    template_name = "workorders/workorder_detail_enhanced.html"
    context_object_name = "work_order"

    def get_queryset(self):
        return WorkOrder.objects.select_related(
            "customer", "drill_bit", "assigned_to", "design", "bom",
            "sales_order", "rig", "well", "procedure", "department",
            "created_by", "evaluated_by", "qc_by", "reviewed_by_eng",
            "approved_by"
        ).prefetch_related(
            "documents", "photos", "materials", "time_logs",
            "cutter_evaluations__entries", "router_entries",
            "lpt_reports", "api_thread_inspections", "repair_boms__lines"
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo = self.object

        context["page_title"] = f"Job Card - {wo.wo_number}"

        # Generate QR code
        base_url = getattr(settings, "SITE_URL", None)
        context["qr_code"] = generate_work_order_qr(wo, base_url)

        # Get applicable instructions
        context["instructions"] = self.get_applicable_instructions(wo)

        # E-Checklist
        try:
            context["e_checklist"] = wo.evaluation_checklist
        except EvaluationChecklist.DoesNotExist:
            context["e_checklist"] = None

        # Cutter evaluations by type (legacy individual variables)
        context["ardt_evaluation"] = wo.cutter_evaluations.filter(
            evaluation_type=CutterEvaluationMatrix.EvaluationType.ARDT
        ).first()
        context["eng_evaluation"] = wo.cutter_evaluations.filter(
            evaluation_type=CutterEvaluationMatrix.EvaluationType.ENGINEER
        ).first()
        context["rework_evaluation"] = wo.cutter_evaluations.filter(
            evaluation_type=CutterEvaluationMatrix.EvaluationType.REWORK
        ).first()

        # Build evaluations list
        account_code = wo.account.code if wo.account else ''
        is_new_bit = wo.wo_type in [WorkOrder.WOType.FC_NEW, WorkOrder.WOType.RC_NEW]
        is_ur = account_code == 'UR'
        is_aramco = account_code == 'ARAMCO'

        # Try to get configured evaluation route for this WO
        # Wrapped in try/except to handle case where evaluation_routes table
        # doesn't exist (migration not applied or database sync issue)
        try:
            evaluation_route = EvaluationRoute.get_route_for_workorder(wo)
        except Exception:
            # Table doesn't exist or other DB error - fall back to legacy
            evaluation_route = None

        evaluations = []

        if evaluation_route:
            # Use configured route steps
            context['evaluation_route'] = evaluation_route
            for step in evaluation_route.steps.all().order_by('order'):
                eval_obj = wo.cutter_evaluations.filter(evaluation_type=step.evaluation_type).first()
                evaluations.append({
                    'type_code': step.evaluation_type,
                    'type_label': step.get_evaluation_type_display(),
                    'evaluation': eval_obj,
                    'exists': eval_obj is not None,
                    'entry_count': len(eval_obj.entries.all()) if eval_obj else 0,
                    'help_text': step.condition_description if step.is_conditional else '',
                    'is_na': False,
                    'is_required': step.is_required,
                    'is_conditional': step.is_conditional,
                    'show_decision_field': step.show_decision_field,
                    'show_cutter_matrix': step.show_cutter_matrix,
                    'show_cutters_details': step.show_cutters_details,
                })
        else:
            # Fallback: Legacy hardcoded flow
            context['evaluation_route'] = None
            eval_flow = [
                # (type_code, type_label, show_condition, help_text)
                ('PDC_EVAL', 'PDC Evaluation', True, 'Starting point - includes Die Check + E-Checklist'),
                ('ARDT', 'ARDT Evaluation (Legacy)', False, 'Legacy - use PDC Evaluation'),
                ('QC', 'QC Evaluation', not is_new_bit, 'N/A for new bits, required for repair'),
                ('ENGINEER', 'Technical Rep. Evaluation', not is_new_bit and not is_ur and not is_aramco, 'N/A for new bits, UR, or Aramco'),
                ('ARAMCO_REP', 'Aramco Rep. Evaluation', is_aramco, 'Aramco inspector evaluation'),
                ('DIE_CHECK', 'Die Check', True, 'Pre-processing die check'),
                ('FINAL_DIE_CHECK', 'Final Die Check', True, 'Post-processing die check'),
                ('FINAL_QC', 'Final QC Evaluation', True, 'Final quality check'),
                ('FINAL_INSPECTION', 'Final Inspection', True, 'Final sign-off'),
                ('REWORK', 'Rework Evaluation', False, 'Only shown when rework is needed'),
                ('RECEIVING', 'Receiving Evaluation', False, 'Component receiving - tracked separately'),
            ]

            for type_code, type_label, show_by_default, help_text in eval_flow:
                eval_obj = wo.cutter_evaluations.filter(evaluation_type=type_code).first()
                show = show_by_default or (eval_obj is not None)
                if show:
                    evaluations.append({
                        'type_code': type_code,
                        'type_label': type_label,
                        'evaluation': eval_obj,
                        'exists': eval_obj is not None,
                        'entry_count': len(eval_obj.entries.all()) if eval_obj else 0,
                        'help_text': help_text,
                        'is_na': not show_by_default and eval_obj is None,
                        'is_required': True,
                        'is_conditional': False,
                        'show_decision_field': True,
                        'show_cutter_matrix': True,
                        'show_cutters_details': True,
                    })

        context['evaluations'] = evaluations
        context['is_new_bit'] = is_new_bit
        context['is_ur'] = is_ur
        context['is_aramco'] = is_aramco

        # Evaluation progress summary
        total_evals = len([e for e in evaluations if not e.get('is_na')])
        complete_evals = len([e for e in evaluations if e.get('exists') and e['evaluation'].is_complete])
        context['eval_progress'] = {
            'total': total_evals,
            'complete': complete_evals,
            'percent': round(complete_evals * 100 / total_evals) if total_evals > 0 else 0,
        }

        # Receiving inspection (from drill bit, if linked)
        if wo.drill_bit_id:
            context['receiving_inspection'] = ReceivingInspection.objects.filter(
                drill_bit_id=wo.drill_bit_id
            ).select_related('inspected_by').first()
        else:
            context['receiving_inspection'] = None

        # Router sheet entries
        context["router_entries"] = wo.router_entries.order_by('step_number')

        # LPT Reports
        context["lpt_reports"] = wo.lpt_reports.order_by('-created_at')

        # API Thread Inspections
        context["api_inspections"] = wo.api_thread_inspections.order_by('-created_at')

        # Cost summary
        try:
            context["cost_summary"] = wo.cost_summary
        except WorkOrderCost.DoesNotExist:
            context["cost_summary"] = None

        # Status transitions — StatusTransitionLog removed (Feb 2026), was never written to
        context["status_history"] = []

        return context

    def get_applicable_instructions(self, work_order):
        """Get all instructions that apply to this work order."""
        instructions = []

        for rule in InstructionRule.objects.filter(is_active=True):
            # Check WO type filter
            if rule.applies_to_wo_types:
                if work_order.wo_type not in rule.applies_to_wo_types:
                    continue

            # Check bit type filter
            if rule.applies_to_bit_types and work_order.drill_bit:
                if work_order.drill_bit.bit_type not in rule.applies_to_bit_types:
                    continue

            # Check conditions
            conditions_met = True
            for condition in rule.conditions.all():
                if not condition.evaluate(work_order):
                    conditions_met = False
                    break

            if conditions_met:
                instructions.append(rule)

        return sorted(instructions, key=lambda x: -x.priority)


# =============================================================================
# DRILL BIT LIFECYCLE VIEWS
# =============================================================================

class DrillBitListEnhancedView(LoginRequiredMixin, ListView):
    """
    Enhanced Drill Bit List with lifecycle tracking.
    """
    model = DrillBit
    template_name = "workorders/drillbit_list_enhanced.html"
    context_object_name = "drill_bits"
    paginate_by = None

    def get_paginate_by(self, queryset):
        page_size = self.request.GET.get('page_size', 'all')
        if page_size == 'all':
            return None
        try:
            page_size = int(page_size)
            if page_size in [25, 50, 100, 200]:
                return page_size
        except (ValueError, TypeError):
            pass
        return None

    def get_queryset(self):
        queryset = DrillBit.objects.select_related(
            "design", "design__size", "design__connection_ref",
            "design__iadc_code_ref", "design__breaker_slot",
            "design__application_ref", "design__formation_type_ref",
            "bom", "bom__smi_type", "brazing_bom", "brazing_bom__smi_type",
            "system_bom", "system_bom__smi_type",
            "account",
            "customer", "rig", "well", "current_location", "bit_location"
        ).prefetch_related(
            "design__special_technologies"
        ).order_by("-created_at")

        # Filters
        status = self.request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)

        condition = self.request.GET.get("condition")
        if condition:
            queryset = queryset.filter(condition=condition)

        ownership = self.request.GET.get("ownership")
        if ownership:
            queryset = queryset.filter(ownership=ownership)

        bit_type = self.request.GET.get("bit_type")
        if bit_type:
            queryset = queryset.filter(bit_type=bit_type)

        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(serial_number__icontains=search)
                | Q(mat_number__icontains=search)
                | Q(customer__name__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Drill Bits"
        context["status_choices"] = DrillBit.Status.choices
        context["condition_choices"] = DrillBit.Condition.choices
        context["ownership_choices"] = DrillBit.Ownership.choices
        context["bit_type_choices"] = DrillBit.BitCategory.choices

        context["current_status"] = self.request.GET.get("status", "")
        context["current_condition"] = self.request.GET.get("condition", "")
        context["current_ownership"] = self.request.GET.get("ownership", "")
        context["current_bit_type"] = self.request.GET.get("bit_type", "")
        context["current_search"] = self.request.GET.get("search", "")

        from apps.sales.models import Account
        context["accounts"] = Account.objects.filter(is_active=True).order_by("sort_order", "code")

        # Bits currently in plan (for hiding Add to Plan button)
        context["bits_in_plan"] = set(
            ProductionPlanEntry.objects.filter(
                status=ProductionPlanEntry.Status.PLANNED
            ).values_list('drill_bit_id', flat=True)
        )

        # Bits with active WOs (set of bit PKs for quick lookup)
        context["bits_active_wo"] = set(
            WorkOrder.objects.filter(
                status__in=DrillBit.ACTIVE_WO_STATUSES,
                drill_bit__isnull=False
            ).values_list('drill_bit_id', flat=True)
        )

        paginate_by = self.get_paginate_by(None)
        context["page_size"] = 'all' if paginate_by is None else paginate_by
        context["total_count"] = self.get_queryset().count()

        return context


class DrillBitDetailEnhancedView(LoginRequiredMixin, DetailView):
    """
    Enhanced Drill Bit Detail with full lifecycle history.
    """
    model = DrillBit
    template_name = "workorders/drillbit_detail_enhanced.html"
    context_object_name = "drill_bit"

    def get_queryset(self):
        return DrillBit.objects.select_related(
            "design", "design__size", "design__connection_ref",
            "design__iadc_code_ref", "design__breaker_slot",
            "design__application_ref", "design__formation_type_ref",
            "design__upper_section_type", "design__connection_type_ref",
            "design__connection_size_ref",
            "customer", "rig", "well", "current_location",
            "bit_location", "bit_size_ref", "product_type", "created_by",
            "bom", "bom__smi_type", "brazing_bom", "brazing_bom__smi_type",
            "system_bom", "system_bom__smi_type",
        ).prefetch_related(
            "work_orders", "evaluations", "bit_events",
            "receiving_inspections", "receiving_inspections__inspected_by",
            "design__special_technologies",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = self.object

        context["page_title"] = f"Drill Bit - {bit.serial_number}"

        # Generate QR code
        base_url = getattr(settings, "SITE_URL", None)
        context["qr_code"] = generate_drill_bit_qr(bit, base_url)

        # Event history
        context["events"] = bit.bit_events.order_by('-event_date')[:20]

        # Work orders
        context["work_orders"] = bit.work_orders.order_by('-created_at')

        # Evaluations
        context["evaluations"] = bit.evaluations.order_by('-evaluation_date')

        # Receiving Inspections
        context["receiving_inspections"] = bit.receiving_inspections.order_by('-inspection_date')

        # ── BOM source_data for inline display ──
        active_bom = bit.brazing_bom or bit.bom or bit.system_bom
        if active_bom and active_bom.source_data:
            source_data = active_bom.source_data
            # Enrich cutter shapes from inventory
            try:
                from apps.cutter_map.views import _enrich_cutter_shapes_from_inventory
                _enrich_cutter_shapes_from_inventory(source_data)
            except Exception:
                pass
            context["bom_source_data"] = source_data
            # Build cutter_shapes dict {int(index): base64_data_uri}
            cutter_shapes = {}
            prefix = "data:image/png;base64,"
            raw_shapes = source_data.get("cutter_shapes", {})
            for k, v in raw_shapes.items():
                try:
                    idx = int(k)
                    data = v.get("data", "") if isinstance(v, dict) else (v if isinstance(v, str) else "")
                    # Fix double data-URI prefix
                    while data.startswith(prefix + prefix):
                        data = data[len(prefix):]
                    cutter_shapes[idx] = data
                except (ValueError, TypeError):
                    pass
            context["cutter_shapes"] = cutter_shapes
        context["active_bom"] = active_bom

        # ── Design pocket layout ──
        design = bit.design
        if design:
            import json as _json
            from apps.technology.models import DesignPocket, DesignPocketConfig
            pockets = DesignPocket.objects.filter(
                design=design
            ).select_related(
                'pocket_config', 'pocket_config__pocket_size', 'pocket_config__pocket_shape'
            ).order_by('blade_number', 'row_number', 'position_in_row')
            context["design_pockets"] = pockets

            pocket_configs = DesignPocketConfig.objects.filter(
                design=design
            ).select_related(
                'pocket_size', 'pocket_shape'
            ).order_by('order')
            context["pocket_configs"] = pocket_configs

            # Build grid data with row-aligned virtual columns
            # Pads shorter blades so R2 cutters always appear after R1 separator
            from collections import defaultdict
            pocket_list = list(pockets)

            blade_row_counts = defaultdict(int)
            for p in pocket_list:
                blade_row_counts[(p.blade_number, p.row_number)] += 1

            all_blades = sorted(set(p.blade_number for p in pocket_list))
            all_rows = sorted(set(p.row_number for p in pocket_list))

            row_max = {}
            for rn in all_rows:
                row_max[rn] = max(
                    blade_row_counts.get((bn, rn), 0) for bn in all_blades
                )

            row_start = {}
            off = 1
            for rn in all_rows:
                row_start[rn] = off
                off += row_max[rn]

            row_separators = []
            cum = 0
            for rn in all_rows[:-1]:
                cum += row_max[rn]
                row_separators.append(cum)

            grid_data = {}
            location_data = {}
            engagement_data = {}
            for p in pocket_list:
                vcol = row_start[p.row_number] + p.position_in_row - 1
                key = f"{p.blade_number}_{vcol}"
                grid_data[key] = p.pocket_config_id
                if p.blade_location:
                    location_data[key] = p.blade_location
                if p.engagement_order:
                    engagement_data[key] = p.engagement_order

            context["pocket_grid_data_json"] = _json.dumps(grid_data)
            context["pocket_row_separators_json"] = _json.dumps(sorted(row_separators))
            context["pocket_location_data_json"] = _json.dumps(location_data)
            context["pocket_engagement_data_json"] = _json.dumps(engagement_data)

            # Generate display colors for configs (same as DesignPocketsView)
            colors = ['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6',
                      '#EC4899', '#06B6D4', '#84CC16', '#F97316', '#6366F1']
            config_data = {}
            for i, cfg in enumerate(pocket_configs):
                dc = cfg.color_code if cfg.color_code else colors[i % len(colors)]
                cfg.display_color = dc
                config_data[cfg.pk] = {
                    "count": cfg.count,
                    "order": cfg.order,
                    "color": dc,
                    "name": cfg.pocket_size.display_name if cfg.pocket_size else '',
                    "sizeCode": cfg.pocket_size.code if cfg.pocket_size else '',
                }
            context["pocket_config_data_json"] = _json.dumps(config_data)
            context["config_total"] = sum(cfg.count for cfg in pocket_configs)

        return context


# =============================================================================
# CUTTER EVALUATION MATRIX VIEWS
# =============================================================================

class CutterEvaluationCreateView(LoginRequiredMixin, CreateView):
    """
    Create a new cutter evaluation matrix for a work order.
    """
    model = CutterEvaluationMatrix
    template_name = "workorders/cutter_evaluation_form.html"
    fields = ['evaluation_type', 'ardt_remark', 'eng_remark', 'general_remark',
              'ardt_matrix_buildup', 'eng_matrix_buildup', 'ncr_ref_no']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        context['work_order'] = wo
        context['page_title'] = f"Cutter Evaluation - {wo.wo_number}"

        # Get pocket layout from design if available
        if wo.design:
            context['pocket_layout'] = wo.design.pockets.order_by('blade_number', 'row_number', 'position_in_row')

        return context

    def form_valid(self, form):
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        form.instance.work_order = wo
        form.instance.evaluated_by = self.request.user
        form.instance.evaluated_at = timezone.now()

        # Set evaluation number
        existing = CutterEvaluationMatrix.objects.filter(
            work_order=wo,
            evaluation_type=form.cleaned_data['evaluation_type']
        ).count()
        form.instance.evaluation_number = existing + 1

        # Apply section defaults based on eval type + account workflow type
        workflow_type = None
        if wo.account and hasattr(wo.account, 'workflow_type'):
            workflow_type = wo.account.workflow_type
        form.instance.apply_section_defaults(workflow_type=workflow_type)

        # Apply user overrides from section checkboxes (override defaults)
        section_fields = [
            'include_checklist', 'include_cutter_grid', 'include_pocket_eval',
            'include_die_check', 'include_pressure_test', 'include_thread_inspection',
        ]
        for field in section_fields:
            setattr(form.instance, field, field in self.request.POST)

        # Generate inspection number (EV-YYYY-NNNN)
        from django.utils import timezone as tz
        year = tz.now().year
        last = CutterEvaluationMatrix.objects.filter(
            inspection_number__startswith=f'EV-{year}-'
        ).order_by('-inspection_number').first()
        if last and last.inspection_number:
            try:
                seq = int(last.inspection_number.split('-')[-1]) + 1
            except (ValueError, IndexError):
                seq = 1
        else:
            seq = 1
        form.instance.inspection_number = f'EV-{year}-{seq:04d}'

        return super().form_valid(form)

    def get_success_url(self):
        # PDC_EVAL uses the new pre-repair evaluation page
        if self.object.evaluation_type == 'PDC_EVAL':
            return reverse('workorders:pre_repair_eval_edit', kwargs={
                'wo_pk': self.kwargs['wo_pk'],
                'pk': self.object.pk
            })
        return reverse('workorders:cutter_evaluation_edit', kwargs={
            'wo_pk': self.kwargs['wo_pk'],
            'pk': self.object.pk
        })


class CutterEvaluationEditView(LoginRequiredMixin, TemplateView):
    """
    Interactive cutter evaluation matrix editor.
    """
    template_name = "workorders/cutter_evaluation_matrix.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        matrix = get_object_or_404(CutterEvaluationMatrix, pk=self.kwargs['pk'])

        context['work_order'] = wo
        context['matrix'] = matrix
        context['page_title'] = f"{matrix.get_evaluation_type_display()} - {wo.wo_number}"

        # Build the blade/cutter grid
        # Determine max blades and cutters from design or default
        max_blades = 12
        max_cutters = 15

        if wo.design:
            # Get from pocket layout
            pockets = wo.design.pockets.all()
            if pockets.exists():
                max_blades = max(p.blade_number for p in pockets)
                max_cutters = max(p.position_in_blade for p in pockets)

        # Build evaluation grid
        grid = {}
        for entry in matrix.entries.all():
            grid[(entry.blade_number, entry.cutter_position)] = entry

        # Build rows for template
        rows = []
        for blade in range(1, max_blades + 1):
            row = {'blade': blade, 'cutters': []}
            for cutter in range(1, max_cutters + 1):
                entry = grid.get((blade, cutter))
                row['cutters'].append({
                    'position': cutter,
                    'entry': entry,
                    'action': entry.action if entry else '',
                })
            rows.append(row)

        context['rows'] = rows
        context['max_cutters'] = max_cutters
        context['cutter_positions'] = list(range(1, max_cutters + 1))
        context['total_grid_height'] = max_blades * 30
        context['action_choices'] = CutterEvaluationEntry.Action.choices
        context['source_choices'] = CutterEvaluationEntry.CutterSource.choices

        # Saved cutters details (from previous save or BOM data)
        # Use json.dumps for safe JS embedding (not Python repr)
        context['saved_cutters_details'] = _json.dumps(matrix.cutters_details or [])

        # BOM lines for pre-populating cutters details (if no saved data)
        bom_lines_json = []
        active_bom = None
        if wo.bom:
            active_bom = wo.bom
        elif wo.drill_bit:
            active_bom = getattr(wo.drill_bit, 'brazing_bom', None) or getattr(wo.drill_bit, 'bom', None)
        if active_bom:
            for line in active_bom.lines.select_related('inventory_item').order_by('order_number', 'line_number'):
                part_no = line.hdbs_code or ''
                desc = ''
                if line.inventory_item:
                    part_no = part_no or line.inventory_item.mat_number or line.inventory_item.code
                    desc = line.inventory_item.name
                else:
                    desc = line.cutter_type or ''
                bom_lines_json.append({
                    'qty': line.quantity,
                    'size_mm': line.cutter_size or '',
                    'part_no': part_no,
                    'description': desc,
                    'remarks': '',
                })
        context['bom_lines_json'] = _json.dumps(bom_lines_json)

        # Dynamic row range for cutters details table
        num_detail_rows = max(10, len(bom_lines_json), len(matrix.cutters_details or []))
        context['cutter_detail_row_range'] = range(1, num_detail_rows + 1)

        # Cutter state history: all prior evaluations for this WO
        prior_evals = wo.cutter_evaluations.exclude(pk=matrix.pk).order_by(
            'evaluation_type', 'evaluation_number'
        ).prefetch_related('entries')
        eval_history = []
        for ev in prior_evals:
            eval_history.append({
                'type': ev.get_evaluation_type_display(),
                'number': ev.evaluation_number,
                'decision': ev.get_decision_display() if ev.decision else '',
                'date': ev.evaluated_at.isoformat() if ev.evaluated_at else '',
                'entry_count': ev.entries.count(),
            })
        context['eval_history'] = eval_history

        # Build cumulative cutter state from all prior evaluations
        # Track per position: last action across all evaluations
        cutter_state = {}
        all_evals_ordered = wo.cutter_evaluations.order_by('created_at').prefetch_related('entries')
        for ev in all_evals_ordered:
            if ev.pk == matrix.pk:
                continue
            for entry in ev.entries.all():
                key = (entry.blade_number, entry.cutter_position)
                if entry.action:
                    prev = cutter_state.get(key, {})
                    history = prev.get('history', [])
                    history.append({
                        'eval_type': ev.get_evaluation_type_display(),
                        'action': entry.action,
                    })
                    cutter_state[key] = {
                        'last_action': entry.action,
                        'history': history,
                    }
        # Convert tuple keys to string for JSON serialization
        cutter_state_serializable = {}
        for (blade, pos), val in cutter_state.items():
            cutter_state_serializable[f"{blade},{pos}"] = val
        context['cutter_state_json'] = _json.dumps(cutter_state_serializable)

        # Section visibility from model flags
        context['include_checklist'] = matrix.include_checklist
        context['include_cutter_grid'] = matrix.include_cutter_grid
        context['include_pocket_eval'] = matrix.include_pocket_eval
        context['include_die_check'] = matrix.include_die_check
        context['include_pressure_test'] = matrix.include_pressure_test
        context['include_thread_inspection'] = matrix.include_thread_inspection

        # Backward compat aliases used in template
        context['has_checklist'] = matrix.include_checklist
        context['has_die_check'] = matrix.include_die_check

        # Section toggle data for dropdown
        context['section_toggles'] = [
            ('include_checklist', 'Checklist'),
            ('include_cutter_grid', 'Cutter Grid'),
            ('include_pocket_eval', 'Pocket Evaluation'),
            ('include_die_check', 'Die Check'),
            ('include_pressure_test', 'LPT (Pressure Test)'),
            ('include_thread_inspection', 'API Thread Inspection'),
        ]
        context['active_sections'] = [
            k for k, _ in context['section_toggles'] if getattr(matrix, k, False)
        ]

        # Checklist items for this evaluation type
        checklist_items = CutterEvaluationMatrix.CHECKLIST_ITEMS.get(
            matrix.evaluation_type, []
        )
        saved_checklist = matrix.checklist_data or []
        saved_map = {}
        for item in saved_checklist:
            saved_map[item.get('item', '')] = item
        checklist_for_template = []
        for idx, item_label in enumerate(checklist_items, 1):
            saved = saved_map.get(item_label, {})
            checklist_for_template.append({
                'number': idx,
                'item': item_label,
                'status': saved.get('status', ''),
                'remarks': saved.get('remarks', ''),
            })
        context['checklist_items'] = checklist_for_template
        context['checklist_json'] = _json.dumps(checklist_for_template)

        # Die check data
        context['die_check_json'] = _json.dumps(matrix.die_check_data or {})

        # Pressure test data (LPT)
        context['pressure_test_json'] = _json.dumps(matrix.pressure_test_data or {})

        # Thread inspection data
        context['thread_inspection_json'] = _json.dumps(matrix.thread_inspection_data or {})

        return context

    def post(self, request, *args, **kwargs):
        """Handle grid updates via AJAX — supports both single-cell and bulk JSON."""
        matrix = get_object_or_404(CutterEvaluationMatrix, pk=kwargs['pk'])

        content_type = request.content_type or ''
        if 'application/json' in content_type:
            data = _json.loads(request.body)

            # Handle section toggle
            if 'toggle_section' in data:
                section_key = data['toggle_section']
                enabled = data.get('enabled', False)
                valid_sections = {
                    'include_checklist', 'include_cutter_grid', 'include_pocket_eval',
                    'include_die_check', 'include_pressure_test', 'include_thread_inspection',
                }
                if section_key in valid_sections:
                    setattr(matrix, section_key, bool(enabled))
                    matrix.save(update_fields=[section_key, 'updated_at'])
                return JsonResponse({'success': True})

            entries_data = data.get('entries', [])
            remarks = data.get('remarks', '')
            decision = data.get('decision', '')
            cutters_details = data.get('cutters_details', None)
            checklist_data = data.get('checklist_data', None)
            die_check_data = data.get('die_check_data', None)
            pressure_test_data = data.get('pressure_test_data', None)
            thread_inspection_data = data.get('thread_inspection_data', None)
            mark_complete = data.get('mark_complete', None)

            with transaction.atomic():
                # Clear existing entries and re-create
                matrix.entries.all().delete()
                for e in entries_data:
                    CutterEvaluationEntry.objects.create(
                        matrix=matrix,
                        blade_number=e['blade'],
                        cutter_position=e['position'],
                        action=e.get('action', ''),
                    )

                # Update matrix fields
                update_fields = ['general_remark', 'decision', 'cutters_details', 'updated_at']
                matrix.general_remark = remarks
                matrix.decision = decision
                if cutters_details is not None:
                    matrix.cutters_details = cutters_details
                if checklist_data is not None:
                    matrix.checklist_data = checklist_data
                    update_fields.append('checklist_data')
                if die_check_data is not None:
                    matrix.die_check_data = die_check_data
                    update_fields.append('die_check_data')
                if pressure_test_data is not None:
                    matrix.pressure_test_data = pressure_test_data
                    update_fields.append('pressure_test_data')
                if thread_inspection_data is not None:
                    matrix.thread_inspection_data = thread_inspection_data
                    update_fields.append('thread_inspection_data')

                # Mark complete / reopen support
                if mark_complete is True:
                    matrix.is_complete = True
                    matrix.qc_by = request.user
                    matrix.qc_at = timezone.now()
                    update_fields.extend(['is_complete', 'qc_by', 'qc_at'])
                elif mark_complete is False:
                    matrix.is_complete = False
                    update_fields.append('is_complete')

                matrix.save(update_fields=update_fields)

            return JsonResponse({
                'success': True,
                'count': len(entries_data),
                'is_complete': matrix.is_complete,
            })

        # Legacy single-cell POST
        blade = int(request.POST.get('blade'))
        position = int(request.POST.get('position'))
        action = request.POST.get('action', '')
        source = request.POST.get('source', '')
        notes = request.POST.get('notes', '')

        entry, created = CutterEvaluationEntry.objects.update_or_create(
            matrix=matrix,
            blade_number=blade,
            cutter_position=position,
            defaults={
                'action': action,
                'cutter_source': source,
                'notes': notes,
            }
        )

        return JsonResponse({
            'success': True,
            'entry_id': entry.pk,
            'action': entry.get_action_display(),
        })


@login_required
@require_POST
def api_evaluation_mark_complete(request, wo_pk, pk):
    """Toggle is_complete on a CutterEvaluationMatrix. Sets qc_by/qc_at on first completion."""
    matrix = get_object_or_404(CutterEvaluationMatrix, pk=pk, work_order_id=wo_pk)
    data = _json.loads(request.body) if request.body else {}
    mark_complete = data.get('mark_complete', True)

    with transaction.atomic():
        matrix.is_complete = bool(mark_complete)
        update_fields = ['is_complete', 'updated_at']
        if mark_complete and not matrix.qc_by:
            matrix.qc_by = request.user
            matrix.qc_at = timezone.now()
            update_fields.extend(['qc_by', 'qc_at'])
        elif not mark_complete:
            # Reopen — keep qc_by/qc_at as audit trail
            pass
        matrix.save(update_fields=update_fields)

    wo = matrix.work_order
    if mark_complete:
        notify(
            actor=request.user,
            verb=f"completed {matrix.get_evaluation_type_display()} evaluation for",
            target=wo.wo_number,
            priority="HIGH",
            action_url=f"/workorders/enhanced/{wo.pk}/",
            entity_type="CutterEvaluationMatrix",
            entity_id=matrix.pk,
        )

    return JsonResponse({
        'success': True,
        'is_complete': matrix.is_complete,
        'qc_by': str(matrix.qc_by) if matrix.qc_by else None,
        'qc_at': matrix.qc_at.isoformat() if matrix.qc_at else None,
    })


# =============================================================================
# ROUTER SHEET VIEWS
# =============================================================================

class RouterSheetView(LoginRequiredMixin, TemplateView):
    """
    Router Sheet with step-by-step tracking and QR scanning.
    """
    template_name = "workorders/router_sheet.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        wo = get_object_or_404(WorkOrder, pk=self.kwargs['pk'])
        context['work_order'] = wo
        context['page_title'] = f"Router Sheet - {wo.wo_number}"

        # Get or create router entries from process route
        entries = list(wo.router_entries.order_by('step_number'))

        # If no entries exist, create from the appropriate ProcessRoute
        if not entries:
            route = self._get_route_for_wo(wo)
            if route:
                for op in route.operations.order_by('sequence'):
                    RouterSheetEntry.objects.get_or_create(
                        work_order=wo,
                        step_number=op.sequence,
                        defaults={
                            'step_description': op.operation_name,
                        }
                    )
                entries = list(wo.router_entries.order_by('step_number'))

        context['entries'] = entries

        # Generate WO QR for scanning
        base_url = getattr(settings, "SITE_URL", None)
        context['wo_qr'] = generate_work_order_qr(wo, base_url)

        return context

    def _get_route_for_wo(self, wo):
        """Find the appropriate ProcessRoute for this work order."""
        # Determine workflow type from account
        is_manufacture = False
        if wo.account:
            if wo.account.workflow_type == 'MANUFACTURE':
                is_manufacture = True
            elif wo.account.code in ('L3', 'L4'):
                is_manufacture = True

        # Also check WO type
        if wo.wo_type in ('FC_NEW', 'RC_NEW'):
            is_manufacture = True

        if is_manufacture:
            route = ProcessRoute.objects.filter(
                route_number='RT-FC-MANUFACTURE', is_active=True
            ).first()
        else:
            route = ProcessRoute.objects.filter(
                route_number='RT-FC-REPAIR', is_active=True
            ).first()

        return route


@login_required
def router_step_scan(request, wo_pk, step_number):
    """
    Handle QR scan for router step start/end.
    """
    wo = get_object_or_404(WorkOrder, pk=wo_pk)
    entry = get_object_or_404(RouterSheetEntry, work_order=wo, step_number=step_number)

    action = request.POST.get('action', 'start')
    station_qr = request.POST.get('station_qr', '')

    if action == 'start':
        entry.qr_scan_start = timezone.now()
        entry.operator = request.user
        entry.station_qr = station_qr
        entry.save()
        messages.success(request, f"Step {step_number} started")
    elif action == 'end':
        entry.qr_scan_end = timezone.now()
        entry.is_complete = True
        entry.save()
        messages.success(request, f"Step {step_number} completed")

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'step_number': step_number,
            'is_complete': entry.is_complete,
            'duration': entry.duration_minutes,
        })

    return redirect('workorders:router_sheet', pk=wo_pk)


@login_required
def api_router_step_scan(request, wo_pk, step_number):
    """
    API endpoint for QR scan - start/complete a router step.
    Returns JSON response for HTMX/fetch usage.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    wo = get_object_or_404(WorkOrder, pk=wo_pk)
    entry = get_object_or_404(RouterSheetEntry, work_order=wo, step_number=step_number)

    action = request.POST.get('action', 'start')
    station_qr = request.POST.get('station_qr', '')

    if action == 'start':
        if entry.qr_scan_start:
            return JsonResponse({'error': 'Step already started', 'success': False})
        entry.qr_scan_start = timezone.now()
        entry.operator = request.user
        entry.station_qr = station_qr
        entry.save()

        return JsonResponse({
            'success': True,
            'step_number': step_number,
            'action': 'started',
            'started_at': entry.qr_scan_start.isoformat(),
            'operator': request.user.get_short_name() or request.user.username,
        })
    elif action == 'end':
        if not entry.qr_scan_start:
            return JsonResponse({'error': 'Step not started yet', 'success': False})
        if entry.is_complete:
            return JsonResponse({'error': 'Step already completed', 'success': False})
        entry.qr_scan_end = timezone.now()
        entry.is_complete = True
        entry.save()

        # Notify only when ALL router steps are now complete
        total = RouterSheetEntry.objects.filter(work_order=wo).count()
        done = RouterSheetEntry.objects.filter(work_order=wo, is_complete=True).count()
        if total > 0 and done >= total:
            # All steps done — move bit to Finished Goods
            if wo.drill_bit:
                wo.drill_bit.move_to('WH-FG', f'All router steps completed — WO {wo.wo_number}', request.user)
            notify(
                actor=request.user,
                verb="completed all router steps for",
                target=wo.wo_number,
                priority="HIGH",
                action_url=f"/workorders/{wo.pk}/router-sheet/",
                entity_type="WorkOrder",
                entity_id=wo.pk,
            )

        return JsonResponse({
            'success': True,
            'step_number': step_number,
            'action': 'completed',
            'completed_at': entry.qr_scan_end.isoformat(),
            'duration_minutes': entry.duration_minutes,
        })
    elif action == 'skip':
        entry.is_complete = True
        entry.save()
        return JsonResponse({
            'success': True,
            'step_number': step_number,
            'action': 'skipped',
        })

    return JsonResponse({'error': f'Unknown action: {action}', 'success': False})


# =============================================================================
# QC FORMS VIEWS
# =============================================================================

class EvaluationChecklistView(LoginRequiredMixin, UpdateView):
    """
    E-Checklist form for FC Bit Evaluation.
    """
    model = EvaluationChecklist
    template_name = "workorders/e_checklist_form.html"
    fields = [
        'bit_cleanliness', 'bit_cleanliness_remarks',
        'paperwork', 'paperwork_remarks',
        'bit_stamping', 'bit_stamping_remarks',
        'die_check', 'die_check_remarks',
        'ring_gauge_go', 'ring_gauge_go_remarks',
        'ring_gauge_no_go', 'ring_gauge_no_go_remarks',
        'nozzle_bore_liner', 'nozzle_bore_liner_remarks',
        'nozzle_threads', 'nozzle_threads_remarks',
        'apex', 'apex_remarks',
        'junk_slot', 'junk_slot_remarks',
        'breaker_slot', 'breaker_slot_remarks',
        'body_condition', 'body_condition_remarks',
        'mud_seal_surface', 'mud_seal_surface_remarks',
        'api_pin', 'api_pin_remarks',
        'inner_diameter', 'inner_diameter_remarks',
        'overall_pass', 'general_remarks',
    ]

    def get_object(self, queryset=None):
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        checklist, created = EvaluationChecklist.objects.get_or_create(
            work_order=wo,
            defaults={'inspector': self.request.user}
        )
        return checklist

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['work_order'] = self.object.work_order
        context['page_title'] = f"E-Checklist - {self.object.work_order.wo_number}"
        context['result_choices'] = EvaluationChecklist.Result.choices
        return context

    def form_valid(self, form):
        form.instance.inspector = self.request.user
        form.instance.inspection_date = timezone.now().date()

        # Track item timestamps for auditing
        old_instance = EvaluationChecklist.objects.filter(pk=form.instance.pk).first()
        item_timestamps = form.instance.item_timestamps or {}
        now = timezone.now().isoformat()

        # Check which fields changed and update their timestamps
        checklist_fields = [
            'bit_cleanliness', 'paperwork', 'bit_stamping', 'die_check',
            'ring_gauge_go', 'ring_gauge_no_go', 'nozzle_bore_liner', 'nozzle_threads',
            'apex', 'junk_slot', 'breaker_slot', 'body_condition',
            'mud_seal_surface', 'api_pin', 'inner_diameter', 'overall_pass'
        ]
        for field in checklist_fields:
            new_value = form.cleaned_data.get(field)
            old_value = getattr(old_instance, field, '') if old_instance else ''
            if new_value and new_value != old_value:
                item_timestamps[field] = now

        form.instance.item_timestamps = item_timestamps
        messages.success(self.request, "E-Checklist saved successfully")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('workorders:workorder_detail_enhanced', kwargs={'pk': self.kwargs['wo_pk']})


class LPTReportCreateView(LoginRequiredMixin, CreateView):
    """
    Create LPT Report.
    """
    model = LPTReport
    template_name = "workorders/lpt_report_form.html"
    fields = [
        'test_type', 'technique', 'procedure_ref',
        'cleaner_product', 'cleaner_batch', 'cleaner_expiry',
        'penetrant_product', 'penetrant_batch', 'penetrant_expiry',
        'developer_product', 'developer_batch', 'developer_expiry',
        'surface_temperature', 'penetrant_dwell_time',
        'light_intensity', 'developer_dwell_time',
        'result', 'disposition',
    ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        context['work_order'] = wo
        context['page_title'] = f"LPT Report - {wo.wo_number}"
        return context

    def form_valid(self, form):
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        form.instance.work_order = wo
        form.instance.lpt_operator = self.request.user
        form.instance.operator_date = timezone.now().date()

        # Generate report number
        count = LPTReport.objects.filter(work_order=wo).count()
        form.instance.report_number = f"LPT-{wo.wo_number}-{count + 1:02d}"

        messages.success(self.request, "LPT Report created successfully")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('workorders:workorder_detail_enhanced', kwargs={'pk': self.kwargs['wo_pk']})


class APIThreadInspectionCreateView(LoginRequiredMixin, CreateView):
    """
    Create API Thread Inspection.
    """
    model = APIThreadInspection
    template_name = "workorders/api_thread_form.html"
    fields = [
        'pin_size',
        'pin_face_ok', 'pin_face_remarks',
        'thread_ok', 'thread_remarks',
        'pitch_gauge_ok', 'pitch_gauge_remarks',
        'mud_seal_ok', 'mud_seal_remarks',
        'other_observation', 'pin_height',
        'thread_repair_required',
        'repair_brush_selected', 'upper_section_replacement',
        'initial_result',
    ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        context['work_order'] = wo
        context['page_title'] = f"API Thread Inspection - {wo.wo_number}"
        return context

    def form_valid(self, form):
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        form.instance.work_order = wo
        form.instance.inspector = self.request.user
        form.instance.inspection_date = timezone.now().date()

        # Generate inspection number
        count = APIThreadInspection.objects.filter(work_order=wo).count()
        form.instance.inspection_number = f"API-{wo.wo_number}-{count + 1:02d}"

        messages.success(self.request, "API Thread Inspection created successfully")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('workorders:workorder_detail_enhanced', kwargs={'pk': self.kwargs['wo_pk']})


# =============================================================================
# INSTRUCTION RULES VIEWS
# =============================================================================

class InstructionRuleListView(LoginRequiredMixin, ListView):
    """
    List all instruction rules.
    """
    model = InstructionRule
    template_name = "workorders/instruction_rule_list.html"
    context_object_name = "rules"
    paginate_by = 25

    def get_queryset(self):
        queryset = InstructionRule.objects.prefetch_related('conditions').order_by('-priority', 'name')

        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(instruction_text__icontains=search)
            )

        active = self.request.GET.get('active')
        if active == 'yes':
            queryset = queryset.filter(is_active=True)
        elif active == 'no':
            queryset = queryset.filter(is_active=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Instruction Rules'
        context['current_search'] = self.request.GET.get('search', '')
        context['current_active'] = self.request.GET.get('active', '')
        return context


class InstructionRuleCreateView(LoginRequiredMixin, CreateView):
    """
    Create a new instruction rule with inline conditions.
    """
    model = InstructionRule
    template_name = "workorders/instruction_rule_form.html"
    fields = ['name', 'description', 'instruction_text', 'priority',
              'applies_to_wo_types', 'applies_to_bit_types', 'is_active']
    success_url = reverse_lazy('workorders:instruction_rule_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create Instruction Rule'
        context['wo_type_choices'] = WorkOrder.WOType.choices
        context['bit_type_choices'] = DrillBit.BitCategory.choices

        # Add condition formset
        from .forms import InstructionRuleConditionFormSet
        if self.request.POST:
            context['condition_formset'] = InstructionRuleConditionFormSet(
                self.request.POST, prefix='conditions'
            )
        else:
            context['condition_formset'] = InstructionRuleConditionFormSet(prefix='conditions')
        return context

    def form_valid(self, form):
        from .forms import InstructionRuleConditionFormSet
        context = self.get_context_data()
        condition_formset = context['condition_formset']

        form.instance.created_by = self.request.user
        self.object = form.save()

        if condition_formset.is_valid():
            condition_formset.instance = self.object
            condition_formset.save()
        else:
            # Re-render form with errors
            return self.render_to_response(self.get_context_data(form=form))

        messages.success(self.request, "Instruction rule created successfully")
        return redirect(self.success_url)


class InstructionRuleUpdateView(LoginRequiredMixin, UpdateView):
    """
    Update an instruction rule with inline conditions.
    """
    model = InstructionRule
    template_name = "workorders/instruction_rule_form.html"
    fields = ['name', 'description', 'instruction_text', 'priority',
              'applies_to_wo_types', 'applies_to_bit_types', 'is_active']
    success_url = reverse_lazy('workorders:instruction_rule_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Rule: {self.object.name}'
        context['wo_type_choices'] = WorkOrder.WOType.choices
        context['bit_type_choices'] = DrillBit.BitCategory.choices

        # Add condition formset
        from .forms import InstructionRuleConditionFormSet
        if self.request.POST:
            context['condition_formset'] = InstructionRuleConditionFormSet(
                self.request.POST, instance=self.object, prefix='conditions'
            )
        else:
            context['condition_formset'] = InstructionRuleConditionFormSet(
                instance=self.object, prefix='conditions'
            )
        return context

    def form_valid(self, form):
        from .forms import InstructionRuleConditionFormSet
        context = self.get_context_data()
        condition_formset = context['condition_formset']

        self.object = form.save()

        if condition_formset.is_valid():
            condition_formset.instance = self.object
            condition_formset.save()
        else:
            # Re-render form with errors
            return self.render_to_response(self.get_context_data(form=form))

        messages.success(self.request, "Instruction rule updated successfully")
        return redirect(self.success_url)


class InstructionRuleDeleteView(LoginRequiredMixin, DeleteView):
    """
    Delete an instruction rule.
    """
    model = InstructionRule
    template_name = "workorders/instruction_rule_confirm_delete.html"
    success_url = reverse_lazy('workorders:instruction_rule_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Instruction rule deleted successfully")
        return super().delete(request, *args, **kwargs)


# =============================================================================
# EXPORT VIEWS
# =============================================================================

@login_required
def export_work_orders_excel(request):
    """
    Export work orders to Excel format.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, "Excel export requires openpyxl library")
        return redirect('workorders:workorder_list_enhanced')

    # Apply same filters as list view
    queryset = WorkOrder.objects.select_related(
        "customer", "drill_bit", "assigned_to", "design"
    ).order_by("-created_at")

    status = request.GET.get("status")
    if status:
        queryset = queryset.filter(status=status)

    wo_type = request.GET.get("wo_type")
    if wo_type:
        queryset = queryset.filter(wo_type=wo_type)

    search = request.GET.get("search")
    if search:
        queryset = queryset.filter(
            Q(wo_number__icontains=search)
            | Q(customer__name__icontains=search)
            | Q(drill_bit__serial_number__icontains=search)
        )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Orders"

    # Headers
    headers = [
        'WO Number', 'Type', 'Status', 'Priority',
        'Serial Number', 'Customer', 'DRSS#', 'Brazing MAT#',
        'Due Date', 'Assigned To', 'Created Date'
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_num, wo in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=wo.wo_number)
        ws.cell(row=row_num, column=2, value=wo.get_wo_type_display())
        ws.cell(row=row_num, column=3, value=wo.get_status_display())
        ws.cell(row=row_num, column=4, value=wo.get_priority_display())
        ws.cell(row=row_num, column=5, value=wo.drill_bit.serial_number if wo.drill_bit else '')
        ws.cell(row=row_num, column=6, value=wo.customer.name if wo.customer else '')
        ws.cell(row=row_num, column=7, value=wo.drss_no)
        ws.cell(row=row_num, column=8, value=wo.brazing_mat_no)
        ws.cell(row=row_num, column=9, value=wo.due_date.strftime('%Y-%m-%d') if wo.due_date else '')
        ws.cell(row=row_num, column=10, value=wo.assigned_to.get_full_name() if wo.assigned_to else '')
        ws.cell(row=row_num, column=11, value=wo.created_at.strftime('%Y-%m-%d'))

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"work_orders_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# =============================================================================
# PRODUCTION PLANNER - WIP Dashboard with Real-time Updates
# =============================================================================

class ProductionPlannerView(LoginRequiredMixin, TemplateView):
    """
    Production Planner Dashboard - V2 dark theme (matches WO list page).
    Shows planned bits, ready-for-planning bits, WIP, and completed items.
    All data serialized to JSON for JS-driven rendering.
    """
    template_name = "workorders/production_planner.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        import json as _j
        from apps.sales.models import Account

        accounts = Account.objects.filter(is_active=True).order_by('sort_order', 'code')
        account_filter = self.request.GET.get('account')

        wip_statuses = [
            WorkOrder.Status.DRAFT, WorkOrder.Status.PLANNED,
            WorkOrder.Status.RELEASED, WorkOrder.Status.IN_PROGRESS,
            WorkOrder.Status.ON_HOLD, WorkOrder.Status.QC_PENDING,
            WorkOrder.Status.QC_FAILED,
        ]
        completed_statuses = [
            WorkOrder.Status.QC_PASSED, WorkOrder.Status.COMPLETED,
        ]

        today = timezone.now().date()

        # ── READY FOR PLANNING (bits with RECEIVED status, no WO, not in plan) ──
        ready_qs = DrillBit.objects.filter(
            status=DrillBit.Status.RECEIVED
        ).select_related('design', 'design__size', 'account', 'brazing_bom', 'system_bom', 'bom').order_by('-updated_at')
        if account_filter:
            ready_qs = ready_qs.filter(account__code=account_filter)

        # Exclude bits with active WO or already in plan
        active_wo_bit_ids = set(
            WorkOrder.objects.filter(status__in=wip_statuses)
            .values_list('drill_bit_id', flat=True)
        )
        in_plan_bit_ids = set(
            ProductionPlanEntry.objects.filter(status=ProductionPlanEntry.Status.PLANNED)
            .values_list('drill_bit_id', flat=True)
        )
        ready_bits_json = []
        for bit in ready_qs:
            if bit.pk in active_wo_bit_ids or bit.pk in in_plan_bit_ids:
                continue
            days_waiting = (today - bit.received_date).days if bit.received_date else 0
            ready_bits_json.append({
                'id': bit.pk,
                'serial': bit.serial_number,
                'size': str(bit.size) if bit.size else '',
                'type': bit.design.hdbs_type if bit.design else '',
                'designMat': bit.design.mat_no if bit.design else '',
                'refMat': bit.design.ref_mat_no if bit.design and bit.design.ref_mat_no else '',
                'systemMat': (bit.brazing_bom.system_mat_no if bit.brazing_bom and bit.brazing_bom.system_mat_no
                              else (bit.bom.system_mat_no if bit.bom and bit.bom.system_mat_no
                              else (bit.system_bom.code if bit.system_bom else ''))),
                'brazingMat': (bit.brazing_bom.code if bit.brazing_bom else (bit.bom.code if bit.bom else '')),
                'account': bit.account.code if bit.account else '',
                'accountName': bit.account.name if bit.account else '',
                'received': bit.received_date.strftime('%Y-%m-%d') if bit.received_date else '',
                'daysWaiting': days_waiting,
                'repair': bit.revision_number or 0,
                'condition': bit.condition or '',
                'level': bit.level or (bit.design.order_level if bit.design else ''),
            })

        # ── PLANNED ENTRIES ──
        planned_qs = ProductionPlanEntry.objects.filter(
            status__in=[ProductionPlanEntry.Status.PLANNED, ProductionPlanEntry.Status.PENDING_RELEASE]
        ).select_related(
            'drill_bit', 'drill_bit__design', 'drill_bit__brazing_bom', 'drill_bit__system_bom',
            'drill_bit__bom', 'drill_bit__account', 'account', 'created_by'
        ).order_by('sequence', '-priority', 'planned_date')
        if account_filter:
            planned_qs = planned_qs.filter(account__code=account_filter)

        planned_json = []
        for entry in planned_qs:
            bit = entry.drill_bit
            _requester = ''
            if entry.created_by:
                _requester = entry.created_by.get_full_name() or entry.created_by.username
            planned_json.append({
                'entryId': entry.pk,
                'bitId': bit.pk,
                'serial': bit.serial_number,
                'size': str(bit.size) if bit.size else '',
                'type': bit.design.hdbs_type if bit.design else '',
                'designMat': bit.design.mat_no if bit.design else '',
                'refMat': bit.design.ref_mat_no if bit.design and bit.design.ref_mat_no else '',
                'systemMat': (bit.brazing_bom.system_mat_no if bit.brazing_bom and bit.brazing_bom.system_mat_no
                              else (bit.bom.system_mat_no if bit.bom and bit.bom.system_mat_no
                              else (bit.system_bom.code if bit.system_bom else ''))),
                'brazingMat': (bit.brazing_bom.code if bit.brazing_bom else (bit.bom.code if bit.bom else '')),
                'account': entry.account.code if entry.account else (bit.account.code if bit.account else ''),
                'priority': entry.priority,
                'priorityDisplay': entry.get_priority_display(),
                'requester': _requester,
                'plannedDate': entry.planned_date.strftime('%Y-%m-%d') if entry.planned_date else '',
                'dueDate': entry.due_date.strftime('%Y-%m-%d') if entry.due_date else '',
                'dueDateHistory': entry.due_date_history or [],
                'overdue': bool(entry.due_date and entry.due_date < today),
                'notes': entry.notes or '',
                'intendedType': entry.get_intended_wo_type_display() if entry.intended_wo_type else '',
                'received': bit.received_date.strftime('%Y-%m-%d') if bit.received_date else '',
                'bitUrl': reverse('workorders:drillbit_detail', args=[bit.pk]),
                'repair': bit.revision_number or 0,
                'condition': bit.condition or '',
                'level': bit.level or (bit.design.order_level if bit.design else ''),
                'planStatus': entry.status,
                'releaseDestination': bit.get_release_destination().name if bit.get_release_destination() else '',
            })

        # ── WIP + COMPLETED ──
        base_qs = WorkOrder.objects.select_related(
            'drill_bit', 'drill_bit__design', 'account', 'assigned_to',
            'plan_entry', 'plan_entry__created_by', 'created_by'
        ).prefetch_related('router_entries').exclude(
            status=WorkOrder.Status.CANCELLED
        ).order_by('-created_at')
        if account_filter:
            base_qs = base_qs.filter(account__code=account_filter)

        key_steps = [
            ('buildup', 'Build Up'), ('braze', 'Braze'),
            ('grinding', 'Final grinding'), ('tip_grinding', 'Tip Grinding'),
            ('qc', '1st check'), ('thread_clean', 'Thread Cleaning'),
            ('body_clean', 'Body Cleaning'), ('usr', 'USR'),
            ('final', 'Final Inspection'),
        ]

        wip_json = []
        for wo in base_qs:
            router_entries = wo.router_entries.all().order_by('step_number')
            total_steps = router_entries.count()
            done_steps = router_entries.filter(qr_scan_end__isnull=False).count()
            cur = router_entries.filter(
                qr_scan_start__isnull=False, qr_scan_end__isnull=True
            ).first()
            pct = int((done_steps / total_steps * 100)) if total_steps > 0 else 0

            # Step statuses
            route = []
            for sk, sn in key_steps:
                e = router_entries.filter(
                    Q(step_description__icontains=sn) | Q(step_description__icontains=sk)
                ).first()
                if e:
                    if e.qr_scan_end:
                        route.append({'key': sk, 'label': sn, 'state': 'done'})
                    elif e.qr_scan_start:
                        route.append({'key': sk, 'label': sn, 'state': 'active'})
                    else:
                        route.append({'key': sk, 'label': sn, 'state': 'pending'})
                else:
                    route.append({'key': sk, 'label': sn, 'state': 'na'})

            # Map status to category
            if wo.status in [s.value for s in completed_statuses]:
                cat = 'dn'
            elif wo.status in [WorkOrder.Status.ON_HOLD]:
                cat = 'hd'
            elif wo.status in [WorkOrder.Status.QC_PENDING, WorkOrder.Status.QC_FAILED]:
                cat = 'rv'
            else:
                cat = 'ip'

            # Estimated finish: if X% done in Y days, total = Y / (X/100)
            est_finish = ''
            if pct > 0 and wo.created_at:
                days_elapsed = (today - wo.created_at.date()).days or 1
                est_total = days_elapsed / (pct / 100)
                est_remaining = max(0, int(est_total - days_elapsed))
                from datetime import timedelta
                est_date = today + timedelta(days=est_remaining)
                est_finish = est_date.strftime('%Y-%m-%d')

            # Get bit pk for timeline
            bit_pk = wo.drill_bit.pk if wo.drill_bit else None

            # Plan entry requester
            plan_entry = getattr(wo, 'plan_entry', None)
            requester = ''
            if plan_entry and plan_entry.created_by:
                requester = plan_entry.created_by.get_full_name() or plan_entry.created_by.username
            elif wo.created_by:
                requester = wo.created_by.get_full_name() or wo.created_by.username

            _bit = wo.drill_bit
            wip_json.append({
                'woId': wo.pk,
                'bitId': bit_pk,
                'woNum': wo.wo_number,
                'serial': _bit.serial_number if _bit else '',
                'size': str(_bit.size) if _bit and _bit.size else '',
                'type': _bit.design.hdbs_type if _bit and _bit.design else '',
                'designMat': _bit.design.mat_no if _bit and _bit.design else '',
                'refMat': (_bit.design.ref_mat_no if _bit and _bit.design and _bit.design.ref_mat_no else ''),
                'systemMat': wo.system_mat_no or '',
                'brazingMat': wo.brazing_mat_no or '',
                'account': wo.account.code if wo.account else '',
                'status': cat,
                'statusDisplay': wo.get_status_display(),
                'prio': 'H' if wo.priority in ['URGENT', 'HIGH', 'CRITICAL'] else ('L' if wo.priority == 'LOW' else 'M'),
                'requester': requester,
                'assigned': wo.assigned_to.get_short_name() if wo.assigned_to else '',
                'releaseDate': wo.created_at.strftime('%Y-%m-%d') if wo.created_at else '',
                'pct': pct,
                'curStep': cur.step_description if cur else '',
                'doneSteps': done_steps,
                'totalSteps': total_steps,
                'route': route,
                'due': wo.due_date.strftime('%Y-%m-%d') if wo.due_date else '',
                'estFinish': est_finish,
                'detailUrl': reverse('workorders:workorder_detail_enhanced', args=[wo.pk]),
                'repair': (_bit.revision_number if _bit else 0) or 0,
                'condition': (_bit.condition if _bit else '') or '',
                'level': (_bit.level if _bit else '') or '',
            })

        # ── STATS ──
        total_planned = len(planned_json)
        total_ready = len(ready_bits_json)
        total_wip = len([w for w in wip_json if w['status'] in ('ip', 'hd', 'rv')])
        completed_today = WorkOrder.objects.filter(
            status=WorkOrder.Status.COMPLETED,
            actual_end__date=today
        ).count()

        account_summary = []
        for acct in accounts:
            wc = WorkOrder.objects.filter(account=acct, status__in=wip_statuses).count()
            pc = ProductionPlanEntry.objects.filter(
                account=acct, status=ProductionPlanEntry.Status.PLANNED
            ).count()
            if wc + pc > 0:
                account_summary.append({'code': acct.code, 'wip': wc, 'plan': pc, 'total': wc + pc})

        context['planner_json'] = _j.dumps({
            'ready': ready_bits_json,
            'planned': planned_json,
            'wip': wip_json,
            'stats': {
                'ready': total_ready,
                'planned': total_planned,
                'wip': total_wip,
                'completedToday': completed_today,
            },
            'accountSummary': account_summary,
        })
        context['accounts'] = accounts
        context['current_account'] = account_filter or ''
        context['page_title'] = 'Production Planner'
        return context


@login_required
def api_production_wip_status(request):
    """
    API endpoint for real-time WIP status updates.
    Returns JSON with current WIP counts and active step for each work order.
    Used for polling/WebSocket updates.
    """
    wip_statuses = [
        WorkOrder.Status.RELEASED,
        WorkOrder.Status.IN_PROGRESS,
        WorkOrder.Status.QC_PENDING,
    ]

    work_orders = WorkOrder.objects.filter(
        status__in=wip_statuses
    ).select_related('drill_bit', 'account').prefetch_related('router_entries')

    data = []
    for wo in work_orders:
        router_entries = wo.router_entries.all()
        total = router_entries.count()
        completed = router_entries.filter(qr_scan_end__isnull=False).count()
        current = router_entries.filter(
            qr_scan_start__isnull=False,
            qr_scan_end__isnull=True
        ).first()

        data.append({
            'wo_id': wo.pk,
            'wo_number': wo.wo_number,
            'serial': wo.drill_bit.serial_number if wo.drill_bit else None,
            'account': wo.account.code if wo.account else None,
            'status': wo.status,
            'progress': int((completed / total * 100)) if total > 0 else 0,
            'current_step': current.step_description if current else None,
            'completed_steps': completed,
            'total_steps': total,
        })

    return JsonResponse({'wip': data, 'timestamp': timezone.now().isoformat()})


class ProductionPlannerCreateWOView(LoginRequiredMixin, TemplateView):
    """
    Quick Work Order creation from Production Planner.
    Lookup by serial number to auto-populate account, design level (L3/L4).
    """
    template_name = "workorders/production_planner_create_wo.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.sales.models import Account

        context['accounts'] = Account.objects.filter(is_active=True).order_by('sort_order', 'code')
        context['page_title'] = 'Create Work Order from Planner'
        return context

    def post(self, request, *args, **kwargs):
        from apps.sales.models import Account

        serial_number = request.POST.get('serial_number', '').strip()
        account_code = request.POST.get('account')

        # Lookup existing drill bit
        drill_bit = DrillBit.objects.filter(serial_number=serial_number).first()

        if drill_bit:
            # Drill bit exists - use its account (first time) or provided account
            account = drill_bit.account
            if not account and account_code:
                account = Account.objects.filter(code=account_code).first()
                # Set account on drill bit for first time
                if account:
                    drill_bit.account = account
                    drill_bit.save(update_fields=['account'])
        else:
            # New drill bit - require account
            if not account_code:
                messages.error(request, 'Account is required for new drill bits.')
                return redirect('workorders:production_planner_create_wo')
            account = Account.objects.filter(code=account_code).first()

        if not account:
            messages.error(request, 'Invalid account selected.')
            return redirect('workorders:production_planner_create_wo')

        # Determine WO type based on design level and workflow
        wo_type = WorkOrder.WOType.FC_REPAIR
        if account.workflow_type == 'MANUFACTURE':
            if drill_bit and drill_bit.design:
                level = getattr(drill_bit.design, 'level', 'L3')
                if level == 'L4':
                    wo_type = WorkOrder.WOType.FC_REWORK
                else:
                    wo_type = WorkOrder.WOType.FC_NEW

        # Generate WO number
        wo_number = account.generate_wo_number()

        # Create work order
        wo = WorkOrder.objects.create(
            wo_number=wo_number,
            wo_type=wo_type,
            drill_bit=drill_bit,
            design=drill_bit.design if drill_bit else None,
            account=account,
            status=WorkOrder.Status.DRAFT,
            created_by=request.user,
        )

        messages.success(request, f'Work Order {wo_number} created successfully.')
        return redirect('workorders:workorder_detail_enhanced', pk=wo.pk)


# =============================================================================
# PRODUCTION PLAN API ENDPOINTS
# =============================================================================

@login_required
def api_add_to_plan(request):
    """
    API endpoint to add a drill bit to the production plan.
    POST body: { serial_number, account, priority, planned_date, intended_wo_type, notes }
    """
    import json
    from apps.sales.models import Account

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    serial_number = data.get('serial_number', '').strip()
    if not serial_number:
        return JsonResponse({'success': False, 'error': 'Serial number required'})

    # Find the drill bit
    drill_bit = DrillBit.objects.select_related('account', 'bom', 'brazing_bom', 'system_bom').filter(serial_number=serial_number).first()
    if not drill_bit:
        return JsonResponse({'success': False, 'error': 'Drill bit not found'})

    # FC bits must have a BOM assigned before planning
    if drill_bit.bit_type == 'FC' and not (drill_bit.bom or drill_bit.brazing_bom or drill_bit.system_bom):
        return JsonResponse({
            'success': False,
            'error': 'Cannot add to planner — this FC bit has no BOM assigned. Create a BOM first.'
        })

    # Status validation — only certain statuses can be added to the planner
    PLANNABLE_STATUSES = [
        DrillBit.Status.RECEIVED,
        DrillBit.Status.IN_STOCK,
        DrillBit.Status.BACKLOADED,
        DrillBit.Status.IN_COMPONENTS,
    ]
    if drill_bit.status not in PLANNABLE_STATUSES:
        status_display = drill_bit.get_status_display()
        return JsonResponse({
            'success': False,
            'error': f'Cannot add to planner — bit status is "{status_display}". '
                     f'Only bits with status Received, In Stock, Backloaded, or In Components can be planned.'
        })

    # Get account — required for planning
    account = None
    account_code = data.get('account', '').strip()
    if account_code:
        account = Account.objects.filter(code=account_code).first()
        if not account:
            return JsonResponse({'success': False, 'error': f'Account "{account_code}" not found'})
    elif drill_bit.account:
        account = drill_bit.account

    if not account:
        return JsonResponse({'success': False, 'error': 'Account is required. Select an account before adding to plan.'})

    # Determine intended WO type from work_type selection
    work_type = data.get('work_type', '').strip()
    intended_wo_type = data.get('intended_wo_type', '')
    if not intended_wo_type and work_type:
        if work_type == 'REPAIR':
            intended_wo_type = 'FC_REPAIR'
        elif work_type == 'NEW':
            intended_wo_type = 'FC_NEW'

    # Parse due_date if provided
    due_date_str = data.get('due_date', '')
    due_date = None
    if due_date_str:
        try:
            from datetime import datetime
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass  # Will auto-calculate

    # Add to plan (returns 4-tuple: entry, created, error_code, error_message)
    try:
        entry, created, error_code, error_message = ProductionPlanEntry.add_to_plan(
            drill_bit=drill_bit,
            account=account,
            priority=data.get('priority', 'NORMAL'),
            planned_date=data.get('planned_date') or None,
            due_date=due_date,
            intended_wo_type=intended_wo_type,
            notes=data.get('notes', ''),
            user=request.user
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Database error: {str(e)}'
        }, status=500)

    if not created:
        return JsonResponse({
            'success': False,
            'error': error_message or 'This drill bit is already in the plan',
            'error_code': error_code or 'IN_PLAN'
        })

    # Update bit's account if different, and store previous status for cancel/restore
    update_fields = ['updated_at']
    if drill_bit.account != account:
        old_acc = drill_bit.account.code if drill_bit.account else ''
        drill_bit.log_change('Account', old_acc, account.code, request.user)
        drill_bit.account = account
        update_fields += ['account', 'change_log']

    # Save previous status on the plan entry so we can restore on cancel
    entry.notes = (entry.notes or '') + (f'\n[prev_status:{drill_bit.status}]' if drill_bit.status else '')
    entry.save(update_fields=['notes'])

    drill_bit.save(update_fields=update_fields)

    return JsonResponse({
        'success': True,
        'entry_id': entry.pk,
        'serial_number': serial_number,
        'message': f'Added {serial_number} to production plan'
    })


@login_required
def api_create_wo_from_plan(request):
    """
    API endpoint to create a Work Order from a production plan entry.
    POST with ?entry_id=X
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    entry_id = request.GET.get('entry_id')
    if not entry_id:
        return JsonResponse({'success': False, 'error': 'entry_id required'})

    try:
        entry = ProductionPlanEntry.objects.select_related('drill_bit', 'account').get(pk=entry_id)
    except ProductionPlanEntry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Plan entry not found'})

    if entry.status != ProductionPlanEntry.Status.PLANNED:
        return JsonResponse({'success': False, 'error': 'Entry already has a Work Order'})

    # Create the work order (returns 4-tuple: wo, success, error_code, error_message)
    try:
        wo, success, error_code, error_message = entry.create_work_order(user=request.user)

        if not success:
            response = {
                'success': False,
                'error': error_message or 'Failed to create work order',
                'error_code': error_code or 'UNKNOWN'
            }
            # If blocked by existing WO, include link to it
            if error_code == 'ACTIVE_WO' and wo:
                response['blocking_wo_id'] = wo.pk
                response['blocking_wo_number'] = wo.wo_number
                response['blocking_wo_url'] = reverse('workorders:workorder_detail_enhanced', args=[wo.pk])
            return JsonResponse(response)

        return JsonResponse({
            'success': True,
            'wo_id': wo.pk,
            'wo_number': wo.wo_number,
            'redirect_url': reverse('workorders:workorder_detail_enhanced', args=[wo.pk])
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_release_plan_entry(request):
    """
    Release a planned entry: marks as RELEASED and auto-creates the Work Order.
    POST body: { entry_id }
    """
    import json
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    entry_id = data.get('entry_id')
    if not entry_id:
        return JsonResponse({'success': False, 'error': 'entry_id required'})

    try:
        entry = ProductionPlanEntry.objects.select_related('drill_bit', 'account').get(pk=entry_id)
    except ProductionPlanEntry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Plan entry not found'})

    if entry.status != ProductionPlanEntry.Status.PLANNED:
        return JsonResponse({'success': False, 'error': f'Cannot release — status is {entry.get_status_display()}'})

    bit = entry.drill_bit
    if not bit:
        return JsonResponse({'success': False, 'error': 'No drill bit linked to this plan entry'})

    # Determine release destination
    dest = bit.get_release_destination()
    dest_name = dest.name if dest else 'Production Floor'
    dest_code = bit.get_release_destination_code()

    # Check if bit is ALREADY at the destination (operator moved it beforehand)
    current_loc = bit.bit_location
    already_there = (current_loc and dest and
                     (current_loc.pk == dest.pk or current_loc.code == dest_code or
                      current_loc.location_type == 'WIP'))

    if already_there:
        # Bit is already at destination — create WO immediately, no pending
        entry.status = ProductionPlanEntry.Status.RELEASED
        entry.save(update_fields=['status', 'updated_at'])

        try:
            wo, success, error_code, error_message = entry.create_work_order(user=request.user)
            if success:
                bit.log_change('Plan Status', 'Planned', 'Released + WO Created', request.user)
                bit.save(update_fields=['change_log', 'updated_at'])
                try:
                    from apps.notifications.services import notify
                    notify(
                        actor=request.user,
                        verb=f"released and created WO {wo.wo_number} for",
                        target=bit.serial_number,
                        priority="HIGH",
                        action_url=reverse('workorders:workorder_detail_enhanced', args=[wo.pk]),
                        entity_type="WorkOrder",
                        entity_id=wo.pk,
                    )
                except Exception:
                    pass
                return JsonResponse({
                    'success': True,
                    'status': 'RELEASED',
                    'wo_created': True,
                    'wo_number': wo.wo_number,
                    'redirect_url': reverse('workorders:workorder_detail_enhanced', args=[wo.pk]),
                    'message': f'Bit already at {current_loc.name}.\nWO {wo.wo_number} created immediately!'
                })
            else:
                entry.status = ProductionPlanEntry.Status.PLANNED
                entry.save(update_fields=['status', 'updated_at'])
                return JsonResponse({'success': False, 'error': f'WO creation failed: {error_message}'})
        except Exception as e:
            entry.status = ProductionPlanEntry.Status.PLANNED
            entry.save(update_fields=['status', 'updated_at'])
            return JsonResponse({'success': False, 'error': str(e)})

    else:
        # Bit is NOT at destination — set PENDING_RELEASE, notify operator
        entry.status = ProductionPlanEntry.Status.PENDING_RELEASE
        entry.save(update_fields=['status', 'updated_at'])

        bit.log_change('Plan Status', 'Planned', 'Pending Release', request.user)
        bit.save(update_fields=['change_log', 'updated_at'])

        try:
            from apps.notifications.services import notify
            notify(
                actor=request.user,
                verb=f"requests release of",
                target=f"{bit.serial_number} to {dest_name}",
                priority="HIGH",
                action_url=reverse('workorders:location_transfers') + f'?serial={bit.serial_number}',
                entity_type="DrillBit",
                entity_id=bit.pk,
            )
        except Exception:
            pass

        return JsonResponse({
            'success': True,
            'status': 'PENDING_RELEASE',
            'destination': dest_name,
            'serial': bit.serial_number,
            'message': f'Release initiated for {bit.serial_number}.\n\nWaiting for physical transfer to {dest_name}.\nWO will be created automatically once the bit arrives.'
        })


@login_required
def api_remove_from_plan(request):
    """
    API endpoint to remove a drill bit from the production plan.
    POST with ?entry_id=X
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    entry_id = request.GET.get('entry_id')
    if not entry_id:
        return JsonResponse({'success': False, 'error': 'entry_id required'})

    try:
        entry = ProductionPlanEntry.objects.get(pk=entry_id)
    except ProductionPlanEntry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Plan entry not found'})

    if entry.status == ProductionPlanEntry.Status.WO_CREATED:
        return JsonResponse({'success': False, 'error': 'Cannot remove - Work Order already created'})

    # Restore previous status on the drill bit if we stored it
    bit = entry.drill_bit
    if bit and entry.notes:
        import re
        m = re.search(r'\[prev_status:(\w+)\]', entry.notes)
        if m:
            prev_status = m.group(1)
            # Validate the status is a real choice
            valid_statuses = [s.value for s in DrillBit.Status]
            if prev_status in valid_statuses:
                bit.status = prev_status
                bit.save(update_fields=['status', 'updated_at'])

    # Mark as removed (soft delete)
    entry.status = ProductionPlanEntry.Status.REMOVED
    entry.save(update_fields=['status', 'updated_at'])

    return JsonResponse({
        'success': True,
        'bit_id': bit.pk if bit else None,
        'serial': bit.serial_number if bit else '',
        'restored_status': bit.status if bit else '',
        'restored_status_display': bit.get_status_display() if bit else '',
        'message': f'Removed from plan. Bit status restored to {bit.get_status_display() if bit else "N/A"}.'
    })


# =============================================================================
# PLANNER — Due Date Update with History
# =============================================================================

@login_required
@login_required
def api_update_plan_entry(request):
    """
    Update editable fields on a plan entry.
    POST body: { entry_id, field, value }
    Supported fields: priority, notes, intended_wo_type, requester_name
    """
    import json
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    entry_id = data.get('entry_id')
    field = data.get('field', '')
    value = data.get('value', '')

    if not entry_id or not field:
        return JsonResponse({'success': False, 'error': 'entry_id and field required'})

    try:
        entry = ProductionPlanEntry.objects.get(pk=entry_id)
    except ProductionPlanEntry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Plan entry not found'})

    ALLOWED_FIELDS = {
        'priority': lambda v: v if v in ['LOW', 'NORMAL', 'HIGH', 'URGENT', 'CRITICAL'] else None,
        'notes': lambda v: v.strip(),
        'intended_wo_type': lambda v: v if v in ['', 'FC_NEW', 'FC_REPAIR', 'RC_NEW', 'RC_REPAIR', 'FC_USED', 'RC_USED'] else None,
    }

    if field not in ALLOWED_FIELDS:
        return JsonResponse({'success': False, 'error': f'Field "{field}" is not editable'})

    clean_value = ALLOWED_FIELDS[field](value)
    if clean_value is None:
        return JsonResponse({'success': False, 'error': f'Invalid value for {field}'})

    setattr(entry, field, clean_value)
    entry.save(update_fields=[field, 'updated_at'])

    return JsonResponse({
        'success': True,
        'field': field,
        'value': clean_value,
        'display': entry.get_priority_display() if field == 'priority' else clean_value,
        'message': f'{field} updated'
    })


def api_update_plan_due_date(request):
    """
    Update due date on a plan entry, storing the old value in history.
    POST body: { entry_id, new_due_date (YYYY-MM-DD), justification }
    """
    import json
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    entry_id = data.get('entry_id')
    new_due_str = data.get('new_due_date', '').strip()
    justification = data.get('justification', '').strip()

    if not entry_id or not new_due_str:
        return JsonResponse({'success': False, 'error': 'entry_id and new_due_date required'})

    try:
        entry = ProductionPlanEntry.objects.get(pk=entry_id)
    except ProductionPlanEntry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Plan entry not found'})

    from datetime import datetime
    try:
        new_due = datetime.strptime(new_due_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format (YYYY-MM-DD)'})

    # Store old value in history
    history = entry.due_date_history or []
    history.append({
        'old': entry.due_date.isoformat() if entry.due_date else None,
        'new': new_due.isoformat(),
        'reason': justification,
        'changed_by': request.user.get_short_name() or request.user.username,
        'changed_at': timezone.now().isoformat(),
    })

    entry.due_date = new_due
    entry.due_date_history = history
    entry.save(update_fields=['due_date', 'due_date_history', 'updated_at'])

    return JsonResponse({
        'success': True,
        'new_due_date': new_due.isoformat(),
        'history_count': len(history),
    })


# =============================================================================
# PLANNER — Update Account (plan entry, drill bit, or WO)
# =============================================================================

@login_required
def api_delete_work_order(request, pk):
    """
    Delete a Work Order only if nothing has been done on it.
    Checks: status must be DRAFT/PLANNED/RELEASED, no started router steps,
    no completed evaluations, no LPT/Thread reports.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        wo = WorkOrder.objects.select_related('drill_bit').get(pk=pk)
    except WorkOrder.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Work Order not found'})

    # Admin force delete — bypasses all checks
    force = request.GET.get('force') == '1' or request.POST.get('force') == '1'
    if force:
        if not request.user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Force delete requires admin privileges.'})
        wo_number = wo.wo_number
        bit = wo.drill_bit
        bit_id = bit.pk if bit else None
        account = wo.account
        # Delete all related data
        wo.cutter_evaluations.all().delete()
        wo.router_entries.all().delete()
        wo.standalone_lpt_reports.all().delete()
        wo.standalone_thread_reports.all().delete()
        wo.delete()
        if bit:
            from apps.workorders.models import BitEvent
            old_status = bit.get_status_display()
            bit.log_change('Status', old_status, 'Received', request.user)
            bit.status = DrillBit.Status.RECEIVED
            bit.save(update_fields=['status', 'change_log', 'updated_at'])
            from django.utils import timezone as _tz
            if bit.current_location:
                BitEvent.objects.create(
                    bit=bit, event_type=BitEvent.EventType.TRANSFER,
                    event_date=_tz.now(),
                    notes=f'WO {wo_number} force-deleted by admin (was {old_status}).',
                    performed_by=request.user, location=bit.current_location,
                )
            has_plan = ProductionPlanEntry.objects.filter(
                drill_bit=bit, status=ProductionPlanEntry.Status.WO_CREATED
            ).exists()
        else:
            has_plan = False
        return JsonResponse({'success': True, 'wo_number': wo_number, 'bit_id': bit_id, 'has_plan_entry': has_plan, 'message': f'Work Order {wo_number} force-deleted.'})

    # Check status — only deletable in early stages
    deletable_statuses = [WorkOrder.Status.DRAFT, WorkOrder.Status.PLANNED, WorkOrder.Status.RELEASED]
    if wo.status not in deletable_statuses:
        return JsonResponse({
            'success': False,
            'error': f'Cannot delete — WO status is "{wo.get_status_display()}". Only Draft, Planned, or Released WOs can be deleted.'
        })

    # Check router steps — any started?
    started_steps = wo.router_entries.filter(qr_scan_start__isnull=False).count()
    if started_steps > 0:
        return JsonResponse({
            'success': False,
            'error': f'Cannot delete — {started_steps} router step(s) have been started.'
        })

    # Check evaluations — any completed or approved?
    completed_evals = wo.cutter_evaluations.filter(is_complete=True).count()
    if completed_evals > 0:
        return JsonResponse({
            'success': False,
            'error': f'Cannot delete — {completed_evals} evaluation(s) have been completed. Delete them first.'
        })

    # Check LPT reports
    lpt_count = wo.standalone_lpt_reports.count()
    if lpt_count > 0:
        return JsonResponse({
            'success': False,
            'error': f'Cannot delete — {lpt_count} LPT report(s) exist. Delete them first.'
        })

    # Check thread inspections
    thread_count = wo.standalone_thread_reports.count()
    if thread_count > 0:
        return JsonResponse({
            'success': False,
            'error': f'Cannot delete — {thread_count} thread inspection(s) exist. Delete them first.'
        })

    # Safe to delete — clean up related data
    wo_number = wo.wo_number
    bit = wo.drill_bit

    # Delete draft/pending evaluations (not completed)
    wo.cutter_evaluations.filter(is_complete=False).delete()

    # Delete empty router entries (not started)
    wo.router_entries.filter(qr_scan_start__isnull=True).delete()

    # Remember info before deleting
    account = wo.account
    bit_id = bit.pk if bit else None

    # Delete the WO
    wo.delete()

    # Restore drill bit + log event
    if bit:
        from apps.workorders.models import BitEvent
        old_status = bit.get_status_display()
        old_location = bit.current_location

        # Restore status — don't force a specific location
        if bit.status in ('IN_PRODUCTION', 'IN_REPAIR'):
            bit.log_change('Status', old_status, 'Received', request.user)
            bit.status = DrillBit.Status.RECEIVED
            bit.save(update_fields=['status', 'change_log', 'updated_at'])

        # Log the event with current location info
        from django.utils import timezone as _tz
        if old_location:
            BitEvent.objects.create(
                bit=bit,
                event_type=BitEvent.EventType.TRANSFER,
                event_date=_tz.now(),
                notes=f'WO {wo_number} deleted (was {old_status}). Bit remains at current location.',
                performed_by=request.user,
                location=old_location,
            )

        # Check if plan entry exists — don't restore yet, let user decide
        has_plan = ProductionPlanEntry.objects.filter(
            drill_bit=bit, status=ProductionPlanEntry.Status.WO_CREATED
        ).exists()
    else:
        has_plan = False

    return JsonResponse({
        'success': True,
        'wo_number': wo_number,
        'bit_id': bit_id,
        'has_plan_entry': has_plan,
        'message': f'Work Order {wo_number} deleted.'
    })


@login_required
@login_required
def api_restore_plan_entry(request, bit_pk):
    """Restore a WO_CREATED plan entry back to PLANNED status for a given bit."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    entry = ProductionPlanEntry.objects.filter(
        drill_bit_id=bit_pk, status=ProductionPlanEntry.Status.WO_CREATED
    ).first()
    if not entry:
        return JsonResponse({'success': False, 'error': 'No plan entry found to restore.'})
    entry.status = ProductionPlanEntry.Status.PLANNED
    entry.save(update_fields=['status', 'updated_at'])
    return JsonResponse({'success': True, 'message': 'Bit returned to planner.'})


@login_required
def api_delete_evaluation(request, pk):
    """
    Delete a CutterEvaluationMatrix only if not completed/approved and
    no dependent data exists (die check reports referencing it, etc.)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        ev = CutterEvaluationMatrix.objects.select_related('work_order').get(pk=pk)
    except CutterEvaluationMatrix.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Evaluation not found'})

    # Check status — cannot delete completed or approved
    if ev.is_complete:
        return JsonResponse({
            'success': False,
            'error': 'Cannot delete — this evaluation is marked as complete. Reopen it first.'
        })
    if hasattr(ev, 'status') and ev.status in ('COMPLETED', 'APPROVED'):
        return JsonResponse({
            'success': False,
            'error': f'Cannot delete — evaluation status is "{ev.get_status_display()}".'
        })

    # Check for dependent die check reports
    die_checks = DieCheckReport.objects.filter(evaluation=ev).count() if hasattr(DieCheckReport, 'evaluation') else 0
    if die_checks > 0:
        return JsonResponse({
            'success': False,
            'error': f'Cannot delete — {die_checks} die check report(s) depend on this evaluation. Delete them first.'
        })

    ev_type = ev.get_evaluation_type_display()
    wo_number = ev.work_order.wo_number if ev.work_order else ''
    ev.delete()

    return JsonResponse({
        'success': True,
        'message': f'{ev_type} evaluation deleted from {wo_number}.'
    })


@login_required
@login_required
@login_required
def api_delete_transfer(request, pk):
    """Admin-only: delete a BitEvent transfer record."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)
    from apps.workorders.models import BitEvent
    try:
        event = BitEvent.objects.get(pk=pk, event_type=BitEvent.EventType.TRANSFER)
    except BitEvent.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transfer record not found'})
    event.delete()
    return JsonResponse({'success': True, 'message': 'Transfer record deleted.'})


@login_required
def api_edit_transfer(request, pk):
    """Admin-only: edit a BitEvent transfer record (change location)."""
    import json
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Admin access required'}, status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    from apps.workorders.models import BitEvent
    try:
        event = BitEvent.objects.select_related('bit').get(pk=pk)
    except BitEvent.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Event not found'})

    new_location_id = data.get('location_id')
    new_notes = data.get('notes')

    if new_location_id:
        new_loc = Location.objects.filter(pk=new_location_id, is_active=True).first()
        if not new_loc:
            return JsonResponse({'success': False, 'error': 'Location not found'})
        old_loc = event.to_location or event.location
        event.to_location = new_loc
        event.location = new_loc
        # Auto-update notes to reflect the location change
        old_name = str(old_loc) if old_loc else '—'
        event.notes = f'Corrected: {old_name} → {new_loc.name}. {new_notes or ""}'.strip()
        # Also update the bit's current location if this is the latest event
        latest = BitEvent.objects.filter(bit=event.bit).order_by('-event_date').first()
        if latest and latest.pk == event.pk:
            event.bit.bit_location = new_loc
            event.bit.save(update_fields=['bit_location', 'updated_at'])
    elif new_notes is not None:
        event.notes = new_notes

    event.save()
    return JsonResponse({
        'success': True,
        'message': 'Transfer record updated.',
        'location': str(event.to_location or event.location),
    })


@login_required
def api_locations_list(request):
    """Return all active locations as JSON for dropdowns."""
    locs = Location.objects.filter(is_active=True).order_by('location_type', 'name')
    return JsonResponse({
        'locations': [{'id': l.pk, 'name': l.name, 'code': l.code, 'type': l.get_location_type_display()} for l in locs]
    })


@login_required
def api_transfer_bit_location(request):
    """
    Transfer a drill bit to a new location.
    POST body: { bit_id, location_id, reason }
    Creates a BitEvent(TRANSFER) and updates DrillBit.current_location + bit_location.
    """
    import json
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    bit_id = data.get('bit_id')
    location_id = data.get('location_id')
    reason = data.get('reason', '').strip()

    if not bit_id or not location_id:
        return JsonResponse({'success': False, 'error': 'bit_id and location_id required'})

    try:
        bit = DrillBit.objects.select_related('current_location', 'bit_location').get(pk=bit_id)
    except DrillBit.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Drill bit not found'})

    try:
        new_location = Location.objects.get(pk=location_id, is_active=True)
    except Location.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Location not found'})

    from_location = bit.bit_location
    if from_location and (from_location.pk == new_location.pk or from_location.name == new_location.name):
        return JsonResponse({'success': False, 'error': 'Bit is already at this location'})

    # Log the change
    from django.utils import timezone as _tz
    from apps.workorders.models import BitEvent

    bit.log_change('Location', str(from_location) if from_location else '—', str(new_location), request.user)

    # Create BitEvent
    BitEvent.objects.create(
        bit=bit,
        event_type=BitEvent.EventType.TRANSFER,
        event_date=_tz.now(),
        location=new_location,
        from_location=from_location,
        to_location=new_location,
        notes=reason or f'Manual transfer to {new_location.name}',
        performed_by=request.user,
    )

    # Update bit location (bit_location = workorders.Location)
    bit.bit_location = new_location
    bit.save(update_fields=['bit_location', 'change_log', 'updated_at'])

    # Check if this bit has a PENDING_RELEASE plan entry — auto-create WO on arrival
    wo_created_msg = ''
    pending_entry = ProductionPlanEntry.objects.filter(
        drill_bit=bit, status=ProductionPlanEntry.Status.PENDING_RELEASE
    ).select_related('account').first()

    if pending_entry:
        # Check if bit arrived at the correct release destination
        expected_dest = bit.get_release_destination_code()
        if new_location.code == expected_dest or new_location.location_type == 'WIP':
            # Bit arrived — create WO automatically
            pending_entry.status = ProductionPlanEntry.Status.RELEASED
            pending_entry.save(update_fields=['status', 'updated_at'])

            try:
                wo, success, error_code, error_message = pending_entry.create_work_order(user=request.user)
                if success:
                    wo_created_msg = f' WO {wo.wo_number} auto-created!'
                    bit.log_change('Plan Status', 'Pending Release', 'Released + WO Created', request.user)
                    bit.save(update_fields=['change_log', 'updated_at'])
                    # Notify planner
                    try:
                        from apps.notifications.services import notify
                        notify(
                            actor=request.user,
                            verb=f"transferred and created WO {wo.wo_number} for",
                            target=bit.serial_number,
                            priority="HIGH",
                            action_url=reverse('workorders:workorder_detail_enhanced', args=[wo.pk]),
                            entity_type="WorkOrder",
                            entity_id=wo.pk,
                        )
                    except Exception:
                        pass
            except Exception as e:
                wo_created_msg = f' (WO creation failed: {str(e)})'

    return JsonResponse({
        'success': True,
        'from': str(from_location) if from_location else '—',
        'to': new_location.name,
        'wo_created': bool(wo_created_msg),
        'message': f'Bit transferred to {new_location.name}.{wo_created_msg}'
    })


@login_required
def api_assign_bit_bom(request):
    """
    Assign a BOM to a drill bit (brazing_bom, system_bom, or default bom).
    POST body: { bit_id, bom_id, field: 'brazing'|'system'|'bom' }
    """
    import json
    from apps.technology.models import BOM

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    bit_id = data.get('bit_id')
    bom_id = data.get('bom_id')
    field = data.get('field', 'bom')  # 'brazing', 'system', or 'bom'

    if not bit_id:
        return JsonResponse({'success': False, 'error': 'bit_id required'})

    try:
        bit = DrillBit.objects.select_related('design', 'bom', 'brazing_bom', 'system_bom').get(pk=bit_id)
    except DrillBit.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Drill bit not found'})

    # Capture old values for audit
    old_sys = bit.system_bom.code if bit.system_bom else ''
    old_brz = bit.brazing_bom.code if bit.brazing_bom else ''
    old_bom = bit.bom.code if bit.bom else ''

    if not bom_id:
        update_fields = ['updated_at', 'change_log']
        if field == 'system':
            bit.log_change('System BOM', old_sys, '', request.user)
            if old_brz: bit.log_change('Brazing BOM', old_brz, '', request.user)
            if old_bom: bit.log_change('BOM', old_bom, '', request.user)
            bit.system_bom = None
            bit.brazing_bom = None
            bit.bom = None
            update_fields += ['system_bom', 'brazing_bom', 'bom']
        elif field == 'brazing':
            bit.log_change('Brazing BOM', old_brz, '', request.user)
            if old_bom: bit.log_change('BOM', old_bom, '', request.user)
            bit.brazing_bom = None
            bit.bom = None
            update_fields += ['brazing_bom', 'bom']
        else:
            bit.log_change('BOM', old_bom, '', request.user)
            bit.bom = None
            update_fields += ['bom']
        bit.save(update_fields=update_fields)
        return JsonResponse({
            'success': True,
            'bom_code': '',
            'system_mat': bit.system_bom.system_mat_no if bit.system_bom else '',
            'brz_code': bit.brazing_bom.code if bit.brazing_bom else '',
            'sys_code': bit.system_bom.code if bit.system_bom else '',
            'message': 'Cleared'
        })

    try:
        bom = BOM.objects.get(pk=bom_id)
    except BOM.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'BOM not found'})

    # Validate BOM belongs to the same design
    if bit.design and bom.design_id != bit.design_id:
        return JsonResponse({'success': False, 'error': 'BOM does not belong to this bit\'s design'})

    update_fields = ['updated_at', 'change_log']
    if field == 'system':
        bit.log_change('System BOM', old_sys, bom.code, request.user)
        bit.system_bom = bom
        bit.brazing_bom = None
        bit.bom = None
        update_fields += ['system_bom', 'brazing_bom', 'bom']
    elif field == 'brazing':
        bit.log_change('Brazing BOM', old_brz, bom.code, request.user)
        bit.brazing_bom = bom
        bit.bom = bom
        if bom.system_mat_no:
            sys_bom = BOM.objects.filter(
                design_id=bit.design_id, system_mat_no=bom.system_mat_no
            ).first()
            if sys_bom:
                bit.log_change('System BOM', old_sys, sys_bom.code, request.user)
                bit.system_bom = sys_bom
                update_fields.append('system_bom')
        update_fields += ['brazing_bom', 'bom']
    else:
        bit.log_change('BOM', old_bom, bom.code, request.user)
        bit.bom = bom
        update_fields += ['bom']
    bit.save(update_fields=update_fields)

    return JsonResponse({
        'success': True,
        'bom_id': bom.pk,
        'bom_code': bom.code or str(bom),
        'system_mat': bom.system_mat_no or '',
        'brz_code': bit.brazing_bom.code if bit.brazing_bom else '',
        'sys_code': bit.system_bom.code if bit.system_bom else '',
        'status': bom.get_status_display(),
        'message': f'{bom.code} assigned'
    })


def api_update_plan_account(request):
    """
    Change account on a plan entry AND its drill bit.
    POST body: { entry_id, account_code }
    """
    import json
    from apps.sales.models import Account

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    entry_id = data.get('entry_id')
    account_code = data.get('account_code', '').strip()
    if not entry_id or not account_code:
        return JsonResponse({'success': False, 'error': 'entry_id and account_code required'})

    try:
        entry = ProductionPlanEntry.objects.select_related('drill_bit').get(pk=entry_id)
    except ProductionPlanEntry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Plan entry not found'})

    if entry.status != ProductionPlanEntry.Status.PLANNED:
        return JsonResponse({'success': False, 'error': 'Cannot change account — WO already created'})

    try:
        account = Account.objects.get(code=account_code, is_active=True)
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'Account "{account_code}" not found'})

    # Update both plan entry and drill bit
    old_acc = entry.drill_bit.account.code if entry.drill_bit.account else ''
    entry.account = account
    entry.save(update_fields=['account', 'updated_at'])

    entry.drill_bit.log_change('Account', old_acc, account.code, request.user)
    entry.drill_bit.account = account
    entry.drill_bit.save(update_fields=['account', 'change_log', 'updated_at'])

    return JsonResponse({
        'success': True,
        'account_code': account.code,
        'account_name': account.name,
    })


@login_required
def api_update_bit_account(request):
    """
    Change account on a drill bit (from Ready view or drill bit inventory).
    POST body: { bit_id, account_code }
    """
    import json
    from apps.sales.models import Account

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    bit_id = data.get('bit_id')
    account_code = data.get('account_code', '').strip()
    if not bit_id or not account_code:
        return JsonResponse({'success': False, 'error': 'bit_id and account_code required'})

    try:
        bit = DrillBit.objects.get(pk=bit_id)
    except DrillBit.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Drill bit not found'})

    try:
        account = Account.objects.get(code=account_code, is_active=True)
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'Account "{account_code}" not found'})

    old_account = bit.account.code if bit.account else ''
    bit.log_change('Account', old_account, account.code, request.user)
    bit.account = account
    bit.save(update_fields=['account', 'change_log', 'updated_at'])

    return JsonResponse({
        'success': True,
        'account_code': account.code,
        'account_name': account.name,
    })


@login_required
def api_change_wo_account(request):
    """
    Change account on a Work Order. This cancels the old WO and creates a new one
    with the new account. If route steps have been started, they are copied to the
    new WO preserving all timestamps, operators, and completion data.

    POST body: { wo_id, account_code }
    """
    import json
    from apps.sales.models import Account

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    wo_id = data.get('wo_id')
    account_code = data.get('account_code', '').strip()
    if not wo_id or not account_code:
        return JsonResponse({'success': False, 'error': 'wo_id and account_code required'})

    try:
        old_wo = WorkOrder.objects.select_related('drill_bit', 'account').get(pk=wo_id)
    except WorkOrder.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Work Order not found'})

    if old_wo.status in [WorkOrder.Status.COMPLETED, WorkOrder.Status.CANCELLED]:
        return JsonResponse({'success': False, 'error': f'Cannot change account on {old_wo.get_status_display()} WO'})

    try:
        new_account = Account.objects.get(code=account_code, is_active=True)
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'Account "{account_code}" not found'})

    if old_wo.account == new_account:
        return JsonResponse({'success': False, 'error': 'Same account — no change needed'})

    # Collect existing router entries BEFORE cancelling
    old_router_entries = list(
        RouterSheetEntry.objects.filter(work_order=old_wo).order_by('step_number')
    )

    # Generate new WO number under new account
    new_wo_number = new_account.generate_wo_number()

    # Create new WO with same fields
    new_wo = WorkOrder.objects.create(
        wo_number=new_wo_number,
        wo_type=old_wo.wo_type,
        drill_bit=old_wo.drill_bit,
        design=old_wo.design,
        account=new_account,
        status=old_wo.status,
        priority=old_wo.priority,
        planned_start=old_wo.planned_start,
        due_date=old_wo.due_date,
        created_by=request.user,
        notes=f"Account changed from {old_wo.account.code if old_wo.account else '—'} "
              f"to {new_account.code}. Original WO: {old_wo.wo_number}",
    )

    # Copy router entries (preserving all timestamps and completion data)
    for old_entry in old_router_entries:
        RouterSheetEntry.objects.create(
            work_order=new_wo,
            step_number=old_entry.step_number,
            step_description=old_entry.step_description,
            qr_scan_start=old_entry.qr_scan_start,
            qr_scan_end=old_entry.qr_scan_end,
            station_qr=old_entry.station_qr,
            manual_date=old_entry.manual_date,
            manual_time_receipt=old_entry.manual_time_receipt,
            operator=old_entry.operator,
            is_complete=old_entry.is_complete,
            remarks=old_entry.remarks,
            cerebro_removal=old_entry.cerebro_removal,
            oring_removal=old_entry.oring_removal,
        )

    # Update plan entry if exists
    plan_entry = getattr(old_wo, 'plan_entry', None)
    if plan_entry:
        plan_entry.work_order = new_wo
        plan_entry.account = new_account
        plan_entry.save(update_fields=['work_order', 'account', 'updated_at'])

    # Update drill bit account
    if old_wo.drill_bit:
        old_wo.drill_bit.account = new_account
        old_wo.drill_bit.save(update_fields=['account', 'updated_at'])

    # Cancel old WO
    old_wo.status = WorkOrder.Status.CANCELLED
    old_wo.notes = (old_wo.notes or '') + f"\nCancelled: Account changed to {new_account.code}. New WO: {new_wo_number}"
    old_wo.save(update_fields=['status', 'notes', 'updated_at'])

    return JsonResponse({
        'success': True,
        'old_wo': old_wo.wo_number,
        'new_wo_id': new_wo.pk,
        'new_wo_number': new_wo_number,
        'account_code': new_account.code,
        'router_entries_copied': len(old_router_entries),
        'redirect_url': reverse('workorders:workorder_detail_enhanced', args=[new_wo.pk]),
    })


# =============================================================================
# PLANNER — Bit Timeline API
# =============================================================================

@login_required
def api_bit_timeline(request, bit_pk):
    """
    Return full chronological timeline of a drill bit for the planner drawer.
    Aggregates: BitEvents, ReceivingInspection, PlanEntries, WorkOrders, RouterSheetEntries.
    """
    from apps.workorders.models import (
        BitEvent, ReceivingInspection, RouterSheetEntry,
        CutterEvaluationMatrix, DieCheckReport, BackloadItem,
    )
    from apps.notifications.models import FormRevision

    try:
        bit = DrillBit.objects.select_related('design', 'account', 'bom', 'brazing_bom', 'system_bom').get(pk=bit_pk)
    except DrillBit.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Drill bit not found'}, status=404)

    events = []

    def _who(user):
        if not user: return ''
        return user.get_full_name() or user.username

    # 1. Bit registered
    events.append({
        'date': bit.created_at.isoformat(),
        'type': 'system',
        'action': 'Registered',
        'detail': f'Serial: {bit.serial_number}, Design: {bit.design.mat_no if bit.design else "—"}',
        'status': bit.get_status_display(),
        'who': _who(bit.created_by),
        'where': '',
        'url': '',
    })

    # 2. BitEvents (lifecycle)
    for ev in BitEvent.objects.filter(bit=bit).select_related('location', 'from_location', 'to_location', 'performed_by').order_by('event_date'):
        where = ''
        if ev.to_location and ev.from_location:
            where = f'{ev.from_location.name} → {ev.to_location.name}'
        elif ev.to_location:
            where = f'→ {ev.to_location.name}'
        elif ev.location:
            where = ev.location.name
        events.append({
            'date': ev.event_date.isoformat(),
            'type': 'event',
            'action': ev.get_event_type_display(),
            'detail': ev.notes or '',
            'status': '',
            'who': _who(ev.performed_by),
            'where': where,
            'url': '',
            'event_id': ev.pk,
        })

    # 3. Backload batch
    for bi in BackloadItem.objects.filter(drill_bit=bit).select_related('batch', 'batch__account'):
        events.append({
            'date': bi.batch.created_at.isoformat(),
            'type': 'receiving',
            'action': f'Backload Batch {bi.batch.batch_number}',
            'detail': f'Match: {bi.get_match_status_display()}, Account: {bi.batch.account.code if bi.batch.account else "—"}',
            'status': '',
            'who': '',
            'where': '',
            'url': reverse('workorders:backload_batch_detail', args=[bi.batch.pk]),
        })

    # 4. Receiving Inspections (start + completion as separate rows)
    for ri in ReceivingInspection.objects.filter(drill_bit=bit).select_related('inspected_by').order_by('created_at'):
        ri_url = reverse('workorders:receiving_inspection_edit', args=[bit.pk, ri.pk])
        events.append({
            'date': ri.created_at.isoformat(),
            'type': 'inspection',
            'action': 'Receiving Inspection',
            'detail': f'Started on {ri.inspection_date}',
            'status': 'Started',
            'who': _who(ri.inspected_by),
            'where': '',
            'url': ri_url,
        })
        if ri.is_complete and ri.updated_at != ri.created_at:
            events.append({
                'date': ri.updated_at.isoformat(),
                'type': 'inspection',
                'action': 'Receiving Inspection',
                'detail': f'Result: {ri.get_result_display()}',
                'status': 'Completed',
                'who': _who(ri.inspected_by),
                'where': '',
                'url': ri_url,
            })

    # 5. Plan entries
    for pe in ProductionPlanEntry.objects.filter(drill_bit=bit).select_related('account', 'created_by').order_by('created_at'):
        events.append({
            'date': pe.created_at.isoformat(),
            'type': 'plan',
            'action': f'Added to Planner',
            'detail': f'Account: {pe.account.code if pe.account else "—"}, Priority: {pe.get_priority_display()}, Due: {pe.due_date or "—"}',
            'status': pe.get_status_display(),
            'who': _who(pe.created_by),
            'where': '',
            'url': reverse('workorders:production_planner'),
        })
        if pe.status == ProductionPlanEntry.Status.RELEASED:
            events.append({
                'date': pe.updated_at.isoformat(),
                'type': 'plan',
                'action': 'Released for Production',
                'detail': '', 'status': 'Released',
                'who': '', 'where': '',
                'url': reverse('workorders:production_planner'),
            })
        elif pe.status == ProductionPlanEntry.Status.REMOVED:
            events.append({
                'date': pe.updated_at.isoformat(),
                'type': 'plan',
                'action': 'Removed from Planner',
                'detail': '', 'status': 'Removed',
                'who': '', 'where': '', 'url': '',
            })
        for h in (pe.due_date_history or []):
            events.append({
                'date': h.get('changed_at', pe.updated_at.isoformat()),
                'type': 'change',
                'action': 'Due Date Changed',
                'detail': f'{h.get("old","?")} → {h.get("new","?")}. {h.get("reason","")}',
                'status': '', 'who': h.get('changed_by', ''), 'where': '', 'url': '',
            })

    # 6. Work Orders (creation + milestones only)
    for wo in WorkOrder.objects.filter(drill_bit=bit).select_related('account', 'created_by').order_by('created_at'):
        wo_url = reverse('workorders:workorder_detail_enhanced', args=[wo.pk])
        events.append({
            'date': wo.created_at.isoformat(),
            'type': 'wo',
            'action': f'WO {wo.wo_number} Created',
            'detail': f'Type: {wo.get_wo_type_display()}, Account: {wo.account.code if wo.account else "—"}',
            'status': wo.get_status_display(),
            'who': _who(wo.created_by),
            'where': '',
            'url': wo_url,
        })
        first_step = wo.router_entries.filter(qr_scan_start__isnull=False).order_by('qr_scan_start').first()
        if first_step:
            events.append({
                'date': first_step.qr_scan_start.isoformat(),
                'type': 'wo',
                'action': f'WO {wo.wo_number} — Production Started',
                'detail': f'First step: {first_step.step_description}',
                'status': 'In Progress',
                'who': _who(first_step.operator),
                'where': '',
                'url': reverse('workorders:router_sheet', args=[wo.pk]),
            })
        if wo.status in ('COMPLETED', 'QC_PASSED', 'CANCELLED'):
            events.append({
                'date': wo.updated_at.isoformat(),
                'type': 'wo',
                'action': f'WO {wo.wo_number} — {wo.get_status_display()}',
                'detail': '', 'status': wo.get_status_display(),
                'who': '', 'where': '', 'url': wo_url,
            })

    # 7. Form Revisions (inspection edits)
    ri_ids = list(ReceivingInspection.objects.filter(drill_bit=bit).values_list('pk', flat=True))
    if ri_ids:
        for fr in FormRevision.objects.filter(
            entity_type='ReceivingInspection', entity_id__in=ri_ids
        ).select_related('revised_by').order_by('revised_at'):
            events.append({
                'date': fr.revised_at.isoformat(),
                'type': 'change',
                'action': f'Inspection Edited (Rev {fr.revision_number})',
                'detail': fr.change_summary or '',
                'status': '', 'who': _who(fr.revised_by), 'where': '', 'url': '',
            })

    # 8. Field change log (from DrillBit.change_log)
    for cl in (bit.change_log or []):
        events.append({
            'date': cl.get('when', bit.updated_at.isoformat()),
            'type': 'change',
            'action': f'{cl.get("field","?")} Changed',
            'detail': f'{cl.get("old","—") or "—"} → {cl.get("new","—") or "—"}',
            'status': '', 'who': cl.get('who', ''), 'where': '', 'url': '',
        })

    # Sort by date
    events.sort(key=lambda e: e['date'])

    return JsonResponse({
        'success': True,
        'serial': bit.serial_number,
        'events': events,
        'total': len(events),
    })


# =============================================================================
# PLANNER SETTINGS - Due Date Configuration
# =============================================================================

class PlannerSettingsView(LoginRequiredMixin, TemplateView):
    """
    Control page for Production Planner settings.
    Manage due date calculation, weekends, and holidays.
    """
    template_name = "workorders/planner_settings.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get singleton settings
        settings_obj = PlannerSettings.get_settings()

        # Get upcoming holidays
        today = timezone.now().date()
        upcoming_holidays = PlannerHoliday.objects.filter(
            date__gte=today,
            is_active=True
        ).order_by('date')[:20]

        # All holidays for editing
        all_holidays = PlannerHoliday.objects.all().order_by('-date')

        context['settings'] = settings_obj
        context['upcoming_holidays'] = upcoming_holidays
        context['all_holidays'] = all_holidays
        context['page_title'] = 'Planner Settings'

        # Day options for weekend selector
        context['day_options'] = [
            (0, 'Monday'),
            (1, 'Tuesday'),
            (2, 'Wednesday'),
            (3, 'Thursday'),
            (4, 'Friday'),
            (5, 'Saturday'),
            (6, 'Sunday'),
        ]

        return context

    def post(self, request, *args, **kwargs):
        """Handle settings update."""
        import json

        settings_obj = PlannerSettings.get_settings()

        # Update due days settings
        default_due_days = request.POST.get('default_due_days')
        ur_due_days = request.POST.get('ur_due_days')

        if default_due_days:
            try:
                settings_obj.default_due_days = int(default_due_days)
            except ValueError:
                pass

        if ur_due_days:
            try:
                settings_obj.ur_due_days = int(ur_due_days)
            except ValueError:
                pass

        # Update weekend days
        weekend_days = request.POST.getlist('weekend_days')
        settings_obj.weekend_days = [int(d) for d in weekend_days]

        settings_obj.updated_by = request.user
        settings_obj.save()

        messages.success(request, 'Planner settings updated successfully.')
        return redirect('workorders:planner_settings')


@login_required
def api_add_holiday(request):
    """API endpoint to add a new holiday."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    date_str = data.get('date', '')
    name = data.get('name', '').strip()

    if not date_str or not name:
        return JsonResponse({'success': False, 'error': 'Date and name are required'})

    try:
        from datetime import datetime
        holiday_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format'})

    # Check for duplicate
    if PlannerHoliday.objects.filter(date=holiday_date).exists():
        return JsonResponse({'success': False, 'error': 'Holiday already exists for this date'})

    holiday = PlannerHoliday.objects.create(
        date=holiday_date,
        name=name,
        created_by=request.user
    )

    return JsonResponse({
        'success': True,
        'holiday_id': holiday.pk,
        'message': f'Added holiday: {name}'
    })


@login_required
def api_delete_holiday(request, pk):
    """API endpoint to delete a holiday."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        holiday = PlannerHoliday.objects.get(pk=pk)
        holiday_name = holiday.name
        holiday.delete()
        return JsonResponse({
            'success': True,
            'message': f'Deleted holiday: {holiday_name}'
        })
    except PlannerHoliday.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Holiday not found'})


@login_required
def api_toggle_holiday(request, pk):
    """API endpoint to toggle holiday active status."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    try:
        holiday = PlannerHoliday.objects.get(pk=pk)
        holiday.is_active = not holiday.is_active
        holiday.save(update_fields=['is_active'])
        return JsonResponse({
            'success': True,
            'is_active': holiday.is_active,
            'message': f'Holiday {"enabled" if holiday.is_active else "disabled"}'
        })
    except PlannerHoliday.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Holiday not found'})


@login_required
def api_preview_due_date(request):
    """API endpoint to preview due date calculation."""
    account_code = request.GET.get('account', '')
    today = timezone.now().date()

    due_date = PlannerSettings.calculate_due_date(
        start_date=today,
        account_code=account_code if account_code else None
    )

    # Get settings for display
    settings_obj = PlannerSettings.get_settings()
    days_used = settings_obj.ur_due_days if account_code == 'UR' else settings_obj.default_due_days

    return JsonResponse({
        'success': True,
        'start_date': today.strftime('%Y-%m-%d'),
        'due_date': due_date.strftime('%Y-%m-%d'),
        'due_date_display': due_date.strftime('%b %d, %Y'),
        'working_days': days_used,
        'account': account_code or 'Default'
    })


# =============================================================================
# EVALUATION ROUTE BUILDER
# =============================================================================

class EvaluationRouteBuilderView(LoginRequiredMixin, TemplateView):
    """
    Dashboard to configure evaluation routes for different account + bit type combinations.
    Allows drag-and-drop arrangement of evaluation steps.
    """
    template_name = "workorders/evaluation_route_builder.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Import Account model here to avoid circular imports
        from apps.sales.models import Account

        # Get all accounts
        accounts = Account.objects.filter(is_active=True).order_by('sort_order', 'code')

        # Build data structure: account -> bit_type -> route
        route_matrix = {}
        for account in accounts:
            route_matrix[account.code] = {
                'account': account,
                'new_route': None,
                'used_route': None,
            }

        # Get all routes
        routes = EvaluationRoute.objects.filter(is_active=True).select_related('account').prefetch_related('steps')
        for route in routes:
            if route.account.code in route_matrix:
                if route.bit_type == EvaluationRoute.BitType.NEW:
                    route_matrix[route.account.code]['new_route'] = route
                else:
                    route_matrix[route.account.code]['used_route'] = route

        context['route_matrix'] = route_matrix
        context['accounts'] = accounts

        # All available evaluation types
        context['evaluation_types'] = CutterEvaluationMatrix.EvaluationType.choices

        context['page_title'] = 'Evaluation Route Builder'
        context['breadcrumbs'] = [
            {'title': 'Work Orders', 'url': reverse('workorders:workorder_list_enhanced')},
            {'title': 'Route Builder', 'url': None}
        ]
        return context


class EvaluationRouteDetailView(LoginRequiredMixin, View):
    """
    View/Edit a specific evaluation route.
    """
    def get(self, request, pk):
        route = get_object_or_404(EvaluationRoute.objects.prefetch_related('steps'), pk=pk)

        # Get steps in order
        steps = route.steps.all().order_by('order')

        # All available evaluation types
        evaluation_types = CutterEvaluationMatrix.EvaluationType.choices

        # Types already in the route
        used_types = set(step.evaluation_type for step in steps)

        # Available types (not yet in route)
        available_types = [(code, label) for code, label in evaluation_types if code not in used_types]

        return render(request, 'workorders/evaluation_route_detail.html', {
            'route': route,
            'steps': steps,
            'evaluation_types': evaluation_types,
            'available_types': available_types,
            'page_title': f'Edit Route: {route.name}',
            'breadcrumbs': [
                {'title': 'Work Orders', 'url': reverse('workorders:workorder_list_enhanced')},
                {'title': 'Route Builder', 'url': reverse('workorders:evaluation_route_builder')},
                {'title': route.name, 'url': None}
            ]
        })


@login_required
def api_create_route(request):
    """API to create a new evaluation route."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})

    account_id = data.get('account_id')
    bit_type = data.get('bit_type')
    name = data.get('name', '')

    if not account_id or not bit_type:
        return JsonResponse({'success': False, 'error': 'account_id and bit_type required'})

    from apps.sales.models import Account
    try:
        account = Account.objects.get(pk=account_id)
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Account not found'})

    # Create default name if not provided
    if not name:
        type_label = 'New Build' if bit_type == 'NEW' else 'Repair'
        name = f"{account.code} {type_label} Standard"

    # Check if route already exists
    existing = EvaluationRoute.objects.filter(account=account, bit_type=bit_type, is_active=True).first()
    if existing:
        return JsonResponse({
            'success': False,
            'error': f'Route already exists for {account.code} - {bit_type}',
            'route_id': existing.id
        })

    # Create the route
    route = EvaluationRoute.objects.create(
        name=name,
        account=account,
        bit_type=bit_type,
        is_default=True,
        created_by=request.user
    )

    return JsonResponse({
        'success': True,
        'route_id': route.id,
        'route_name': route.name,
        'message': f'Route created: {route.name}'
    })


@login_required
def api_update_route(request, pk):
    """API to update a route (name, description, etc.)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})

    try:
        route = EvaluationRoute.objects.get(pk=pk)
    except EvaluationRoute.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Route not found'})

    # Update fields
    if 'name' in data:
        route.name = data['name']
    if 'description' in data:
        route.description = data['description']
    if 'is_active' in data:
        route.is_active = data['is_active']
    if 'is_default' in data:
        # If setting as default, unset others
        if data['is_default']:
            EvaluationRoute.objects.filter(
                account=route.account,
                bit_type=route.bit_type,
                is_default=True
            ).exclude(pk=route.pk).update(is_default=False)
        route.is_default = data['is_default']

    route.save()

    return JsonResponse({
        'success': True,
        'message': 'Route updated'
    })


@login_required
def api_delete_route(request, pk):
    """API to delete (deactivate) a route."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        route = EvaluationRoute.objects.get(pk=pk)
    except EvaluationRoute.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Route not found'})

    # Soft delete - deactivate
    route.is_active = False
    route.save()

    return JsonResponse({
        'success': True,
        'message': f'Route "{route.name}" deleted'
    })


@login_required
def api_add_route_step(request, pk):
    """API to add a step to a route."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})

    try:
        route = EvaluationRoute.objects.get(pk=pk)
    except EvaluationRoute.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Route not found'})

    evaluation_type = data.get('evaluation_type')
    if not evaluation_type:
        return JsonResponse({'success': False, 'error': 'evaluation_type required'})

    # Check if already exists
    if route.steps.filter(evaluation_type=evaluation_type).exists():
        return JsonResponse({'success': False, 'error': 'Step already exists in route'})

    # Get next order number
    max_order = route.steps.aggregate(max_order=models.Max('order'))['max_order'] or 0

    # Create step
    step = EvaluationRouteStep.objects.create(
        route=route,
        evaluation_type=evaluation_type,
        order=max_order + 1,
        is_required=data.get('is_required', True),
        is_conditional=data.get('is_conditional', False),
        condition_description=data.get('condition_description', ''),
        show_decision_field=data.get('show_decision_field', True),
        show_cutter_matrix=data.get('show_cutter_matrix', True),
        show_cutters_details=data.get('show_cutters_details', True),
    )

    return JsonResponse({
        'success': True,
        'step_id': step.id,
        'message': f'Step added: {step.get_evaluation_type_display()}'
    })


@login_required
def api_update_route_step(request, pk, step_pk):
    """API to update a step in a route."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})

    try:
        step = EvaluationRouteStep.objects.get(pk=step_pk, route_id=pk)
    except EvaluationRouteStep.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Step not found'})

    # Update fields
    if 'is_required' in data:
        step.is_required = data['is_required']
    if 'is_conditional' in data:
        step.is_conditional = data['is_conditional']
    if 'condition_description' in data:
        step.condition_description = data['condition_description']
    if 'show_decision_field' in data:
        step.show_decision_field = data['show_decision_field']
    if 'show_cutter_matrix' in data:
        step.show_cutter_matrix = data['show_cutter_matrix']
    if 'show_cutters_details' in data:
        step.show_cutters_details = data['show_cutters_details']
    if 'order' in data:
        step.order = data['order']

    step.save()

    return JsonResponse({
        'success': True,
        'message': 'Step updated'
    })


@login_required
def api_delete_route_step(request, pk, step_pk):
    """API to delete a step from a route."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    try:
        step = EvaluationRouteStep.objects.get(pk=step_pk, route_id=pk)
    except EvaluationRouteStep.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Step not found'})

    step_name = step.get_evaluation_type_display()
    step.delete()

    return JsonResponse({
        'success': True,
        'message': f'Step "{step_name}" removed'
    })


@login_required
def api_reorder_route_steps(request, pk):
    """API to reorder steps in a route."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})

    import json
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})

    step_order = data.get('step_order', [])  # List of step IDs in new order

    try:
        route = EvaluationRoute.objects.get(pk=pk)
    except EvaluationRoute.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Route not found'})

    # Update order for each step
    for index, step_id in enumerate(step_order):
        EvaluationRouteStep.objects.filter(pk=step_id, route=route).update(order=index)

    return JsonResponse({
        'success': True,
        'message': 'Steps reordered'
    })


@login_required
def api_get_route(request, pk):
    """API to get route data including all steps."""
    try:
        route = EvaluationRoute.objects.prefetch_related('steps').get(pk=pk)
    except EvaluationRoute.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Route not found'})

    steps = []
    for step in route.steps.all().order_by('order'):
        steps.append({
            'id': step.id,
            'evaluation_type': step.evaluation_type,
            'evaluation_type_display': step.get_evaluation_type_display(),
            'order': step.order,
            'is_required': step.is_required,
            'is_conditional': step.is_conditional,
            'condition_description': step.condition_description,
            'show_decision_field': step.show_decision_field,
            'show_cutter_matrix': step.show_cutter_matrix,
            'show_cutters_details': step.show_cutters_details,
        })

    return JsonResponse({
        'success': True,
        'route': {
            'id': route.id,
            'name': route.name,
            'description': route.description,
            'bit_type': route.bit_type,
            'bit_type_display': route.get_bit_type_display(),
            'account_code': route.account.code,
            'is_default': route.is_default,
            'is_active': route.is_active,
            'steps': steps,
        }
    })


@login_required
def api_get_evaluation_types(request):
    """API to get all available evaluation types."""
    types = []
    for code, label in CutterEvaluationMatrix.EvaluationType.choices:
        # Skip legacy types
        if code == 'ARDT':
            continue
        types.append({
            'code': code,
            'label': label,
        })

    return JsonResponse({
        'success': True,
        'evaluation_types': types
    })


# =============================================================================
# RECEIVING INSPECTION (QAS/005-1) — Linked to DrillBit
# =============================================================================

def _get_bom_blade_data(drill_bit):
    """
    Extract blade data from a drill bit's BOM source_data for the evaluation grid.
    Returns (blade_data_list, bom_summary_list, cutter_config_list, has_data, cutter_grid_ctx) tuple.
    blade_data_list: [{name, rows: [{row_key, positions: [{pos_name, cutters: [{type, group, chamfer}]}]}]}]
    cutter_config_list: [{order, color, count, type, group, chamfer}]
    cutter_grid_ctx: dict of JSON strings for pocket-style cutter grid
    """
    bom = drill_bit.brazing_bom or drill_bit.system_bom or getattr(drill_bit, 'bom', None)
    if not bom or not bom.source_data:
        return [], [], [], False, {}

    source_data = bom.source_data if isinstance(bom.source_data, dict) else {}
    raw_blades = source_data.get("blades", [])
    # Key is "summary" in current data; fall back to "bom" for legacy records
    bom_summary = source_data.get("summary", []) or source_data.get("bom", [])

    if not raw_blades:
        return [], bom_summary, [], False, {}

    POSITIONS = ["CONE", "NOSE", "SHOULDER", "GAUGE", "PAD"]
    ROW_KEYS = ["r1", "r2", "r3", "r4"]

    blade_data = []
    for blade in raw_blades:
        blade_name = blade.get("name", "")
        rows = []
        for rk in ROW_KEYS:
            row_data = blade.get(rk, {})
            if not row_data:
                continue
            positions = []
            for pos in POSITIONS:
                cutters = row_data.get(pos, [])
                if cutters:
                    positions.append({
                        "pos_name": pos,
                        "cutters": cutters,  # [{type, group, chamfer}, ...]
                    })
            if positions:
                rows.append({"row_key": rk, "positions": positions})
        if rows:
            blade_data.append({"name": blade_name, "rows": rows})

    # Build BOM summary lookup by index (the "group" field on each cutter
    # is a 1-based index into this summary array — the definitive link)
    summary_by_index = {}
    for bom_row in bom_summary:
        summary_by_index[bom_row.get('index')] = bom_row

    # Cutter shape images from source_data and InventoryItem fallback
    cutter_shapes_sd = source_data.get("cutter_shapes", {})
    # Pre-load InventoryItem shapes keyed by mat_number for fallback
    mat_numbers = [r.get('mat_number', '') for r in bom_summary if r.get('mat_number')]
    _item_shapes = {}
    _item_shape_attrs = {}  # mat_number -> cutter shape text attribute
    if mat_numbers:
        from apps.inventory.models import InventoryItem, ItemAttributeValue
        for item in InventoryItem.objects.filter(
            mat_number__in=mat_numbers
        ).exclude(shape_image_base64__isnull=True).exclude(shape_image_base64=''):
            _item_shapes[item.mat_number] = item.shape_image_base64
        # Load "Cutter Shape" text attributes (substrate_shape, CategoryAttribute pk=8)
        for av in ItemAttributeValue.objects.filter(
            attribute__attribute__name__in=['Cutter Shape', 'Substrate Shape'],
            item__mat_number__in=mat_numbers,
        ).select_related('item'):
            if av.text_value:
                _item_shape_attrs[av.item.mat_number] = av.text_value

    # Build cutter config list (unique cutter types with colors for the grid)
    CUTTER_COLORS = ['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6',
                     '#EC4899', '#06B6D4', '#84CC16', '#F97316', '#6366F1',
                     '#14B8A6', '#F43F5E', '#A855F7', '#22D3EE']
    seen_types = {}
    cutter_config_list = []
    type_idx = 0
    for blade in blade_data:
        for row in blade['rows']:
            for pos in row['positions']:
                for cutter in pos['cutters']:
                    ct = cutter.get('type', '')
                    cg = cutter.get('group', '')
                    cc = cutter.get('chamfer', '')
                    # Fall back to BOM summary chamfer when blade cutter chamfer is empty
                    bom_match = summary_by_index.get(cg, {})
                    if not cc:
                        cc = bom_match.get('chamfer', '')
                    key = f"{ct}|{cg}|{cc}"
                    if key not in seen_types:
                        color = CUTTER_COLORS[type_idx % len(CUTTER_COLORS)]
                        mat_no = bom_match.get('mat_number', '')
                        # Resolve shape: source_data cutter_shapes first, then InventoryItem
                        shape_data = cutter_shapes_sd.get(str(cg), cutter_shapes_sd.get(cg))
                        if isinstance(shape_data, dict):
                            shape_uri = shape_data.get('data', '')
                        elif isinstance(shape_data, str):
                            shape_uri = shape_data
                        else:
                            shape_uri = _item_shapes.get(mat_no, '')
                        seen_types[key] = {
                            'order': type_idx + 1, 'color': color, 'count': 0,
                            'type': ct, 'group': cg, 'chamfer': cc,
                            'size': bom_match.get('size', ''),
                            'mat_number': mat_no,
                            'shape': shape_uri,
                            'cutter_shape': _item_shape_attrs.get(mat_no, ''),
                        }
                        type_idx += 1
                    seen_types[key]['count'] += 1
    cutter_config_list = sorted(seen_types.values(), key=lambda x: x['order'])

    # ── Build cutter grid context (pocket-style flat layout) ──
    # Count cutters per row per blade
    blade_row_counts = {}
    for blade in blade_data:
        for row in blade['rows']:
            count = sum(len(pos['cutters']) for pos in row['positions'])
            blade_row_counts[(blade['name'], row['row_key'])] = count

    all_blade_names = [b['name'] for b in blade_data]
    # Determine which rows exist across all blades
    all_row_keys = [rk for rk in ROW_KEYS
                    if any(blade_row_counts.get((bn, rk), 0) > 0 for bn in all_blade_names)]

    # Max cutters per row across all blades (for vertical alignment)
    row_max = {}
    for rk in all_row_keys:
        row_max[rk] = max(blade_row_counts.get((bn, rk), 0) for bn in all_blade_names)

    # Row start offsets (virtual column where each row begins)
    row_start = {}
    col_offset = 1
    for rk in all_row_keys:
        row_start[rk] = col_offset
        col_offset += row_max[rk]

    # Row separators (cumulative end of each row except last)
    row_separators = []
    cumulative = 0
    for rk in all_row_keys[:-1]:
        cumulative += row_max[rk]
        row_separators.append(cumulative)

    max_col = col_offset - 1  # total virtual columns

    # Row-suffixed header labels (1a, 2a, 3a | 1b, 2b, ...)
    _row_suffixes = {rk: chr(ord('a') + idx) for idx, rk in enumerate(all_row_keys)}
    cutter_header_labels = {}
    for rk in all_row_keys:
        suffix = _row_suffixes[rk]
        for i in range(row_max[rk]):
            vcol = row_start[rk] + i
            cutter_header_labels[vcol] = f"{i + 1}{suffix}"

    # Build grid data: maps "blade_vcol" -> config_order / cell_ref / cutter_number
    cutter_grid_data = {}
    cutter_cell_ref = {}
    cutter_number_data = {}

    for blade in blade_data:
        bn = blade['name']
        cutter_seq = 0
        for rk in all_row_keys:
            row_obj = next((r for r in blade['rows'] if r['row_key'] == rk), None)
            if not row_obj:
                continue
            pos_offset = 0
            for pos in row_obj['positions']:
                pn = pos['pos_name'].upper()
                for i in range(len(pos['cutters'])):
                    cutter_seq += 1
                    vcol = row_start[rk] + pos_offset
                    key = f"{bn}_{vcol}"

                    c = pos['cutters'][i]
                    ct = c.get('type', '')
                    cg = c.get('group', '')
                    cc = c.get('chamfer', '')
                    # Fall back to BOM summary chamfer (same as config building loop)
                    if not cc:
                        cc = summary_by_index.get(cg, {}).get('chamfer', '')
                    type_key = f"{ct}|{cg}|{cc}"
                    cfg = seen_types.get(type_key)

                    cutter_grid_data[key] = cfg['order'] if cfg else 0
                    cutter_cell_ref[key] = {'b': bn, 'r': rk, 'p': pn, 'i': i}
                    cutter_number_data[key] = cutter_seq
                    pos_offset += 1

    cutter_grid_ctx = {
        'cutter_grid_data_json': _json.dumps(cutter_grid_data),
        'cutter_cell_ref_json': _json.dumps(cutter_cell_ref),
        'cutter_number_data_json': _json.dumps(cutter_number_data),
        'cutter_header_labels_json': _json.dumps(cutter_header_labels),
        'cutter_blade_names_json': _json.dumps(all_blade_names),
        'cutter_row_separators_json': _json.dumps(sorted(row_separators)),
        'cutter_max_col': max_col,
        'cutter_blade_names': all_blade_names,
    }

    return blade_data, bom_summary, cutter_config_list, bool(blade_data), cutter_grid_ctx


def _get_pocket_grid_context(drill_bit):
    """
    Build pocket config + grid data for the receiving inspection template.
    Reuses the same pattern as DrillBitDetailEnhancedView (lines 562-621).
    Returns dict of context keys to add.
    """
    design = drill_bit.design if drill_bit else None
    if not design:
        return {'has_pocket_data': False}

    from apps.technology.models import DesignPocket, DesignPocketConfig

    pockets = DesignPocket.objects.filter(
        design=design
    ).select_related(
        'pocket_config', 'pocket_config__pocket_size', 'pocket_config__pocket_shape'
    ).order_by('blade_number', 'row_number', 'position_in_row')

    if not pockets.exists():
        return {'has_pocket_data': False}

    pocket_configs = DesignPocketConfig.objects.filter(
        design=design
    ).select_related('pocket_size', 'pocket_shape').order_by('order')

    # Build grid data with row-aligned virtual columns
    # Different blades have different cutter counts per row. We pad shorter
    # blades so all R2 cutters appear after the R1 separator, etc.
    from collections import defaultdict
    pocket_list = list(pockets)

    blade_row_counts = defaultdict(int)
    for p in pocket_list:
        blade_row_counts[(p.blade_number, p.row_number)] += 1

    all_blades = sorted(set(p.blade_number for p in pocket_list))
    all_rows = sorted(set(p.row_number for p in pocket_list))

    # Max cutters per row across ALL blades
    row_max = {}
    for row in all_rows:
        row_max[row] = max(
            blade_row_counts.get((blade, row), 0) for blade in all_blades
        )

    # Row start offsets (virtual column where each row begins)
    row_start = {}
    offset = 1
    for row in all_rows:
        row_start[row] = offset
        offset += row_max[row]

    # Row separators — cumulative end of each row except last
    row_separators = []
    cumulative = 0
    for row in all_rows[:-1]:
        cumulative += row_max[row]
        row_separators.append(cumulative)

    max_pos = offset - 1  # total virtual columns

    # Row-suffixed header labels (1a, 2a, 3a | 1b, 2b, ...)
    _row_suffixes = {row: chr(ord('a') + idx) for idx, row in enumerate(all_rows)}
    pocket_header_labels = {}
    for row in all_rows:
        suffix = _row_suffixes[row]
        for i in range(row_max[row]):
            vcol = row_start[row] + i
            pocket_header_labels[vcol] = f"{i + 1}{suffix}"

    # Build grid data using virtual columns
    grid_data = {}
    location_data = {}
    for p in pocket_list:
        vcol = row_start[p.row_number] + p.position_in_row - 1
        key = f"{p.blade_number}_{vcol}"
        grid_data[key] = p.pocket_config_id
        if p.blade_location:
            location_data[key] = p.blade_location

    # Config display colors — always use vibrant palette (ignore DB gray values)
    colors = ['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6',
              '#EC4899', '#06B6D4', '#84CC16', '#F97316', '#6366F1']
    config_data = {}
    config_list = []
    for i, cfg in enumerate(pocket_configs):
        dc = colors[i % len(colors)]
        cfg.display_color = dc
        config_data[cfg.pk] = {
            "count": cfg.count,
            "order": cfg.order,
            "color": dc,
            "name": cfg.pocket_size.display_name if cfg.pocket_size else '',
            "sizeCode": cfg.pocket_size.code if cfg.pocket_size else '',
        }
        config_list.append({
            'order': cfg.order,
            'size': cfg.pocket_size.display_name if cfg.pocket_size else '—',
            'shape': cfg.pocket_shape.name if cfg.pocket_shape else '—',
            'length_type': cfg.length_code if cfg.length_code else cfg.get_length_type_display(),
            'count': cfg.count,
            'color': dc,
        })

    # Determine blade count
    blade_nums = sorted(set(p.blade_number for p in pockets))

    # Sequential pocket numbers per blade (for B2P6 format labels)
    pocket_number_data = {}
    blade_counters = {}
    for p in pocket_list:
        bn = p.blade_number
        if bn not in blade_counters:
            blade_counters[bn] = 0
        blade_counters[bn] += 1
        vcol = row_start[p.row_number] + p.position_in_row - 1
        key = f"{bn}_{vcol}"
        pocket_number_data[key] = blade_counters[bn]

    return {
        'has_pocket_data': True,
        'pocket_grid_data_json': _json.dumps(grid_data),
        'pocket_row_separators_json': _json.dumps(sorted(row_separators)),
        'pocket_location_data_json': _json.dumps(location_data),
        'pocket_config_data_json': _json.dumps(config_data),
        'pocket_number_data_json': _json.dumps(pocket_number_data),
        'pocket_header_labels_json': _json.dumps(pocket_header_labels),
        'pocket_blade_nums_json': _json.dumps(blade_nums),
        'pocket_config_list': config_list,
        'pocket_blade_nums': blade_nums,
        'pocket_max_pos': max_pos,
        'pocket_config_total': sum(cfg.count for cfg in pocket_configs),
    }


def _apply_inspection_result_to_bit(bit, result, user=None):
    """
    Set drill bit status + location based on inspection result.
    Routes to different locations based on bit type:
    - New bits (L3/L4/L5.5) → Components Warehouse (needs production)
    - Ready bits (L5) → Finished Goods Warehouse (dispatch only)
    - Repair bits (condition REPAIRED/USED/etc.) → Evaluation Area
    - Rejected → stays in Receiving Area
    """
    if result == ReceivingInspection.InspectionResult.REJECTED:
        bit.log_change('Status', bit.get_status_display(), 'Rejected', user)
        bit.status = DrillBit.Status.REJECTED
        bit.save(update_fields=['status', 'change_log', 'updated_at'])
        bit.move_to('RCV-AREA', 'Inspection result: Rejected — stays in Receiving', user)
    else:
        # ACCEPTED or CONDITIONAL
        bit.log_change('Status', bit.get_status_display(), 'Received', user)
        bit.status = DrillBit.Status.RECEIVED
        bit.save(update_fields=['status', 'change_log', 'updated_at'])

        # Route based on bit type
        level = bit.level or (bit.design.order_level if bit.design else '')
        is_repair = bit.condition in ('REPAIRED', 'RERUN', 'USED', 'NOT_USED', 'RETROFITTED')

        if level == '5':
            # L5 Ready bits → Finished Goods (dispatch only, no production)
            bit.move_to('WH-FG', f'Inspection {result}: L5 ready bit → Finished Goods', user)
        elif is_repair:
            # Repair bits → Evaluation Area (needs evaluation before planning)
            bit.move_to('EVALUATION', f'Inspection {result}: Repair bit → Evaluation', user)
        elif level in ('3', '4', '5.5'):
            # New manufacturing bits → Components Warehouse
            bit.move_to('WH-COMP', f'Inspection {result}: L{level} new bit → Components Warehouse', user)
        else:
            # Fallback → Evaluation Area
            bit.move_to('EVALUATION', f'Inspection {result}', user)


class ReceivingInspectionCreateView(LoginRequiredMixin, CreateView):
    """Create a new Receiving Inspection for a drill bit."""
    model = ReceivingInspection
    template_name = "workorders/receiving_inspection_form.html"

    def get_form_class(self):
        from .forms import ReceivingInspectionForm
        return ReceivingInspectionForm

    def get_drill_bit(self):
        return get_object_or_404(
            DrillBit.objects.select_related(
                'design', 'design__size', 'bom', 'system_bom', 'brazing_bom',
            ),
            pk=self.kwargs['bit_pk'],
        )

    def dispatch(self, request, *args, **kwargs):
        bit = self.get_drill_bit()
        if bit.status == DrillBit.Status.UNREGISTERED:
            from django.contrib import messages
            messages.error(request, f"Bit {bit.serial_number} has status Unregistered and cannot be inspected. Please register it first.")
            return redirect('workorders:receiving_inspection_list')
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        bit = self.get_drill_bit()
        # Fix 2: Default inspection_date to today
        initial['inspection_date'] = timezone.now().date()
        # Fix 1: Auto-fill date_of_receipt from drill bit's received_date
        if bit.received_date:
            initial['date_of_receipt'] = bit.received_date
        else:
            # Fallback: check BitEvent for RECEIVED event
            evt = BitEvent.objects.filter(
                bit=bit, event_type=BitEvent.EventType.RECEIVED
            ).order_by('event_date').first()
            if evt:
                initial['date_of_receipt'] = evt.event_date.date() if evt.event_date else None
            else:
                # Second fallback: check BackloadItem
                from .models import BackloadItem
                bl_item = BackloadItem.objects.filter(
                    drill_bit=bit, received_date__isnull=False
                ).order_by('-received_date').first()
                if bl_item:
                    initial['date_of_receipt'] = bl_item.received_date.date() if bl_item.received_date else None
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = self.get_drill_bit()
        context['drill_bit'] = bit
        context['page_title'] = f"New Receiving Inspection — {bit.serial_number}"
        context['is_new'] = True
        context['report_number'] = 'RI-NEW'
        # QAS/005-1 checklist items (all NA for new) — 11 items
        context['checklist_items'] = [
            (1, "Bit Cleanliness", "vi_gauge_pads", "NA"),
            (2, "Ring Gage GO", "vi_bit_face", "NA"),
            (3, "Ring Gage NO GO", "vi_general", "NA"),
            (4, "Nozzle Threads", "vi_nozzles", "NA"),
            (5, "Breaker Slot", "vi_bit_breaker", "NA"),
            (6, "Junk Slot", "vi_junk_slot", "NA"),
            (7, "API Pin Connection", "vi_pin_connection", "NA"),
            (8, "Cutters", "vi_blades", "NA"),
            (9, "No Body Damage", "vi_bit_body", "NA"),
            (10, "Nozzle Liner Fit", "vi_nozzle_liner", "NA"),
            (11, "Q-Note from Vendor", "vi_vendor_note", "NA"),
        ]
        context['checklist_remarks'] = {}
        context['is_l5'] = (bit.level == '5')

        # BOM blade data for cutter evaluation grid
        blade_data, bom_summary, cutter_config_list, has_bom, cutter_grid_ctx = _get_bom_blade_data(bit)
        context['has_bom_data'] = has_bom
        context['blade_data_json'] = _json.dumps(blade_data)
        context['bom_summary_json'] = _json.dumps(bom_summary)
        context['eval_data_json'] = '{}'
        context['cutter_config_list'] = cutter_config_list
        context['cutter_config_json'] = _json.dumps({
            cfg['order']: {'color': cfg['color'], 'type': cfg['type'], 'group': cfg['group']}
            for cfg in cutter_config_list
        })
        context.update(cutter_grid_ctx)

        # Pocket grid data
        context.update(_get_pocket_grid_context(bit))
        context['pocket_eval_data_json'] = '{}'
        context['pocket_remark_annotations_json'] = '{}'
        context['cutter_remark_annotations_json'] = '{}'
        return context

    def _save_json_fields(self, form):
        """Parse and save JSON hidden inputs for eval data, pocket data, and checklist remarks."""
        for field, post_key in [
            ('cutter_evaluation_data', 'cutter_evaluation_data'),
            ('pocket_evaluation_data', 'pocket_evaluation_data'),
            ('checklist_remarks', 'checklist_remarks'),
            ('pocket_remark_annotations', 'pocket_remark_annotations'),
            ('cutter_remark_annotations', 'cutter_remark_annotations'),
        ]:
            raw = self.request.POST.get(post_key, '{}')
            try:
                setattr(form.instance, field, _json.loads(raw))
            except (ValueError, TypeError):
                setattr(form.instance, field, {})
        # Plain-text auto-remarks
        form.instance.pocket_auto_remarks = self.request.POST.get('pocket_auto_remarks', '')
        form.instance.cutter_auto_remarks = self.request.POST.get('cutter_auto_remarks', '')

    def form_valid(self, form):
        form.instance.drill_bit = self.get_drill_bit()
        form.instance.inspected_by = self.request.user
        if not form.instance.inspection_date:
            form.instance.inspection_date = timezone.now().date()
        self._save_json_fields(form)

        # Handle "Save & Complete" on create — same as edit path
        mark_complete = self.request.POST.get('mark_complete')
        if mark_complete == 'true':
            if form.instance.result == ReceivingInspection.InspectionResult.PENDING:
                messages.error(self.request,
                    "Cannot complete inspection — please set the Result "
                    "(Accepted / Rejected / Conditional) before completing.")
                response = super().form_valid(form)
                return response
            form.instance.is_complete = True
            form.instance.qc_approved_by = self.request.user
            form.instance.qc_approved_at = timezone.now()

        response = super().form_valid(form)

        # Update drill bit status + location on completion
        if mark_complete == 'true' and form.instance.is_complete:
            bit = form.instance.drill_bit
            _apply_inspection_result_to_bit(bit, form.instance.result, self.request.user)
            messages.success(self.request, "Receiving inspection created and completed.")
            notify(
                actor=self.request.user,
                verb="completed receiving inspection for",
                target=f"SN {bit.serial_number}",
                priority="HIGH",
                action_url=reverse('workorders:receiving_inspection_edit',
                                   kwargs={'bit_pk': bit.pk, 'pk': self.object.pk}),
                entity_type="ReceivingInspection",
                entity_id=self.object.pk,
            )
        else:
            messages.success(self.request, "Receiving inspection created.")

        return response

    def get_success_url(self):
        return reverse('workorders:receiving_inspection_edit',
                        kwargs={'bit_pk': self.kwargs['bit_pk'], 'pk': self.object.pk})


class ReceivingInspectionEditView(LoginRequiredMixin, UpdateView):
    """Edit an existing Receiving Inspection."""
    model = ReceivingInspection
    template_name = "workorders/receiving_inspection_form.html"

    def get_form_class(self):
        from .forms import ReceivingInspectionForm
        return ReceivingInspectionForm

    def get_queryset(self):
        return ReceivingInspection.objects.select_related(
            'drill_bit', 'drill_bit__design', 'drill_bit__design__size',
            'drill_bit__bom', 'drill_bit__system_bom', 'drill_bit__brazing_bom',
            'work_order', 'inspected_by', 'qc_approved_by'
        ).filter(drill_bit__pk=self.kwargs['bit_pk'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = self.object.drill_bit
        context['drill_bit'] = bit
        context['page_title'] = f"Receiving Inspection — {bit.serial_number}"
        context['is_new'] = False
        context['report_number'] = self.object.report_number
        context['checklist_items'] = self.object.checklist_items
        context['checklist_remarks'] = self.object.checklist_remarks or {}
        context['is_l5'] = (bit.level == '5')
        # QR code for print header
        from apps.workorders.utils import generate_drill_bit_qr
        context['bit_qr_base64'] = generate_drill_bit_qr(bit)
        # BOM blade data for cutter evaluation grid
        blade_data, bom_summary, cutter_config_list, has_bom, cutter_grid_ctx = _get_bom_blade_data(bit)
        context['has_bom_data'] = has_bom
        context['blade_data_json'] = _json.dumps(blade_data)
        context['bom_summary_json'] = _json.dumps(bom_summary)
        context['eval_data_json'] = _json.dumps(self.object.cutter_evaluation_data or {})
        context['cutter_config_list'] = cutter_config_list
        context['cutter_config_json'] = _json.dumps({
            cfg['order']: {'color': cfg['color'], 'type': cfg['type'], 'group': cfg['group']}
            for cfg in cutter_config_list
        })
        context.update(cutter_grid_ctx)
        # Pocket grid data
        context.update(_get_pocket_grid_context(bit))
        context['pocket_eval_data_json'] = _json.dumps(self.object.pocket_evaluation_data or {})
        # Remark annotations (for restoring clickable annotations on edit)
        context['pocket_remark_annotations_json'] = _json.dumps(self.object.pocket_remark_annotations or {})
        context['cutter_remark_annotations_json'] = _json.dumps(self.object.cutter_remark_annotations or {})
        # Attachments
        context['attachments'] = self.object.attachments.all()
        # Version history
        from apps.notifications.models import FormRevision
        context['revisions'] = FormRevision.objects.filter(
            entity_type="ReceivingInspection", entity_id=self.object.pk
        ).select_related('revised_by').order_by('-revision_number')[:20]
        # Reopen guard: determine if reopening is allowed
        if self.object.is_complete and bit:
            from apps.workorders.models import WorkOrder
            has_active_wo = bit.work_orders.exclude(
                status__in=[WorkOrder.Status.CANCELLED, WorkOrder.Status.COMPLETED]
            ).exists()
            bit_in_receiving = bit.status in (
                DrillBit.Status.RECEIVED, DrillBit.Status.RECEIVING,
                DrillBit.Status.IN_EVALUATION, DrillBit.Status.IN_STOCK,
                DrillBit.Status.BACKLOADED, DrillBit.Status.REJECTED,
            )
            context['can_reopen'] = not has_active_wo and bit_in_receiving
        else:
            context['can_reopen'] = False
        return context

    def _snapshot_inspection(self, obj):
        """Build a dict snapshot of key inspection fields for revision tracking."""
        snapshot = {}
        for f in ['inspection_date', 'po_number', 'date_of_receipt', 'result', 'remarks',
                   'vi_pin_connection', 'vi_bit_body', 'vi_bit_breaker', 'vi_blades',
                   'vi_nozzles', 'vi_junk_slot', 'vi_gauge_pads', 'vi_bit_face',
                   'vi_general', 'vi_nozzle_liner', 'vi_vendor_note',
                   'is_complete']:
            val = getattr(obj, f, None)
            snapshot[f] = str(val) if val is not None else ""
        # JSON fields — store as-is
        for f in ['cutter_evaluation_data', 'pocket_evaluation_data', 'checklist_remarks']:
            snapshot[f] = getattr(obj, f, {}) or {}
        return snapshot

    def form_valid(self, form):
        # Capture pre-save state for revision tracking
        old_obj = ReceivingInspection.objects.filter(pk=form.instance.pk).first()
        snapshot_old = self._snapshot_inspection(old_obj) if old_obj else None

        # Save JSON fields from hidden inputs
        for field, post_key in [
            ('cutter_evaluation_data', 'cutter_evaluation_data'),
            ('pocket_evaluation_data', 'pocket_evaluation_data'),
            ('checklist_remarks', 'checklist_remarks'),
            ('pocket_remark_annotations', 'pocket_remark_annotations'),
            ('cutter_remark_annotations', 'cutter_remark_annotations'),
        ]:
            raw = self.request.POST.get(post_key, '{}')
            try:
                setattr(form.instance, field, _json.loads(raw))
            except (ValueError, TypeError):
                setattr(form.instance, field, {})
        # Plain-text auto-remarks
        form.instance.pocket_auto_remarks = self.request.POST.get('pocket_auto_remarks', '')
        form.instance.cutter_auto_remarks = self.request.POST.get('cutter_auto_remarks', '')
        # Check if "mark_complete" was submitted
        mark_complete = self.request.POST.get('mark_complete')
        bit = form.instance.drill_bit
        serial = bit.serial_number if bit else "Unknown"

        if mark_complete == 'true' and not form.instance.is_complete:
            # Block completion if result is still PENDING
            if form.instance.result == ReceivingInspection.InspectionResult.PENDING:
                messages.error(self.request,
                    "Cannot complete inspection — please set the Result "
                    "(Accepted / Rejected / Conditional) before completing.")
                return super().form_valid(form)
            form.instance.is_complete = True
            form.instance.qc_approved_by = self.request.user
            form.instance.qc_approved_at = timezone.now()
            messages.success(self.request, "Receiving inspection marked as complete.")
        elif mark_complete == 'false' and form.instance.is_complete:
            # Reopen guard: only if bit still in receiving/evaluation area and no active WO
            can_reopen = True
            if bit:
                has_active_wo = bit.work_orders.exclude(
                    status__in=[WorkOrder.Status.CANCELLED, WorkOrder.Status.COMPLETED]
                ).exists()
                if has_active_wo:
                    messages.error(self.request,
                        "Cannot reopen — a work order already exists for this drill bit.")
                    can_reopen = False
                elif bit.status not in (
                    DrillBit.Status.RECEIVED, DrillBit.Status.RECEIVING,
                    DrillBit.Status.IN_EVALUATION, DrillBit.Status.IN_STOCK,
                    DrillBit.Status.BACKLOADED, DrillBit.Status.REJECTED,
                ):
                    messages.error(self.request,
                        f"Cannot reopen — drill bit has moved to '{bit.get_status_display()}'. "
                        "Contact admin for changes.")
                    can_reopen = False
            if can_reopen:
                form.instance.is_complete = False
                form.instance.qc_approved_by = None
                form.instance.qc_approved_at = None
                messages.success(self.request, "Receiving inspection reopened.")
            else:
                return super().form_valid(form)
        else:
            messages.success(self.request, "Receiving inspection saved.")

        # Update drill bit status + location on completion/reopen
        if mark_complete == 'true' and form.instance.is_complete and bit:
            _apply_inspection_result_to_bit(bit, form.instance.result, self.request.user)
        elif mark_complete == 'false' and not form.instance.is_complete and bit:
            bit.status = DrillBit.Status.RECEIVING
            rcv_loc = Location.objects.filter(code='RCV-AREA').first()
            if rcv_loc:
                bit.bit_location = rcv_loc
            bit.save(update_fields=['status', 'bit_location', 'updated_at'])

        # Notify on completion
        if mark_complete == 'true' and form.instance.is_complete:
            notify(
                actor=self.request.user,
                verb="completed receiving inspection for",
                target=f"SN {serial}",
                priority="HIGH",
                action_url=reverse('workorders:receiving_inspection_edit',
                                   kwargs={'bit_pk': bit.pk, 'pk': form.instance.pk}) if bit else "",
                entity_type="ReceivingInspection",
                entity_id=form.instance.pk,
            )

        response = super().form_valid(form)

        # Create form revision after save
        snapshot_new = self._snapshot_inspection(form.instance)
        if snapshot_old is None or snapshot_old != snapshot_new:
            create_form_revision(
                entity_type="ReceivingInspection",
                entity_id=form.instance.pk,
                document_code="QAS/005-1",
                snapshot_new=snapshot_new,
                snapshot_old=snapshot_old,
                revised_by=self.request.user,
            )

        return response

    def get_success_url(self):
        return reverse('workorders:receiving_inspection_edit',
                        kwargs={'bit_pk': self.kwargs['bit_pk'], 'pk': self.object.pk})


@login_required
@require_POST
def api_receiving_inspection_complete(request, bit_pk, pk):
    """Toggle is_complete on a ReceivingInspection. POST-only."""
    inspection = get_object_or_404(ReceivingInspection, pk=pk, drill_bit__pk=bit_pk)

    bit = inspection.drill_bit
    serial = bit.serial_number if bit else "Unknown"
    if inspection.is_complete:
        # Reopen guard: only if bit still in receiving/evaluation area and no active WO
        if bit:
            has_active_wo = bit.work_orders.exclude(
                status__in=[WorkOrder.Status.CANCELLED, WorkOrder.Status.COMPLETED]
            ).exists()
            if has_active_wo:
                return JsonResponse({'success': False,
                    'message': 'Cannot reopen — a work order already exists for this drill bit.'})
            if bit.status not in (
                DrillBit.Status.RECEIVED, DrillBit.Status.RECEIVING,
                DrillBit.Status.IN_EVALUATION, DrillBit.Status.IN_STOCK,
                DrillBit.Status.BACKLOADED, DrillBit.Status.REJECTED,
            ):
                return JsonResponse({'success': False,
                    'message': f"Cannot reopen — drill bit has moved to '{bit.get_status_display()}'."})
        # Reopen
        inspection.is_complete = False
        inspection.qc_approved_by = None
        inspection.qc_approved_at = None
        inspection.save(update_fields=['is_complete', 'qc_approved_by', 'qc_approved_at', 'updated_at'])
        if bit:
            bit.status = DrillBit.Status.RECEIVING
            rcv_loc = Location.objects.filter(code='RCV-AREA').first()
            if rcv_loc:
                bit.bit_location = rcv_loc
            bit.save(update_fields=['status', 'bit_location', 'updated_at'])
        return JsonResponse({'success': True, 'is_complete': False, 'message': 'Inspection reopened.'})
    else:
        # Block completion if result is still PENDING
        if inspection.result == ReceivingInspection.InspectionResult.PENDING:
            return JsonResponse({'success': False,
                'message': 'Cannot complete — please set the Result first.'})
        # Complete
        inspection.is_complete = True
        inspection.qc_approved_by = request.user
        inspection.qc_approved_at = timezone.now()
        inspection.save(update_fields=['is_complete', 'qc_approved_by', 'qc_approved_at', 'updated_at'])
        if bit:
            _apply_inspection_result_to_bit(bit, inspection.result)
        notify(
            actor=request.user,
            verb="completed receiving inspection for",
            target=f"SN {serial}",
            priority="HIGH",
            action_url=reverse('workorders:receiving_inspection_edit',
                               kwargs={'bit_pk': bit_pk, 'pk': pk}),
            entity_type="ReceivingInspection",
            entity_id=pk,
        )
        return JsonResponse({'success': True, 'is_complete': True, 'message': 'Inspection marked complete.'})


@login_required
@require_POST
def api_receiving_inspection_upload(request, bit_pk, pk):
    """Upload an attachment to a receiving inspection."""
    inspection = get_object_or_404(ReceivingInspection, pk=pk, drill_bit__pk=bit_pk)
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)
    name = request.POST.get('name', 'Q-Note').strip() or 'Q-Note'
    attachment = ReceivingInspectionAttachment.objects.create(
        inspection=inspection,
        file=file,
        name=name,
        uploaded_by=request.user,
    )
    return JsonResponse({
        'success': True,
        'attachment': {
            'id': attachment.pk,
            'name': attachment.name,
            'file_url': attachment.file.url,
            'file_extension': attachment.file_extension,
            'is_image': attachment.is_image,
            'uploaded_at': attachment.uploaded_at.strftime('%b %d, %Y %H:%M'),
            'uploaded_by': request.user.get_full_name() or request.user.username,
        }
    })


@login_required
@require_POST
def api_receiving_inspection_delete_attachment(request, bit_pk, pk, att_pk):
    """Delete an attachment from a receiving inspection."""
    attachment = get_object_or_404(ReceivingInspectionAttachment, pk=att_pk,
                                  inspection__pk=pk, inspection__drill_bit__pk=bit_pk)
    attachment.file.delete(save=False)
    attachment.delete()
    return JsonResponse({'success': True})


# =============================================================================
# EVALUATION AUTO-CREATE (Standalone per-type URLs)
# =============================================================================

class EvaluationAutoCreateView(LoginRequiredMixin, View):
    """
    GET /workorders/<wo_pk>/evaluation/<type_code>/
    Auto-creates CutterEvaluationMatrix for the given type if not exists,
    then redirects to the matrix editor.
    """
    def get(self, request, wo_pk, type_code):
        wo = get_object_or_404(WorkOrder, pk=wo_pk)
        # Validate type_code
        valid_codes = [c[0] for c in CutterEvaluationMatrix.EvaluationType.choices]
        if type_code not in valid_codes:
            messages.error(request, f"Invalid evaluation type: {type_code}")
            return redirect('workorders:workorder_detail_enhanced', pk=wo.pk)

        matrix, created = CutterEvaluationMatrix.objects.get_or_create(
            work_order=wo,
            evaluation_type=type_code,
            evaluation_number=1,
            defaults={'evaluated_by': request.user}
        )
        if created:
            messages.info(request, f"Created new {matrix.get_evaluation_type_display()}")
        # PDC_EVAL uses the new pre-repair evaluation page
        if type_code == 'PDC_EVAL':
            return redirect('workorders:pre_repair_eval_edit', wo_pk=wo.pk, pk=matrix.pk)
        return redirect('workorders:cutter_evaluation_edit', wo_pk=wo.pk, pk=matrix.pk)


# ========================================================================
# PRE-REPAIR EVALUATION (PDC_EVAL) — Cloned from Receiving Inspection
# ========================================================================

class PreRepairEvalEditView(LoginRequiredMixin, TemplateView):
    """
    Pre-repair evaluation page (QAS/1002) — cloned from ReceivingInspection.
    Uses CutterEvaluationMatrix model with JSON-based checklist.
    Features: checklist, pocket grid, cutter grid, photos, decision,
    plus 3 icon buttons for standalone Die Check / LPT / Thread pages.
    """
    template_name = "workorders/pre_repair_evaluation.html"

    def _get_wo_and_matrix(self):
        wo = get_object_or_404(
            WorkOrder.objects.select_related(
                'drill_bit', 'drill_bit__design', 'drill_bit__design__size',
                'drill_bit__bom', 'drill_bit__system_bom', 'drill_bit__brazing_bom',
                'account',
            ),
            pk=self.kwargs['wo_pk'],
        )
        matrix = get_object_or_404(CutterEvaluationMatrix, pk=self.kwargs['pk'])
        return wo, matrix

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo, matrix = self._get_wo_and_matrix()
        bit = wo.drill_bit

        context['work_order'] = wo
        context['matrix'] = matrix
        context['drill_bit'] = bit
        context['page_title'] = f"Internal Evaluation Sheet — {wo.wo_number}"
        context['is_new'] = False
        context['report_number'] = matrix.inspection_number or f'EV-{matrix.pk:04d}'

        # Checklist items for PDC_EVAL (18 items from model constant)
        checklist_items_labels = CutterEvaluationMatrix.CHECKLIST_ITEMS.get('PDC_EVAL', [])
        saved_checklist = matrix.checklist_data or []
        saved_map = {}
        for item in saved_checklist:
            saved_map[item.get('item', '')] = item

        # Build tuples matching receiving inspection format: (num, label, field_key, current_value)
        checklist_for_template = []
        for idx, label in enumerate(checklist_items_labels, 1):
            saved = saved_map.get(label, {})
            field_key = f'cl_{idx}'
            status = saved.get('status', '')
            checklist_for_template.append((idx, label, field_key, status))
        context['checklist_items'] = checklist_for_template
        context['checklist_remarks'] = {}
        # Rebuild remarks dict from saved checklist data
        for idx, label in enumerate(checklist_items_labels, 1):
            saved = saved_map.get(label, {})
            field_key = f'cl_{idx}'
            if saved.get('remarks') or saved.get('reason'):
                context['checklist_remarks'][field_key] = {
                    'reason': saved.get('reason', ''),
                    'remarks': saved.get('remarks', ''),
                }

        # Decision choices
        context['decision_choices'] = CutterEvaluationMatrix.Decision.choices

        # Standalone test reports linked to this evaluation
        context['die_check_reports'] = DieCheckReport.objects.filter(
            evaluation=matrix
        ).order_by('-created_at')
        context['lpt_reports'] = StandaloneLPTReport.objects.filter(
            evaluation=matrix
        ).order_by('-created_at')
        context['thread_reports'] = StandaloneThreadReport.objects.filter(
            evaluation=matrix
        ).order_by('-created_at')

        # BOM blade data for cutter evaluation grid
        if bit:
            blade_data, bom_summary, cutter_config_list, has_bom, cutter_grid_ctx = _get_bom_blade_data(bit)
            context['has_bom_data'] = has_bom
            context['blade_data_json'] = _json.dumps(blade_data)
            context['bom_summary_json'] = _json.dumps(bom_summary)
            context['cutter_config_list'] = cutter_config_list
            context['cutter_config_json'] = _json.dumps({
                cfg['order']: {'color': cfg['color'], 'type': cfg['type'], 'group': cfg['group']}
                for cfg in cutter_config_list
            })
            context.update(cutter_grid_ctx)

            # Pocket grid data
            context.update(_get_pocket_grid_context(bit))
        else:
            context['has_bom_data'] = False
            context['has_pocket_data'] = False

        # Existing eval data from matrix JSONFields
        context['eval_data_json'] = _json.dumps(matrix.die_check_data or {})
        # Use pocket_evaluation_data on CutterEvaluationMatrix
        context['pocket_eval_data_json'] = _json.dumps(matrix.pocket_evaluation_data or {})
        context['pocket_remark_annotations_json'] = _json.dumps(matrix.pocket_remark_annotations or {})
        context['cutter_remark_annotations_json'] = _json.dumps(matrix.cutter_remark_annotations or {})

        # Version history
        from apps.notifications.models import FormRevision
        context['revisions'] = FormRevision.objects.filter(
            entity_type="CutterEvaluationMatrix", entity_id=matrix.pk
        ).select_related('revised_by').order_by('-revision_number')[:20]

        return context

    def post(self, request, *args, **kwargs):
        """Handle saves via AJAX JSON or form POST."""
        wo, matrix = self._get_wo_and_matrix()

        content_type = request.content_type or ''
        if 'application/json' in content_type:
            data = _json.loads(request.body)
            return self._handle_json_save(request, matrix, data)

        # Standard form POST (Save / Mark Complete)
        return self._handle_form_save(request, wo, matrix)

    def _handle_json_save(self, request, matrix, data):
        """Handle AJAX JSON save (grid updates, checklist, etc.)."""
        update_fields = ['updated_at']

        # Checklist data: convert from field-based format to list-of-dicts
        checklist_post = data.get('checklist_data')
        if checklist_post is not None:
            matrix.checklist_data = checklist_post
            update_fields.append('checklist_data')

        # Cutter eval entries (blade/row/pos/idx format)
        cutter_eval_data = data.get('cutter_eval_data')
        if cutter_eval_data is not None:
            matrix.die_check_data = cutter_eval_data
            update_fields.append('die_check_data')

        # Pocket eval data
        pocket_eval_data = data.get('pocket_eval_data')
        if pocket_eval_data is not None:
            matrix.pocket_evaluation_data = pocket_eval_data
            update_fields.append('pocket_evaluation_data')

        # Decision
        decision = data.get('decision')
        if decision is not None:
            matrix.decision = decision
            update_fields.append('decision')

        # General remark
        remarks = data.get('remarks')
        if remarks is not None:
            matrix.general_remark = remarks
            update_fields.append('general_remark')

        # Remark annotations
        for field in ('pocket_remark_annotations', 'cutter_remark_annotations'):
            val = data.get(field)
            if val is not None:
                setattr(matrix, field, val)
                update_fields.append(field)

        # Mark complete
        mark_complete = data.get('mark_complete')
        if mark_complete is True:
            matrix.is_complete = True
            matrix.qc_by = request.user
            matrix.qc_at = timezone.now()
            update_fields.extend(['is_complete', 'qc_by', 'qc_at'])
        elif mark_complete is False:
            matrix.is_complete = False
            update_fields.append('is_complete')

        matrix.save(update_fields=update_fields)

        return JsonResponse({
            'success': True,
            'is_complete': matrix.is_complete,
        })

    def _handle_form_save(self, request, wo, matrix):
        """Handle standard form POST (Save button)."""
        # Build checklist_data from POST radio buttons and remarks
        checklist_labels = CutterEvaluationMatrix.CHECKLIST_ITEMS.get('PDC_EVAL', [])
        checklist_data = []
        checklist_remarks_raw = {}
        try:
            checklist_remarks_raw = _json.loads(request.POST.get('checklist_remarks', '{}'))
        except (ValueError, TypeError):
            pass

        for idx, label in enumerate(checklist_labels, 1):
            field_key = f'cl_{idx}'
            status = request.POST.get(field_key, '')
            remark_data = checklist_remarks_raw.get(field_key, {})
            checklist_data.append({
                'item': label,
                'status': status,
                'reason': remark_data.get('reason', ''),
                'remarks': remark_data.get('remarks', ''),
            })

        update_fields = ['checklist_data', 'updated_at']
        matrix.checklist_data = checklist_data

        # Cutter eval data from hidden input
        raw_cutter = request.POST.get('cutter_evaluation_data', '{}')
        try:
            matrix.die_check_data = _json.loads(raw_cutter)
            update_fields.append('die_check_data')
        except (ValueError, TypeError):
            pass

        # Pocket eval data from hidden input
        raw_pocket = request.POST.get('pocket_evaluation_data', '{}')
        try:
            matrix.pocket_evaluation_data = _json.loads(raw_pocket)
            update_fields.append('pocket_evaluation_data')
        except (ValueError, TypeError):
            pass

        # Decision
        decision = request.POST.get('decision', '')
        if decision:
            matrix.decision = decision
            update_fields.append('decision')

        # Remarks
        remarks = request.POST.get('general_remark', '')
        matrix.general_remark = remarks
        update_fields.append('general_remark')

        # Auto-remarks
        pocket_auto = request.POST.get('pocket_auto_remarks', '')
        cutter_auto = request.POST.get('cutter_auto_remarks', '')

        # Remark annotations
        for field, post_key in [
            ('pocket_remark_annotations', 'pocket_remark_annotations'),
            ('cutter_remark_annotations', 'cutter_remark_annotations'),
        ]:
            raw = request.POST.get(post_key, '{}')
            try:
                setattr(matrix, field, _json.loads(raw))
                update_fields.append(field)
            except (ValueError, TypeError):
                pass

        # Mark complete
        mark_complete = request.POST.get('mark_complete')
        if mark_complete == 'true' and not matrix.is_complete:
            matrix.is_complete = True
            matrix.qc_by = request.user
            matrix.qc_at = timezone.now()
            update_fields.extend(['is_complete', 'qc_by', 'qc_at'])
            messages.success(request, "Evaluation marked as complete.")
            # Notify
            serial = wo.drill_bit.serial_number if wo.drill_bit else wo.wo_number
            notify(
                actor=request.user,
                verb="completed PDC evaluation for",
                target=f"SN {serial}",
                priority="HIGH",
                action_url=reverse('workorders:pre_repair_eval_edit',
                                   kwargs={'wo_pk': wo.pk, 'pk': matrix.pk}),
                entity_type="CutterEvaluationMatrix",
                entity_id=matrix.pk,
            )
        elif mark_complete == 'false' and matrix.is_complete:
            matrix.is_complete = False
            update_fields.append('is_complete')
            messages.success(request, "Evaluation reopened.")
        else:
            messages.success(request, "Evaluation saved.")

        matrix.save(update_fields=update_fields)

        # Version tracking
        create_form_revision(
            entity_type="CutterEvaluationMatrix",
            entity_id=matrix.pk,
            document_code="QAS/1002",
            snapshot_new={'checklist_data': matrix.checklist_data, 'decision': matrix.decision,
                          'is_complete': matrix.is_complete, 'general_remark': matrix.general_remark},
            snapshot_old=None,
            revised_by=request.user,
        )

        return redirect('workorders:pre_repair_eval_edit', wo_pk=wo.pk, pk=matrix.pk)


# ── Standalone Test Page Views (Placeholders) ──

class DieCheckReportView(LoginRequiredMixin, TemplateView):
    """Standalone Die Check test page with LPT materials, photos, barcode scanning."""
    template_name = "workorders/die_check_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo = get_object_or_404(WorkOrder.objects.select_related(
            'drill_bit', 'drill_bit__design', 'drill_bit__design__size',
            'drill_bit__system_bom', 'drill_bit__brazing_bom',
        ), pk=self.kwargs['wo_pk'])
        context['work_order'] = wo
        context['drill_bit'] = wo.drill_bit
        context['page_title'] = f"Die Check — {wo.wo_number}"

        eval_pk = self.kwargs.get('eval_pk')
        evaluation = get_object_or_404(CutterEvaluationMatrix, pk=eval_pk) if eval_pk else None
        context['evaluation'] = evaluation

        report_pk = self.kwargs.get('pk')
        if report_pk:
            report = get_object_or_404(DieCheckReport, pk=report_pk)
            context['report'] = report
            context['is_new'] = False
        else:
            report = None
            context['report'] = None
            context['is_new'] = True

        # Auto-detect stage for new reports
        if report:
            context['stage'] = report.stage
        else:
            context['stage'] = DieCheckReport.auto_detect_stage(wo)

        # Materials data for pre-population
        context['materials_data_json'] = _json.dumps(
            report.materials_data if report and report.materials_data else {}
        )

        # Stage choices for dropdown
        context['stage_choices'] = DieCheckReport.Stage.choices

        # Grid data (saved cutter evaluation) for pre-population
        context['grid_data_json'] = _json.dumps(
            report.grid_data if report and report.grid_data else {}
        )

        # Cutter grid context from BOM (same as receiving inspection)
        bit = wo.drill_bit
        has_bom = False
        if bit:
            blade_data, bom_summary, cutter_config_list, has_bom, cutter_grid_ctx = _get_bom_blade_data(bit)
            if has_bom:
                context['cutter_config_list'] = cutter_config_list
                context['cutter_config_json'] = _json.dumps([
                    {k: v for k, v in cfg.items() if k != 'shape'}
                    for cfg in cutter_config_list
                ])
                context.update(cutter_grid_ctx)
        context['has_bom_data'] = has_bom

        return context

    def post(self, request, *args, **kwargs):
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        eval_pk = self.kwargs.get('eval_pk')
        evaluation = get_object_or_404(CutterEvaluationMatrix, pk=eval_pk) if eval_pk else None

        report_pk = self.kwargs.get('pk')
        if report_pk:
            report = get_object_or_404(DieCheckReport, pk=report_pk)
        else:
            report = DieCheckReport(
                work_order=wo,
                drill_bit=wo.drill_bit,
                evaluation=evaluation,
                performed_by=request.user,
            )

        # Stage
        report.stage = request.POST.get('stage', DieCheckReport.Stage.BEFORE_BRAZE)

        # Grid data
        try:
            report.grid_data = _json.loads(request.POST.get('grid_data', '{}'))
        except (ValueError, TypeError):
            report.grid_data = {}

        # Materials data (penetrant, developer)
        try:
            report.materials_data = _json.loads(request.POST.get('materials_data', '{}'))
        except (ValueError, TypeError):
            report.materials_data = {}

        report.result = request.POST.get('result', '')
        report.remarks = request.POST.get('remarks', '')

        if request.POST.get('mark_complete') == 'true':
            report.is_complete = True
            report.performed_at = timezone.now()

        report.save()

        # Auto-fill checklist when completed
        if report.is_complete and evaluation:
            self._auto_fill_checklist(evaluation, report)

        # Send notification for "Waiting Quality Decision" cutters
        decisions = (report.grid_data or {}).get('decisions', {})
        waiting_cutters = [
            k for k, v in decisions.items()
            if isinstance(v, dict) and v.get('decision') == 'WAITING_QD'
        ]
        if waiting_cutters:
            cutter_list = ', '.join(waiting_cutters[:5])
            if len(waiting_cutters) > 5:
                cutter_list += f' (+{len(waiting_cutters) - 5} more)'
            notify(
                actor=request.user,
                verb=f'requested quality decision on die check for',
                target=wo.wo_number,
                recipients='all',
                priority='HIGH',
                action_url=reverse('workorders:die_check_edit',
                                   kwargs={'wo_pk': wo.pk, 'pk': report.pk, 'eval_pk': eval_pk}) if eval_pk else
                            reverse('workorders:workorder_detail_enhanced', kwargs={'pk': wo.pk}),
                message=f'Cutters needing decision: {cutter_list}',
                entity_type='DieCheckReport',
                entity_id=report.pk,
            )

        messages.success(request, "Die check report saved.")
        if eval_pk:
            return redirect('workorders:pre_repair_eval_edit', wo_pk=wo.pk, pk=eval_pk)
        return redirect('workorders:workorder_detail_enhanced', pk=wo.pk)

    def _auto_fill_checklist(self, evaluation, report):
        """Auto-fill checklist item #4 (Die Check) with result."""
        checklist = evaluation.checklist_data or []
        for item in checklist:
            if item.get('item') == 'Die Check':
                item['status'] = 'OK' if report.result == 'PASS' else 'NOT_OK'
                item['remarks'] = f"Die Check ({report.get_stage_display()}): {report.get_result_display()} (Report #{report.pk})"
                break
        evaluation.checklist_data = checklist
        evaluation.save(update_fields=['checklist_data', 'updated_at'])


class StandaloneLPTReportView(LoginRequiredMixin, TemplateView):
    """Standalone LPT (Pressure Test) page."""
    template_name = "workorders/standalone_lpt_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo = get_object_or_404(WorkOrder.objects.select_related(
            'drill_bit', 'drill_bit__design',
        ), pk=self.kwargs['wo_pk'])
        context['work_order'] = wo
        context['drill_bit'] = wo.drill_bit
        context['page_title'] = f"Pressure Test (LPT) — {wo.wo_number}"
        eval_pk = self.kwargs.get('eval_pk')
        context['evaluation'] = get_object_or_404(CutterEvaluationMatrix, pk=eval_pk) if eval_pk else None
        report_pk = self.kwargs.get('pk')
        if report_pk:
            context['report'] = get_object_or_404(StandaloneLPTReport, pk=report_pk)
            context['is_new'] = False
        else:
            context['report'] = None
            context['is_new'] = True
        return context

    def post(self, request, *args, **kwargs):
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        eval_pk = self.kwargs.get('eval_pk')
        evaluation = get_object_or_404(CutterEvaluationMatrix, pk=eval_pk) if eval_pk else None

        report_pk = self.kwargs.get('pk')
        if report_pk:
            report = get_object_or_404(StandaloneLPTReport, pk=report_pk)
        else:
            report = StandaloneLPTReport(
                work_order=wo, drill_bit=wo.drill_bit,
                evaluation=evaluation, performed_by=request.user,
            )
        try:
            report.test_data = _json.loads(request.POST.get('test_data', '{}'))
        except (ValueError, TypeError):
            report.test_data = {}
        report.result = request.POST.get('result', '')
        report.remarks = request.POST.get('remarks', '')
        if request.POST.get('mark_complete') == 'true':
            report.is_complete = True
            report.performed_at = timezone.now()
        report.save()

        # Auto-fill checklist item #16 (Pressure Test)
        if report.is_complete and evaluation:
            checklist = evaluation.checklist_data or []
            for item in checklist:
                if 'Pressure Test' in item.get('item', ''):
                    item['status'] = 'OK' if report.result == 'PASS' else 'NOT_OK'
                    item['remarks'] = f"LPT: {report.get_result_display()} (Report #{report.pk})"
                    break
            evaluation.checklist_data = checklist
            evaluation.save(update_fields=['checklist_data', 'updated_at'])

        messages.success(request, "LPT report saved.")
        if eval_pk:
            return redirect('workorders:pre_repair_eval_edit', wo_pk=wo.pk, pk=eval_pk)
        return redirect('workorders:workorder_detail_enhanced', pk=wo.pk)


class StandaloneThreadReportView(LoginRequiredMixin, TemplateView):
    """Standalone API Thread Inspection page."""
    template_name = "workorders/standalone_thread_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        wo = get_object_or_404(WorkOrder.objects.select_related(
            'drill_bit', 'drill_bit__design',
        ), pk=self.kwargs['wo_pk'])
        context['work_order'] = wo
        context['drill_bit'] = wo.drill_bit
        context['page_title'] = f"API Thread Inspection — {wo.wo_number}"
        eval_pk = self.kwargs.get('eval_pk')
        context['evaluation'] = get_object_or_404(CutterEvaluationMatrix, pk=eval_pk) if eval_pk else None
        report_pk = self.kwargs.get('pk')
        if report_pk:
            context['report'] = get_object_or_404(StandaloneThreadReport, pk=report_pk)
            context['is_new'] = False
        else:
            context['report'] = None
            context['is_new'] = True
        return context

    def post(self, request, *args, **kwargs):
        wo = get_object_or_404(WorkOrder, pk=self.kwargs['wo_pk'])
        eval_pk = self.kwargs.get('eval_pk')
        evaluation = get_object_or_404(CutterEvaluationMatrix, pk=eval_pk) if eval_pk else None
        report_pk = self.kwargs.get('pk')
        if report_pk:
            report = get_object_or_404(StandaloneThreadReport, pk=report_pk)
        else:
            report = StandaloneThreadReport(
                work_order=wo, drill_bit=wo.drill_bit,
                evaluation=evaluation, performed_by=request.user,
            )
        try:
            report.inspection_data = _json.loads(request.POST.get('inspection_data', '{}'))
        except (ValueError, TypeError):
            report.inspection_data = {}
        report.result = request.POST.get('result', '')
        report.remarks = request.POST.get('remarks', '')
        if request.POST.get('mark_complete') == 'true':
            report.is_complete = True
            report.performed_at = timezone.now()
        report.save()

        # Auto-fill checklist item #14 (API Pin) — optional
        if report.is_complete and evaluation:
            checklist = evaluation.checklist_data or []
            for item in checklist:
                if 'API Pin' in item.get('item', ''):
                    item['status'] = 'OK' if report.result == 'PASS' else 'NOT_OK'
                    item['remarks'] = f"API Thread: {report.get_result_display()} (Report #{report.pk})"
                    break
            evaluation.checklist_data = checklist
            evaluation.save(update_fields=['checklist_data', 'updated_at'])

        messages.success(request, "API Thread inspection saved.")
        if eval_pk:
            return redirect('workorders:pre_repair_eval_edit', wo_pk=wo.pk, pk=eval_pk)
        return redirect('workorders:workorder_detail_enhanced', pk=wo.pk)


# ═══════════════════════════════════════════════════════════════════════════
# LOCATION TRANSFERS
# ═══════════════════════════════════════════════════════════════════════════

class LocationTransferView(LoginRequiredMixin, TemplateView):
    """Central page for transferring drill bits between locations."""
    template_name = "workorders/location_transfers.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recent_transfers'] = BitEvent.objects.filter(
            event_type=BitEvent.EventType.TRANSFER
        ).select_related(
            'bit', 'location', 'from_location', 'to_location', 'performed_by'
        ).order_by('-event_date')[:50]
        context['locations'] = Location.objects.filter(
            is_active=True
        ).order_by('location_type', 'name')
        context['pre_serial'] = self.request.GET.get('serial', '')
        return context


class AllLocationsView(LoginRequiredMixin, TemplateView):
    """Unified view showing all locations from all 3 sources."""
    template_name = "workorders/all_locations.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.sales.models import Warehouse
        from apps.inventory.models import InventoryLocation
        from django.db.models import Count

        # Bit Locations (workorders.Location) — with bit count
        bit_locations = Location.objects.filter(is_active=True).annotate(
            bit_count=Count('current_bits')
        ).order_by('location_type', 'name')

        # Warehouses (sales.Warehouse)
        warehouses = Warehouse.objects.all().order_by('name')

        # Stock Locations (inventory.InventoryLocation)
        stock_locations = InventoryLocation.objects.filter(is_active=True).order_by('warehouse__name', 'name')

        # Build unified list
        all_locs = []
        for loc in bit_locations:
            all_locs.append({
                'source': 'Bit Location',
                'code': loc.code,
                'name': loc.name,
                'type': loc.get_location_type_display(),
                'parent': '',
                'bits': loc.bit_count,
                'active': loc.is_active,
                'edit_url': f'/work-orders/locations/{loc.pk}/edit/',
            })
        for wh in warehouses:
            all_locs.append({
                'source': 'Warehouse',
                'code': wh.code,
                'name': wh.name,
                'type': 'Warehouse',
                'parent': '',
                'bits': '',
                'active': wh.is_active if hasattr(wh, 'is_active') else True,
                'edit_url': '',
            })
        for sl in stock_locations:
            all_locs.append({
                'source': 'Stock Bin',
                'code': sl.code,
                'name': sl.name,
                'type': 'Stock Location',
                'parent': sl.warehouse.name if sl.warehouse else '',
                'bits': '',
                'active': sl.is_active,
                'edit_url': '',
            })

        context['all_locations'] = all_locs
        context['total_bit_locs'] = bit_locations.count()
        context['total_warehouses'] = warehouses.count()
        context['total_stock_locs'] = stock_locations.count()
        return context
