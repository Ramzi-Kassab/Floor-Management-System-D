"""
ARDT FMS - Drill Bit Inventory Views
Comprehensive inventory management for drill bits.

Features:
- Dashboard with aggregations by status, type, customer, location
- Full CRUD operations
- Action views for lifecycle events (receive, ship, transfer, scrap)
- Location management
- Excel export
"""

from datetime import datetime, timedelta
from decimal import Decimal

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)

from .models import BackloadItem, BitEvent, BOMPendingRequest, DrillBit, Location, WorkOrder
from .forms import DrillBitUpdateForm

# Import Customer model for First Event Wizard
try:
    from apps.sales.models import Customer
except ImportError:
    Customer = None

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# DRILL BIT INVENTORY DASHBOARD
# =============================================================================


class DrillBitInventoryDashboardView(LoginRequiredMixin, TemplateView):
    """
    Drill Bit Inventory Dashboard with summary cards and aggregations.
    """

    template_name = "workorders/drillbit_inventory_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Drill Bit Inventory"

        # Get all drill bits for aggregations
        bits = DrillBit.objects.all()

        # Summary cards
        context["total_bits"] = bits.count()
        context["fc_bits"] = bits.filter(bit_type="FC").count()
        context["rc_bits"] = bits.filter(bit_type="RC").count()

        # Status counts
        context["status_counts"] = list(
            bits.values("status")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Lifecycle status counts
        context["lifecycle_counts"] = list(
            bits.values("lifecycle_status")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Physical status counts
        context["physical_counts"] = list(
            bits.values("physical_status")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # By customer
        context["customer_counts"] = list(
            bits.filter(customer__isnull=False)
            .values("customer__code", "customer__name")
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        )

        # By location
        context["location_counts"] = list(
            bits.filter(bit_location__isnull=False)
            .values("bit_location__code", "bit_location__name")
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        )

        # By size
        context["size_counts"] = list(
            bits.values("size")
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        )

        # Cost summary
        context["total_original_cost"] = (
            bits.aggregate(total=Sum("original_cost"))["total"] or Decimal("0")
        )
        context["total_repair_cost"] = (
            bits.aggregate(total=Sum("total_repair_cost"))["total"] or Decimal("0")
        )
        context["total_book_value"] = (
            bits.aggregate(total=Sum("current_book_value"))["total"] or Decimal("0")
        )

        # Recent events
        context["recent_events"] = BitEvent.objects.select_related(
            "bit", "location", "performed_by"
        ).order_by("-event_date")[:10]

        # Bits received in last 30 days
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        context["recent_received"] = bits.filter(
            received_date__gte=thirty_days_ago
        ).count()

        # Bits scrapped in last 30 days
        context["recent_scrapped"] = bits.filter(
            scrap_date__gte=thirty_days_ago
        ).count()

        # Bits currently in production/repair
        context["in_production"] = bits.filter(
            lifecycle_status__in=["IN_REPAIR", "EVALUATION"]
        ).count()

        # Bits available for deployment
        context["available"] = bits.filter(
            status__in=["IN_STOCK", "READY"],
            lifecycle_status__in=["NEW", "REPAIRED", "RERUN"],
        ).count()

        # Get choices for display
        context["status_display"] = dict(DrillBit.Status.choices)
        context["lifecycle_display"] = dict(DrillBit.LifecycleStatus.choices)
        context["physical_display"] = dict(DrillBit.PhysicalStatus.choices)

        return context


# =============================================================================
# DRILL BIT CRUD VIEWS
# =============================================================================


class DrillBitCreateView(LoginRequiredMixin, CreateView):
    """
    Register a new drill bit in the system - IDENTITY ONLY.

    Captures only what defines the bit:
    - Serial Number
    - Design (L3/L4)
    - BOM (L5, optional)

    After registration, redirects to "First Event" page where user can:
    - Record RECEIVED event (bit arrived)
    - Record CUSTOMER_INTAKE event (customer brought for service)
    - Skip (just register, no event yet - for bits still in production)
    """

    model = DrillBit
    template_name = "workorders/drillbit_form.html"

    def get_form_class(self):
        from .forms import DrillBitCreateForm
        return DrillBitCreateForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Register New Drill Bit"
        context["action"] = "create"

        from apps.technology.models import BOM, Design

        # Designs for table display
        designs = Design.objects.select_related(
            "size", "iadc_code_ref"
        ).prefetch_related("boms").order_by("mat_no")
        context["designs"] = designs

        # BOMs for table display
        boms = BOM.objects.filter(
            design__isnull=False
        ).select_related(
            "design", "design__size", "design__iadc_code_ref", "smi_type"
        ).order_by("design__mat_no", "code")
        context["boms"] = boms

        # Get Designs for JavaScript (legacy — keep for form fallback)
        designs_data = []
        for design in designs:
            designs_data.append({
                "id": design.id,
                "mat_no": design.mat_no,
                "hdbs_type": design.hdbs_type or "",
                "size": str(design.size) if design.size else "",
                "category": design.category or "",
            })
        context["designs_json"] = designs_data

        # Get BOMs with design_id for JavaScript filtering
        boms_data = []
        for bom in boms.filter(status="ACTIVE"):
            boms_data.append({
                "id": bom.id,
                "code": bom.code,
                "name": bom.name or "",
                "design_id": bom.design.id if bom.design else None,
                "design_mat_no": bom.design.mat_no if bom.design else "",
                "system_mat_no": bom.system_mat_no or "",
            })
        context["boms_json"] = boms_data

        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)

        # No event created here - First Event wizard will capture location and create event
        # The bit exists in the system but has no location/status until first event

        design_info = self.object.design.mat_no if self.object.design else "Unknown"
        bom_info = f" with BOM {self.object.bom.code}" if self.object.bom else ""
        messages.success(
            self.request,
            f'Drill bit "{self.object.serial_number}" registered for design {design_info}{bom_info}.',
        )

        # ── Auto-link: Check for UNMATCHED backload items with this serial ──
        unmatched = BackloadItem.objects.filter(
            serial_number=self.object.serial_number,
            match_status=BackloadItem.MatchStatus.UNMATCHED,
        ).select_related("batch")

        if unmatched.exists():
            now = timezone.now()
            receiving_loc = Location.objects.filter(
                location_type=Location.LocationType.RECEIVING
            ).first()

            linked_batches = []
            for item in unmatched:
                item.drill_bit = self.object
                item.match_status = BackloadItem.MatchStatus.NEW_REGISTERED
                item.status = BackloadItem.ItemStatus.RECEIVED
                item.received_date = now
                item.received_by = self.request.user

                # Create backloaded event
                event = BitEvent.objects.create(
                    bit=self.object,
                    event_type=BitEvent.EventType.BACKLOADED,
                    event_date=now,
                    location=receiving_loc,
                    notes=f"Auto-confirmed from batch {item.batch.batch_number}",
                    performed_by=self.request.user,
                )
                item.bit_event = event
                item.save(update_fields=[
                    "drill_bit", "match_status", "status",
                    "received_date", "received_by", "bit_event",
                ])

                # Update batch counts
                item.batch.update_counts()
                item.batch.auto_update_status()
                linked_batches.append(item.batch.batch_number)

            # Update bit status
            self.object.lifecycle_status = DrillBit.LifecycleStatus.BACKLOADED
            self.object.status = DrillBit.Status.BACKLOADED
            self.object.condition = DrillBit.Condition.USED
            self.object.physical_status = DrillBit.PhysicalStatus.AT_ARDT
            self.object.backload_count = (self.object.backload_count or 0) + 1
            self.object.last_backload_date = now.date()
            self.object.derive_ownership()
            self.object.save(update_fields=[
                "lifecycle_status", "status", "condition", "ownership",
                "physical_status",
                "backload_count", "last_backload_date",
            ])

            batch_list = ", ".join(linked_batches)
            messages.info(
                self.request,
                f"Also confirmed in backload batch(es): {batch_list}.",
            )

        return response

    def get_success_url(self):
        # Redirect to first event page
        return reverse("workorders:drillbit_first_event", kwargs={"pk": self.object.pk})


class DrillBitFirstEventView(LoginRequiredMixin, TemplateView):
    """
    First Event wizard - capture what's happening with a newly registered bit.

    Options:
    1. RECEIVED - Bit physically arrived at ARDT warehouse
    2. CUSTOMER_INTAKE - Customer brought bit for service
    3. IN_PRODUCTION - Bit is still being manufactured (no location yet)
    4. Skip - Just registered, will add event later
    """
    template_name = "workorders/drillbit_first_event.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = get_object_or_404(DrillBit, pk=self.kwargs['pk'])
        context['bit'] = bit

        # Filtered location lists for each option
        # "Received at ARDT" — warehouse, receiving, WIP areas
        context['received_locations'] = Location.objects.filter(
            is_active=True,
            location_type__in=[
                Location.LocationType.WAREHOUSE,
                Location.LocationType.RECEIVING,
                Location.LocationType.WIP,
            ]
        ).order_by('name')

        # "Customer Intake" — receiving, evaluation, backload areas
        context['intake_locations'] = Location.objects.filter(
            is_active=True,
            location_type__in=[
                Location.LocationType.RECEIVING,
                Location.LocationType.EVALUATION,
                Location.LocationType.WAREHOUSE,
            ]
        ).order_by('name')

        if Customer:
            context['customers'] = Customer.objects.filter(is_active=True).order_by('name') if hasattr(Customer, 'is_active') else Customer.objects.all().order_by('name')
        else:
            context['customers'] = []

        # Default USA location for "In Production" option (Woodlands, Texas)
        context['usa_location'] = Location.objects.filter(
            location_type=Location.LocationType.USA, is_active=True
        ).first()
        # Default locations for Received and Customer Intake
        context['default_received_location'] = Location.objects.filter(
            code='RCV-AREA', is_active=True
        ).first()
        context['default_intake_location'] = Location.objects.filter(
            code='BACKLOAD', is_active=True
        ).first()
        return context

    def post(self, request, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=self.kwargs['pk'])
        event_type = request.POST.get('event_type')

        if event_type == 'skip':
            messages.info(request, f'Drill bit "{bit.serial_number}" registered. You can add events later from the detail page.')
            return redirect('workorders:drillbit_detail', pk=bit.pk)

        location_id = request.POST.get('location')
        customer_id = request.POST.get('customer')
        notes = request.POST.get('notes', '')

        if not location_id:
            messages.error(request, 'Please select a location.')
            return redirect('workorders:drillbit_first_event', pk=bit.pk)

        location = get_object_or_404(Location, pk=location_id)

        # Condition based on level:
        #   L5 = cutters brazed = finished good
        #   L3/L4 = no cutters = components
        if bit.level == '5':
            smart_condition = DrillBit.Condition.FINISHED_GOOD
        else:
            smart_condition = DrillBit.Condition.COMPONENTS

        # Update bit fields based on event type
        if event_type == 'received':
            # ARDT received new bit
            bit.bit_location = location
            bit.status = DrillBit.Status.RECEIVING
            bit.condition = smart_condition
            bit.lifecycle_status = DrillBit.LifecycleStatus.NEW
            bit.physical_status = DrillBit.PhysicalStatus.AT_ARDT
            bit.accounting_status = DrillBit.AccountingStatus.ARDT_OWNED
            bit.derive_ownership()
            event_type_choice = BitEvent.EventType.RECEIVED
            event_notes = f"Received at {location.name}. {notes}".strip()

        elif event_type == 'customer_intake':
            # Customer brought bit for service
            bit.bit_location = location
            bit.status = DrillBit.Status.IN_EVALUATION
            bit.condition = DrillBit.Condition.USED
            bit.lifecycle_status = DrillBit.LifecycleStatus.EVALUATION
            bit.physical_status = DrillBit.PhysicalStatus.AT_ARDT
            bit.accounting_status = DrillBit.AccountingStatus.CUSTOMER_OWNED
            bit.derive_ownership()
            if customer_id and Customer:
                bit.customer = get_object_or_404(Customer, pk=customer_id)
            event_type_choice = BitEvent.EventType.BACKLOADED  # Using BACKLOADED for customer intake
            event_notes = f"Customer intake at {location.name}. Customer: {bit.customer.name if bit.customer else 'Not specified'}. {notes}".strip()

        elif event_type == 'in_production':
            # Bit is being manufactured externally (USA) — not at ARDT yet
            bit.status = DrillBit.Status.IN_PRODUCTION_USA
            bit.condition = smart_condition
            bit.lifecycle_status = DrillBit.LifecycleStatus.NEW
            bit.physical_status = DrillBit.PhysicalStatus.IN_TRANSIT
            # No physical location yet, but we need one for the event
            # Use the selected location as "pending delivery to"
            bit.bit_location = location
            bit.derive_ownership()
            event_type_choice = BitEvent.EventType.RECEIVED
            event_notes = f"Registered - In production (USA), pending delivery to {location.name}. {notes}".strip()

        else:
            messages.error(request, 'Invalid event type.')
            return redirect('workorders:drillbit_first_event', pk=bit.pk)

        bit.save()

        # Create the event
        BitEvent.objects.create(
            bit=bit,
            event_type=event_type_choice,
            event_date=timezone.now(),
            location=location,
            notes=event_notes,
            performed_by=request.user,
        )

        messages.success(request, f'Drill bit "{bit.serial_number}" is now tracked at {location.name}.')

        # Auto-create BOM pending request for MANUFACTURE bits without BOM
        if (
            event_type == 'received'
            and bit.account
            and bit.account.workflow_type in ('MANUFACTURE', 'BOTH')
            and not bit.bom
            and not bit.brazing_bom
            and not bit.system_bom
        ):
            existing = BOMPendingRequest.objects.filter(
                drill_bit=bit,
                status=BOMPendingRequest.RequestStatus.OPEN,
            ).exists()
            if not existing:
                BOMPendingRequest.objects.create(
                    drill_bit=bit,
                    requested_by=request.user,
                    notes="Auto-created: BOM not assigned at registration.",
                )
                messages.info(request, "BOM not assigned. Added to tech team queue.")

        return redirect('workorders:drillbit_detail', pk=bit.pk)


class DrillBitUpdateView(LoginRequiredMixin, UpdateView):
    """
    Edit drill bit identity - Design and BOMs only.

    Uses the same table-based design/BOM selection as the create page.
    """

    model = DrillBit
    template_name = "workorders/drillbit_form.html"
    form_class = DrillBitUpdateForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit Drill Bit - {self.object.serial_number}"
        context["action"] = "edit"

        from apps.technology.models import BOM, Design

        # Same querysets as create view for table-based selection
        designs = Design.objects.select_related(
            "size", "iadc_code_ref"
        ).prefetch_related("boms").order_by("mat_no")
        context["designs"] = designs

        boms = BOM.objects.filter(
            design__isnull=False
        ).select_related(
            "design", "design__size", "design__iadc_code_ref", "smi_type"
        ).order_by("design__mat_no", "code")
        context["boms"] = boms

        # JSON data for JS filtering (legacy)
        designs_data = []
        for design in designs:
            designs_data.append({
                "id": design.id,
                "mat_no": design.mat_no,
                "hdbs_type": design.hdbs_type or "",
                "size": str(design.size) if design.size else "",
                "category": design.category or "",
            })
        context["designs_json"] = designs_data

        boms_data = []
        for bom in boms.filter(status="ACTIVE"):
            boms_data.append({
                "id": bom.id,
                "code": bom.code,
                "name": bom.name or "",
                "design_id": bom.design.id if bom.design else None,
                "design_mat_no": bom.design.mat_no if bom.design else "",
                "system_mat_no": bom.system_mat_no or "",
            })
        context["boms_json"] = boms_data

        # Pre-selected IDs for edit mode
        context["current_design_id"] = self.object.design_id
        context["current_brazing_bom_id"] = self.object.brazing_bom_id
        context["current_system_bom_id"] = self.object.system_bom_id
        context["current_bom_id"] = self.object.bom_id
        context["current_level"] = self.object.level or ""

        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Drill bit "{self.object.serial_number}" updated successfully.',
        )
        return response

    def get_success_url(self):
        return reverse("workorders:drillbit_detail_enhanced", kwargs={"pk": self.object.pk})


class DrillBitDeleteView(LoginRequiredMixin, DeleteView):
    """
    Delete a drill bit record permanently from the system.
    This is a data cleanup action, NOT a physical scrap (use ScrapView for that).
    """

    model = DrillBit
    template_name = "workorders/drillbit_confirm_delete.html"
    success_url = reverse_lazy("workorders:drillbit_list_enhanced")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Delete Drill Bit - {self.object.serial_number}"
        context["work_orders_count"] = self.object.work_orders.count()
        context["events_count"] = self.object.bit_events.count()
        return context

    def delete(self, request, *args, **kwargs):
        bit = self.get_object()
        serial = bit.serial_number
        bit.delete()
        messages.success(request, f'Drill bit "{serial}" has been permanently deleted.')
        return redirect(self.success_url)


# =============================================================================
# DRILL BIT ACTION VIEWS (Lifecycle Events)
# =============================================================================


class DrillBitReceiveView(LoginRequiredMixin, TemplateView):
    """
    Record receipt of a drill bit.
    """

    template_name = "workorders/drillbit_action_receive.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])
        context["drill_bit"] = bit
        context["page_title"] = f"Receive Drill Bit - {bit.serial_number}"
        context["locations"] = Location.objects.filter(is_active=True)
        return context

    def post(self, request, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])

        location_id = request.POST.get("location")
        received_date = request.POST.get("received_date")
        notes = request.POST.get("notes", "")

        # Validate location
        if not location_id:
            messages.error(request, "Please select a receiving location.")
            return redirect("workorders:drillbit_receive", pk=bit.pk)

        try:
            location = get_object_or_404(Location, pk=location_id)

            if received_date:
                bit.received_date = datetime.strptime(received_date, "%Y-%m-%d").date()

            bit.bit_location = location
            bit.physical_status = DrillBit.PhysicalStatus.AT_ARDT
            bit.status = DrillBit.Status.IN_STOCK
            bit.lifecycle_status = DrillBit.LifecycleStatus.NEW
            bit.save()

            # Create event
            BitEvent.objects.create(
                bit=bit,
                event_type=BitEvent.EventType.RECEIVED,
                event_date=timezone.now(),
                location=location,
                notes=notes or "Received into inventory.",
                performed_by=request.user,
            )

            messages.success(request, f'Drill bit "{bit.serial_number}" received successfully.')
        except Exception as e:
            messages.error(request, f"Error receiving drill bit: {str(e)}")
            return redirect("workorders:drillbit_receive", pk=bit.pk)

        return redirect("workorders:drillbit_detail", pk=bit.pk)


class DrillBitShipView(LoginRequiredMixin, TemplateView):
    """
    Record shipment of a drill bit to customer/rig.
    """

    template_name = "workorders/drillbit_action_ship.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])
        context["drill_bit"] = bit
        context["page_title"] = f"Ship Drill Bit - {bit.serial_number}"
        context["locations"] = Location.objects.filter(
            is_active=True, location_type__in=["RIG", "WAREHOUSE"]
        )
        # Get customers and rigs
        from apps.sales.models import Customer, Rig

        context["customers"] = Customer.objects.filter(is_active=True)
        context["rigs"] = Rig.objects.filter(is_active=True)
        return context

    def post(self, request, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])

        destination_id = request.POST.get("destination")
        customer_id = request.POST.get("customer")
        rig_id = request.POST.get("rig")
        shipping_ref = request.POST.get("shipping_ref", "")
        notes = request.POST.get("notes", "")

        from_location = bit.bit_location
        to_location = get_object_or_404(Location, pk=destination_id) if destination_id else None

        # Update bit
        if customer_id:
            from apps.sales.models import Customer

            bit.customer = get_object_or_404(Customer, pk=customer_id)

        if rig_id:
            from apps.sales.models import Rig

            bit.rig = get_object_or_404(Rig, pk=rig_id)

        if to_location:
            bit.bit_location = to_location

        bit.status = DrillBit.Status.DISPATCHED
        # Condition stays as-is (Finished Good, Repaired, Rerun, Retrofitted)
        bit.physical_status = DrillBit.PhysicalStatus.IN_TRANSIT
        bit.lifecycle_status = DrillBit.LifecycleStatus.DEPLOYED
        bit.deployment_count += 1
        bit.last_deployed_date = timezone.now().date()
        bit.save()

        # Create event
        BitEvent.objects.create(
            bit=bit,
            event_type=BitEvent.EventType.DEPLOYED,
            event_date=timezone.now(),
            location=to_location or from_location,
            from_location=from_location,
            to_location=to_location,
            rig=bit.rig,
            notes=f"Shipped. Ref: {shipping_ref}. {notes}".strip(),
            performed_by=request.user,
        )

        messages.success(request, f'Drill bit "{bit.serial_number}" shipped successfully.')
        return redirect("workorders:drillbit_detail", pk=bit.pk)


class DrillBitTransferView(LoginRequiredMixin, TemplateView):
    """
    Transfer drill bit between locations.
    """

    template_name = "workorders/drillbit_action_transfer.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])
        context["drill_bit"] = bit
        context["page_title"] = f"Transfer Drill Bit - {bit.serial_number}"
        context["locations"] = Location.objects.filter(is_active=True)
        return context

    def post(self, request, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])

        to_location_id = request.POST.get("to_location")
        transfer_reason = request.POST.get("reason", "")
        notes = request.POST.get("notes", "")

        # Validate destination location
        if not to_location_id:
            messages.error(request, "Please select a destination location.")
            return redirect("workorders:drillbit_transfer", pk=bit.pk)

        try:
            from_location = bit.bit_location
            to_location = get_object_or_404(Location, pk=to_location_id)

            # Prevent transfer to same location
            if from_location and from_location.pk == to_location.pk:
                messages.warning(request, "Bit is already at the selected location.")
                return redirect("workorders:drillbit_transfer", pk=bit.pk)

            # Update bit
            bit.bit_location = to_location
            bit.save()

            # Create event
            BitEvent.objects.create(
                bit=bit,
                event_type=BitEvent.EventType.TRANSFER,
                event_date=timezone.now(),
                location=to_location,
                from_location=from_location,
                to_location=to_location,
                notes=f"Transfer reason: {transfer_reason}. {notes}".strip(),
                performed_by=request.user,
            )

            messages.success(
                request,
                f'Drill bit "{bit.serial_number}" transferred to {to_location.name}.',
            )
        except Exception as e:
            messages.error(request, f"Error transferring drill bit: {str(e)}")
            return redirect("workorders:drillbit_transfer", pk=bit.pk)

        return redirect("workorders:drillbit_detail", pk=bit.pk)


class DrillBitReturnView(LoginRequiredMixin, TemplateView):
    """
    Record return of drill bit from field/customer.
    """

    template_name = "workorders/drillbit_action_return.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])
        context["drill_bit"] = bit
        context["page_title"] = f"Return Drill Bit - {bit.serial_number}"
        context["locations"] = Location.objects.filter(
            is_active=True, location_type__in=["WAREHOUSE", "EVALUATION", "REPAIR_SHOP"]
        )
        return context

    def post(self, request, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])

        location_id = request.POST.get("location")
        condition = request.POST.get("condition", "GOOD")
        notes = request.POST.get("notes", "")

        from_location = bit.bit_location
        to_location = get_object_or_404(Location, pk=location_id) if location_id else None

        # Update bit
        if to_location:
            bit.bit_location = to_location

        bit.status = DrillBit.Status.BACKLOADED
        bit.condition = DrillBit.Condition.USED  # returned from field = used
        bit.physical_status = DrillBit.PhysicalStatus.AT_ARDT
        bit.lifecycle_status = DrillBit.LifecycleStatus.BACKLOADED
        bit.backload_count += 1
        bit.last_backload_date = timezone.now().date()
        bit.save()

        # Create event
        BitEvent.objects.create(
            bit=bit,
            event_type=BitEvent.EventType.BACKLOADED,
            event_date=timezone.now(),
            location=to_location or from_location,
            from_location=from_location,
            to_location=to_location,
            notes=f"Returned. Condition: {condition}. {notes}".strip(),
            performed_by=request.user,
        )

        messages.success(request, f'Drill bit "{bit.serial_number}" returned successfully.')
        return redirect("workorders:drillbit_detail", pk=bit.pk)


class DrillBitScrapView(LoginRequiredMixin, TemplateView):
    """
    Mark drill bit as scrapped.
    """

    template_name = "workorders/drillbit_action_scrap.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])
        context["drill_bit"] = bit
        context["page_title"] = f"Scrap Drill Bit - {bit.serial_number}"
        context["locations"] = Location.objects.filter(
            is_active=True, location_type="SCRAP"
        )
        return context

    def post(self, request, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])

        scrap_reason = request.POST.get("reason", "")
        notes = request.POST.get("notes", "")
        scrap_location_id = request.POST.get("location")

        from_location = bit.bit_location
        scrap_location = (
            get_object_or_404(Location, pk=scrap_location_id)
            if scrap_location_id
            else from_location
        )

        # Update bit
        bit.status = DrillBit.Status.SCRAPPED
        bit.condition = DrillBit.Condition.SCRAPPED
        bit.lifecycle_status = DrillBit.LifecycleStatus.SCRAP
        bit.scrap_date = timezone.now().date()
        if scrap_location:
            bit.bit_location = scrap_location
        bit.save()

        # Create event
        BitEvent.objects.create(
            bit=bit,
            event_type=BitEvent.EventType.SCRAPPED,
            event_date=timezone.now(),
            location=scrap_location or from_location,
            notes=f"Scrap reason: {scrap_reason}. {notes}".strip(),
            performed_by=request.user,
        )

        messages.success(
            request, f'Drill bit "{bit.serial_number}" marked as scrapped.'
        )
        return redirect("workorders:drillbit_detail", pk=bit.pk)


class DrillBitStartRepairView(LoginRequiredMixin, TemplateView):
    """
    Record start of repair for a drill bit.
    """

    template_name = "workorders/drillbit_action_start_repair.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])
        context["drill_bit"] = bit
        context["page_title"] = f"Start Repair - {bit.serial_number}"
        context["work_orders"] = WorkOrder.objects.filter(
            status__in=["DRAFT", "PLANNED", "RELEASED"]
        ).order_by("-created_at")[:20]
        context["locations"] = Location.objects.filter(
            is_active=True, location_type="REPAIR_SHOP"
        )
        return context

    def post(self, request, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=self.kwargs["pk"])

        work_order_id = request.POST.get("work_order")
        location_id = request.POST.get("location")
        notes = request.POST.get("notes", "")

        work_order = (
            get_object_or_404(WorkOrder, pk=work_order_id) if work_order_id else None
        )
        location = (
            get_object_or_404(Location, pk=location_id)
            if location_id
            else bit.bit_location
        )

        # Update bit
        bit.status = DrillBit.Status.IN_REPAIR
        # Condition stays as-is (Used, Finished Good for defect repair, etc.)
        bit.lifecycle_status = DrillBit.LifecycleStatus.IN_REPAIR
        if location:
            bit.bit_location = location
        bit.save()

        # Create event
        BitEvent.objects.create(
            bit=bit,
            event_type=BitEvent.EventType.REPAIR_START,
            event_date=timezone.now(),
            location=location,
            work_order=work_order,
            notes=notes or "Repair started.",
            performed_by=request.user,
        )

        messages.success(
            request, f'Repair started for drill bit "{bit.serial_number}".'
        )
        return redirect("workorders:drillbit_detail", pk=bit.pk)


# =============================================================================
# LOCATION MANAGEMENT VIEWS
# =============================================================================


class LocationListView(LoginRequiredMixin, ListView):
    """
    List all locations.
    """

    model = Location
    template_name = "workorders/location_list.html"
    context_object_name = "locations"

    def get_queryset(self):
        queryset = Location.objects.annotate(bit_count=Count("current_bits"))

        # Filter by type
        location_type = self.request.GET.get("type")
        if location_type:
            queryset = queryset.filter(location_type=location_type)

        # Filter by active
        is_active = self.request.GET.get("active")
        if is_active == "1":
            queryset = queryset.filter(is_active=True)
        elif is_active == "0":
            queryset = queryset.filter(is_active=False)

        return queryset.order_by("location_type", "name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Locations"
        context["location_types"] = Location.LocationType.choices
        context["current_type"] = self.request.GET.get("type", "")
        context["current_active"] = self.request.GET.get("active", "")
        return context


class LocationCreateView(LoginRequiredMixin, CreateView):
    """
    Create a new location.
    """

    model = Location
    template_name = "workorders/location_form.html"
    fields = ["code", "name", "location_type", "address", "is_active"]
    success_url = reverse_lazy("workorders:location_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Add Location"
        context["action"] = "create"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, f'Location "{self.object.name}" created successfully.'
        )
        return response


class LocationUpdateView(LoginRequiredMixin, UpdateView):
    """
    Edit a location.
    """

    model = Location
    template_name = "workorders/location_form.html"
    fields = ["code", "name", "location_type", "address", "is_active"]
    success_url = reverse_lazy("workorders:location_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Edit Location - {self.object.name}"
        context["action"] = "edit"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, f'Location "{self.object.name}" updated successfully.'
        )
        return response


class LocationDeleteView(LoginRequiredMixin, DeleteView):
    """
    Delete a location.
    """

    model = Location
    template_name = "workorders/location_confirm_delete.html"
    success_url = reverse_lazy("workorders:location_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = f"Delete Location - {self.object.name}"
        context["bit_count"] = self.object.current_bits.count()
        return context

    def delete(self, request, *args, **kwargs):
        location = self.get_object()
        if location.current_bits.exists():
            messages.error(
                request,
                f'Cannot delete location "{location.name}" - it has drill bits assigned.',
            )
            return redirect("workorders:location_list")

        messages.success(request, f'Location "{location.name}" deleted.')
        return super().delete(request, *args, **kwargs)


# =============================================================================
# BIT EVENT VIEWS
# =============================================================================


class BitEventListView(LoginRequiredMixin, ListView):
    """
    View all bit events (lifecycle history).
    """

    model = BitEvent
    template_name = "workorders/bitevent_list.html"
    context_object_name = "events"
    paginate_by = 50

    def get_queryset(self):
        queryset = BitEvent.objects.select_related(
            "bit", "location", "from_location", "to_location", "work_order", "performed_by"
        ).order_by("-event_date")

        # Filter by event type
        event_type = self.request.GET.get("event_type")
        if event_type:
            queryset = queryset.filter(event_type=event_type)

        # Filter by bit
        bit_id = self.request.GET.get("bit")
        if bit_id:
            queryset = queryset.filter(bit_id=bit_id)

        # Filter by date range
        date_from = self.request.GET.get("date_from")
        if date_from:
            queryset = queryset.filter(event_date__date__gte=date_from)

        date_to = self.request.GET.get("date_to")
        if date_to:
            queryset = queryset.filter(event_date__date__lte=date_to)

        # Search
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(bit__serial_number__icontains=search)
                | Q(notes__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Bit Events"
        context["event_types"] = BitEvent.EventType.choices
        context["current_event_type"] = self.request.GET.get("event_type", "")
        context["current_search"] = self.request.GET.get("search", "")
        context["current_date_from"] = self.request.GET.get("date_from", "")
        context["current_date_to"] = self.request.GET.get("date_to", "")
        return context


# =============================================================================
# EXCEL EXPORT
# =============================================================================


class DrillBitQRLabelsView(LoginRequiredMixin, View):
    """Generate multi-size QR labels for one or more drill bits."""

    def get(self, request):
        from .utils import generate_drill_bit_qr

        bit_ids_raw = request.GET.get("bits", "")
        bit_ids = [x.strip() for x in bit_ids_raw.split(",") if x.strip().isdigit()]

        bits = DrillBit.objects.filter(pk__in=bit_ids).select_related(
            "design", "design__size", "design__connection_type_ref",
            "design__connection_size_ref", "account",
            "brazing_bom", "brazing_bom__smi_type",
            "bom", "bom__smi_type",
            "system_bom", "system_bom__smi_type",
        )

        base_url = getattr(settings, "SITE_URL", None) or request.build_absolute_uri("/")[:-1]

        labels = []
        for bit in bits:
            active_bom = bit.brazing_bom or bit.bom or bit.system_bom
            smi_display = str(active_bom.smi_type) if active_bom and active_bom.smi_type else ""
            size_display = ""
            if bit.design and bit.design.size:
                size_display = str(bit.design.size.size_display)
            elif bit.size:
                size_display = f'{bit.size}"'
            design_mat = ""
            if bit.design:
                design_mat = str(bit.design)
            # Connection display from design
            connection_display = ""
            if bit.design:
                if bit.design.connection_type_ref and bit.design.connection_size_ref:
                    connection_display = f"{bit.design.connection_type_ref.code} {bit.design.connection_size_ref.size_inches}"
                elif bit.design.connection_type_ref:
                    connection_display = str(bit.design.connection_type_ref.code)

            labels.append({
                "bit": bit,
                "qr_base64": generate_drill_bit_qr(bit, base_url),
                "smi_display": smi_display,
                "size_display": size_display,
                "design_mat": design_mat,
                "account_name": str(bit.account) if bit.account else "",
                "bit_type_display": bit.get_bit_type_display() if bit.bit_type else "",
                "connection_display": connection_display,
            })

        copies = int(request.GET.get("copies", "1"))
        if copies < 1:
            copies = 1
        if copies > 50:
            copies = 50

        from django.shortcuts import render
        return render(request, "workorders/drillbit_qr_labels.html", {
            "labels": labels,
            "copies": copies,
            "copy_range": range(copies),
            "size": request.GET.get("size", "medium"),
            "page_title": f"QR Labels ({len(labels)} bit{'s' if len(labels) != 1 else ''})",
        })


# =============================================================================


class DrillBitExportExcelView(LoginRequiredMixin, View):
    """
    Export drill bits to Excel with column/record selection.

    Query params:
      columns: 'visible' | 'all'
      records: 'filtered' | 'all'
      visible_cols: CSV of column keys (when columns=visible)
      bit_ids: CSV of bit PKs (when records=filtered)
      status, condition, ownership, bit_type, search: standard filters
    """

    # All available columns (key matches template data-column attributes)
    ALL_COLUMNS = [
        {"key": "serial", "name": "Serial Number"},
        {"key": "level", "name": "Level"},
        {"key": "type", "name": "FC/RC"},
        {"key": "design", "name": "Design MAT"},
        {"key": "refmat", "name": "Ref Design MAT"},
        {"key": "systembom", "name": "System MAT"},
        {"key": "brazingbom", "name": "Brazing BOM"},
        {"key": "hdbs", "name": "HDBS Type"},
        {"key": "smi", "name": "SMI Type"},
        {"key": "size", "name": "Size"},
        {"key": "connection", "name": "Connection"},
        {"key": "iadc", "name": "IADC"},
        {"key": "breaker", "name": "Breaker Slot"},
        {"key": "specialtech", "name": "Special Tech"},
        {"key": "application", "name": "Application"},
        {"key": "account", "name": "Business Unit"},
        {"key": "customer", "name": "Customer"},
        {"key": "location", "name": "Location"},
        {"key": "status", "name": "Status"},
        {"key": "condition", "name": "Condition"},
        {"key": "ownership", "name": "Ownership"},
        {"key": "created", "name": "Registered"},
    ]

    def _get_cell_value(self, bit, key):
        """Return the export value for a given column key."""
        if key == "serial":
            return bit.serial_number
        elif key == "level":
            return f"L{bit.level}" if bit.level else ""
        elif key == "type":
            return bit.get_bit_type_display()
        elif key == "design":
            return bit.design.mat_no if bit.design else ""
        elif key == "refmat":
            return bit.design.ref_mat_no if bit.design and bit.design.ref_mat_no else ""
        elif key == "systembom":
            if bit.brazing_bom and bit.brazing_bom.system_mat_no:
                return bit.brazing_bom.system_mat_no
            elif bit.bom and bit.bom.system_mat_no:
                return bit.bom.system_mat_no
            elif bit.system_bom:
                return bit.system_bom.code
            return ""
        elif key == "brazingbom":
            if bit.brazing_bom:
                return bit.brazing_bom.code
            elif bit.bom:
                return bit.bom.code
            return ""
        elif key == "hdbs":
            return bit.design.hdbs_type if bit.design and bit.design.hdbs_type else ""
        elif key == "smi":
            for bom_ref in [bit.brazing_bom, bit.system_bom, bit.bom]:
                if bom_ref and hasattr(bom_ref, "smi_type") and bom_ref.smi_type:
                    return bom_ref.smi_type.smi_name
            return bit.design.smi_type if bit.design and bit.design.smi_type else ""
        elif key == "size":
            return bit.design.size.size_display if bit.design and bit.design.size else ""
        elif key == "connection":
            return bit.design.connection_ref.mat_no if bit.design and bit.design.connection_ref else ""
        elif key == "iadc":
            return bit.design.iadc_code_ref.code if bit.design and bit.design.iadc_code_ref else ""
        elif key == "breaker":
            return bit.design.breaker_slot.mat_no if bit.design and bit.design.breaker_slot else ""
        elif key == "specialtech":
            if bit.design:
                techs = bit.design.special_technologies.all()
                return ", ".join(t.code for t in techs) if techs else ""
            return ""
        elif key == "application":
            parts = []
            if bit.design:
                if bit.design.application_ref:
                    parts.append(bit.design.application_ref.name)
                if bit.design.formation_type_ref:
                    parts.append(bit.design.formation_type_ref.name)
            return " / ".join(parts) if parts else ""
        elif key == "account":
            return bit.account.code if bit.account else ""
        elif key == "customer":
            return bit.customer.name if bit.customer else ""
        elif key == "location":
            return bit.bit_location.name if bit.bit_location else ""
        elif key == "status":
            return bit.get_status_display()
        elif key == "condition":
            return bit.get_condition_display()
        elif key == "ownership":
            return bit.get_ownership_display()
        elif key == "created":
            return bit.created_at.strftime("%Y-%m-%d") if bit.created_at else ""
        return ""

    def get(self, request, *args, **kwargs):
        if not HAS_OPENPYXL:
            messages.error(request, "Excel export requires openpyxl. Please install it.")
            return redirect("workorders:drillbit_list_enhanced")

        # ── Determine which columns to export ──
        columns_option = request.GET.get("columns", "all")
        if columns_option == "visible":
            visible_cols = request.GET.get("visible_cols", "")
            visible_keys = [c.strip() for c in visible_cols.split(",") if c.strip()]
            if visible_keys:
                export_columns = [c for c in self.ALL_COLUMNS if c["key"] in visible_keys]
            else:
                export_columns = list(self.ALL_COLUMNS)
        else:
            export_columns = list(self.ALL_COLUMNS)

        # Ensure serial is always first
        serial_col = {"key": "serial", "name": "Serial Number"}
        if not any(c["key"] == "serial" for c in export_columns):
            export_columns.insert(0, serial_col)

        # ── Determine which records to export ──
        records_option = request.GET.get("records", "all")

        # Build queryset with rich select_related
        queryset = DrillBit.objects.select_related(
            "design", "design__size", "design__connection_ref",
            "design__iadc_code_ref", "design__breaker_slot",
            "design__application_ref", "design__formation_type_ref",
            "account", "customer", "rig", "bit_location",
            "brazing_bom", "brazing_bom__smi_type",
            "system_bom", "system_bom__smi_type",
            "bom", "bom__smi_type",
        ).prefetch_related(
            "design__special_technologies",
        ).order_by("-created_at")

        if records_option == "filtered":
            bit_ids = request.GET.get("bit_ids", "")
            id_list = [int(x) for x in bit_ids.split(",") if x.strip().isdigit()]
            if id_list:
                queryset = queryset.filter(pk__in=id_list)

        # Apply standard filters (from list page URL params)
        status_filter = request.GET.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        condition_filter = request.GET.get("condition")
        if condition_filter:
            queryset = queryset.filter(condition=condition_filter)

        ownership_filter = request.GET.get("ownership")
        if ownership_filter:
            queryset = queryset.filter(ownership=ownership_filter)

        bit_type_filter = request.GET.get("bit_type")
        if bit_type_filter:
            queryset = queryset.filter(bit_type=bit_type_filter)

        search = request.GET.get("search")
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(serial_number__icontains=search)
                | Q(design__mat_no__icontains=search)
                | Q(design__hdbs_type__icontains=search)
            )

        # ── Create workbook ──
        wb = Workbook()
        ws = wb.active
        ws.title = "Drill Bits"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Row number column
        cell = ws.cell(row=1, column=1, value="#")
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        # Write headers
        for col_idx, col_def in enumerate(export_columns, 2):
            cell = ws.cell(row=1, column=col_idx, value=col_def["name"])
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, bit in enumerate(queryset, 2):
            # Row number
            cell = ws.cell(row=row_num, column=1, value=row_num - 1)
            cell.border = thin_border

            for col_idx, col_def in enumerate(export_columns, 2):
                value = self._get_cell_value(bit, col_def["key"])
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-width
        total_cols = len(export_columns) + 1
        for col in range(1, total_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions["A"].width = 6  # Row number column

        # Freeze header row
        ws.freeze_panes = "A2"

        # Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="drill_bits_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )
        wb.save(response)
        return response


# =============================================================================
# API VIEWS FOR AJAX
# =============================================================================


class DrillBitSearchAPIView(LoginRequiredMixin, View):
    """
    API endpoint for drill bit search (for autocomplete).
    Supports filtering by design_id for BOM creation.
    """

    def get(self, request, *args, **kwargs):
        query = request.GET.get("q", "")
        design_id = request.GET.get("design_id", "")

        # If design_id is provided, filter by design (for BOM creation)
        if design_id:
            bits = DrillBit.objects.filter(design_id=design_id).select_related(
                "customer", "bit_location", "design"
            ).order_by("-created_at")[:50]
        elif len(query) >= 2:
            # Standard text search
            bits = DrillBit.objects.filter(
                Q(serial_number__icontains=query) | Q(mat_number__icontains=query)
            ).select_related("customer", "bit_location")[:20]
        else:
            return JsonResponse({"results": []})

        results = [
            {
                "id": bit.pk,
                "serial_number": bit.serial_number,
                "mat_number": bit.mat_number,
                "type": bit.get_bit_type_display(),
                "status": bit.get_status_display(),
                "lifecycle_status": bit.get_lifecycle_status_display() if hasattr(bit, 'lifecycle_status') else None,
                "location": bit.bit_location.name if bit.bit_location else None,
                "customer": bit.customer.name if bit.customer else None,
                # BOM linkage info for the linking dialog
                "brazing_bom_id": bit.brazing_bom_id if hasattr(bit, 'brazing_bom_id') else None,
                "system_bom_id": bit.system_bom_id if hasattr(bit, 'system_bom_id') else None,
            }
            for bit in bits
        ]

        return JsonResponse({"results": results})


class DrillBitQuickEventAPIView(LoginRequiredMixin, View):
    """
    API endpoint to quickly create an event for a drill bit.
    """

    def post(self, request, pk, *args, **kwargs):
        bit = get_object_or_404(DrillBit, pk=pk)

        event_type = request.POST.get("event_type")
        notes = request.POST.get("notes", "")

        if not event_type:
            return JsonResponse({"error": "Event type required"}, status=400)

        # Validate event type
        valid_types = [choice[0] for choice in BitEvent.EventType.choices]
        if event_type not in valid_types:
            return JsonResponse({"error": "Invalid event type"}, status=400)

        # Create event
        event = BitEvent.objects.create(
            bit=bit,
            event_type=event_type,
            event_date=timezone.now(),
            location=bit.bit_location,
            notes=notes,
            performed_by=request.user,
        )

        return JsonResponse({
            "success": True,
            "event_id": event.pk,
            "message": f"Event created: {event.get_event_type_display()}",
        })
