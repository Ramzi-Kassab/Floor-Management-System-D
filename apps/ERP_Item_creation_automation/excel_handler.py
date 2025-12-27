# excel_handler.py (only the route below needs replacing)

import os, json
import pandas as pd
from flask import Blueprint, render_template, request, flash, session


def _dedupe_selection_session():
    """Ensure all_selected_data and selected_rows_by_sheet are consistent, de-duplicated,
    and keyed by (__sheet__, __excel_row__)."""
    by_sheet = session.get("selected_rows_by_sheet", {}) or {}
    all_sel = session.get("all_selected_data", []) or []

    seen = set()
    new_all = []
    for r in all_sel:
        key = (r.get("__sheet__"), r.get("__excel_row__"))
        if key in seen:
            continue
        seen.add(key)
        new_all.append(r)

    # rebuild by_sheet from new_all
    new_by = {}
    for r in new_all:
        sh = r.get("__sheet__")
        new_by.setdefault(sh, []).append(r)

    session["all_selected_data"] = new_all
    session["selected_rows_by_sheet"] = new_by


bp = Blueprint("excel_handler", __name__)

# ---------- helpers (assume you already have these) ----------
def find_excel_file(folder: str, partial: str):
    import os, glob
    raw_folder = (folder or "").strip()
    folder = os.path.normpath(raw_folder.strip('"').strip("'"))
    partial_lc = (partial or "").strip().lower()

    info = []
    info.append(f"FOLDER(raw)={raw_folder!r}")
    info.append(f"FOLDER(norm)={folder!r} exists={os.path.isdir(folder)}")
    info.append(f"PARTIAL(lc)={partial_lc!r}")

    if not folder or not os.path.isdir(folder):
        return None, "\n".join(info + ["-> Folder is missing or not accessible."])

    pats = ["*.xlsx", "*.xlsm", "*.xls"]
    all_hits = []
    for pat in pats:
        all_hits.extend(glob.glob(os.path.join(folder, pat)))
    info.append(f"ALL_EXCELS={len(all_hits)}")
    for p in all_hits[:10]:
        info.append(f"  - {os.path.basename(p)}")

    hits = [p for p in all_hits if (not partial_lc) or (partial_lc in os.path.basename(p).lower())]
    info.append(f"MATCHED={len(hits)}")
    for p in hits[:10]:
        info.append(f"  * {os.path.basename(p)}")

    if not hits:
        return None, "\n".join(info + ["-> No files matched the partial. Try leaving it blank or adjust it."])

    hits.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    picked = hits[0]
    info.append(f"PICKED={picked}")
    return picked, "\n".join(info)

def list_sheets(file_path: str) -> list[str]:
    try:
        return pd.ExcelFile(file_path).sheet_names
    except Exception:
        return []

def read_sheet(file_path: str, sheet_name: str, rows_to_display: str = "all", header_row: int = 0) -> pd.DataFrame:
    rtd = (rows_to_display or "all").strip().lower()
    nrows = None if rtd in ("", "all") else int(rtd)
    return pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, nrows=nrows)

def write_to_excel(file_path: str, sheet_name: str, row_index: int, column_name: str, value, header_row: int = 1) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[header_row]]
    col_idx = headers.index(column_name) + 1
    target_row = header_row + 1 + row_index
    ws.cell(row=target_row, column=col_idx, value=value)
    wb.save(file_path)

# ---------- load config ----------
def load_config():
    here = os.path.dirname(os.path.abspath(__file__))
    cfg_path = os.path.join(here, "config.json")
    try:
        with open(cfg_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

# ---------- ROUTE (REPLACE THIS WHOLE FUNCTION) ----------
@bp.route("/excel_handler", methods=["GET", "POST"])
def excel_handler():
    # Prefill from config.json on GET
    cfg = load_config()
    default_path = cfg.get("default_excel_folder", "")
    default_partial = cfg.get("default_excel_filename", "")

    # Build context (form overrides config when present)
    ctx = {
        "excel_path": request.form.get("excel_path") or request.args.get("excel_path") or default_path,
        "excel_name_partial": request.form.get("excel_name_partial") or request.args.get("excel_name_partial") or default_partial,
        "file_found_path": None,
        "sheets": [],
        "selected_sheet": request.form.get("sheet_name") or "",
        "rows_to_display": request.form.get("rows_to_display") or "all",
        "columns": None,
        "rows": None,
    }

    action = request.form.get("action")
    if request.method == "POST":
        print("[POST] form ->", dict(request.form))  # <- debug: must show excel_path + excel_name_partial

    # ---------- FIND ----------
    if request.method == "POST" and action == "find":
        file_path, dbg = find_excel_file(ctx["excel_path"], ctx["excel_name_partial"])
        print("[find] DEBUG\n" + dbg)

        if not file_path:
            flash("No Excel file matched that folder/name. Check the path and partial.", "danger")
            flash(dbg.replace("\n", "<br>"), "warning")
            return render_template("excel_handler.html", **ctx)

        # reset selection if user switched to a different file
        if session.get("_last_file") != file_path:
            session.pop("all_selected_data", None)
            session.pop("selected_rows_by_sheet", None)
        session["_last_file"] = file_path

        ctx["file_found_path"] = file_path

        ctx["sheets"] = list_sheets(file_path)
        if ctx["sheets"]:
            ctx["selected_sheet"] = ctx["sheets"][0]
            flash(f"File found: {file_path}", "success")
        else:
            flash("Found the file, but couldn’t read any sheets.", "danger")
        return render_template("excel_handler.html", **ctx)

    # ---------- READ ----------
    if request.method == "POST" and action == "read":
        file_path, dbg = find_excel_file(ctx["excel_path"], ctx["excel_name_partial"])
        print("[read] DEBUG\n" + dbg)

        if not file_path:
            flash("Please click Find File first (file could not be resolved).", "danger")
            flash(dbg.replace("\n", "<br>"), "warning")
            return render_template("excel_handler.html", **ctx)

        # reset selection if user switched to a different file
        if session.get("_last_file") != file_path:
            session.pop("all_selected_data", None)
            session.pop("selected_rows_by_sheet", None)
        session["_last_file"] = file_path

        ctx["file_found_path"] = file_path

        ctx["sheets"] = list_sheets(file_path)
        sheet_name = request.form.get("sheet_name") or (ctx["sheets"][0] if ctx["sheets"] else "")
        ctx["selected_sheet"] = sheet_name

        if not sheet_name:
            flash("Select a sheet, then click Read Sheet.", "warning")
            return render_template("excel_handler.html", **ctx)

        try:
            df = read_sheet(file_path, sheet_name, ctx["rows_to_display"])
            ctx["columns"] = list(df.columns)
            ctx["rows"] = df.to_dict(orient="records")
            session['last_column_order'] = ctx['columns']
            flash(f"Loaded {sheet_name}.", "success")
        except Exception as e:
            flash(f"Error reading sheet: {e}", "danger")
        return render_template("excel_handler.html", **ctx)

    # ---------- WRITE ----------
    if request.method == "POST" and action == "write":
        file_path, dbg = find_excel_file(ctx["excel_path"], ctx["excel_name_partial"])
        print("[write] DEBUG\n" + dbg)

        if not file_path:
            flash("Please click Find File first (file could not be resolved).", "danger")
            flash(dbg.replace("\n", "<br>"), "warning")
            return render_template("excel_handler.html", **ctx)

        if session.get("_last_file") != file_path:
            session.pop("all_selected_data", None)
            session.pop("selected_rows_by_sheet", None)
        session["_last_file"] = file_path

        ctx["file_found_path"] = file_path
        ctx["sheets"] = list_sheets(file_path)
        try:
            sheet_name = request.form["sheet_name"]
            row_index = int(request.form["row_index"])
            column_name = request.form["column_name"]
            value = request.form.get("value", "")
            write_to_excel(file_path, sheet_name, row_index, column_name, value, header_row=1)

            ctx["selected_sheet"] = sheet_name
            flash(f"Wrote value to {sheet_name} – row {row_index}, column '{column_name}'.", "success")

            df = read_sheet(file_path, sheet_name, ctx["rows_to_display"])
            ctx["columns"] = list(df.columns)
            ctx["rows"] = df.to_dict(orient="records")
        except Exception as e:
            flash(f"Error writing to Excel: {e}", "danger")
        return render_template("excel_handler.html", **ctx)

    # ---------- SELECT ROWS ----------
    if request.method == "POST" and action == "select_rows":
        # Expect: sheet_name, selected_rows[] (indices from current preview)
        sheet_name = request.form.get("sheet_name", "")
        # reconcile context from config again so hidden fields work after redirect
        cfg = load_config()
        ctx["excel_path"] = request.form.get("excel_path") or ctx["excel_path"] or cfg.get("default_excel_folder", "")
        ctx["excel_name_partial"] = request.form.get("excel_name_partial") or ctx["excel_name_partial"] or cfg.get("default_excel_filename", "")

        file_path, dbg = find_excel_file(ctx["excel_path"], ctx["excel_name_partial"])
        if not file_path or not sheet_name:
            flash("Please find a file and select a sheet first.", "warning")
            return render_template("excel_handler.html", **ctx)

        if session.get("_last_file") != file_path:
            session.pop("all_selected_data", None)
            session.pop("selected_rows_by_sheet", None)
        session["_last_file"] = file_path

        # Re-read full sheet (we need all rows to pull the picked indices safely)
        try:
            df_full = read_sheet(file_path, sheet_name, "all")
        except Exception as e:
            flash(f"Error reading sheet for selection: {e}", "danger")
            return render_template("excel_handler.html", **ctx)

        selected_idxs = request.form.getlist("selected_rows")
        picked_rows = []
        for idx_s in selected_idxs:
            try:
                i = int(idx_s)
            except ValueError:
                continue
            if 0 <= i < len(df_full):
                row = df_full.iloc[i].to_dict()
                row["__excel_row__"] = i + 2            # Excel 1-based (header at row 1)
                row["__sheet__"] = sheet_name            # track owning sheet
                picked_rows.append(row)

        # --- SESSION STRUCTURE & DEDUPE ---
        if "selected_rows_by_sheet" not in session:
            session["selected_rows_by_sheet"] = {}

        if "all_selected_data" not in session:
            session["all_selected_data"] = []

        by_sheet = session["selected_rows_by_sheet"]
        all_sel = session["all_selected_data"]
        _dedupe_selection_session()  # <-- ensure no duplicates leak through


        # Build a set of existing keys to avoid duplicates: (sheet, excel_row)
        existing_keys = {(r.get("__sheet__"), r.get("__excel_row__")) for r in all_sel}

        actually_added = []
        for r in picked_rows:
            key = (r["__sheet__"], r["__excel_row__"])
            if key in existing_keys:
                continue
            actually_added.append(r)
            existing_keys.add(key)
            all_sel.append(r)

        # Update per-sheet bucket
        sheet_bucket = by_sheet.get(sheet_name, [])
        # Also dedupe inside the sheet bucket
        sheet_keys = {(r.get("__sheet__"), r.get("__excel_row__")) for r in sheet_bucket}
        for r in actually_added:
            key = (r["__sheet__"], r["__excel_row__"])
            if key not in sheet_keys:
                sheet_bucket.append(r)
                sheet_keys.add(key)
        by_sheet[sheet_name] = sheet_bucket

        session["selected_rows_by_sheet"] = by_sheet
        session["all_selected_data"] = all_sel
        _dedupe_selection_session()  # <-- ensure no duplicates leak through


        # Refresh preview context
        ctx["file_found_path"] = file_path
        ctx["sheets"] = list_sheets(file_path)
        ctx["selected_sheet"] = sheet_name
        df_preview = read_sheet(file_path, sheet_name, ctx["rows_to_display"])
        ctx["columns"] = list(df_preview.columns)
        session['last_column_order'] = ctx['columns']  # keep order consistent with workflows.html
        ctx["rows"] = df_preview.to_dict(orient="records")

        flash(f"Added {len(actually_added)} new row(s) from '{sheet_name}'.", "success")
        return render_template("excel_handler.html", **ctx)

    # ---------- REMOVE SELECTED (UNSELECT) ----------
    if request.method == "POST" and action == "remove_selected":
        remove_keys = request.form.getlist("remove_keys")  # values like "SHEET|ROWNUM"
        if not remove_keys:
            flash("No rows selected to remove.", "warning")
            return render_template("excel_handler.html", **ctx)

        # Parse keys into tuples
        to_remove = set()
        for k in remove_keys:
            if "|" in k:
                sh, rn = k.split("|", 1)
                try:
                    to_remove.add((sh, int(rn)))
                except ValueError:
                    continue

        by_sheet = session.get("selected_rows_by_sheet", {}) or {}
        all_sel = session.get("all_selected_data", []) or []

        # Filter all selections
        all_sel = [r for r in all_sel if (r.get("__sheet__"), r.get("__excel_row__")) not in to_remove]
        session["all_selected_data"] = all_sel

        # Filter per-sheet buckets
        for sh in list(by_sheet.keys()):
            rows = by_sheet.get(sh, []) or []
            rows = [r for r in rows if (r.get("__sheet__"), r.get("__excel_row__")) not in to_remove]
            if rows:
                by_sheet[sh] = rows
            else:
                by_sheet.pop(sh, None)
        session["selected_rows_by_sheet"] = by_sheet

        # FINAL: normalize + dedupe
        _dedupe_selection_session()

        flash(f"Removed {len(to_remove)} row(s) from selection.", "info")
        return render_template("excel_handler.html", **ctx)


    # ---------- RESET SELECTIONS ----------
    if request.method == "POST" and action == "reset_selections":
        session.pop("all_selected_data", None)
        session.pop("selected_rows_by_sheet", None)
        flash("All selections cleared.", "info")
        return render_template("excel_handler.html", **ctx)

    # ---------- GET ----------
    return render_template("excel_handler.html", **ctx)
